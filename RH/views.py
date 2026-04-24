import os
import shutil
import zipfile
from collections import defaultdict
import re
import time
import threading
import logging
import tempfile
import io
from datetime import datetime, date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Q, Avg, F
from django.http import HttpResponseForbidden, HttpResponse, FileResponse, JsonResponse
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.views.generic import CreateView, DetailView, UpdateView, DeleteView, ListView
from django.forms import inlineformset_factory
from django.db import transaction
from django.conf import settings
from django.utils import timezone
from django.core.files.storage import default_storage
import locale
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib import messages
from django.db.models import Sum, Count, Avg
from decimal import Decimal
from .models import Vacacion, Prestamo, PagoPrestamo, HistoricoVacaciones
from .forms import (
    VacacionForm, PrestamoForm, AprobarVacacionForm, PagoPrestamoForm,
    ControlVacanteForm
)
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
import json
import traceback
# Importa todos tus modelos de RH
from .models import (
    Empleado, Salario, Contrato, HistorialLaboral, BajaEmpleado, 
    Vacacion, Prestamo, PagoPrestamo, DocumentoOperador, Hijo,
    Departamento, Puesto, MotivoBaja, TipoDocumentoOperador,
    MotivoInactivacion, DivisionOperativa, TipoCarga, TipoViaje,
    ControlVacante
)

# Otras importaciones de librerías
from xhtml2pdf import pisa
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML
from .models import Empleado
from django.conf import settings

# Importar generador de documentos
from .utils.document_generator import DocumentGenerator

# Importar modelos y forms
from .models import (
    Empleado, Departamento, Puesto, MotivoInactivacion,
    TipoDocumentoOperador, DocumentoOperador, HistorialLaboral, Salario, Contrato, Hijo,
    TipoViaje, DivisionOperativa, TipoCarga, BajaEmpleado, MotivoBaja
)
from .forms import (
    EmpleadoForm, DepartamentoForm, PuestoForm, MotivoInactivacionForm, 
    TipoDocumentoOperadorForm, DocumentoOperadorForm, HistorialLaboralForm, 
    SalarioForm, ContratoForm, HijoForm, BajaEmpleadoForm, RecontratacionForm
)

# Otras importaciones de librerías
from xhtml2pdf import pisa
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML
from .models import Empleado
from django.conf import settings

logger = logging.getLogger(__name__)

# ============================================
# FUNCIONES PARA CALCULAR ANTIGÜEDAD
# ============================================

def calcular_antiguedad_años_vista(fecha_contratacion):
    """
    Calcula solo los años de antigüedad - para usar en views
    """
    if not fecha_contratacion or not isinstance(fecha_contratacion, date):
        return 0
    
    hoy = date.today()
    años = hoy.year - fecha_contratacion.year
    
    # Ajustar si aún no ha cumplido años este año
    if (hoy.month, hoy.day) < (fecha_contratacion.month, fecha_contratacion.day):
        años -= 1
    
    return años


def calcular_antiguedad_completa_vista(fecha_contratacion):
    """
    Calcula años y meses de antigüedad - para usar en views
    """
    if not fecha_contratacion or not isinstance(fecha_contratacion, date):
        return {'años': 0, 'meses': 0, 'texto': 'Sin antigüedad'}
    
    hoy = date.today()
    
    # Calcular años
    años = hoy.year - fecha_contratacion.year
    meses = hoy.month - fecha_contratacion.month
    
    # Ajustar si aún no ha cumplido años este año
    if (hoy.month, hoy.day) < (fecha_contratacion.month, fecha_contratacion.day):
        años -= 1
        meses += 12
    
    # Ajustar meses si es negativo
    if meses < 0:
        meses += 12
    
    # Ajustar por día del mes
    if hoy.day < fecha_contratacion.day:
        meses -= 1
        if meses < 0:
            meses += 12
    
    # Generar texto descriptivo
    if años > 0:
        texto = f"{años} año{'s' if años != 1 else ''}"
        if meses > 0:
            texto += f" {meses} mes{'es' if meses != 1 else ''}"
    elif meses > 0:
        texto = f"{meses} mes{'es' if meses != 1 else ''}"
    else:
        texto = "Menos de 1 mes"
    
    return {
        'años': años,
        'meses': meses,
        'texto': texto
    }


def calcular_dias_vacaciones_vista(fecha_contratacion):
    """
    Calcula días de vacaciones según antigüedad - para usar en views
    """
    if not fecha_contratacion or not isinstance(fecha_contratacion, date):
        return 0
    
    antiguedad = calcular_antiguedad_años_vista(fecha_contratacion)
    
    # Días según antigüedad
    if antiguedad < 1:
        return 0
    elif antiguedad == 1:
        return 12
    elif antiguedad == 2:
        return 14
    elif antiguedad == 3:
        return 16
    elif antiguedad == 4:
        return 18
    elif antiguedad == 5:
        return 20
    else:
        # 20 días base + 2 días por cada 5 años adicionales
        años_extra = antiguedad - 5
        dias_extra = (años_extra // 5) * 2
        return 20 + dias_extra

# En RH/views.py - actualizar la función inicio_rh
def inicio_rh(request):
    today = date.today()
    
    # --- 1. KPIs Generales ---
    total_empleados = Empleado.objects.no_eliminados().count()
    empleados_activos = Empleado.objects.no_eliminados().filter(activo=True).count()
    empleados_inactivos = Empleado.objects.no_eliminados().filter(activo=False).count()
    total_departamentos = Departamento.objects.count()
    total_eliminados = Empleado.objects.filter(eliminado=True).count()

    # Cálculo de Porcentajes
    porcentaje_activos = round((empleados_activos / total_empleados * 100), 1) if total_empleados > 0 else 0
    porcentaje_inactivos = round((empleados_inactivos / total_empleados * 100), 1) if total_empleados > 0 else 0

    # --- 1.5 KPIs Operativos ---
    operadores_migmar = Empleado.objects.no_eliminados().filter(
        activo=True, 
        empresa='MIGMAR', 
        puesto__nombre__icontains='Operador'
    ).count()
    
    operadores_marco = Empleado.objects.no_eliminados().filter(
        activo=True, 
        empresa='MARCO_MORALES', 
        puesto__nombre__icontains='Operador'
    ).count()

    # --- 2. Cumpleaños ---
    cumpleanos_hoy = []
    empleados_cumple = Empleado.objects.no_eliminados().filter(
        fecha_nacimiento__month=today.month,
        fecha_nacimiento__day=today.day,
        activo=True
    )
    for emp in empleados_cumple:
        edad = today.year - emp.fecha_nacimiento.year
        cumpleanos_hoy.append({'empleado': emp, 'edad_a_cumplir': edad})

    # --- 3. ALERTAS ---
    alertas_rh = []
    
    # A. Alerta de Contratos por Vencer
    fecha_limite_contrato = today + timedelta(days=30)
    contratos_por_vencer = Contrato.objects.filter(
        tipo_contrato='DETERMINADO',
        fecha_fin__range=[today, fecha_limite_contrato],
        empleado__activo=True,
        empleado__eliminado=False
    ).select_related('empleado', 'empleado__puesto', 'empleado__departamento').prefetch_related('empleado__division_operativa')

    for c in contratos_por_vencer:
        dias = (c.fecha_fin - today).days
        divisiones = ", ".join([div.nombre for div in c.empleado.division_operativa.all()])
        tipo_contrato_str = c.get_tipo_contrato_display() if hasattr(c, 'get_tipo_contrato_display') else 'Contrato'
        
        alertas_rh.append({
            'titulo': f'Vencimiento de Contrato ({dias} días)',
            'descripcion': f'{c.empleado.nombre_completo}',
            'empleado_id': c.empleado.id,  # <-- ID AÑADIDO PARA EL LINK
            'empresa': c.empleado.get_empresa_display() if c.empleado.empresa else "",
            'division': divisiones,
            'puesto': c.empleado.puesto.nombre if c.empleado.puesto else "",
            'departamento': c.empleado.departamento.nombre if c.empleado.departamento else "",
            'categoria': f'Contrato {tipo_contrato_str}',
            'tipo': 'warning',
            'icono': 'file-contract',
            'fecha': c.fecha_fin
        })

    # B. Documentos Vencidos
    fecha_limite_docs = today + timedelta(days=15)
    docs_vencidos = DocumentoOperador.objects.filter(
        fecha_vencimiento__lte=fecha_limite_docs,
        empleado__activo=True,
        empleado__eliminado=False
    ).select_related('empleado', 'tipo_documento', 'empleado__puesto', 'empleado__departamento').prefetch_related('empleado__division_operativa')

    for d in docs_vencidos:
        if d.fecha_vencimiento < today:
            tipo_alerta = 'danger'
            texto_dias = "VENCIDO"
        else:
            tipo_alerta = 'warning'
            texto_dias = f"Vence: {(d.fecha_vencimiento - today).days} días"

        divisiones = ", ".join([div.nombre for div in d.empleado.division_operativa.all()])

        alertas_rh.append({
            'titulo': f'{d.tipo_documento.nombre}',
            'descripcion': f'{texto_dias} - {d.empleado.nombre_completo}',
            'empleado_id': d.empleado.id,  # <-- ID AÑADIDO PARA EL LINK
            'empresa': d.empleado.get_empresa_display() if d.empleado.empresa else "",
            'division': divisiones,
            'puesto': d.empleado.puesto.nombre if d.empleado.puesto else "",
            'departamento': d.empleado.departamento.nombre if d.empleado.departamento else "",
            'categoria': 'Documento Operativo',
            'tipo': 'tipo_alerta',
            'icono': 'id-card',
            'fecha': d.fecha_vencimiento
        })
    
    # C. Vacantes activas (Omitidas del Dashboard de alertas)

    # D. Vacaciones próximas
    vacaciones_proximas = Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__range=[today, today + timedelta(days=7)],
        empleado__eliminado=False
    ).select_related('empleado', 'empleado__puesto', 'empleado__departamento').prefetch_related('empleado__division_operativa')
    
    for v in vacaciones_proximas:
        dias = (v.fecha_inicio - today).days
        divisiones = ", ".join([div.nombre for div in v.empleado.division_operativa.all()])
        
        alertas_rh.append({
            'titulo': 'Vacaciones próximas',
            'descripcion': f'{v.empleado.nombre_completo} - {v.dias_solicitados} días',
            'empleado_id': v.empleado.id,  # <-- ID AÑADIDO PARA EL LINK
            'empresa': v.empleado.get_empresa_display() if v.empleado.empresa else "",
            'division': divisiones,
            'puesto': v.empleado.puesto.nombre if v.empleado.puesto else "",
            'departamento': v.empleado.departamento.nombre if v.empleado.departamento else "",
            'categoria': 'Vacaciones',
            'tipo': 'info',
            'icono': 'umbrella-beach',
            'fecha': v.fecha_inicio
        })

    alertas_rh.sort(key=lambda x: (x['tipo'] != 'danger', x['fecha']))

    # --- 4. Gráficos ---
    departamento_distribucion = Empleado.objects.no_eliminados().filter(
        activo=True
    ).values('departamento__nombre').annotate(count=Count('id')).order_by('-count')
    
    departamento_distribucion_list = [{'nombre': item['departamento__nombre'] or 'Sin Asignar', 'count': item['count']} for item in departamento_distribucion]

    # --- 5. VACACIONES ---
    vacaciones_activas = Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__lte=today,
        fecha_fin__gte=today,
        empleado__eliminado=False
    ).count()
    
    vacaciones_pendientes_aprobacion = Vacacion.objects.filter(
        estado='PENDIENTE',
        empleado__eliminado=False
    ).count()
    
    año_actual = today.year
    historico_actual = HistoricoVacaciones.objects.filter(
        empleado__eliminado=False
    ).filter(año=año_actual)
    
    if historico_actual.exists():
        total_dias_correspondientes = historico_actual.aggregate(total=Sum('dias_correspondientes'))['total'] or 0
        total_dias_tomados = historico_actual.aggregate(total=Sum('dias_tomados'))['total'] or 0
        vacaciones_progreso = (total_dias_tomados / total_dias_correspondientes) * 100 if total_dias_correspondientes > 0 else 0
    else:
        vacaciones_progreso = 0
    
    proximas_vacaciones = Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__gte=today,
        fecha_inicio__lte=today + timedelta(days=30),
        empleado__eliminado=False
    ).select_related('empleado').order_by('fecha_inicio')[:5]
    
    proximas_vacaciones_formatted = []
    for v in proximas_vacaciones:
        proximas_vacaciones_formatted.append({
            'empleado': v.empleado,
            'fecha_inicio': v.fecha_inicio,
            'fecha_fin': v.fecha_fin,
            'dias': v.dias_solicitados,
            'dias_restantes': v.dias_restantes,
            'estado': v.estado.lower()
        })

    # --- 6. PRÉSTAMOS ---
    prestamos_activos = Prestamo.objects.filter(
        estado='EN_CURSO',
        empleado__eliminado=False
    ).count()
    
    total_prestamos_activos = Prestamo.objects.filter(
        estado='EN_CURSO',
        empleado__eliminado=False
    ).aggregate(total=Sum('saldo_pendiente'))['total'] or Decimal('0')
    
    total_prestamos_valor = Prestamo.objects.filter(
        estado='EN_CURSO',
        empleado__eliminado=False
    ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0')
    
    prestamos_progreso = (float(total_prestamos_activos) / float(total_prestamos_valor)) * 100 if total_prestamos_valor > 0 else 0
    
    prestamos_pendientes = Prestamo.objects.filter(
        estado='EN_CURSO',
        empleado__eliminado=False
    ).select_related('empleado').order_by('fecha_primer_pago')[:5]
    
    prestamos_pendientes_formatted = []
    for p in prestamos_pendientes:
        prestamos_pendientes_formatted.append({
            'empleado': p.empleado,
            'monto_total': p.monto_total,
            'monto_pagado': p.monto_pagado,
            'saldo_pendiente': p.saldo_pendiente,
            'plazo_meses': round(p.plazo_semanas / 4.33, 1),
            'progreso_pago': p.progreso_pago
        })

    context = {
        'total_empleados': total_empleados,
        'empleados_activos': empleados_activos,
        'total_departamentos': total_departamentos,
        'empleados_inactivos': empleados_inactivos,
        'empleados_activos_porcentaje': porcentaje_activos,
        'empleados_inactivos_porcentaje': porcentaje_inactivos,
        'operadores_migmar': operadores_migmar,
        'operadores_marco': operadores_marco,
        'cumpleanos_hoy': cumpleanos_hoy,
        'departamento_distribucion': departamento_distribucion_list,
        'alertas_rh': alertas_rh,
        'today': today,
        'total_eliminados': total_eliminados,
        'vacaciones_activas': vacaciones_activas,
        'vacaciones_pendientes': vacaciones_pendientes_aprobacion,
        'vacaciones_progreso': vacaciones_progreso,
        'proximas_vacaciones': proximas_vacaciones_formatted,
        'prestamos_activos': prestamos_activos,
        'total_prestamos_activos': total_prestamos_activos,
        'prestamos_progreso': prestamos_progreso,
        'prestamos_pendientes': prestamos_pendientes_formatted,
    }
    return render(request, 'rh/home.html', context)

class EmpleadoListView(ListView):
    model = Empleado
    template_name = 'rh/lista_empleados.html' 
    context_object_name = 'empleados'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().no_eliminados().select_related('puesto', 'departamento')
        user = self.request.user

        if not user.is_superuser and not user.groups.filter(name='RH_FULL').exists():
            if user.groups.filter(name='GERENTE_MIGMAR').exists():
                queryset = queryset.filter(empresa='MIGMAR')
            elif user.groups.filter(name='GERENTE_MARCO_MORALES').exists():
                queryset = queryset.filter(empresa='MARCO_MORALES')
            elif user.groups.filter(name='GERENTE_MARCO_AUTOZONE').exists():
                queryset = queryset.filter(empresa='MARCO_MORALES')
            elif user.groups.filter(name='GERENTE_TALLER').exists():
                queryset = queryset.filter(departamento__nombre__icontains='TALLER')
            elif user.groups.filter(name='GERENTE_CHIHUAHUA').exists():
                queryset = queryset.filter(division_operativa__nombre__icontains='CHIHUAHUA')
            else:
                return Empleado.objects.none()

        nombre = self.request.GET.get('nombre', '')
        depto_id = self.request.GET.get('departamento', '')
        puesto_id = self.request.GET.get('puesto', '')
        estado = self.request.GET.get('estado', '')
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')
        tipo_viaje_id = self.request.GET.get('tipo_viaje', '')
        empresa_param = self.request.GET.get('empresa', '')

        # --- BÚSQUEDA PROFESIONAL POR NOMBRE ---
        if nombre:
            terminos = nombre.split()
            for termino in terminos:
                queryset = queryset.filter(Q(nombre__icontains=termino) | Q(apellido__icontains=termino))
                
        if depto_id:
            queryset = queryset.filter(departamento__id=depto_id)
        if puesto_id:
            queryset = queryset.filter(puesto__id=puesto_id)
        if estado in ['0', '1']:
            queryset = queryset.filter(activo=(estado == '1'))
        if fecha_inicio:
            queryset = queryset.filter(fecha_contratacion__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_contratacion__lte=fecha_fin)
        if tipo_viaje_id:
            queryset = queryset.filter(tipo_viaje__id=tipo_viaje_id)
        if empresa_param:
            queryset = queryset.filter(empresa=empresa_param)

        sort_by = self.request.GET.get('sort', 'fecha_contratacion')
        direction = self.request.GET.get('direction', 'desc')

        if direction == 'desc':
            if not sort_by.startswith('-'): sort_by = f'-{sort_by}'
        else:
            if sort_by.startswith('-'): sort_by = sort_by[1:]

        valid_sort_fields = ['id', 'apellido', 'puesto__nombre', 'departamento__nombre', 'fecha_contratacion', 'empresa']
        if sort_by.replace('-', '') == 'nombre':
            sort_by = sort_by.replace('nombre', 'apellido')

        if sort_by.replace('-', '') in valid_sort_fields:
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('-fecha_contratacion')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        user = self.request.user
        
        context['puede_ver_sueldos'] = user.is_superuser or user.groups.filter(name='RH_FULL').exists()

        # === NUMERACIÓN DINÁMICA DE MAYOR A MENOR ===
        total_filtrados = context['paginator'].count
        start_index = context['page_obj'].start_index() if total_filtrados > 0 else 1

        for i, empleado in enumerate(context['object_list']):
            # Asigna el número en reversa exacto para esta fila (Ej: 150, 149, 148...)
            empleado.indice_descendente = total_filtrados - (start_index - 1) - i
            
            empleado.age = empleado.edad if empleado.edad is not None else 0
            if empleado.fecha_contratacion:
                start_date = empleado.fecha_contratacion
                end_date = empleado.fecha_inactivacion if not empleado.activo and empleado.fecha_inactivacion else today
                empleado.dias_laborados = (end_date - start_date).days
            else:
                empleado.dias_laborados = 0

        context['departamentos'] = Departamento.objects.all().order_by('nombre')
        context['puestos'] = Puesto.objects.all().order_by('nombre')
        context['tipos_viaje'] = TipoViaje.objects.all().order_by('nombre')
        context['empresas'] = Empleado.EMPRESA_CHOICES
        
        qs_seguro = self.get_queryset()
        context['total_empleados'] = qs_seguro.count()
        context['empleados_activos'] = qs_seguro.filter(activo=True).count()
        context['empleados_inactivos'] = qs_seguro.filter(activo=False).count()

        current_get_params = self.request.GET.copy()
        if 'page' in current_get_params:
            del current_get_params['page']
        context['query_string'] = current_get_params.urlencode()
        
        return context

def cumpleanos_rh(request):
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        pass 

    today = date.today()
    
    # 1. Obtener todos los empleados activos QUE TENGAN fecha de nacimiento registrada
    empleados = Empleado.objects.filter(activo=True, fecha_nacimiento__isnull=False)

    cumpleanos_hoy = []
    
    # 2. Preparar la estructura para los 12 meses
    meses_nombres = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    cumpleanos_por_mes = []
    for i, nombre in enumerate(meses_nombres, 1):
        cumpleanos_por_mes.append({
            'numero_mes': i,
            'nombre_mes': nombre,
            'es_mes_actual': (i == today.month),
            'empleados': []
        })

    # 3. Agrupar a los empleados en su mes correspondiente
    for emp in empleados:
        # Calcular fecha de cumpleaños en el año actual
        try:
            cumple_este_ano = date(today.year, emp.fecha_nacimiento.month, emp.fecha_nacimiento.day)
        except ValueError:
            # Manejo para 29 de febrero en años no bisiestos -> pasa al 28 de febrero
            cumple_este_ano = date(today.year, 2, 28)

        # Calcular edad a cumplir (o cumplida)
        edad = today.year - emp.fecha_nacimiento.year
        
        # Calcular días faltantes
        dias_faltantes = (cumple_este_ano - today).days
        es_hoy = (dias_faltantes == 0)
        ya_paso = (dias_faltantes < 0)

        datos_empleado = {
            'empleado': emp,
            'dia_nacimiento': emp.fecha_nacimiento.day,
            'fecha_cumple': cumple_este_ano,
            'edad_a_cumplir': edad,
            'dias_faltantes': dias_faltantes,
            'dias_faltantes_abs': abs(dias_faltantes),
            'es_hoy': es_hoy,
            'ya_paso': ya_paso
        }

        # Insertarlo en la lista del mes que le corresponde (índice es mes - 1)
        indice_mes = emp.fecha_nacimiento.month - 1
        cumpleanos_por_mes[indice_mes]['empleados'].append(datos_empleado)

        if es_hoy:
            cumpleanos_hoy.append(datos_empleado)

    # 4. Ordenar los empleados dentro de cada mes por su día de nacimiento
    for mes in cumpleanos_por_mes:
        mes['empleados'] = sorted(mes['empleados'], key=lambda x: x['dia_nacimiento'])

    context = {
        'cumpleanos_hoy': cumpleanos_hoy,
        'cumpleanos_por_mes': cumpleanos_por_mes, # Enviamos la lista completa de meses
        'today': today,
    }
    return render(request, 'rh/cumpleanos.html', context)

# Formsets para Documentos, Historial, Salario, Contrato, Hijos
DocumentoOperadorFormSet = inlineformset_factory(
    Empleado, DocumentoOperador, form=DocumentoOperadorForm,
    extra=0, can_delete=True # Changed to extra=0
)
HistorialLaboralFormSet = inlineformset_factory(
    Empleado, HistorialLaboral, form=HistorialLaboralForm,
    fk_name='empleado',
    extra=0, can_delete=True # Changed to extra=0
)
SalarioFormSet = inlineformset_factory(
    Empleado, Salario, form=SalarioForm,
    extra=0, can_delete=True # Changed to extra=0
)
ContratoFormSet = inlineformset_factory(
    Empleado, Contrato, form=ContratoForm,
    extra=0, can_delete=True # Changed to extra=0
)
HijoFormSet = inlineformset_factory( # New Formset for Hijo
    Empleado, Hijo, fields=['nombre', 'fecha_nacimiento'],
    extra=0, can_delete=True # Set extra=0
)

class EmpleadoCreateView(CreateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'rh/empleado_form.html'
    success_url = reverse_lazy('rh:lista_empleados')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['documentos_operador_formset'] = DocumentoOperadorFormSet(self.request.POST, self.request.FILES, prefix='documentos_operador')
            data['historial_laboral_formset'] = HistorialLaboralFormSet(self.request.POST, self.request.FILES, prefix='historial_laboral')
            data['salario_formset'] = SalarioFormSet(self.request.POST, prefix='salarios')
            data['contrato_formset'] = ContratoFormSet(self.request.POST, self.request.FILES, prefix='contratos')
            data['hijos_formset'] = HijoFormSet(self.request.POST, prefix='hijos') # Added HijoFormSet
        else:
            data['documentos_operador_formset'] = DocumentoOperadorFormSet(prefix='documentos_operador')
            data['historial_laboral_formset'] = HistorialLaboralFormSet(prefix='historial_laboral')
            data['salario_formset'] = SalarioFormSet(prefix='salarios')
            data['contrato_formset'] = ContratoFormSet(prefix='contratos')
            data['hijos_formset'] = HijoFormSet(prefix='hijos') # Added HijoFormSet
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        formsets = {
            'documentos_operador': context['documentos_operador_formset'],
            'historial_laboral': context['historial_laboral_formset'],
            'salarios': context['salario_formset'],
            'contratos': context['contrato_formset'],
            'hijos': context['hijos_formset'] # Added HijoFormSet
        }
        
        if all(fs.is_valid() for fs in formsets.values()):
            with transaction.atomic():
                self.object = form.save()
                for fs in formsets.values():
                    fs.instance = self.object
                    fs.save()
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def form_invalid(self, form):
        context = self.get_context_data(form=form)
        
        all_errors = []
        if form.errors:
            for field, error_list in form.errors.items():
                field_label = form.fields.get(field).label if form.fields.get(field) else field
                all_errors.append(f"Error en '{field_label}': {error_list[0]}")

        formsets = {
            "Documentos de Operador": context['documentos_operador_formset'],
            "Contratos": context['contrato_formset'],
            "Historial Laboral": context['historial_laboral_formset'],
            "Salarios": context['salario_formset'],
            "Hijos": context['hijos_formset'] # Added HijoFormSet for error reporting
        }

        for name, fs in formsets.items():
            if fs.errors:
                for form_errors in fs.errors:
                    if form_errors:
                        all_errors.append(f"Por favor, revisa la sección '{name}'. Contiene datos incompletos o incorrectos.")

        context['all_errors'] = all_errors
        return self.render_to_response(context)


class EmpleadoUpdateView(UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'rh/empleado_form.html'
    context_object_name = 'empleado'
    success_url = reverse_lazy('rh:lista_empleados')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['documentos_operador_formset'] = DocumentoOperadorFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix='documentos_operador')
            data['historial_laboral_formset'] = HistorialLaboralFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix='historial_laboral')
            data['salario_formset'] = SalarioFormSet(self.request.POST, instance=self.object, prefix='salarios')
            data['contrato_formset'] = ContratoFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix='contratos')
            data['hijos_formset'] = HijoFormSet(self.request.POST, instance=self.object, prefix='hijos') # Added HijoFormSet
        else:
            data['documentos_operador_formset'] = DocumentoOperadorFormSet(instance=self.object, prefix='documentos_operador')
            data['historial_laboral_formset'] = HistorialLaboralFormSet(instance=self.object, prefix='historial_laboral')
            data['salario_formset'] = SalarioFormSet(instance=self.object, prefix='salarios')
            data['contrato_formset'] = ContratoFormSet(instance=self.object, prefix='contratos')
            data['hijos_formset'] = HijoFormSet(instance=self.object, prefix='hijos') # Added HijoFormSet
        
        if self.object:
            employee = self.object
            historial = employee.historial_laboral_eventos.all()
            
            tenure_days = (date.today() - employee.fecha_contratacion).days if employee.fecha_contratacion else 0
            # Calculation for vacation days based on years of service.
            # This is a simplified example; often vacation accrual is more complex.
            years_service = tenure_days / 365.25
            vacation_days = 0
            if years_service >= 1 and years_service < 2:
                vacation_days = 12
            elif years_service >= 2 and years_service < 3:
                vacation_days = 14
            elif years_service >= 3 and years_service < 4:
                vacation_days = 16
            elif years_service >= 4 and years_service < 5:
                vacation_days = 18
            elif years_service >= 5:
                # After 5 years, add 2 days every 5 years
                vacation_days = 20 + ( (int(years_service) - 5) // 5 ) * 2
            
            latest_salary = employee.salarios.order_by('-fecha_efectiva').first()
            monthly_salary = latest_salary.sueldo_mensual if latest_salary else 0
            
            actas_count = historial.filter(tipo_evento='ACTA_ADMINISTRATIVA').count()
            suspensiones_count = historial.filter(tipo_evento='SUSPENSION').count()
            recontrataciones_count = historial.filter(tipo_evento='RECONTRATACION').count()
            permisos_count = historial.filter(tipo_evento='PERMISO').count()

            data['dashboard_stats'] = {
                'vacaciones_disponibles': vacation_days,
                'sueldo_mensual': f"${monthly_salary:,.2f}" if monthly_salary else "N/A",
                'actas_administrativas': actas_count,
                'suspensiones': suspensiones_count,
                'recontrataciones': recontrataciones_count,
                'permisos': permisos_count
            }
            
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        formsets = {
            'documentos_operador': context['documentos_operador_formset'],
            'historial_laboral': context['historial_laboral_formset'],
            'salarios': context['salario_formset'],
            'contratos': context['contrato_formset'],
            'hijos': context['hijos_formset'] # Added HijoFormSet
        }
        
        if all(fs.is_valid() for fs in formsets.values()):
            original_activo = self.get_object().activo
            new_activo = form.cleaned_data['activo']
            
            with transaction.atomic():
                self.object = form.save(commit=False)
                if original_activo and not new_activo:
                    self.object.fecha_inactivacion = date.today()
                elif not original_activo and new_activo:
                    self.object.motivo_inactivacion = None
                    self.object.fecha_inactivacion = None
                self.object.save()
                form.save_m2m()

                for fs in formsets.values():
                    fs.instance = self.object
                    fs.save()
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def form_invalid(self, form):
        context = self.get_context_data(form=form)
        
        all_errors = []
        if form.errors:
            for field, error_list in form.errors.items():
                for error in error_list:
                    field_label = form.fields.get(field).label if form.fields.get(field) else field
                    all_errors.append(f"Error en '{field_label}': {error}")

        formsets = {
            "Documentos de Operador": context['documentos_operador_formset'],
            "Contratos": context['contrato_formset'],
            "Historial Laboral": context['historial_laboral_formset'],
            "Salarios": context['salario_formset'],
            "Hijos": context['hijos_formset'] # Added HijoFormSet for error reporting
        }

        for name, fs in formsets.items():
            if fs.errors:
                for form_errors in fs.errors:
                    if form_errors:
                        all_errors.append(f"Por favor, revisa la sección '{name}'. Contiene datos incompletos o incorrectos.")

        context['all_errors'] = all_errors
        return self.render_to_response(context)


class EmpleadoDetailView(DetailView):
    model = Empleado
    template_name = 'rh/empleado_detail.html'
    context_object_name = 'empleado'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Cargar los documentos del operador relacionados, optimizando la consulta
        context['documentos_operador'] = self.object.documentos_operador.all().select_related('tipo_documento')
        return context

# En RH/views.py - modificar o reemplazar EmpleadoDeleteView
@login_required
def eliminar_empleado(request, pk):
    """
    Soft delete - mueve el empleado a la papelera
    """
    empleado = get_object_or_404(Empleado, pk=pk, eliminado=False)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo_eliminacion', '')
        
        # Usar el soft delete en lugar de eliminar permanentemente
        empleado.soft_delete(usuario=request.user, motivo=motivo)
        
        messages.success(
            request,
            f'✅ {empleado.nombre_completo} ha sido movido a la papelera de reciclaje.'
        )
        
        return redirect('rh:lista_empleados')
    
    context = {
        'empleado': empleado,
        'page_title': f'Mover a papelera - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/empleado_eliminar_confirm.html', context)


# Vistas CRUD para el Modelo Departamento
class DepartamentoListView(ListView):
    model = Departamento
    template_name = 'rh/lista_departamentos.html'
    context_object_name = 'departamentos'
    ordering = ['nombre']

class DepartamentoCreateView(CreateView):
    model = Departamento
    form_class = DepartamentoForm
    template_name = 'rh/departamento_form.html'
    success_url = reverse_lazy('rh:lista_departamentos')

class DepartamentoDetailView(DetailView):
    model = Departamento
    template_name = 'rh/departamento_detail.html'
    context_object_name = 'departamento'

class DepartamentoUpdateView(UpdateView):
    model = Departamento
    form_class = DepartamentoForm
    template_name = 'rh/departamento_form.html'
    success_url = reverse_lazy('rh:lista_departamentos')

class DepartamentoDeleteView(DeleteView):
    model = Departamento
    template_name = 'rh/departamento_confirm_delete.html'
    context_object_name = 'departamento'
    success_url = reverse_lazy('rh:lista_departamentos')

# Vistas CRUD para el Modelo Puesto
class PuestoListView(ListView):
    model = Puesto
    template_name = 'rh/lista_puestos.html'
    context_object_name = 'puestos'
    ordering = ['nombre']

class PuestoCreateView(CreateView):
    model = Puesto
    form_class = PuestoForm
    template_name = 'rh/puesto_form.html'
    success_url = reverse_lazy('rh:lista_puestos')

class PuestoDetailView(DetailView):
    model = Puesto
    template_name = 'rh/puesto_detail.html'
    context_object_name = 'puesto'

class PuestoUpdateView(UpdateView):
    model = Puesto
    form_class = PuestoForm
    template_name = 'rh/puesto_form.html'
    success_url = reverse_lazy('rh:lista_puestos')

class PuestoDeleteView(DeleteView):
    model = Puesto
    template_name = 'rh/puesto_confirm_delete.html'
    context_object_name = 'puesto'
    success_url = reverse_lazy('rh:lista_puestos')

# Vistas CRUD para el Modelo MotivoInactivacion
class MotivoInactivacionListView(ListView):
    model = MotivoInactivacion
    template_name = 'rh/lista_motivos_inactivacion.html'
    context_object_name = 'motivos'
    ordering = ['motivo']

class MotivoInactivacionCreateView(CreateView):
    model = MotivoInactivacion
    form_class = MotivoInactivacionForm
    template_name = 'rh/motivo_inactivacion_form.html'
    success_url = reverse_lazy('rh:lista_motivos_inactivacion')

class MotivoInactivacionUpdateView(UpdateView):
    model = MotivoInactivacion
    form_class = MotivoInactivacionForm
    template_name = 'rh/motivo_inactivacion_form.html'
    success_url = reverse_lazy('rh:lista_motivos_inactivacion')

class MotivoInactivacionDeleteView(DeleteView):
    model = MotivoInactivacion
    template_name = 'rh/motivo_inactivacion_confirm_delete.html'
    context_object_name = 'motivo'
    success_url = reverse_lazy('rh:lista_motivos_inactivacion')

# Vistas CRUD para el Modelo TipoDocumentoOperador
class TipoDocumentoOperadorListView(ListView):
    model = TipoDocumentoOperador
    template_name = 'rh/lista_tipos_documento_operador.html'
    context_object_name = 'tipos_documento'
    ordering = ['nombre']

class TipoDocumentoOperadorCreateView(CreateView):
    model = TipoDocumentoOperador
    form_class = TipoDocumentoOperadorForm
    template_name = 'rh/tipo_documento_operador_form.html'
    success_url = reverse_lazy('rh:lista_tipos_documento_operador')

class TipoDocumentoOperadorUpdateView(UpdateView):
    model = TipoDocumentoOperador
    form_class = TipoDocumentoOperadorForm
    template_name = 'rh/tipo_documento_operador_form.html'
    success_url = reverse_lazy('rh:lista_tipos_documento_operador')

class TipoDocumentoOperadorDeleteView(DeleteView):
    model = TipoDocumentoOperador
    template_name = 'rh/tipo_documento_operador_confirm_delete.html'
    context_object_name = 'tipo_documento'
    success_url = reverse_lazy('rh:lista_tipos_documento_operador')
    
    
def generar_pdf_empleado(request, pk):
    """
    Genera un PDF con la ficha completa de un empleado, usando una consulta optimizada.
    """
    # Consulta optimizada para precargar todos los datos necesarios
    empleado = get_object_or_404(
        Empleado.objects.select_related(
            'puesto', 'departamento', 'supervisor', 'motivo_inactivacion'
        ).prefetch_related(
            'hijos', 'contratos', 'documentos_operador__tipo_documento',
            'salarios', 'historial_laboral_eventos', 'tipo_viaje',
            'tipo_carga', 'division_operativa'
        ),
        pk=pk
    )
    
    template_path = 'rh/empleado_pdf_template.html'
    context = {'empleado': empleado}
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ficha-empleado-{empleado.nombre}-{empleado.apellido}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF <pre>' + html + '</pre>')
    return response


def semaforo_documentos_view(request):
    """
    Vista para mostrar un semáforo global de todos los documentos de operador y contratos.
    """
    today = date.today()
    expirable_items = []

    # 1. Obtenemos todos los documentos de operador con fecha de vencimiento
    documentos = DocumentoOperador.objects.filter(
        fecha_vencimiento__isnull=False,
        empleado__activo=True
    ).select_related('empleado', 'tipo_documento')

    for doc in documentos:
        expirable_items.append({
            'empleado': doc.empleado,
            'nombre_item': f"Doc. Operador: {doc.tipo_documento.nombre}",
            'fecha_vencimiento': doc.fecha_vencimiento,
            'comentarios': doc.observaciones
        })

    # 2. Obtenemos todos los contratos determinados con fecha de fin
    contratos_determinados = Contrato.objects.filter(
        tipo_contrato='DETERMINADO',
        fecha_fin__isnull=False,
        empleado__activo=True
    ).select_related('empleado')

    for contrato in contratos_determinados:
        expirable_items.append({
            'empleado': contrato.empleado,
            'nombre_item': 'Contrato Determinado',
            'fecha_vencimiento': contrato.fecha_fin,
            'comentarios': contrato.comentarios
        })

    # 3. Ordenamos la lista combinada por fecha de vencimiento
    expirable_items.sort(key=lambda x: x['fecha_vencimiento'])
    
    docs_rojo = []
    docs_amarillo = []
    docs_verde = []

    # Clasificamos cada item en su categoría correspondiente
    for item in expirable_items:
        days_remaining = (item['fecha_vencimiento'] - today).days
        item['dias_restantes'] = days_remaining

        if days_remaining <= 10:
            docs_rojo.append(item)
        elif days_remaining <= 30:
            docs_amarillo.append(item)
        else:
            docs_verde.append(item)

    context = {
        'docs_rojo': docs_rojo,
        'docs_amarillo': docs_amarillo,
        'docs_verde': docs_verde,
        'today': today,
        'page_title': 'Semáforo de Vencimientos'
    }
    return render(request, 'rh/semaforo_documentos.html', context)


def export_empleados_excel(request):
    # 1. SE AGREGA 'documentos_operador__tipo_documento' al prefetch_related para optimizar la carga
    queryset = Empleado.objects.no_eliminados().select_related(
        'puesto', 'departamento', 'motivo_inactivacion', 'supervisor'
    ).prefetch_related('division_operativa', 'tipo_carga', 'tipo_viaje', 'documentos_operador__tipo_documento')

    user = request.user

    if not user.is_superuser and not user.groups.filter(name='RH_FULL').exists():
        if user.groups.filter(name='GERENTE_MIGMAR').exists():
            queryset = queryset.filter(empresa='MIGMAR')
        elif user.groups.filter(name='GERENTE_MARCO_MORALES').exists():
            queryset = queryset.filter(empresa='MARCO_MORALES')
        elif user.groups.filter(name='GERENTE_MARCO_AUTOZONE').exists():
            queryset = queryset.filter(empresa='MARCO_MORALES')
        elif user.groups.filter(name='GERENTE_TALLER').exists():
            queryset = queryset.filter(departamento__nombre__icontains='TALLER')
        elif user.groups.filter(name='GERENTE_CHIHUAHUA').exists():
            queryset = queryset.filter(division_operativa__nombre__icontains='CHIHUAHUA')
        else:
            queryset = Empleado.objects.none()

    nombre = request.GET.get('nombre', '')
    depto_id = request.GET.get('departamento', '')
    puesto_id = request.GET.get('puesto', '')
    estado = request.GET.get('estado', '')
    tipo_viaje_id = request.GET.get('tipo_viaje', '')
    empresa_param = request.GET.get('empresa', '')

    # --- BÚSQUEDA PROFESIONAL POR NOMBRE ---
    if nombre:
        terminos = nombre.split()
        for termino in terminos:
            queryset = queryset.filter(Q(nombre__icontains=termino) | Q(apellido__icontains=termino))
            
    if depto_id: queryset = queryset.filter(departamento__id=depto_id)
    if puesto_id: queryset = queryset.filter(puesto__id=puesto_id)
    if estado in ['0', '1']: queryset = queryset.filter(activo=(estado == '1'))
    if tipo_viaje_id: queryset = queryset.filter(tipo_viaje__id=tipo_viaje_id)
    if empresa_param: queryset = queryset.filter(empresa=empresa_param)

    sort_by = request.GET.get('sort', 'fecha_contratacion')
    direction = request.GET.get('direction', 'desc')

    if direction == 'desc':
        if not sort_by.startswith('-'): sort_by = f'-{sort_by}'
    else:
        if sort_by.startswith('-'): sort_by = sort_by[1:]

    valid_sort_fields = ['id', 'apellido', 'puesto__nombre', 'departamento__nombre', 'fecha_contratacion', 'empresa']
    if sort_by.replace('-', '') == 'nombre':
        sort_by = sort_by.replace('nombre', 'apellido')

    if sort_by.replace('-', '') in valid_sort_fields:
        empleados = queryset.order_by(sort_by)
    else:
        empleados = queryset.order_by('-fecha_contratacion')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Empleados'

    # 2. SE AGREGA LA COLUMNA 'LICENCIA' A LOS ENCABEZADOS
    headers = [
        'No.', 'ID (UUID)', 'ID Origen', 'Número Empleado', 'Nombre', 'Apellido', 
        'Puesto', 'Puesto Anterior', 'Departamento', 'Supervisor', 'Email', 
        'Fecha Ingreso', 'Fecha Contratación', 'Estatus', 'Motivo Inactivación', 'Fecha Inactivación',
        'Fecha Nacimiento', 'Edad', 'Antigüedad (Años)', 'Dirección', 'Colonia', 'Código Postal', 
        'Ciudad', 'Estado', 'País', 'Teléfono Personal', 'Estado Civil', 'Nacionalidad',
        'CURP', 'RFC', 'NSS', 'Empresa', 'Tipo de Viaje', 'Tipo de Carga', 'Divisiones Operativas',
        'LICENCIA', # <--- NUEVA COLUMNA
        'Nombre Cónyuge', 'Teléfono Cónyuge',
        'Banco', 'CLABE Interbancaria', 'Número de Cuenta', 'Número de Tarjeta',
        'Nombre Referencia 1', 'Teléfono Referencia 1', 'Relación Referencia 1',
        'Nombre Referencia 2', 'Teléfono Referencia 2', 'Relación Referencia 2'
    ]
    worksheet.append(headers)

    # === CONTADOR DESCENDENTE ===
    total_empleados = empleados.count()
    contador = total_empleados 

    for empleado in empleados:
        fecha_contratacion = empleado.fecha_contratacion.strftime('%Y-%m-%d') if empleado.fecha_contratacion else ''
        fecha_ingreso = empleado.fecha_ingreso.strftime('%Y-%m-%d') if empleado.fecha_ingreso else ''
        fecha_inactivacion = empleado.fecha_inactivacion.strftime('%Y-%m-%d') if empleado.fecha_inactivacion else ''
        fecha_nacimiento = empleado.fecha_nacimiento.strftime('%Y-%m-%d') if empleado.fecha_nacimiento else ''
        divisiones = ", ".join([div.nombre for div in empleado.division_operativa.all()])
        tipos_viaje_str = ", ".join([tv.nombre for tv in empleado.tipo_viaje.all()])
        tipos_carga_str = ", ".join([tc.nombre for tc in empleado.tipo_carga.all()])
        nombre_empresa = empleado.get_empresa_display() or 'Sin Empresa'

        # 3. VERIFICAR SI TIENE LICENCIA
        tiene_licencia = "No"
        for doc in empleado.documentos_operador.all():
            if 'licencia' in doc.tipo_documento.nombre.lower():
                # Si quieres que además muestre el número de licencia pon esto:
                # tiene_licencia = f"Sí ({doc.numero_documento})" if doc.numero_documento else "Sí"
                tiene_licencia = "Sí"
                break

        row_data = [
            contador,
            str(empleado.id),
            empleado.origen_id or '',
            empleado.numero_empleado or '',
            empleado.nombre,
            empleado.apellido,
            empleado.puesto.nombre if empleado.puesto else 'N/A',
            empleado.puesto_anterior or '',
            empleado.departamento.nombre if empleado.departamento else 'N/A',
            empleado.supervisor.nombre_completo if empleado.supervisor else 'N/A',
            empleado.email or '',
            fecha_ingreso,
            fecha_contratacion,
            'Activo' if empleado.activo else 'Inactivo',
            empleado.motivo_inactivacion.motivo if empleado.motivo_inactivacion else '',
            fecha_inactivacion,
            fecha_nacimiento,
            empleado.edad if empleado.edad is not None else '',
            empleado.antiguedad if empleado.antiguedad is not None else '',
            empleado.direccion or '',
            empleado.colonia or '',
            empleado.codigo_postal or '',
            empleado.ciudad or '',
            empleado.estado or '',
            empleado.pais or '',
            empleado.telefono_personal or '',
            empleado.get_estado_civil_display() if empleado.estado_civil else '',
            empleado.nacionalidad or '',
            empleado.curp or '',
            empleado.rfc or '',
            empleado.nss or '',
            nombre_empresa,
            tipos_viaje_str,
            tipos_carga_str,
            divisiones,
            tiene_licencia, # <--- SE INSERTA EL DATO DE LA LICENCIA AQUÍ
            empleado.nombre_conyuge or '',
            empleado.telefono_conyuge or '',
            empleado.banco or '',
            empleado.clabe_interbancaria or '',
            empleado.numero_cuenta or '',
            empleado.numero_tarjeta or '',
            empleado.nombre_referencia_1 or '',
            empleado.telefono_referencia_1 or '',
            empleado.relacion_referencia_1 or '',
            empleado.nombre_referencia_2 or '',
            empleado.telefono_referencia_2 or '',
            empleado.relacion_referencia_2 or '',
        ]
        worksheet.append(row_data)
        contador -= 1 # Resta 1 para la siguiente fila

    full_range = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    tabla = Table(displayName="TablaEmpleados", ref=full_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = style
    worksheet.add_table(tabla)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        worksheet.column_dimensions[column].width = min((max_length + 4), 60) # Limitamos un poco el ancho máximo para no deformar tanto el Excel

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_empleados_filtrado.xlsx"'
    workbook.save(response)
    return response
def reporte_bajas(request):
    """
    Vista para mostrar un reporte de bajas (Renuncias, Bajas, Abandonos)
    y permitir la exportación a un archivo Excel con datos y gráficas.
    """
    eventos_baja = HistorialLaboral.objects.filter(
        tipo_evento__in=['RENUNCIA', 'BAJA', 'ABANDONO']
    ).select_related('empleado').order_by('-fecha_inicio')

    # If the request asks for Excel export
    if request.GET.get('export') == 'excel':
        
        # --- 1. DATA PREPARATION WITH PANDAS ---
        data = []
        # Re-create the flat choices dictionary as it's needed here
        MOTIVO_SALIDA_CHOICES_FLAT = [choice for group in HistorialLaboral.MOTIVO_SALIDA_CHOICES for choice in group[1]]
        evento_choices_dict = dict(HistorialLaboral.EVENT_CHOICES)
        motivo_choices_dict = dict(MOTIVO_SALIDA_CHOICES_FLAT) # Use the flat list for the dict

        for evento in eventos_baja:
            data.append({
                'ID Empleado': evento.empleado.id,
                'Nombre Empleado': f"{evento.empleado.nombre} {evento.empleado.apellido}",
                'Número de Empleado': evento.empleado.numero_empleado,
                'Tipo de Evento': evento_choices_dict.get(evento.tipo_evento, evento.tipo_evento),
                'Motivo de Salida': motivo_choices_dict.get(evento.motivo_salida, evento.motivo_salida),
                'Fecha del Evento': evento.fecha_inicio,
                'Comentario': evento.descripcion,
            })
        
        df = pd.DataFrame(data)

        # --- 2. EXCEL WORKBOOK AND SHEET CREATION ---
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Reporte de Bajas"
        
        # --- 3. DATA SHEET (REPORT) ---
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)

        # Styles and formatting for the data table
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, col in enumerate(ws_data.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx) # Get column letter for setting width
            for cell_idx, cell in enumerate(col):
                # Apply styles
                if cell_idx == 0: # Header row (0-indexed for cells in column iteration)
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = center_align
                
                # Find max length
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_data.column_dimensions[column_letter].width = adjusted_width

        # --- 4. CHARTS SHEET ---
        if not df.empty:
            ws_charts = wb.create_sheet("Graficas")
            df_charts = df.copy()
            df_charts['Fecha del Evento'] = pd.to_datetime(df_charts['Fecha del Evento'])

            # Monthly data
            monthly_counts = df_charts.groupby(df_charts['Fecha del Evento'].dt.strftime('%Y-%m')).size().reset_index(name='Total')
            monthly_counts = monthly_counts.sort_values(by='Fecha del Evento', ascending=True) # Sort by date for chart

            # Weekly data (format: Year-Week)
            weekly_counts = df_charts.groupby(df_charts['Fecha del Evento'].dt.strftime('%Y-W%U')).size().reset_index(name='Total')
            weekly_counts = weekly_counts.sort_values(by='Fecha del Evento', ascending=True) # Sort by date for chart
            
            # Write data for charts on the sheet
            # Monthly Chart Data
            ws_charts.append(['Mes', 'Total Bajas'])
            for _, row in monthly_counts.iterrows(): ws_charts.append(list(row))
            
            ws_charts.append([]) # Blank space
            
            # Weekly Chart Data
            start_row_weekly_data = ws_charts.max_row + 1
            ws_charts.append(['Semana', 'Total Bajas'])
            for _, row in weekly_counts.iterrows(): ws_charts.append(list(row))

            # Create Monthly Chart
            chart_monthly = BarChart()
            chart_monthly.title = "Bajas por Mes"
            chart_monthly.y_axis.title = "Cantidad"
            chart_monthly.x_axis.title = "Mes"
            
            # Data references for monthly chart
            monthly_data_range = Reference(ws_charts, min_col=2, min_row=1, max_row=len(monthly_counts)+1, max_col=2)
            monthly_cats_range = Reference(ws_charts, min_col=1, min_row=2, max_row=len(monthly_counts)+1) # Categories starting from row 2

            chart_monthly.add_data(monthly_data_range, titles_from_data=True)
            chart_monthly.set_categories(monthly_cats_range)
            chart_monthly.dataLabels = openpyxl.chart.label.DataLabelList()
            chart_monthly.dataLabels.showVal = True # Show value on bar
            ws_charts.add_chart(chart_monthly, "D2") # Position the chart

            # Create Weekly Chart
            chart_weekly = BarChart()
            chart_weekly.title = "Bajas por Semana"
            chart_weekly.y_axis.title = "Cantidad"
            chart_weekly.x_axis.title = "Semana"

            # Data references for weekly chart
            weekly_data_range = Reference(ws_charts, min_col=2, min_row=start_row_weekly_data, max_row=ws_charts.max_row, max_col=2)
            weekly_cats_range = Reference(ws_charts, min_col=1, min_row=start_row_weekly_data + 1, max_row=ws_charts.max_row) # Categories starting from row after header

            chart_weekly.add_data(weekly_data_range, titles_from_data=True)
            chart_weekly.set_categories(weekly_cats_range)
            chart_weekly.dataLabels = openpyxl.chart.label.DataLabelList()
            chart_weekly.dataLabels.showVal = True
            ws_charts.add_chart(chart_weekly, "D20") # Position the chart below monthly chart

        # --- 5. RESPONSE AND DOWNLOAD ---
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_bajas_personal.xlsx"'
        wb.save(response)
        return response

    # Render the HTML template if not exporting
    context = {
        'eventos_baja': eventos_baja,
    }
    return render(request, 'rh/reportes.html', context)

def dashboard_view(request):
    """
    Vista para el dashboard de análisis de personal,
    con segmentación administrativa vs operativa y por empresa/división.
    """
    from django.db.models import Count, Q
    
    # Parámetros de filtro
    selected_company = request.GET.get('empresa', 'MIGMAR')
    selected_division_id = request.GET.get('division', '')

    # Query base: Todos los empleados activos de la empresa seleccionada
    base_empleados = Empleado.objects.filter(activo=True, empresa=selected_company)
    
    # Aplicar filtro de división si se proporciona
    if selected_division_id:
        base_empleados = base_empleados.filter(division_operativa__id=selected_division_id)
    
    # 1. Segmentación Administrativo vs Operativo
    # Un "Operador" se identifica por su puesto
    palabras_operador = ['operador', 'chofer', 'conductor', 'motorista', 'driver', 'operadora']
    q_operadores = Q()
    for palabra in palabras_operador:
        q_operadores |= Q(puesto__nombre__icontains=palabra)
        
    operadores_activos = base_empleados.filter(q_operadores).distinct()
    total_operadores = operadores_activos.count()
    
    administrativos_activos = base_empleados.exclude(id__in=operadores_activos.values_list('id', flat=True))
    total_administrativos = administrativos_activos.count()
    
    total_headcount = total_operadores + total_administrativos

    # 2. Análisis de Operadores (Locales vs Foráneos)
    # Identificamos a los que son AMBOS
    operadores_ambos_list = operadores_activos.filter(tipo_viaje__nombre='Local').filter(tipo_viaje__nombre='Foraneo').distinct()
    total_ambos = operadores_ambos_list.count()
    ids_ambos = set(operadores_ambos_list.values_list('id', flat=True))

    operadores_locales_list = operadores_activos.filter(tipo_viaje__nombre='Local').exclude(id__in=ids_ambos)
    total_locales = operadores_locales_list.count()

    operadores_foraneos_list = operadores_activos.filter(tipo_viaje__nombre='Foraneo').exclude(id__in=ids_ambos)
    total_foraneos = operadores_foraneos_list.count()

    # 3. Frecuencia por División (para gráfico de barras)
    divisiones = DivisionOperativa.objects.all()
    frecuencia_por_division = []
    for div in divisiones:
        count = base_empleados.filter(division_operativa=div).count()
        if count > 0 or div.nombre in ['Autozone', 'Walmart', 'Chihuahua']:
            frecuencia_por_division.append({'grupo': div.nombre, 'total': count})

    # 4. Pertenencia Múltiple (Diagrama de Venn)
    # Solo para los 3 grupos principales mencionados en la solicitud original
    ops_autozone = set(operadores_activos.filter(division_operativa__nombre='Autozone').values_list('id', flat=True))
    ops_walmart = set(operadores_activos.filter(division_operativa__nombre='Walmart').values_list('id', flat=True))
    ops_chihuahua = set(operadores_activos.filter(division_operativa__nombre='Chihuahua').values_list('id', flat=True))

    venn_data = {
        'group1_total': len(ops_autozone),
        'group2_total': len(ops_walmart),
        'group3_total': len(ops_chihuahua),
        'g1_g2': len(ops_autozone.intersection(ops_walmart)),
        'g1_g3': len(ops_autozone.intersection(ops_chihuahua)),
        'g2_g3': len(ops_walmart.intersection(ops_chihuahua)),
        'g1_g2_g3': len(ops_autozone.intersection(ops_walmart).intersection(ops_chihuahua)),
    }

    context = {
        'page_title': 'Dashboard de Recursos Humanos',
        'selected_company': selected_company,
        'total_headcount': total_headcount,
        'total_administrativos': total_administrativos,
        'total_operadores': total_operadores,
        'operadores_locales': total_locales,
        'operadores_foraneos': total_foraneos,
        'operadores_ambos': total_ambos,
        'operadores_locales_list': operadores_locales_list,
        'operadores_foraneos_list': operadores_foraneos_list,
        'operadores_ambos_list': operadores_ambos_list,
        'frecuencia_por_grupo': frecuencia_por_division,
        'frecuencia_por_grupo_json': json.dumps(frecuencia_por_division),
        'venn_data': venn_data,
        'venn_data_json': json.dumps(venn_data),
        'groups_names': ['Autozone', 'Walmart', 'Chihuahua'], 
        'groups_names_json': json.dumps(['Autozone', 'Walmart', 'Chihuahua']),
        'divisiones_disponibles': DivisionOperativa.objects.all(),
        'selected_division_id': int(selected_division_id) if selected_division_id else None,
    }
    return render(request, 'rh/dashboard.html', context)


def vacantes_dashboard_view(request):
    today = date.today()

    # 1. GET VACANCIES (TERMINATIONS WITHOUT REPLACEMENT AND REPLACED ONES)
    tipos_baja = ['RENUNCIA', 'BAJA', 'ABANDONO']
    vacantes = HistorialLaboral.objects.filter(tipo_evento__in=tipos_baja)\
        .select_related('empleado__departamento', 'empleado__puesto', 'reemplazo')\
        .order_by('estatus', '-fecha_inicio')

    # Calculate elapsed days and get potential replacements
    for vacante in vacantes:
        if vacante.estatus == 'BUSCANDO':
            vacante.dias_transcurridos = (today - vacante.fecha_inicio).days
            # Logic to find replacements: active employees with the same position and department
            if vacante.empleado and vacante.empleado.puesto and vacante.empleado.departamento:
                vacante.potenciales_reemplazos = Empleado.objects.filter(
                    activo=True,
                    puesto=vacante.empleado.puesto,
                    departamento=vacante.empleado.departamento
                ).exclude(id=vacante.empleado.id)
            else:
                vacante.potenciales_reemplazos = Empleado.objects.none()
        else:
            if vacante.fecha_reemplazo:
                vacante.dias_transcurridos = (vacante.fecha_reemplazo - vacante.fecha_inicio).days
            else:
                vacante.dias_transcurridos = 0

    # 2. PREPARE DATA FOR CHARTS
    pie_chart_data = {}
    bar_chart_data = {}

    for empresa in ['MIGMAR', 'MARCO_MORALES']:
        activos = Empleado.objects.filter(activo=True, empresa=empresa).count()
        pendientes = HistorialLaboral.objects.filter(
            tipo_evento__in=tipos_baja,
            empleado__empresa=empresa,
            estatus='BUSCANDO'
        ).count()
        pie_chart_data[empresa] = {'activos': activos, 'pendientes': pendientes}

        avg_days_data = HistorialLaboral.objects.filter(
            empleado__empresa=empresa,
            estatus='REMPLAZADO',
            fecha_reemplazo__isnull=False
        ).values('empleado__departamento__nombre')\
         .annotate(
            avg_days=Avg(F('fecha_reemplazo') - F('fecha_inicio'))
         ).values('empleado__departamento__nombre', 'avg_days')
        
        bar_chart_data[empresa] = {
            'departamento': [
                {'name': item['empleado__departamento__nombre'], 'avg_days': item['avg_days'].days if item['avg_days'] else 0}
                for item in avg_days_data if item['empleado__departamento__nombre']
            ]
        }

    # 3. CONTROL DE VACANTES (FALTANTES POR PRESUPUESTO)
    controles = ControlVacante.objects.select_related('puesto', 'division').all()
    faltantes_data = []
    
    for c in controles:
        # Filtrar empleados activos que coincidan con la empresa, puesto y división
        actual_query = Empleado.objects.filter(
            activo=True,
            empresa=c.empresa,
            puesto=c.puesto
        )
        if c.division:
            actual_query = actual_query.filter(division_operativa=c.division)
            
        count_actual = actual_query.count()
        faltantes = c.cantidad_presupuestada - count_actual
        
        if faltantes > 0:
            faltantes_data.append({
                'empresa': c.get_empresa_display(),
                'puesto': c.puesto.nombre,
                'division': c.division.nombre if c.division else 'N/A',
                'objetivo': c.cantidad_presupuestada,
                'actual': count_actual,
                'faltantes': faltantes
            })

    context = {
        'vacantes': vacantes,
        'pie_chart_data': pie_chart_data,
        'bar_chart_data': bar_chart_data,
        'faltantes_data': faltantes_data,
    }

    return render(request, 'rh/vacantes_dashboard.html', context)


def asignar_reemplazo(request, pk):
    if request.method == 'POST':
        vacante = get_object_or_404(HistorialLaboral, pk=pk)
        reemplazo_id = request.POST.get('reemplazo_id')
        if reemplazo_id:
            reemplazo = get_object_or_404(Empleado, pk=reemplazo_id)
            vacante.reemplazo = reemplazo
            vacante.estatus = 'REMPLAZADO'
            vacante.fecha_reemplazo = date.today()
            vacante.save()
    return redirect('rh:vacantes_dashboard')

def reporte_documentacion_operador(request):
    """
    Genera un reporte que muestra los operadores activos a quienes les falta
    uno o más de los tipos de documento de operador requeridos.
    El estatus 'Incompleto' se determina dinámicamente para este reporte.
    """
    # 1. Obtener todos los tipos de documento que se consideran obligatorios.
    # Esto hace que el reporte sea dinámico a cambios en la lista de tipos.
    todos_los_tipos_requeridos = TipoDocumentoOperador.objects.all()
    mapa_tipos_requeridos = {tipo.id: tipo.nombre for tipo in todos_los_tipos_requeridos}
    ids_tipos_requeridos = set(mapa_tipos_requeridos.keys())

    # 2. Obtener todos los operadores activos y precargar sus documentos
    # para optimizar las consultas a la base de datos (evita el problema N+1).
    operadores_activos = Empleado.objects.filter(
        activo=True,
        puesto__nombre__icontains='Operador'
    ).prefetch_related('documentos_operador')

    # 3. Identificar a los operadores con documentación incompleta.
    operadores_incompletos = []
    for operador in operadores_activos:
        # Obtener los IDs de los tipos de documento que el operador ya ha entregado.
        ids_documentos_entregados = set(
            operador.documentos_operador.values_list('tipo_documento_id', flat=True)
        )

        # Encontrar la diferencia entre los documentos requeridos y los entregados.
        ids_documentos_faltantes = ids_tipos_requeridos - ids_documentos_entregados

        if ids_documentos_faltantes:
            # Si faltan documentos, se construye la lista de nombres de esos documentos.
            nombres_documentos_faltantes = [mapa_tipos_requeridos[id_faltante] for id_faltante in ids_documentos_faltantes]
            
            # Se añade el operador y su lista de documentos faltantes al resultado.
            operadores_incompletos.append({
                'operador': operador,
                'documentos_faltantes': nombres_documentos_faltantes
            })

    # 4. Preparar el contexto para la plantilla.
    context = {
        'operadores_incompletos': operadores_incompletos,
    }

    # 5. Renderizar la plantilla con los datos.
    return render(request, 'rh/reporte_documentacion_operador.html', context)


def descargar_documentos_empleado(request, pk):
    """
    ============================================================================
    DESCARGA DE DOCUMENTOS COMPLETA PARA EMPLEADO (S3 COMPATIBLE)
    ============================================================================
    """
    from django.http import FileResponse
    import io
    
    # ========== 1. OBTENER EMPLEADO ==========
    empleado = get_object_or_404(Empleado, pk=pk)
    
    # ========== 2. CREAR DOCUMENT GENERATOR ==========
    try:
        generator = DocumentGenerator(empleado)
    except Exception as e:
        logger.error(f"❌ Error creando DocumentGenerator: {str(e)}")
        return HttpResponse(f"Error interno al crear generador de documentos: {str(e)}", status=500)
    
    # ========== 3. VERIFICAR ESTRUCTURA EN STORAGE ==========
    # La lógica de existencia (S3 vs Local) ahora la maneja internamente
    # el DocumentGenerator capa por capa para evitar fallos de default_storage.
    
    # ========== 4. GENERAR ZIP EN MEMORIA ==========
    try:
        zip_buffer = generator.create_zip_file()
        
        if not zip_buffer:
            return HttpResponse("No se pudieron generar documentos para este empleado.", status=404)
        
        # ========== 5. RETORNAR ARCHIVO ==========
        from datetime import datetime as _dt
        timestamp = _dt.now().strftime('%Y%m%d_%H%M')
        tipo = 'OPERADOR' if generator.es_operador else 'ADMIN'
        zip_filename = f"documentos_{empleado.nombre_completo}_{tipo}_{timestamp}.zip".replace(' ', '_')
        
        return FileResponse(
            zip_buffer,
            as_attachment=True,
            filename=zip_filename,
            content_type='application/zip'
        )
        
    except Exception as e:
        logger.error(f"❌ Error en la generación de ZIP: {str(e)}")
        return HttpResponse(f"Error crítico al generar el paquete de documentos: {str(e)}", status=500)

    

def descargar_documento_individual(request, pk, tipo_documento):
    """
    Genera y descarga un único documento (S3 COMPATIBLE)
    """
    from django.http import FileResponse
    
    empleado = get_object_or_404(Empleado, pk=pk)
    generator = DocumentGenerator(empleado)
    
    # Mapeo de slugs de URL a claves internas
    slug_map = {
        'aviso-privacidad': 'aviso_privacidad',
        'convenio-confidencialidad': 'convenio_confidencialidad',
        'autorizacion-correo': 'autorizacion_correo',
        'carta-adhesion': 'carta_adhesion',
        'descripcion-puesto': 'descripcion_puesto',
        'contrato-indeterminado': 'contrato_indeterminado',
        'contrato-determinado': 'contrato_determinado',
        'carta-renuncia': 'carta_renuncia',
        'formato-desfase': 'formato_desfase',
    }
    
    doc_key = slug_map.get(tipo_documento, tipo_documento)
    
    try:
        # Generar el documento en memoria
        output = generator.generate_document(doc_key, doc_key.upper())
        
        if not output:
            return HttpResponse(f"No se pudo generar el documento {tipo_documento}.", status=404)
        
        filename = f"{doc_key.upper()} - {empleado.nombre_completo}.docx".replace(' ', '_')
        
        return FileResponse(
            output,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        logger.error(f"Error generando documento individual {tipo_documento}: {e}")
        return HttpResponse(f"Error: {str(e)}", status=500)
    

def generar_documentos_todos(request):
    """Genera documentos para todos los empleados activos (S3 COMPATIBLE)"""
    if request.method == 'POST':
        empleados = Empleado.objects.filter(activo=True)
        
        # Usar un BytesIO para el ZIP final
        zip_buffer = io.BytesIO()
        
        empleados_con_documentos = []
        empleados_sin_datos = []
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for empleado in empleados:
                try:
                    if not empleado.nombre or not empleado.apellido:
                        empleados_sin_datos.append(f"{empleado.id} - Sin nombre/apellido")
                        continue
                    
                    # Generar documentos en memoria
                    generator = DocumentGenerator(empleado)
                    documents = generator.generate_all_documents()
                    
                    if documents:
                        empleado_folder = f"{empleado.nombre}_{empleado.apellido}".replace(" ", "_")
                        
                        # Mapeo de nombres
                        nombres = {
                            'aviso_privacidad': "AVISO DE PRIVACIDAD.docx",
                            'convenio_confidencialidad': "CONVENIO CONFIDENCIALIDAD.docx",
                            'autorizacion_correo': "AUTORIZACION CORREO SAT.docx",
                            'carta_adhesion': "CARTA ADHESION PLAN.docx",
                            'contrato_trabajo': "CONTRATO TRABAJO.docx",
                            'descripcion_puesto': "DESCRIPCION PUESTO.docx",
                            'carta_renuncia': "CARTA DE RENUNCIA OPERADOR.docx",
                            'formato_desfase': "FORMATO DE DESFASE OPERADOR.docx",
                            'contrato_indeterminado': "CONTRATO INDETERMINADO.docx",
                            'contrato_determinado': "CONTRATO DETERMINADO.docx",
                        }
                        
                        for doc_type, doc_content in documents.items():
                            if doc_content:
                                filename = nombres.get(doc_type, f"{doc_type}.docx")
                                arcname = f"{empleado_folder}/{filename}"
                                zipf.writestr(arcname, doc_content.getvalue())
                        
                        empleados_con_documentos.append(f"{empleado.nombre} {empleado.apellido}")
                    
                except Exception as e:
                    empleados_sin_datos.append(f"{empleado.id} - Error: {str(e)}")
        
        zip_buffer.seek(0)
        
        # Como este view retorna un JSON con una URL de descarga en el original, 
        # pero ahora no queremos guardar archivos en media/temp permanently si es S3,
        # lo ideal sería retornar el FileResponse directamente o guardar temporalmente en storage.
        # Por simplicidad y consistencia con el flujo original, guardaremos el ZIP temporalmente en storage.
        
        zip_filename = f"temp/documentos_todos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        default_storage.save(zip_filename, zip_buffer)
        
        return JsonResponse({
            'success': True,
            'zip_url': default_storage.url(zip_filename),
            'empleados_procesados': len(empleados_con_documentos),
            'empleados_sin_datos': len(empleados_sin_datos),
            'total_empleados': empleados.count()
        })
        
        return response
    
    # GET request - mostrar formulario
    total_empleados = Empleado.objects.filter(activo=True).count()
    
    return render(request, 'rh/generar_documentos_todos.html', {
        'total_empleados': total_empleados,
        'page_title': 'Generar Documentos para Todos los Empleados'
    })


def descargar_documentos_todos(request):
    """Página para descargar documentos de todos los empleados"""
    # Filtrar solo empleados activos con departamento
    empleados = Empleado.objects.filter(
        activo=True, 
        departamento__isnull=False
    ).select_related('departamento', 'puesto').order_by('apellido', 'nombre')
    
    # Agrupar por departamento en la nueva estructura
    departamentos = {}
    for empleado in empleados:
        if empleado.departamento:  # Verificar que tiene departamento
            depto_nombre = empleado.departamento.nombre
            depto_id = empleado.departamento.id
            
            if depto_nombre not in departamentos:
                departamentos[depto_nombre] = {
                    'id': depto_id,
                    'empleados': []
                }
            
            # Solo agregar si tiene PK
            if empleado.pk:
                # Determinar si es operador de manera simple
                es_operador = False
                if empleado.puesto and empleado.puesto.nombre:
                    puesto_nombre = empleado.puesto.nombre.lower()
                    palabras_clave = ['operador', 'chofer', 'conductor', 'motorista', 'driver']
                    es_operador = any(palabra in puesto_nombre for palabra in palabras_clave)
                
                departamentos[depto_nombre]['empleados'].append({
                    'empleado': empleado,
                    'es_operador': es_operador,
                    'tipo_empleado': 'Operador' if es_operador else 'Personal Administrativo'
                })
    
    return render(request, 'rh/descargar_documentos_todos.html', {
        'departamentos': departamentos,
        'total_empleados': empleados.count(),
        'page_title': 'Documentos de Todos los Empleados'
    })

def generar_documentos_por_departamento(request, departamento_id):
    """Genera documentos para todos los empleados de un departamento (S3 COMPATIBLE)"""
    departamento = get_object_or_404(Departamento, pk=departamento_id)
    empleados = Empleado.objects.filter(departamento=departamento, activo=True)
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for empleado in empleados:
            try:
                generator = DocumentGenerator(empleado)
                documents = generator.generate_all_documents()
                
                if documents:
                    empleado_folder = f"{empleado.nombre}_{empleado.apellido}".replace(" ", "_")
                    
                    nombres = {
                        'aviso_privacidad': "AVISO DE PRIVACIDAD.docx",
                        'convenio_confidencialidad': "CONVENIO CONFIDENCIALIDAD.docx",
                        'autorizacion_correo': "AUTORIZACION CORREO SAT.docx",
                        'carta_adhesion': "CARTA ADHESION PLAN.docx",
                        'contrato_indeterminado': "CONTRATO INDETERMINADO.docx",
                        'contrato_determinado': "CONTRATO DETERMINADO.docx",
                        'descripcion_puesto': "DESCRIPCION PUESTO.docx",
                        'carta_renuncia': "CARTA DE RENUNCIA OPERADOR.docx",
                        'formato_desfase': "FORMATO DE DESFASE OPERADOR.docx",
                    }
                    
                    for doc_type, doc_content in documents.items():
                        if doc_content:
                            filename = nombres.get(doc_type, f"{doc_type}.docx")
                            arcname = f"{empleado_folder}/{filename}"
                            zipf.writestr(arcname, doc_content.getvalue())
                                
            except Exception as e:
                print(f"Error con {empleado.nombre}: {e}")
    
    zip_buffer.seek(0)
    zip_filename = f"documentos_{departamento.nombre}_{datetime.now().strftime('%Y%m%d')}.zip".replace(' ', '_')
    
    return FileResponse(zip_buffer, as_attachment=True, filename=zip_filename, content_type='application/zip')
    
    return response

def estado_generacion_documentos(request, pk):
    """
    ============================================================================
    ENDPOINT PARA VERIFICAR ESTADO DE GENERACIÓN
    ============================================================================
    Útil para interfaces que quieren mostrar progreso o verificar
    si la generación está en curso/completada.
    ============================================================================
    """
    # Esta función podría expandirse para monitorear generación asíncrona
    # Por ahora solo verifica si el empleado existe
    
    try:
        empleado = get_object_or_404(Empleado, pk=pk)
        
        return JsonResponse({
            'status': 'ready',
            'empleado': {
                'id': empleado.id,
                'nombre_completo': empleado.nombre_completo,
                'puesto': empleado.puesto.nombre if empleado.puesto else None,
                'es_operador': 'operador' in (empleado.puesto.nombre.lower() if empleado.puesto else '')
            },
            'documentos_disponibles': [
                'aviso-privacidad',
                'convenio-confidencialidad',
                'autorizacion-correo',
                'carta-adhesion',
                'descripcion-puesto',
                'contrato-indeterminado',
                'contrato-determinado',
            ] + (['carta-renuncia', 'formato-desfase'] 
                 if 'operador' in (empleado.puesto.nombre.lower() if empleado.puesto else '') else [])
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_GET
def diagnostico_storage(request):
    """
    Vista de diagnóstico para verificar el contenido del storage (S3 o Local)
    y confirmar la existencia de las carpetas de plantillas.
    """
    if not request.user.is_staff:
        return HttpResponse("Acceso denegado: Se requiere ser administrador.", status=403)
        
    results = {
        'storage_backend': str(default_storage.__class__),
        'base_dir': 'plantillas',
        'carpetas_verificadas': [],
        'archivos_encontrados': [],
        'errores': []
    }
    
    carpetas_a_revisar = ['plantillas', 'plantillas/operadores', 'plantillas/administrativos', 'media', 'media/plantillas']
    
    for folder in carpetas_a_revisar:
        try:
            exists = default_storage.exists(folder)
            results['carpetas_verificadas'].append({
                'path': folder,
                'existe': exists
            })
            
            if exists:
                try:
                    # Intentar listar el contenido
                    dirs, files = default_storage.listdir(folder)
                    for f in files:
                        results['archivos_encontrados'].append(f"{folder}/{f}")
                except Exception as le:
                    results['errores'].append(f"Error listando {folder}: {str(le)}")
        except Exception as e:
            results['errores'].append(f"Error revisando {folder}: {str(e)}")
            
    return render(request, 'rh/diagnostico_storage.html', {'results': results})

def inicializar_datos_operativos(request):
    """
    Inicializa datos por defecto para los modelos operativos si están vacíos.
    """
    if not request.user.is_staff:
        return HttpResponse("Acceso denegado.", status=403)
        
    created_counts = {
        'DivisionOperativa': 0,
        'TipoCarga': 0,
        'TipoViaje': 0
    }
    
    # 1. Divisiones Operativas
    divisiones = ['Autozone', 'Walmart', 'Bafar', 'Femsa', 'Sams', 'Bodega Aurrera', 'Chihuahua']
    for d in divisiones:
        obj, created = DivisionOperativa.objects.get_or_create(nombre=d)
        if created: created_counts['DivisionOperativa'] += 1
        
    # 2. Tipos de Carga
    cargas = ['Seco', 'Refrigerado', 'Peligroso', 'Paquetería']
    for c in cargas:
        obj, created = TipoCarga.objects.get_or_create(nombre=c)
        if created: created_counts['TipoCarga'] += 1
        
    # 3. Tipos de Viaje
    viajes = ['Local', 'Foraneo', 'Cruce']
    for v in viajes:
        obj, created = TipoViaje.objects.get_or_create(nombre=v)
        if created: created_counts['TipoViaje'] += 1
        
    return render(request, 'rh/inicializacion_datos.html', {'counts': created_counts})

    
# En RH/views.py - Agrega estas vistas

@login_required
def dar_baja_empleado(request, pk):
    """
    Vista para dar de baja a un empleado con todos los motivos jerarquizados
    y gestión de conciliación y arbitraje
    """
    empleado = get_object_or_404(Empleado, pk=pk)
    
    # Verificar si ya está inactivo
    if not empleado.activo:
        messages.warning(request, f"{empleado.nombre_completo} ya se encuentra inactivo.")
        return redirect('rh:detalle_empleado', pk=pk)
    
    if request.method == 'POST':
        form = BajaEmpleadoForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Guardar la baja
                    baja = form.save(commit=False)
                    baja.empleado = empleado
                    baja.registrado_por = request.user
                    
                    # Si fue a conciliación pero no se especificó fecha, usar fecha de baja
                    if baja.fue_conciliacion_arbitraje and not baja.fecha_conciliacion_arbitraje:
                        baja.fecha_conciliacion_arbitraje = baja.fecha_baja
                    
                    baja.save()
                    
                    # Dar de baja al empleado
                    empleado.activo = False
                    empleado.fecha_inactivacion = baja.fecha_baja
                    
                    # Obtener el modelo MotivoInactivacion existente o crear uno
                    motivo_nombre = baja.cadena_motivos[:200]  # Limitar longitud
                    
                    # Agregar información de conciliación al motivo si aplica
                    descripcion_extra = ""
                    if baja.fue_conciliacion_arbitraje:
                        descripcion_extra = f" (CONCILIACIÓN Y ARBITRAJE - {baja.fecha_conciliacion_arbitraje})"
                    
                    motivo_inactivacion, _ = MotivoInactivacion.objects.get_or_create(
                        motivo=motivo_nombre + descripcion_extra,
                        defaults={'descripcion': baja.comentario_baja[:500]}
                    )
                    
                    empleado.motivo_inactivacion = motivo_inactivacion
                    empleado.save()
                    
                    # Crear registro en historial laboral
                    descripcion_historial = f"Baja: {baja.cadena_motivos}"
                    
                    # Agregar información de conciliación al historial
                    if baja.fue_conciliacion_arbitraje:
                        conciliacion_info = f" - Proceso de Conciliación y Arbitraje ({baja.fecha_conciliacion_arbitraje})"
                        descripcion_historial += conciliacion_info
                        
                        if baja.observaciones_conciliacion:
                            descripcion_historial += f" - Obs: {baja.observaciones_conciliacion[:100]}"
                    
                    HistorialLaboral.objects.create(
                        empleado=empleado,
                        tipo_evento='BAJA',
                        fecha_inicio=baja.fecha_baja,
                        descripcion=descripcion_historial,
                        motivo_salida=baja.motivo_principal.nombre if baja.motivo_principal else 'BAJA'
                    )
                    
                    # Mensaje de éxito principal
                    messages.success(
                        request,
                        f"✅ {empleado.nombre_completo} ha sido dado de baja exitosamente."
                    )
                    
                    # Mensaje adicional si fue a conciliación
                    if baja.fue_conciliacion_arbitraje:
                        messages.info(
                            request,
                            f"📋 Se registró proceso de Conciliación y Arbitraje con fecha: {baja.fecha_conciliacion_arbitraje}"
                        )
                        
                        if baja.documento_conciliacion:
                            messages.success(
                                request,
                                f"📄 Documento de conciliación guardado correctamente."
                            )
                    
                    # Mensaje si es recontratable
                    if baja.es_recontratable:
                        messages.warning(
                            request,
                            f"⚠️ Este empleado ha sido marcado como RECONTRATABLE. "
                            f"Fecha posible: {baja.fecha_posible_recontratacion if baja.fecha_posible_recontratacion else 'No especificada'}"
                        )
                    
                    # Redirigir al detalle de la baja
                    return redirect('rh:detalle_baja', pk=baja.id)
                    
            except Exception as e:
                messages.error(
                    request,
                    f"❌ Error al dar de baja: {str(e)}"
                )
                logger.error(f"Error dando de baja a {empleado.nombre_completo}: {str(e)}")
        else:
            messages.error(
                request,
                "❌ Por favor corrige los errores en el formulario."
            )
            # Debug: mostrar errores específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"• {field}: {error}")
    else:
        # Formulario inicial con fecha de hoy
        initial_data = {
            'fecha_baja': date.today()
        }
        form = BajaEmpleadoForm(initial=initial_data)
    
    context = {
        'empleado': empleado,
        'form': form,
        'page_title': f'Dar de Baja - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/dar_baja.html', context)


@login_required
def historial_bajas(request):
    """
    Vista para ver el historial de bajas
    """
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    motivo_principal_id = request.GET.get('motivo_principal')
    recontratable = request.GET.get('recontratable')
    recontratado = request.GET.get('recontratado')
    
    bajas = BajaEmpleado.objects.all().select_related(
        'empleado',
        'motivo_principal',
        'motivo_secundario',
        'motivo_detalle',
        'registrado_por'
    ).order_by('-fecha_baja')
    
    # Aplicar filtros
    if fecha_inicio:
        bajas = bajas.filter(fecha_baja__gte=fecha_inicio)
    
    if fecha_fin:
        bajas = bajas.filter(fecha_baja__lte=fecha_fin)
    
    if motivo_principal_id:
        bajas = bajas.filter(motivo_principal_id=motivo_principal_id)
    
    if recontratable == 'si':
        bajas = bajas.filter(es_recontratable=True)
    elif recontratable == 'no':
        bajas = bajas.filter(es_recontratable=False)
    
    if recontratado == 'si':
        bajas = bajas.filter(recontratado=True)
    elif recontratado == 'no':
        bajas = bajas.filter(recontratado=False)
    
    # Estadísticas
    total_bajas = bajas.count()
    recontratables = bajas.filter(es_recontratable=True).count()
    recontratados = bajas.filter(recontratado=True).count()
    
    # Distribución por motivo principal
    distribucion_motivos = bajas.values(
        'motivo_principal__nombre'
    ).annotate(
        total=Count('id')
    ).order_by('-total')
    
    context = {
        'bajas': bajas,
        'total_bajas': total_bajas,
        'recontratables': recontratables,
        'recontratados': recontratados,
        'distribucion_motivos': distribucion_motivos,
        'motivos_principales': MotivoBaja.objects.filter(tipo_motivo='PRINCIPAL', activo=True),
        'page_title': 'Historial de Bajas',
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    
    return render(request, 'rh/historial_bajas.html', context)


@login_required
def detalle_baja(request, pk):
    """
    Vista para ver los detalles de una baja específica
    """
    baja = get_object_or_404(
        BajaEmpleado.objects.select_related(
            'empleado',
            'motivo_principal',
            'motivo_secundario',
            'motivo_detalle',
            'registrado_por'
        ),
        pk=pk
    )
    
    context = {
        'baja': baja,
        'page_title': f'Detalle de Baja - {baja.empleado.nombre_completo}',
    }
    
    return render(request, 'rh/detalle_baja.html', context)


@login_required
def recontratar_empleado(request, pk):
    """
    Vista para recontratar a un empleado dado de baja
    """
    baja = get_object_or_404(
        BajaEmpleado.objects.select_related('empleado'),
        pk=pk,
        es_recontratable=True,
        recontratado=False
    )
    
    if request.method == 'POST':
        form = RecontratacionForm(request.POST, instance=baja)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Recontratar al empleado
                    baja = form.save(commit=False)
                    if baja.recontratar():
                        messages.success(
                            request,
                            f"✅ {baja.empleado.nombre_completo} ha sido recontratado exitosamente."
                        )
                        return redirect('rh:detalle_empleado', pk=baja.empleado.id)
                    else:
                        messages.error(
                            request,
                            "❌ No se pudo recontratar al empleado."
                        )
            except Exception as e:
                messages.error(
                    request,
                    f"❌ Error al recontratar: {str(e)}"
                )
                logger.error(f"Error recontratando a {baja.empleado.nombre_completo}: {str(e)}")
        else:
            messages.error(
                request,
                "❌ Por favor corrige los errores en el formulario."
            )
    else:
        form = RecontratacionForm(instance=baja)
    
    context = {
        'baja': baja,
        'form': form,
        'page_title': f'Recontratar - {baja.empleado.nombre_completo}',
    }

    # DEBUG: Verificar si el template existe
    template_path = 'rh/recontratar.html'
    try:
        template = get_template(template_path)
        print(f"✅ Template encontrado en: {template.origin.name}")
    except:
        # Listar todos los templates disponibles
        from django.conf import settings
        print("🔍 Buscando template en:")
        for dir in settings.TEMPLATES[0]['DIRS']:
            print(f"  - {dir}")
        if settings.TEMPLATES[0].get('APP_DIRS'):
            print("  - En todas las apps instaladas")
    
    return render(request, 'rh/recontratar.html', context)


@login_required
def reporte_bajas_excel(request):
    """
    Exportar historial de bajas a Excel
    """
    bajas = BajaEmpleado.objects.all().select_related(
        'empleado',
        'motivo_principal',
        'motivo_secundario',
        'motivo_detalle'
    ).order_by('-fecha_baja')
    
    # Crear DataFrame
    data = []
    for baja in bajas:
        data.append({
            'ID': baja.id,
            'Empleado': baja.empleado.nombre_completo,
            'Número Empleado': baja.empleado.numero_empleado,
            'Puesto': baja.empleado.puesto.nombre if baja.empleado.puesto else '',
            'Departamento': baja.empleado.departamento.nombre if baja.empleado.departamento else '',
            'Fecha Baja': baja.fecha_baja.strftime('%d/%m/%Y') if baja.fecha_baja else '',
            'Motivo Principal': baja.motivo_principal.nombre if baja.motivo_principal else '',
            'Motivo Secundario': baja.motivo_secundario.nombre if baja.motivo_secundario else '',
            'Motivo Detalle': baja.motivo_detalle.nombre if baja.motivo_detalle else '',
            'Comentario': baja.comentario_baja[:100] if baja.comentario_baja else '',
            'Recontratable': 'Sí' if baja.es_recontratable else 'No',
            'Recontratado': 'Sí' if baja.recontratado else 'No',
            'Días desde Baja': baja.dias_desde_baja,
            'Registrado por': baja.registrado_por.get_full_name() if baja.registrado_por else '',
            'Fecha Registro': baja.fecha_registro.strftime('%d/%m/%Y %H:%M') if baja.fecha_registro else '',
        })
    
    df = pd.DataFrame(data)
    
    # Crear Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Historial de Bajas"
    
    # Agregar datos
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_bajas.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def get_submotivos_ajax(request):
    """
    Vista AJAX para obtener submotivos según el motivo principal seleccionado
    """
    motivo_principal_id = request.GET.get('motivo_principal_id')
    
    if motivo_principal_id:
        # Obtener submotivos (secundarios) del motivo principal
        submotivos = MotivoBaja.objects.filter(
            padre_id=motivo_principal_id,
            tipo_motivo='SECUNDARIO',
            activo=True
        ).values('id', 'nombre')
        
        return JsonResponse(list(submotivos), safe=False)
    
    return JsonResponse([], safe=False)


@login_required
def get_detalles_ajax(request):
    """
    Vista AJAX para obtener detalles según el motivo secundario seleccionado
    """
    motivo_secundario_id = request.GET.get('motivo_secundario_id')
    
    if motivo_secundario_id:
        # Obtener detalles del motivo secundario
        detalles = MotivoBaja.objects.filter(
            padre_id=motivo_secundario_id,
            tipo_motivo='DETALLE',
            activo=True
        ).values('id', 'nombre')
        
        return JsonResponse(list(detalles), safe=False)
    
    return JsonResponse([], safe=False)

# ==============================================
# FUNCIONES AUXILIARES PARA VACACIONES
# ==============================================

def calcular_dias_vacaciones_por_antiguedad(antiguedad_anios):
    """
    Calcula los días de vacaciones según la antigüedad:
    - 1 año: 12 días
    - 2 años: 14 días
    - 3 años: 16 días
    - 4 años: 18 días
    - 5 años: 20 días
    - +5 años: 20 días + 2 días por cada 5 años adicionales
    """
    if antiguedad_anios < 1:
        return 0
    elif antiguedad_anios == 1:
        return 12
    elif antiguedad_anios == 2:
        return 14
    elif antiguedad_anios == 3:
        return 16
    elif antiguedad_anios == 4:
        return 18
    elif antiguedad_anios == 5:
        return 20
    else:
        # 20 días base + 2 días por cada 5 años adicionales
        años_extra = antiguedad_anios - 5
        dias_extra = (años_extra // 5) * 2
        return 20 + dias_extra


def actualizar_historico_vacaciones(empleado, año=None):
    """
    Actualiza el histórico de vacaciones para un empleado
    """
    if año is None:
        año = date.today().year
    
    # Calcular antigüedad
    if empleado.fecha_contratacion:
        antiguedad = (date(año, 12, 31) - empleado.fecha_contratacion).days / 365.25
        antiguedad_anios = int(antiguedad)
    else:
        antiguedad_anios = 0
    
    # Calcular días correspondientes
    dias_correspondientes = calcular_dias_vacaciones_por_antiguedad(antiguedad_anios)
    
    # Calcular días ya tomados en este año
    vacaciones_año = Vacacion.objects.filter(
        empleado=empleado,
        fecha_inicio__year=año,
        estado__in=['APROBADO', 'GOZADO']
    )
    dias_tomados = sum(vacacion.dias_reales or 0 for vacacion in vacaciones_año)
    
    # Obtener o crear histórico
    historico, created = HistoricoVacaciones.objects.get_or_create(
        empleado=empleado,
        año=año,
        defaults={
            'dias_correspondientes': dias_correspondientes,
            'dias_tomados': dias_tomados,
            'dias_pendientes': dias_correspondientes - dias_tomados,
            'dias_antiguedad': antiguedad_anios
        }
    )
    
    if not created:
        historico.dias_correspondientes = dias_correspondientes
        historico.dias_tomados = dias_tomados
        historico.dias_pendientes = dias_correspondientes - dias_tomados
        historico.dias_antiguedad = antiguedad_anios
        historico.save()
    
    return historico


# Agregar propiedad al modelo Empleado (en views o como método)
def get_dias_vacaciones_disponibles(self):
    """Obtiene días de vacaciones disponibles para el empleado"""
    año_actual = date.today().year
    historico = HistoricoVacaciones.objects.filter(
        empleado=self,
        año=año_actual
    ).first()
    
    if historico:
        return historico.dias_pendientes
    
    # Si no existe histórico, calcular
    return calcular_dias_vacaciones_por_antiguedad(self.antiguedad)

# Agregar al modelo Empleado dinámicamente (opcional)
Empleado.dias_vacaciones_disponibles = property(get_dias_vacaciones_disponibles)


# ==============================================
# VISTAS PARA VACACIONES
# ==============================================

@login_required
def gestion_vacaciones(request):
    """
    Vista principal de gestión de vacaciones
    """
    # Filtros
    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    empleado_id = request.GET.get('empleado', '')
    departamento_id = request.GET.get('departamento', '')
    
    # Query base
    vacaciones = Vacacion.objects.all().select_related(
        'empleado', 'empleado__departamento', 'aprobado_por'
    ).order_by('-fecha_solicitud')
    
    # Aplicar filtros
    if estado:
        vacaciones = vacaciones.filter(estado=estado)
    
    if fecha_inicio:
        vacaciones = vacaciones.filter(fecha_inicio__gte=fecha_inicio)
    
    if fecha_fin:
        vacaciones = vacaciones.filter(fecha_inicio__lte=fecha_fin)
    
    if empleado_id:
        vacaciones = vacaciones.filter(empleado_id=empleado_id)
    
    if departamento_id:
        vacaciones = vacaciones.filter(empleado__departamento_id=departamento_id)
    
    # Estadísticas
    total_vacaciones = vacaciones.count()
    aprobadas = vacaciones.filter(estado='APROBADO').count()
    pendientes = vacaciones.filter(estado='PENDIENTE').count()
    gozadas = vacaciones.filter(estado='GOZADO').count()
    
    # Vacaciones próximas (para el dashboard)
    hoy = date.today()
    proximas_vacaciones = Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__gte=hoy,
        fecha_inicio__lte=hoy + timedelta(days=30)
    ).select_related('empleado').order_by('fecha_inicio')[:10]
    
    # Vacaciones activas (actualmente en curso)
    vacaciones_activas = Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__lte=hoy,
        fecha_fin__gte=hoy
    ).count()
    
    context = {
        'vacaciones': vacaciones,
        'total_vacaciones': total_vacaciones,
        'aprobadas': aprobadas,
        'pendientes': pendientes,
        'gozadas': gozadas,
        'vacaciones_activas': vacaciones_activas,
        'proximas_vacaciones': proximas_vacaciones,
        'estados_vacacion': Vacacion.ESTADO_CHOICES,
        'empleados': Empleado.objects.filter(activo=True).order_by('apellido', 'nombre'),
        'departamentos': Departamento.objects.all(),
        'filtros_aplicados': {
            'estado': estado,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'empleado_id': empleado_id,
            'departamento_id': departamento_id,
        },
        'page_title': 'Gestión de Vacaciones',
    }
    
    return render(request, 'rh/gestion_vacaciones.html', context)


@login_required
def solicitar_vacacion(request, empleado_id=None):
    """
    Vista MEJORADA para solicitar vacaciones con modo histórico y 
    soporte para días proporcionales antes de cumplir el año.
    """
    # Obtener parámetros
    empleado_id_get = request.GET.get('empleado_id')
    if empleado_id_get:
        empleado_id = empleado_id_get
    
    # Obtener parámetros de situación especial
    motivo_especial = request.GET.get('motivo_especial')
    descripcion_especial = request.GET.get('descripcion_especial')
    
    # Obtener parámetro para modo histórico
    modo_historico = request.GET.get('modo_historico') == 'true'
    
    if empleado_id:
        empleado = get_object_or_404(Empleado, pk=empleado_id, activo=True)
    else:
        empleado = None
    
    # OBTENER TODOS LOS EMPLEADOS ACTIVOS
    empleados_lista = Empleado.objects.filter(activo=True).order_by('apellido', 'nombre')
    
    # Calcular estadísticas
    total_empleados = empleados_lista.count()
    
    # Variables para el contexto
    dias_disponibles = 0
    historico = None
    dias_tomados_este_año = 0
    
    if empleado:
        # Calcular días disponibles para el empleado
        hoy = date.today()
        año_actual = hoy.year
        
        # Calcular antigüedad
        if empleado.fecha_contratacion:
            # === CÁLCULO MODIFICADO: Días proporcionales vs Antigüedad normal ===
            dias_laborados = (hoy - empleado.fecha_contratacion).days
            
            if dias_laborados < 365 and dias_laborados >= 0:
                # Parte proporcional para menos de 1 año (basado en 12 días)
                dias_disponibles = int((dias_laborados / 365.0) * 12)
            else:
                # Cálculo normal para 1 año o más
                dias_disponibles = calcular_dias_vacaciones_por_antiguedad(
                    calcular_antiguedad_años_vista(empleado.fecha_contratacion)
                )
            
            # Calcular días ya tomados este año
            dias_tomados_este_año = Vacacion.objects.filter(
                empleado=empleado,
                fecha_inicio__year=año_actual,
                estado__in=['APROBADO', 'GOZADO']
            ).aggregate(
                total=Sum('dias_reales')
            )['total'] or 0
        else:
            dias_disponibles = 0
        
        # Obtener histórico de vacaciones si existe
        historico = HistoricoVacaciones.objects.filter(empleado=empleado).order_by('-año')[:3]
    
    if request.method == 'POST':
        # Pasar el modo_historico al formulario
        form = VacacionForm(request.POST, request.FILES, 
                           empleado=empleado, 
                           modo_historico=modo_historico)
        
        if form.is_valid():
            vacacion = form.save(commit=False)
            
            if empleado:
                vacacion.empleado = empleado
            
            vacacion.creado_por = request.user
            
            # Si es modo histórico, marcar como aprobado y gozado automáticamente
            if modo_historico:
                vacacion.estado = 'GOZADO'
                vacacion.aprobado_por = request.user
                vacacion.fecha_aprobacion = date.today()
                
                # Agregar observación de registro histórico
                obs_actual = vacacion.observaciones or ""
                nueva_obs = f"[REGISTRO HISTÓRICO - {date.today().strftime('%d/%m/%Y')}]\n"
                nueva_obs += "Registro de vacaciones históricas del año anterior.\n\n"
                nueva_obs += obs_actual
                vacacion.observaciones = nueva_obs
            
            # Verificar si hay parámetros GET de situación especial
            if motivo_especial and descripcion_especial:
                obs_actual = vacacion.observaciones or ""
                nueva_obs = f"[SITUACIÓN ESPECIAL - {motivo_especial}]\n"
                nueva_obs += f"Motivo: {descripcion_especial}\n\n"
                nueva_obs += obs_actual
                vacacion.observaciones = nueva_obs
            
            vacacion.save()
            
            # Actualizar histórico si es registro histórico
            if modo_historico and vacacion.fecha_inicio:
                año_vacaciones = vacacion.fecha_inicio.year
                actualizar_historico_vacaciones(vacacion.empleado, año_vacaciones)
            
            messages.success(
                request,
                f'✅ {"Registro histórico" if modo_historico else "Solicitud"} de vacaciones '
                f'{"registrada" if modo_historico else "registrada exitosamente"} para '
                f'{vacacion.empleado.nombre_completo}.'
            )
            
            return redirect('rh:gestion_vacaciones')
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"• {field}: {error}")
    else:
        # Pasar el modo_historico al formulario
        form = VacacionForm(empleado=empleado, modo_historico=modo_historico)
    
    # Contexto para el template
    context = {
        'form': form,
        'empleado': empleado,
        'empleados': empleados_lista,
        'dias_disponibles': dias_disponibles,
        'dias_tomados_este_año': dias_tomados_este_año,
        'dias_disponibles_reales': dias_disponibles - dias_tomados_este_año,
        'total_empleados': total_empleados,
        'historico': historico,
        'motivo_especial': motivo_especial,
        'descripcion_especial': descripcion_especial,
        'modo_historico': modo_historico,
        'hoy': date.today(),
        'page_title': 'Registro Histórico de Vacaciones' if modo_historico else 'Solicitar Vacaciones',
    }
    
    return render(request, 'rh/solicitar_vacacion.html', context)


@login_required
def aprobar_vacacion(request, pk):
    """
    Vista para aprobar/rechazar vacaciones
    """
    vacacion = get_object_or_404(
        Vacacion.objects.select_related('empleado'),
        pk=pk
    )
    
    if request.method == 'POST':
        form = AprobarVacacionForm(request.POST, request.FILES, instance=vacacion)
        
        if form.is_valid():
            vacacion = form.save(commit=False)
            vacacion.aprobado_por = request.user
            vacacion.fecha_aprobacion = date.today()
            
            if vacacion.estado == 'APROBADO':
                # Actualizar histórico cuando se aprueba
                año = vacacion.fecha_inicio.year
                actualizar_historico_vacaciones(vacacion.empleado, año)
            
            vacacion.save()
            
            estado_display = vacacion.get_estado_display()
            messages.success(
                request,
                f'✅ Vacaciones de {vacacion.empleado.nombre_completo} '
                f'marcadas como {estado_display}.'
            )
            
            return redirect('rh:gestion_vacaciones')
    else:
        form = AprobarVacacionForm(instance=vacacion)
    
    context = {
        'vacacion': vacacion,
        'form': form,
        'page_title': f'Aprobar Vacaciones - {vacacion.empleado.nombre_completo}',
    }
    
    return render(request, 'rh/aprobar_vacacion.html', context)


@login_required
def detalle_vacacion(request, pk):
    """
    Vista para ver detalles de una vacación
    """
    vacacion = get_object_or_404(
        Vacacion.objects.select_related(
            'empleado',
            'empleado__departamento',
            'aprobado_por',
            'creado_por'
        ),
        pk=pk
    )
    
    context = {
        'vacacion': vacacion,
        'page_title': f'Detalle de Vacación - {vacacion.empleado.nombre_completo}',
    }
    
    return render(request, 'rh/detalle_vacacion.html', context)


@login_required
def historico_vacaciones_empleado(request, empleado_id):
    """
    Vista para ver el histórico de vacaciones de un empleado
    """
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    # Obtener histórico
    historico = HistoricoVacaciones.objects.filter(
        empleado=empleado
    ).order_by('-año')
    
    # Obtener todas las vacaciones del empleado
    vacaciones = Vacacion.objects.filter(
        empleado=empleado
    ).order_by('-fecha_inicio')
    
    # Calcular totales
    total_dias_correspondientes = historico.aggregate(
        total=Sum('dias_correspondientes')
    )['total'] or 0
    
    total_dias_tomados = historico.aggregate(
        total=Sum('dias_tomados')
    )['total'] or 0
    
    total_dias_pendientes = historico.aggregate(
        total=Sum('dias_pendientes')
    )['total'] or 0
    
    context = {
        'empleado': empleado,
        'historico': historico,
        'vacaciones': vacaciones,
        'total_dias_correspondientes': total_dias_correspondientes,
        'total_dias_tomados': total_dias_tomados,
        'total_dias_pendientes': total_dias_pendientes,
        'dias_disponibles_actual': empleado.dias_vacaciones_disponibles,
        'page_title': f'Histórico de Vacaciones - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/historico_vacaciones.html', context)


# ==============================================
# VISTAS PARA PRÉSTAMOS
# ==============================================

@login_required
def gestion_prestamos(request):
    """
    Vista principal de gestión de préstamos
    """
    # Filtros
    estado = request.GET.get('estado', '')
    tipo_prestamo = request.GET.get('tipo_prestamo', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    empleado_id = request.GET.get('empleado', '')
    
    # Query base
    prestamos = Prestamo.objects.all().select_related(
        'empleado', 'empleado__departamento', 'aprobado_por'
    ).order_by('-fecha_solicitud')
    
    # Aplicar filtros
    if estado:
        prestamos = prestamos.filter(estado=estado)
    
    if tipo_prestamo:
        prestamos = prestamos.filter(tipo_prestamo=tipo_prestamo)
    
    if fecha_inicio:
        prestamos = prestamos.filter(fecha_solicitud__gte=fecha_inicio)
    
    if fecha_fin:
        prestamos = prestamos.filter(fecha_solicitud__lte=fecha_fin)
    
    if empleado_id:
        prestamos = prestamos.filter(empleado_id=empleado_id)
    
    # Estadísticas
    total_prestamos = prestamos.count()
    prestamos_activos = prestamos.filter(estado='EN_CURSO').count()
    prestamos_pagados = prestamos.filter(estado='PAGADO').count()
    prestamos_morosos = prestamos.filter(estado='MOROSO').count()
    
    # Total de saldo pendiente
    total_saldo_pendiente = prestamos.filter(estado='EN_CURSO').aggregate(
        total=Sum('saldo_pendiente')
    )['total'] or Decimal('0')
    
    # Total de préstamos activos (para dashboard)
    total_prestamos_activos_valor = prestamos.filter(estado='EN_CURSO').aggregate(
        total=Sum('monto_total')
    )['total'] or Decimal('0')
    
    # Préstamos pendientes de pago (para dashboard)
    prestamos_pendientes_dashboard = prestamos.filter(
        estado='EN_CURSO'
    ).select_related('empleado').order_by('fecha_primer_pago')[:10]
    
    context = {
        'prestamos': prestamos,
        'total_prestamos': total_prestamos,
        'prestamos_activos': prestamos_activos,
        'prestamos_pagados': prestamos_pagados,
        'prestamos_morosos': prestamos_morosos,
        'total_saldo_pendiente': total_saldo_pendiente,
        'total_prestamos_activos_valor': total_prestamos_activos_valor,
        'prestamos_pendientes': prestamos_pendientes_dashboard,
        'estados_prestamo': Prestamo.ESTADO_CHOICES,
        'tipos_prestamo': Prestamo.TIPO_PRESTAMO_CHOICES,
        'empleados': Empleado.objects.filter(activo=True).order_by('apellido', 'nombre'),
        'filtros_aplicados': {
            'estado': estado,
            'tipo_prestamo': tipo_prestamo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'empleado_id': empleado_id,
        },
        'page_title': 'Gestión de Préstamos',
    }
    
    return render(request, 'rh/gestion_prestamos.html', context)


@login_required
def solicitar_prestamo(request, empleado_id=None):
    """
    Vista para solicitar préstamos con plazos en SEMANAS (1-52)
    """
    # Primero verificar si viene por GET parameter
    empleado_id_get = request.GET.get('empleado_id')
    if empleado_id_get:
        empleado_id = empleado_id_get
    
    if empleado_id:
        empleado = get_object_or_404(Empleado, pk=empleado_id, activo=True)
    else:
        empleado = None
    
    # Calcular límites
    limite_maximo = Decimal('0')
    pago_maximo_mensual = Decimal('0')
    salario_mensual = Decimal('0')
    
    if empleado:
        # Obtener salario actual del empleado
        salario_actual = empleado.salario_actual
        if salario_actual and salario_actual.sueldo_mensual:
            salario_mensual = salario_actual.sueldo_mensual
            # Límite: 3 meses de sueldo
            limite_maximo = salario_mensual * Decimal('3')
            # Pago máximo: 40% del sueldo mensual
            pago_maximo_mensual = salario_mensual * Decimal('0.4')
        else:
            # Si no tiene salario registrado, usar un valor por defecto
            salario_mensual = Decimal('5000')  # Valor por defecto
            limite_maximo = salario_mensual * Decimal('3')
            pago_maximo_mensual = salario_mensual * Decimal('0.4')
    
    if request.method == 'POST':
        form = PrestamoForm(request.POST, request.FILES, empleado=empleado)
        
        if form.is_valid():
            prestamo = form.save(commit=False)
            
            if empleado:
                prestamo.empleado = empleado
            else:
                # Permitir seleccionar empleado si no se especificó
                prestamo.empleado_id = request.POST.get('empleado')
            
            prestamo.creado_por = request.user
            prestamo.saldo_pendiente = prestamo.monto_total
            
            # Validar límite máximo de monto
            if prestamo.monto_total > limite_maximo:
                form.add_error('monto_total', 
                    f'El monto máximo permitido es ${limite_maximo:,.2f} (3 meses de sueldo)')
                return render(request, 'rh/solicitar_prestamo.html', {
                    'form': form,
                    'empleado': empleado,
                    'empleados': Empleado.objects.filter(activo=True).order_by('apellido', 'nombre'),
                    'salario_actual': salario_actual,
                    'limite_maximo': limite_maximo,
                    'pago_maximo_mensual': pago_maximo_mensual,
                    'prestamos_activos': empleado.prestamos.filter(estado__in=['APROBADO', 'EN_CURSO']).count() if empleado else 0,
                    'page_title': 'Solicitar Préstamo',
                })
            
            # Obtener información de cálculo del formulario
            info_calculo = form.get_info_calculo()
            if info_calculo:
                pago_semanal = info_calculo.get('pago_semanal', 0)
                pago_mensual_aproximado = info_calculo.get('pago_mensual_aproximado', 0)
                
                # Validar que el pago equivalente mensual no exceda el 40% del sueldo
                if pago_mensual_aproximado > pago_maximo_mensual:
                    # Calcular semanas mínimas sugeridas
                    semanas_minimas = 0
                    if pago_semanal > 0:
                        # Calcular cuántas semanas se necesitan para que el pago mensual sea <= 40% del sueldo
                        pago_semanal_maximo = pago_maximo_mensual / Decimal('4.33')
                        if pago_semanal_maximo > 0:
                            semanas_minimas = int((prestamo.monto_total / pago_semanal_maximo) + 1)
                            # Limitar a máximo 52 semanas
                            semanas_minimas = min(semanas_minimas, 52)
                    
                    form.add_error('plazo_semanas',
                        f'El pago semanal sería de ${pago_semanal:,.2f} '
                        f'(equivalente a ${pago_mensual_aproximado:,.2f} mensual). '
                        f'Esto excede el 40% del sueldo (${pago_maximo_mensual:,.2f}). '
                        f'Recomendación: aumente el plazo a al menos {semanas_minimas} semanas.')
                    
                    return render(request, 'rh/solicitar_prestamo.html', {
                        'form': form,
                        'empleado': empleado,
                        'empleados': Empleado.objects.filter(activo=True).order_by('apellido', 'nombre'),
                        'salario_actual': salario_actual,
                        'limite_maximo': limite_maximo,
                        'pago_maximo_mensual': pago_maximo_mensual,
                        'prestamos_activos': empleado.prestamos.filter(estado__in=['APROBADO', 'EN_CURSO']).count() if empleado else 0,
                        'info_calculo': info_calculo,
                        'page_title': 'Solicitar Préstamo',
                    })
            
            # Si está aprobado automáticamente
            if 'aprobar_ahora' in request.POST and request.POST['aprobar_ahora'] == 'true':
                prestamo.estado = 'APROBADO'
                prestamo.aprobado_por = request.user
                prestamo.fecha_aprobacion = date.today()
            
            prestamo.save()
            
            # Crear plan de pagos si está aprobado
            if prestamo.estado == 'APROBADO':
                crear_plan_pagos(prestamo)
            
            # Mostrar información del préstamo creado
            plazo_meses = prestamo.get_plazo_equivalente_meses()
            messages.success(
                request,
                f'✅ Solicitud de préstamo registrada exitosamente para '
                f'{prestamo.empleado.nombre_completo}. '
                f'Plazo: {prestamo.plazo_semanas} semanas '
                f'({plazo_meses:.1f} meses aproximadamente).'
            )
            
            return redirect('rh:gestion_prestamos')
    else:
        form = PrestamoForm(empleado=empleado)
    
    # Obtener información de cálculo inicial si existe
    info_calculo = None
    if form.is_bound and not form.errors:
        info_calculo = form.get_info_calculo()
    
    context = {
        'form': form,
        'empleado': empleado,
        'empleados': Empleado.objects.filter(activo=True).order_by('apellido', 'nombre'),
        'page_title': 'Solicitar Préstamo',
        'salario_actual': salario_actual if empleado else None,
        'limite_maximo': limite_maximo,
        'pago_maximo_mensual': pago_maximo_mensual,
        'prestamos_activos': empleado.prestamos.filter(estado__in=['APROBADO', 'EN_CURSO']).count() if empleado else 0,
        'info_calculo': info_calculo,
        'semanas_por_mes': 4.33,  # Para mostrar en template
        'plazo_minimo_semanas': 1,
        'plazo_maximo_semanas': 52,
    }
    
    return render(request, 'rh/solicitar_prestamo.html', context)


@csrf_exempt
@require_GET
def api_empleado_info(request, empleado_id):
    """
    API para obtener información del empleado para préstamos
    """
    try:
        # Verificar autenticación
        if not request.user.is_authenticated:
            return JsonResponse({
                'error': 'No autenticado',
                'login_url': '/accounts/login/'
            }, status=401)
        
        # Buscar empleado
        empleado = Empleado.objects.filter(pk=empleado_id, activo=True).first()
        if not empleado:
            return JsonResponse({
                'error': f'Empleado {empleado_id} no encontrado o inactivo',
                'empleado_id': empleado_id
            }, status=404)
        
        # Obtener salario
        salario_actual = empleado.salario_actual
        if salario_actual and salario_actual.sueldo_mensual:
            sueldo_mensual = salario_actual.sueldo_mensual
        else:
            sueldo_mensual = Decimal('5000')  # Valor por defecto
        
        # Calcular límites
        limite_maximo = sueldo_mensual * Decimal('3')
        pago_maximo_mensual = sueldo_mensual * Decimal('0.4')
        
        # Contar préstamos activos
        prestamos_activos = empleado.prestamos.filter(
            estado__in=['APROBADO', 'EN_CURSO']
        ).count()
        
        # Preparar respuesta
        data = {
            'success': True,
            'id': empleado.id,
            'nombre_completo': empleado.nombre_completo,
            'puesto': empleado.puesto.nombre if empleado.puesto else 'Sin puesto',
            'departamento': empleado.departamento.nombre if empleado.departamento else 'Sin departamento',
            'sueldo_mensual': float(sueldo_mensual),
            'limite_maximo': float(limite_maximo),
            'pago_maximo_mensual': float(pago_maximo_mensual),
            'prestamos_activos': prestamos_activos,
            'timestamp': timezone.now().isoformat(),
        }
        
        # Crear respuesta con headers CORS
        response = JsonResponse(data)
        response['Access-Control-Allow-Origin'] = '*'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Error interno del servidor'
        }, status=500)

def crear_plan_pagos(prestamo):
    """
    Crea el plan de pagos SEMANALES para un préstamo aprobado
    """
    # Eliminar pagos existentes si los hay
    prestamo.pagos.all().delete()
    
    # Obtener pago semanal del préstamo
    pago_semanal = prestamo.pago_semanal
    fecha_pago = prestamo.fecha_primer_pago
    
    # Crear un pago por cada semana
    for i in range(prestamo.plazo_semanas):
        PagoPrestamo.objects.create(
            prestamo=prestamo,
            numero_pago=i + 1,
            fecha_pago=fecha_pago,
            monto_programado=pago_semanal,
            estado='PENDIENTE'
        )
        
        # Avanzar UNA SEMANA
        fecha_pago += timedelta(weeks=1)


@login_required
def aprobar_prestamo(request, pk):
    """
    Vista para aprobar o rechazar un préstamo
    """
    prestamo = get_object_or_404(Prestamo.objects.select_related('empleado'), pk=pk)
    
    if request.method == 'POST':
        estado = request.POST.get('estado_final')
        observaciones = request.POST.get('observaciones', '')
        
        if estado in ['APROBADO', 'RECHAZADO']:
            prestamo.estado = estado
            prestamo.aprobado_por = request.user
            prestamo.fecha_aprobacion = date.today()
            prestamo.observaciones = observaciones
            
            if estado == 'APROBADO':
                # Cambiar a "En curso" después de aprobar
                prestamo.estado = 'EN_CURSO'
                # Crear plan de pagos SEMANALES
                crear_plan_pagos(prestamo)
            
            prestamo.save()
            
            messages.success(
                request,
                f'✅ Préstamo de {prestamo.empleado.nombre_completo} '
                f'marcado como {prestamo.get_estado_display()}. '
                f'Plazo: {prestamo.plazo_semanas} semanas.'
            )
            
            return redirect('rh:gestion_prestamos')
    
    # Calcular información para mostrar
    pago_semanal = prestamo.pago_semanal
    pago_mensual_aprox = prestamo.pago_mensual_aproximado if pago_semanal else Decimal('0')
    
    context = {
        'prestamo': prestamo,
        'pago_semanal': pago_semanal,
        'pago_mensual_aprox': pago_mensual_aprox,
        'plazo_meses_aproximado': prestamo.get_plazo_equivalente_meses(),
        'page_title': f'Aprobar Préstamo - {prestamo.empleado.nombre_completo}',
    }
    
    return render(request, 'rh/aprobar_prestamo.html', context)


@login_required
def detalle_prestamo(request, pk):
    """
    Vista para ver detalles de un préstamo
    """
    prestamo = get_object_or_404(
        Prestamo.objects.select_related(
            'empleado',
            'empleado__departamento',
            'aprobado_por',
            'creado_por'
        ),
        pk=pk
    )
    
    # Obtener pagos ordenados
    pagos = prestamo.pagos.all().order_by('numero_pago')
    
    context = {
        'prestamo': prestamo,
        'pagos': pagos,
        'page_title': f'Detalle de Préstamo - {prestamo.empleado.nombre_completo}',
    }
    
    return render(request, 'rh/detalle_prestamo.html', context)


@login_required
def registrar_pago(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)
    
    # Obtener pagos pendientes (asumiendo que tienes este método)
    from datetime import date, timedelta
    pagos_pendientes = []
    if prestamo.estado == 'ACTIVO' and prestamo.saldo_pendiente > 0:
        # Lógica para generar pagos pendientes
        pass
    
    if request.method == 'POST':
        form = PagoPrestamoForm(request.POST, request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.prestamo = prestamo
            
            # CORRECCIÓN: Asignar monto_programado (usando el pago semanal del préstamo)
            pago.monto_programado = prestamo.pago_semanal  # <--- CAMBIO IMPORTANTE
            
            # Asignar número de pago automáticamente
            ultimo_pago = prestamo.pagos.order_by('-numero_pago').first()
            pago.numero_pago = (ultimo_pago.numero_pago + 1) if ultimo_pago else 1
            
            pago.save()
            
            # Actualizar saldo del préstamo
            prestamo.monto_pagado += pago.monto_pagado
            prestamo.save()
            
            messages.success(request, '✅ Pago registrado correctamente')
            return redirect('rh:detalle_prestamo', pk=prestamo.id)
        else:
            # Mostrar errores específicos
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            messages.error(
                request, 
                f'❌ Error al registrar el pago: ' + ' | '.join(error_messages)
            )
    else:
        # CORRECCIÓN: Usar 'fecha_pago' en lugar de 'fecha_pago_real'
        form = PagoPrestamoForm(initial={'fecha_pago': date.today()})
    
    return render(request, 'rh/registrar_pago.html', {
        'form': form,
        'prestamo': prestamo,
        'pagos_pendientes': pagos_pendientes
    })

@login_required
def historico_prestamos_empleado(request, empleado_id):
    """
    Vista para ver el histórico de préstamos de un empleado
    """
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    # Obtener todos los préstamos del empleado
    prestamos = Prestamo.objects.filter(
        empleado=empleado
    ).order_by('-fecha_solicitud')
    
    # Calcular estadísticas
    total_prestamos = prestamos.count()
    total_monto_prestado = prestamos.aggregate(
        total=Sum('monto_total')
    )['total'] or Decimal('0')
    
    total_monto_pagado = prestamos.aggregate(
        total=Sum('monto_pagado')
    )['total'] or Decimal('0')
    
    prestamos_activos = prestamos.filter(estado='EN_CURSO')
    saldo_pendiente = prestamos_activos.aggregate(
        total=Sum('saldo_pendiente')
    )['total'] or Decimal('0')
    
    context = {
        'empleado': empleado,
        'prestamos': prestamos,
        'total_prestamos': total_prestamos,
        'total_monto_prestado': total_monto_prestado,
        'total_monto_pagado': total_monto_pagado,
        'saldo_pendiente': saldo_pendiente,
        'prestamos_activos': prestamos_activos.count(),
        'page_title': f'Histórico de Préstamos - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/historico_prestamos.html', context)


# ==============================================
# VISTAS PARA REPORTES Y DASHBOARD
# ==============================================

@login_required
def reporte_vacaciones_excel(request):
    """
    Exportar reporte de vacaciones a Excel
    """
    # Filtros
    estado = request.GET.get('estado', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    vacaciones = Vacacion.objects.all().select_related(
        'empleado', 'empleado__departamento'
    ).order_by('-fecha_solicitud')
    
    if estado:
        vacaciones = vacaciones.filter(estado=estado)
    
    if fecha_inicio:
        vacaciones = vacaciones.filter(fecha_inicio__gte=fecha_inicio)
    
    if fecha_fin:
        vacaciones = vacaciones.filter(fecha_inicio__lte=fecha_fin)
    
    # Crear DataFrame
    data = []
    for vacacion in vacaciones:
        data.append({
            'ID': vacacion.id,
            'Empleado': vacacion.empleado.nombre_completo,
            'Departamento': vacacion.empleado.departamento.nombre if vacacion.empleado.departamento else '',
            'Tipo': vacacion.get_tipo_vacacion_display(),
            'Fecha Inicio': vacacion.fecha_inicio.strftime('%d/%m/%Y'),
            'Fecha Fin': vacacion.fecha_fin.strftime('%d/%m/%Y'),
            'Días Solicitados': vacacion.dias_solicitados,
            'Días Reales': vacacion.dias_reales or vacacion.dias_laborables,
            'Periodo': vacacion.periodo_correspondiente,
            'Estado': vacacion.get_estado_display(),
            'Fecha Solicitud': vacacion.fecha_solicitud.strftime('%d/%m/%Y'),
            'Aprobado por': vacacion.aprobado_por.get_full_name() if vacacion.aprobado_por else '',
            'Observaciones': vacacion.observaciones[:100] if vacacion.observaciones else '',
        })
    
    df = pd.DataFrame(data)
    
    # Crear Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Vacaciones"
    
    # Agregar datos
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2ecc71", end_color="27ae60", fill_type="solid")
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_vacaciones.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def reporte_prestamos_excel(request):
    """
    Exportar reporte de préstamos a Excel
    """
    # Filtros
    estado = request.GET.get('estado', '')
    tipo_prestamo = request.GET.get('tipo_prestamo', '')
    
    prestamos = Prestamo.objects.all().select_related(
        'empleado', 'empleado__departamento'
    ).order_by('-fecha_solicitud')
    
    if estado:
        prestamos = prestamos.filter(estado=estado)
    
    if tipo_prestamo:
        prestamos = prestamos.filter(tipo_prestamo=tipo_prestamo)
    
    # Crear DataFrame
    data = []
    for prestamo in prestamos:
        data.append({
            'ID': prestamo.id,
            'Empleado': prestamo.empleado.nombre_completo,
            'Departamento': prestamo.empleado.departamento.nombre if prestamo.empleado.departamento else '',
            'Tipo': prestamo.get_tipo_prestamo_display(),
            'Monto Total': float(prestamo.monto_total),
            'Monto Pagado': float(prestamo.monto_pagado),
            'Saldo Pendiente': float(prestamo.saldo_pendiente),
            'Tasa Interés': f"{float(prestamo.tasa_interes)}%",
            'Plazo (meses)': prestamo.plazo_meses,
            'Fecha Solicitud': prestamo.fecha_solicitud.strftime('%d/%m/%Y'),
            'Fecha Aprobación': prestamo.fecha_aprobacion.strftime('%d/%m/%Y') if prestamo.fecha_aprobacion else '',
            'Estado': prestamo.get_estado_display(),
            'Concepto': prestamo.concepto[:100],
            'Pago Mensual': float(prestamo.pago_mensual),
            'Progreso Pago': f"{prestamo.progreso_pago:.1f}%",
        })
    
    df = pd.DataFrame(data)
    
    # Crear Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Préstamos"
    
    # Agregar datos
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9b59b6", end_color="8e44ad", fill_type="solid")
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_prestamos.xlsx"'
    workbook.save(response)
    
    return response

# En RH/views.py - agregar nuevas vistas
@login_required
def papelera_empleados(request):
    """
    Vista para ver empleados eliminados (papelera)
    """
    # Obtener empleados eliminados
    empleados_eliminados = Empleado.objects.eliminados().select_related(
        'puesto', 'departamento', 'eliminado_por'
    ).order_by('-fecha_eliminacion')
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    if fecha_inicio:
        empleados_eliminados = empleados_eliminados.filter(
            fecha_eliminacion__gte=fecha_inicio
        )
    
    if fecha_fin:
        empleados_eliminados = empleados_eliminados.filter(
            fecha_eliminacion__lte=fecha_fin
        )
    
    context = {
        'empleados_eliminados': empleados_eliminados,
        'total_eliminados': empleados_eliminados.count(),
        'page_title': 'Papelera de Reciclaje - Empleados',
    }
    
    return render(request, 'rh/papelera_empleados.html', context)


@login_required
def restaurar_empleado(request, pk):
    """
    Restaurar empleado desde la papelera
    """
    empleado = get_object_or_404(Empleado, pk=pk, eliminado=True)
    
    if request.method == 'POST':
        empleado.restaurar()
        
        messages.success(
            request,
            f'✅ {empleado.nombre_completo} ha sido restaurado desde la papelera.'
        )
        
        return redirect('rh:lista_empleados')
    
    context = {
        'empleado': empleado,
        'page_title': f'Restaurar empleado - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/empleado_restaurar_confirm.html', context)


@login_required
def eliminar_permanentemente(request, pk):
    """
    Eliminar permanentemente un empleado (Solo para administradores)
    """
    empleado = get_object_or_404(Empleado, pk=pk, eliminado=True)
    
    if not request.user.is_superuser:
        messages.error(request, '❌ Solo los administradores pueden eliminar permanentemente.')
        return redirect('rh:papelera_empleados')
    
    if request.method == 'POST':
        # Eliminación permanente
        nombre_empleado = empleado.nombre_completo
        empleado.eliminar_permanentemente()
        
        messages.success(
            request,
            f'⚠️ {nombre_empleado} ha sido eliminado permanentemente del sistema.'
        )
        
        return redirect('rh:papelera_empleados')
    
    context = {
        'empleado': empleado,
        'page_title': f'Eliminar permanentemente - {empleado.nombre_completo}',
    }
    
    return render(request, 'rh/empleado_eliminar_permanentemente.html', context)


@login_required
def vaciar_papelera(request):
    """
    Vaciar toda la papelera (Solo para administradores)
    """
    if not request.user.is_superuser:
        messages.error(request, '❌ Solo los administradores pueden vaciar la papelera.')
        return redirect('rh:papelera_empleados')
    
    if request.method == 'POST':
        empleados_eliminados = Empleado.objects.eliminados()
        total = empleados_eliminados.count()
        
        for empleado in empleados_eliminados:
            empleado.eliminar_permanentemente()
        
        messages.warning(
            request,
            f'⚠️ Se han eliminado permanentemente {total} empleados de la papelera.'
        )
        
        return redirect('rh:papelera_empleados')
    
    total_eliminados = Empleado.objects.eliminados().count()
    
    context = {
        'total_eliminados': total_eliminados,
        'page_title': 'Vaciar papelera de reciclaje',
    }
    
    return render(request, 'rh/vaciar_papelera_confirm.html', context)

# Añadir esta función a tu views.py, después de las funciones existentes

@login_required
def exportar_rh_completo(request):
    """
    Exporta TODOS los datos del módulo RH a un archivo Excel con múltiples pestañas
    Incluye: Empleados activos, en papelera, bajas históricas y todos los registros
    """
    try:
        # Crear un nuevo workbook
        wb = openpyxl.Workbook()
        
        # ========== PESTAÑA 1: EMPLEADOS ACTIVOS (DIRECTORIO) ==========
        ws_activos = wb.active
        ws_activos.title = "Empleados_Activos"
        
        headers_empleados = [
            'ID', 'Número Empleado', 'Nombre', 'Apellido', 'Puesto', 'Departamento',
            'Fecha Contratación', 'Activo', 'Email', 'Fecha Nacimiento',
            'CURP', 'RFC', 'NSS', 'Empresa', 'Dirección', 'Colonia', 'Código Postal',
            'Ciudad', 'Estado', 'País', 'Teléfono Personal', 'Estado Civil',
            'Nacionalidad', 'Supervisor', 'Nombre Cónyuge', 'Teléfono Cónyuge',
            'Banco', 'CLABE Interbancaria', 'Número Cuenta', 'Número Tarjeta',
            'Motivo Inactivación', 'Fecha Inactivación',
            'Nombre Referencia 1', 'Teléfono Referencia 1', 'Relación Referencia 1',
            'Nombre Referencia 2', 'Teléfono Referencia 2', 'Relación Referencia 2',
            'Eliminado', 'Fecha Eliminación', 'Motivo Eliminación', 'Estatus'
        ]
        
        ws_activos.append(headers_empleados)
        
        # Empleados activos (no eliminados, sin baja registrada)
        empleados_activos = Empleado.objects.filter(
            eliminado=False
        ).exclude(
            bajas__isnull=False
        ).select_related(
            'puesto', 'departamento', 'supervisor', 'motivo_inactivacion'
        ).distinct()
        
        print(f"📊 Exportando {empleados_activos.count()} empleados activos")
        
        for emp in empleados_activos:
            row = [
                export_origen_id(emp),
                emp.numero_empleado,
                emp.nombre,
                emp.apellido,
                emp.puesto.nombre if emp.puesto else '',
                emp.departamento.nombre if emp.departamento else '',
                emp.fecha_contratacion.strftime('%Y-%m-%d') if emp.fecha_contratacion else '',
                'Sí' if emp.activo else 'No',
                emp.email,
                emp.fecha_nacimiento.strftime('%Y-%m-%d') if emp.fecha_nacimiento else '',
                emp.curp,
                emp.rfc,
                emp.nss,
                emp.get_empresa_display() if emp.empresa else '',
                emp.direccion,
                emp.colonia,
                emp.codigo_postal,
                emp.ciudad,
                emp.estado,
                emp.pais,
                emp.telefono_personal,
                emp.get_estado_civil_display() if emp.estado_civil else '',
                emp.nacionalidad,
                emp.supervisor.nombre_completo if emp.supervisor else '',
                emp.nombre_conyuge,
                emp.telefono_conyuge,
                emp.banco,
                emp.clabe_interbancaria,
                emp.numero_cuenta,
                emp.numero_tarjeta,
                emp.motivo_inactivacion.motivo if emp.motivo_inactivacion else '',
                emp.fecha_inactivacion.strftime('%Y-%m-%d') if emp.fecha_inactivacion else '',
                emp.nombre_referencia_1,
                emp.telefono_referencia_1,
                emp.relacion_referencia_1,
                emp.nombre_referencia_2,
                emp.telefono_referencia_2,
                emp.relacion_referencia_2,
                'No',
                '',
                '',
                'ACTIVO'
            ]
            ws_activos.append(row)
        
        # ========== PESTAÑA 2: EMPLEADOS EN PAPELERA ==========
        ws_papelera = wb.create_sheet("Empleados_Papelera")
        ws_papelera.append(headers_empleados)
        
        empleados_papelera = Empleado.objects.filter(eliminado=True).select_related(
            'puesto', 'departamento', 'supervisor', 'motivo_inactivacion'
        )
        
        print(f"🗑️  Exportando {empleados_papelera.count()} empleados en papelera")
        
        for emp in empleados_papelera:
            row = [
                export_origen_id(emp),
                emp.numero_empleado,
                emp.nombre,
                emp.apellido,
                emp.puesto.nombre if emp.puesto else '',
                emp.departamento.nombre if emp.departamento else '',
                emp.fecha_contratacion.strftime('%Y-%m-%d') if emp.fecha_contratacion else '',
                'Sí' if emp.activo else 'No',
                emp.email,
                emp.fecha_nacimiento.strftime('%Y-%m-%d') if emp.fecha_nacimiento else '',
                emp.curp,
                emp.rfc,
                emp.nss,
                emp.get_empresa_display() if emp.empresa else '',
                emp.direccion,
                emp.colonia,
                emp.codigo_postal,
                emp.ciudad,
                emp.estado,
                emp.pais,
                emp.telefono_personal,
                emp.get_estado_civil_display() if emp.estado_civil else '',
                emp.nacionalidad,
                emp.supervisor.nombre_completo if emp.supervisor else '',
                emp.nombre_conyuge,
                emp.telefono_conyuge,
                emp.banco,
                emp.clabe_interbancaria,
                emp.numero_cuenta,
                emp.numero_tarjeta,
                emp.motivo_inactivacion.motivo if emp.motivo_inactivacion else '',
                emp.fecha_inactivacion.strftime('%Y-%m-%d') if emp.fecha_inactivacion else '',
                emp.nombre_referencia_1,
                emp.telefono_referencia_1,
                emp.relacion_referencia_1,
                emp.nombre_referencia_2,
                emp.telefono_referencia_2,
                emp.relacion_referencia_2,
                'Sí',
                emp.fecha_eliminacion.strftime('%Y-%m-%d %H:%M:%S') if emp.fecha_eliminacion else '',
                emp.motivo_eliminacion or '',
                'PAPELERA'
            ]
            ws_papelera.append(row)
        
        # ========== PESTAÑA 3: BAJAS HISTÓRICAS ==========
        ws_bajas_historicas = wb.create_sheet("Bajas_Historicas")
        ws_bajas_historicas.append(headers_empleados)
        
        empleados_con_baja = Empleado.objects.filter(
            bajas__isnull=False
        ).select_related(
            'puesto', 'departamento', 'supervisor', 'motivo_inactivacion'
        ).distinct()
        
        print(f"📉 Exportando {empleados_con_baja.count()} bajas históricas")
        
        for emp in empleados_con_baja:
            row = [
                export_origen_id(emp),
                emp.numero_empleado,
                emp.nombre,
                emp.apellido,
                emp.puesto.nombre if emp.puesto else '',
                emp.departamento.nombre if emp.departamento else '',
                emp.fecha_contratacion.strftime('%Y-%m-%d') if emp.fecha_contratacion else '',
                'No',
                emp.email,
                emp.fecha_nacimiento.strftime('%Y-%m-%d') if emp.fecha_nacimiento else '',
                emp.curp,
                emp.rfc,
                emp.nss,
                emp.get_empresa_display() if emp.empresa else '',
                emp.direccion,
                emp.colonia,
                emp.codigo_postal,
                emp.ciudad,
                emp.estado,
                emp.pais,
                emp.telefono_personal,
                emp.get_estado_civil_display() if emp.estado_civil else '',
                emp.nacionalidad,
                emp.supervisor.nombre_completo if emp.supervisor else '',
                emp.nombre_conyuge,
                emp.telefono_conyuge,
                emp.banco,
                emp.clabe_interbancaria,
                emp.numero_cuenta,
                emp.numero_tarjeta,
                emp.motivo_inactivacion.motivo if emp.motivo_inactivacion else '',
                emp.fecha_inactivacion.strftime('%Y-%m-%d') if emp.fecha_inactivacion else '',
                emp.nombre_referencia_1,
                emp.telefono_referencia_1,
                emp.relacion_referencia_1,
                emp.nombre_referencia_2,
                emp.telefono_referencia_2,
                emp.relacion_referencia_2,
                'No',
                '',
                '',
                'BAJA_HISTORICA'
            ]
            ws_bajas_historicas.append(row)
        
        # ========== PESTAÑA 4: REGISTROS DE BAJAS (DETALLES) ==========
        ws_bajas_detalles = wb.create_sheet("Registros_Bajas")
        
        headers_bajas_detalles = [
            'ID', 'Empleado ID', 'Empleado Nombre', 'Motivo Principal',
            'Motivo Secundario', 'Motivo Detalle', 'Fecha Baja',
            'Comentario Baja', 'Es Reconstratable', 'Motivo Reconstratable',
            'Fecha Posible Reconstratación', 'Recontratado',
            'Fecha Reconstratación', 'Fue Conciliación Arbitraje',
            'Fecha Conciliación Arbitraje', 'Fecha Registro', 'Registrado Por'
        ]
        ws_bajas_detalles.append(headers_bajas_detalles)
        
        bajas = BajaEmpleado.objects.all().select_related(
            'empleado', 'motivo_principal', 'motivo_secundario', 
            'motivo_detalle', 'registrado_por'
        ).order_by('-fecha_baja')
        
        for baja in bajas:
            row = [
                baja.id,
                export_origen_id(baja.empleado),
                baja.empleado.nombre_completo,
                baja.motivo_principal.nombre if baja.motivo_principal else '',
                baja.motivo_secundario.nombre if baja.motivo_secundario else '',
                baja.motivo_detalle.nombre if baja.motivo_detalle else '',
                baja.fecha_baja.strftime('%Y-%m-%d') if baja.fecha_baja else '',
                baja.comentario_baja[:200] + '...' if baja.comentario_baja and len(baja.comentario_baja) > 200 else baja.comentario_baja or '',
                'Sí' if baja.es_recontratable else 'No',
                baja.motivo_recontratable or '',
                baja.fecha_posible_recontratacion.strftime('%Y-%m-%d') if baja.fecha_posible_recontratacion else '',
                'Sí' if baja.recontratado else 'No',
                baja.fecha_recontratacion.strftime('%Y-%m-%d') if baja.fecha_recontratacion else '',
                'Sí' if baja.fue_conciliacion_arbitraje else 'No',
                baja.fecha_conciliacion_arbitraje.strftime('%Y-%m-%d') if baja.fecha_conciliacion_arbitraje else '',
                baja.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if baja.fecha_registro else '',
                baja.registrado_por.get_full_name() if baja.registrado_por else ''
            ]
            ws_bajas_detalles.append(row)
        
        # ========== PESTAÑA 5: SALARIOS ==========
        ws_salarios = wb.create_sheet("Salarios")
        headers_salarios = ['ID', 'Empleado ID', 'Empleado Nombre', 'Sueldo Diario', 
                           'Fecha Efectiva', 'Observaciones']
        ws_salarios.append(headers_salarios)
        
        salarios = Salario.objects.all().select_related('empleado')
        for sal in salarios:
            row = [
                sal.id,
                export_origen_id(sal.empleado),
                sal.empleado.nombre_completo,
                float(sal.sueldo_diario) if sal.sueldo_diario else 0,
                sal.fecha_efectiva.strftime('%Y-%m-%d'),
                sal.observaciones or ''
            ]
            ws_salarios.append(row)
        
        # ========== PESTAÑA 6: CONTRATOS ==========
        ws_contratos = wb.create_sheet("Contratos")
        headers_contratos = ['ID', 'Empleado ID', 'Empleado Nombre', 'Tipo Contrato',
                            'Fecha Inicio', 'Fecha Fin', 'Comentarios']
        ws_contratos.append(headers_contratos)
        
        contratos = Contrato.objects.all().select_related('empleado')
        for con in contratos:
            row = [
                con.id,
                export_origen_id(con.empleado),
                con.empleado.nombre_completo,
                con.get_tipo_contrato_display(),
                con.fecha_inicio.strftime('%Y-%m-%d'),
                con.fecha_fin.strftime('%Y-%m-%d') if con.fecha_fin else '',
                con.comentarios or ''
            ]
            ws_contratos.append(row)
        
        # ========== PESTAÑA 7: HISTORIAL LABORAL ==========
        ws_historial = wb.create_sheet("Historial_Laboral")
        headers_historial = ['ID', 'Empleado ID', 'Empleado Nombre', 'Tipo Evento',
                            'Fecha Inicio', 'Fecha Fin', 'Puesto', 'Departamento',
                            'Motivo Salida', 'Descripción', 'Estatus', 'Reemplazo ID',
                            'Reemplazo Nombre', 'Fecha Reemplazo']
        ws_historial.append(headers_historial)
        
        historial = HistorialLaboral.objects.all().select_related('empleado', 'reemplazo')
        for hist in historial:
            row = [
                hist.id,
                export_origen_id(hist.empleado),
                hist.empleado.nombre_completo,
                hist.get_tipo_evento_display(),
                hist.fecha_inicio.strftime('%Y-%m-%d'),
                hist.fecha_fin.strftime('%Y-%m-%d') if hist.fecha_fin else '',
                hist.puesto or '',
                hist.departamento or '',
                hist.get_motivo_salida_display() if hist.motivo_salida else '',
                hist.descripcion or '',
                hist.get_estatus_display(),
                export_origen_id(hist.reemplazo) if hist.reemplazo else '',
                hist.reemplazo.nombre_completo if hist.reemplazo else '',
                hist.fecha_reemplazo.strftime('%Y-%m-%d') if hist.fecha_reemplazo else ''
            ]
            ws_historial.append(row)
        
        # ========== PESTAÑA 8: VACACIONES ==========
        ws_vacaciones = wb.create_sheet("Vacaciones")
        headers_vacaciones = ['ID', 'Empleado ID', 'Empleado Nombre', 'Tipo Vacación',
                             'Fecha Solicitud', 'Fecha Inicio', 'Fecha Fin',
                             'Días Solicitados', 'Días Reales', 'Periodo Correspondiente',
                             'Estado', 'Aprobado Por', 'Fecha Aprobación', 'Observaciones']
        ws_vacaciones.append(headers_vacaciones)
        
        vacaciones = Vacacion.objects.all().select_related('empleado', 'aprobado_por')
        for vac in vacaciones:
            row = [
                vac.id,
                export_origen_id(vac.empleado),
                vac.empleado.nombre_completo,
                vac.get_tipo_vacacion_display(),
                vac.fecha_solicitud.strftime('%Y-%m-%d'),
                vac.fecha_inicio.strftime('%Y-%m-%d'),
                vac.fecha_fin.strftime('%Y-%m-%d'),
                vac.dias_solicitados,
                vac.dias_reales or vac.dias_laborables,
                vac.periodo_correspondiente,
                vac.get_estado_display(),
                vac.aprobado_por.get_full_name() if vac.aprobado_por else '',
                vac.fecha_aprobacion.strftime('%Y-%m-%d') if vac.fecha_aprobacion else '',
                vac.observaciones or ''
            ]
            ws_vacaciones.append(row)
        
        # ========== PESTAÑA 9: PRÉSTAMOS ==========
        ws_prestamos = wb.create_sheet("Préstamos")
        headers_prestamos = ['ID', 'Empleado ID', 'Empleado Nombre', 'Tipo Préstamo',
                            'Fecha Solicitud', 'Fecha Aprobación', 'Monto Total',
                            'Monto Pagado', 'Saldo Pendiente', 'Tasa Interés',
                            'Plazo Semanal', 'Fecha Primer Pago', 'Estado', 'Concepto',
                            'Aprobado Por', 'Observaciones']
        ws_prestamos.append(headers_prestamos)
        
        prestamos = Prestamo.objects.all().select_related('empleado', 'aprobado_por')
        for pre in prestamos:
            row = [
                pre.id,
                export_origen_id(pre.empleado),
                pre.empleado.nombre_completo,
                pre.get_tipo_prestamo_display(),
                pre.fecha_solicitud.strftime('%Y-%m-%d'),
                pre.fecha_aprobacion.strftime('%Y-%m-%d') if pre.fecha_aprobacion else '',
                float(pre.monto_total) if pre.monto_total else 0,
                float(pre.monto_pagado) if pre.monto_pagado else 0,
                float(pre.saldo_pendiente) if pre.saldo_pendiente else 0,
                float(pre.tasa_interes) if pre.tasa_interes else 0,
                pre.plazo_semanas,
                pre.fecha_primer_pago.strftime('%Y-%m-%d'),
                pre.get_estado_display(),
                pre.concepto,
                pre.aprobado_por.get_full_name() if pre.aprobado_por else '',
                pre.observaciones or ''
            ]
            ws_prestamos.append(row)
        
        # ========== PESTAÑA 10: PAGOS PRÉSTAMOS ==========
        ws_pagos = wb.create_sheet("Pagos_Préstamos")
        headers_pagos = ['ID', 'Préstamo ID', 'Empleado ID', 'Empleado Nombre',
                        'Número Pago', 'Fecha Pago Programada', 'Fecha Pago Real',
                        'Monto Programado', 'Monto Pagado', 'Estado', 'Observaciones']
        ws_pagos.append(headers_pagos)
        
        pagos = PagoPrestamo.objects.all().select_related('prestamo', 'prestamo__empleado')
        for pago in pagos:
            row = [
                pago.id,
                pago.prestamo.id,
                export_origen_id(pago.prestamo.empleado),
                pago.prestamo.empleado.nombre_completo,
                pago.numero_pago,
                pago.fecha_pago.strftime('%Y-%m-%d'),
                pago.fecha_pago_real.strftime('%Y-%m-%d') if pago.fecha_pago_real else '',
                float(pago.monto_programado) if pago.monto_programado else 0,
                float(pago.monto_pagado) if pago.monto_pagado else 0,
                pago.get_estado_display(),
                pago.observaciones or ''
            ]
            ws_pagos.append(row)
        
        # ========== PESTAÑA 11: DOCUMENTOS OPERADOR ==========
        ws_docs = wb.create_sheet("Documentos_Operador")
        headers_docs = ['ID', 'Empleado ID', 'Empleado Nombre', 'Tipo Documento',
                       'Número Documento', 'Fecha Expedición', 'Fecha Vencimiento',
                       'Observaciones']
        ws_docs.append(headers_docs)
        
        documentos = DocumentoOperador.objects.all().select_related('empleado', 'tipo_documento')
        for doc in documentos:
            row = [
                doc.id,
                export_origen_id(doc.empleado),
                doc.empleado.nombre_completo,
                doc.tipo_documento.nombre,
                doc.numero_documento or '',
                doc.fecha_expedicion.strftime('%Y-%m-%d') if doc.fecha_expedicion else '',
                doc.fecha_vencimiento.strftime('%Y-%m-%d') if doc.fecha_vencimiento else '',
                doc.observaciones or ''
            ]
            ws_docs.append(row)
        
        # ========== PESTAÑA 12: HIJOS ==========
        ws_hijos = wb.create_sheet("Hijos")
        headers_hijos = ['ID', 'Empleado ID', 'Empleado Nombre', 'Nombre Hijo',
                        'Fecha Nacimiento', 'Edad']
        ws_hijos.append(headers_hijos)
        
        hijos = Hijo.objects.all().select_related('empleado')
        for hijo in hijos:
            edad = hijo.edad if hasattr(hijo, 'edad') else 'N/A'
            row = [
                hijo.id,
                export_origen_id(hijo.empleado),
                hijo.empleado.nombre_completo,
                hijo.nombre,
                hijo.fecha_nacimiento.strftime('%Y-%m-%d') if hijo.fecha_nacimiento else '',
                edad
            ]
            ws_hijos.append(row)
        
        # ========== PESTAÑA 13: CATÁLOGOS ==========
        ws_catalogos = wb.create_sheet("Catálogos")
        headers_catalogos = ['Tipo', 'ID', 'Nombre', 'Descripción', 'Activo']
        ws_catalogos.append(headers_catalogos)
        
        # Departamentos
        for depto in Departamento.objects.all():
            ws_catalogos.append(['DEPARTAMENTO', depto.id, depto.nombre, depto.descripcion or '', 'Sí'])
        
        # Puestos
        for puesto in Puesto.objects.all():
            ws_catalogos.append(['PUESTO', puesto.id, puesto.nombre, puesto.descripcion or '', 'Sí'])
        
        # Motivos de baja
        for motivo in MotivoBaja.objects.all():
            ws_catalogos.append(['MOTIVO_BAJA', motivo.id, motivo.nombre, motivo.descripcion or '', 'Sí' if motivo.activo else 'No'])
        
        # Tipos documento operador
        for tipo in TipoDocumentoOperador.objects.all():
            ws_catalogos.append(['TIPO_DOC_OPERADOR', tipo.id, tipo.nombre, tipo.descripcion or '', 'Sí'])
        
        # Motivos inactivación
        for motivo in MotivoInactivacion.objects.all():
            ws_catalogos.append(['MOTIVO_INACTIVACION', motivo.id, motivo.motivo, motivo.descripcion or '', 'Sí'])
        
        # Divisiones operativas
        for div in DivisionOperativa.objects.all():
            ws_catalogos.append(['DIVISION_OPERATIVA', div.id, div.nombre, '', 'Sí'])
        
        # Tipos de carga
        for tipo in TipoCarga.objects.all():
            ws_catalogos.append(['TIPO_CARGA', tipo.id, tipo.nombre, '', 'Sí'])
        
        # Tipos de viaje
        for tipo in TipoViaje.objects.all():
            ws_catalogos.append(['TIPO_VIAJE', tipo.id, tipo.nombre, '', 'Sí'])
        
        # ========== APLICAR ESTILOS A TODAS LAS HOJAS ==========
        for ws in wb.worksheets:
            # Estilos de encabezado
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar encabezado
            ws.freeze_panes = 'A2'
        
        # ========== PREPARAR RESPUESTA ==========
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="backup_rh_completo_%s.xlsx"' % datetime.now().strftime('%Y%m%d_%H%M%S')
        wb.save(response)
        
        return response
        
    except Exception as e:
        print(f"❌ Error en exportación: {str(e)}")
        print(traceback.format_exc())
        messages.error(request, f"❌ Error al exportar: {str(e)[:200]}")
        return redirect('rh:inicio_rh')


import datetime

def parse_date(date_value):
    """Parsea fechas de Excel de forma robusta"""
    if date_value is None:
        return None
    
    # Si ya es fecha/datetime
    if isinstance(date_value, (datetime.date, datetime.datetime)):
        if isinstance(date_value, datetime.datetime):
            return date_value.date()
        return date_value
    
    # Si es string
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str or date_str.lower() in ('none', 'null', 'nan', '', 'na', 'n/a'):
            return None
        
        # Intentar parsear como fecha
        for fmt in (
            '%Y-%m-%d',
            '%d/%m/%Y', 
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%Y-%m-%d %H:%M:%S',
            '%d/%m/%Y %H:%M:%S',
        ):
            try:
                return datetime.datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    
    # Si es número (fecha serial de Excel)
    try:
        # Convertir a float primero
        if isinstance(date_value, (int, float, str)):
            float_value = float(date_value)
            # Excel usa 1900-01-01 como día 1, con bug del 29/feb/1900
            base_date = datetime.datetime(1899, 12, 30)
            return (base_date + datetime.timedelta(days=float_value)).date()
    except (ValueError, TypeError):
        pass
    
    return None


def safe_float(value, default=0.0):
    """Convierte a float de forma segura"""
    if value is None:
        return default
    try:
        return float(value)
    except:
        return default

def safe_int(value, default=None):
    """Convierte a int de forma segura"""
    if value is None:
        return default
    try:
        return int(value)
    except:
        return default

def safe_str(value, default=''):
    """Convierte a string de forma segura"""
    if value is None:
        return default
    try:
        return str(value).strip()
    except:
        return default
    
def get_empleado_by_origen(origen_id):
    if not origen_id:
        return None
    return Empleado.objects.filter(origen_id=str(origen_id)).first()

from django.utils import timezone

@login_required
def importar_rh_completo(request):
    """Vista para importar datos desde un archivo Excel con distribución por estado"""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            
            # Inicializar resultados
            resultados = {
                'empleados_activos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'empleados_papelera': {'creados': 0, 'actualizados': 0, 'errores': []},
                'bajas_historicas': {'creados': 0, 'actualizados': 0, 'errores': []},
                'registros_bajas': {'creados': 0, 'actualizados': 0, 'errores': []},
                'salarios': {'creados': 0, 'actualizados': 0, 'errores': []},
                'contratos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'historial': {'creados': 0, 'actualizados': 0, 'errores': []},
                'vacaciones': {'creados': 0, 'actualizados': 0, 'errores': []},
                'prestamos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'pagos_prestamos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'documentos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'hijos': {'creados': 0, 'actualizados': 0, 'errores': []},
                'catalogos': {'creados': 0, 'actualizados': 0, 'errores': []},
            }
            
            print(f"=== INICIANDO IMPORTACIÓN COMPLETA ===")
            print(f"📋 Hojas disponibles: {wb.sheetnames}")
            
            # ========== 1. IMPORTAR CATÁLOGOS (PRIMERO) ==========
            if 'Catálogos' in wb.sheetnames:
                ws_catalogos = wb['Catálogos']
                print("📚 Importando catálogos...")
                
                for row_idx, row in enumerate(ws_catalogos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            tipo = safe_str(row[0])
                            cat_id = safe_int(row[1])
                            nombre = safe_str(row[2])
                            descripcion = safe_str(row[3]) if len(row) > 3 else ''
                            activo = safe_str(row[4]) == 'Sí' if len(row) > 4 else True
                            
                            if not nombre:
                                continue
                            
                            if tipo == 'DEPARTAMENTO':
                                depto, created = Departamento.objects.update_or_create(
                                    id=cat_id,
                                    defaults={
                                        'nombre': nombre,
                                        'descripcion': descripcion,
                                        'creado_por': request.user
                                    }
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                    
                            elif tipo == 'PUESTO':
                                puesto, created = Puesto.objects.update_or_create(
                                    id=cat_id,
                                    defaults={
                                        'nombre': nombre,
                                        'descripcion': descripcion,
                                        'creado_por': request.user
                                    }
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                    
                            elif tipo == 'MOTIVO_BAJA':
                                # Determinar tipo de motivo
                                tipo_motivo = 'PRINCIPAL'
                                if 'secundario' in nombre.lower():
                                    tipo_motivo = 'SECUNDARIO'
                                elif 'detalle' in nombre.lower():
                                    tipo_motivo = 'DETALLE'
                                
                                motivo, created = MotivoBaja.objects.update_or_create(
                                    id=cat_id,
                                    defaults={
                                        'nombre': nombre,
                                        'descripcion': descripcion,
                                        'tipo_motivo': tipo_motivo,
                                        'activo': activo
                                    }
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                
                            elif tipo == 'TIPO_DOC_OPERADOR':
                                tipo_doc, created = TipoDocumentoOperador.objects.update_or_create(
                                    id=cat_id,
                                    defaults={
                                        'nombre': nombre,
                                        'descripcion': descripcion,
                                        'activo': activo
                                    }
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                
                            elif tipo == 'MOTIVO_INACTIVACION':
                                motivo, created = MotivoInactivacion.objects.update_or_create(
                                    id=cat_id,
                                    defaults={
                                        'motivo': nombre,
                                        'descripcion': descripcion
                                    }
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                
                            elif tipo == 'DIVISION_OPERATIVA':
                                division, created = DivisionOperativa.objects.update_or_create(
                                    id=cat_id,
                                    defaults={'nombre': nombre}
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                
                            elif tipo == 'TIPO_CARGA':
                                tipo_carga, created = TipoCarga.objects.update_or_create(
                                    id=cat_id,
                                    defaults={'nombre': nombre}
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                
                            elif tipo == 'TIPO_VIAJE':
                                tipo_viaje, created = TipoViaje.objects.update_or_create(
                                    id=cat_id,
                                    defaults={'nombre': nombre}
                                )
                                resultados['catalogos']['creados' if created else 'actualizados'] += 1
                                    
                    except Exception as e:
                        resultados['catalogos']['errores'].append(f"Fila {row_idx}: {str(e)[:100]}")
                        print(f"  ❌ Error en catálogo fila {row_idx}: {str(e)}")
            
            print(f"  ✅ Catálogos: {resultados['catalogos']['creados']} creados, {resultados['catalogos']['actualizados']} actualizados")
            
            # ========== 2. IMPORTAR EMPLEADOS ACTIVOS ==========
            if 'Empleados_Activos' in wb.sheetnames:
                ws_activos = wb['Empleados_Activos']
                print("👥 Importando empleados activos...")
                
                for row_idx, row in enumerate(ws_activos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row[:4]):
                                continue
                            
                            # DEBUG: Para primeras 3 filas
                            if row_idx <= 3:
                                print(f"\n=== DEBUG Fila {row_idx} ===")
                                for i, cell in enumerate(row[:10]):
                                    print(f"Col {i}: {repr(cell)} (Tipo: {type(cell)})")
                            
                            origen_id = safe_str(row[0])
                            if not origen_id:
                                resultados['empleados_activos']['errores'].append(
                                    f"Fila {row_idx}: origen_id vacío"
                                )
                                continue
                            
                            numero_empleado = safe_str(row[1])
                            nombre = safe_str(row[2])
                            apellido = safe_str(row[3])
                            
                            if not nombre and not apellido:
                                continue
                            
                            # Preparar datos básicos del empleado
                            empleado_data = {
                                'numero_empleado': numero_empleado or None,
                                'nombre': nombre,
                                'apellido': apellido,
                                'activo': True,
                                'eliminado': False,
                                'fecha_eliminacion': None,
                                'motivo_eliminacion': '',
                                'puesto_anterior': '',
                            }
                            
                            # ⚠️ CRÍTICO: Validar y asegurar fecha_ingreso ⚠️
                            fecha_contratacion_raw = row[6] if len(row) > 6 else None
                            
                            # DEBUG de fecha
                            if row_idx <= 3:
                                print(f"DEBUG Fila {row_idx}: fecha_contratacion_raw = {repr(fecha_contratacion_raw)}")
                                print(f"DEBUG Fila {row_idx}: parse_date(fecha_contratacion_raw) = {parse_date(fecha_contratacion_raw)}")
                            
                            fecha_contratacion = parse_date(fecha_contratacion_raw)
                            
                            # GARANTIZAR que siempre haya una fecha válida
                            if not fecha_contratacion:
                                fecha_contratacion = timezone.now().date()
                                print(f"⚠️  Fila {row_idx}: Fecha de contratación inválida. Usando fecha actual: {fecha_contratacion}")
                            
                            # ASIGNAR AMBOS CAMPOS (AMBOS SON OBLIGATORIOS)
                            empleado_data['fecha_contratacion'] = fecha_contratacion
                            empleado_data['fecha_ingreso'] = fecha_contratacion
                            
                            if row_idx <= 3:
                                print(f"DEBUG: fecha_contratacion = {fecha_contratacion}")
                                print(f"DEBUG: fecha_ingreso = {fecha_contratacion}")
                            
                            # Campos adicionales si existen
                            if len(row) > 5 and safe_str(row[5]):  # Departamento
                                depto = Departamento.objects.filter(nombre=safe_str(row[5])).first()
                                if depto:
                                    empleado_data['departamento'] = depto
                            
                            if len(row) > 4 and safe_str(row[4]):  # Puesto
                                puesto = Puesto.objects.filter(nombre=safe_str(row[4])).first()
                                if puesto:
                                    empleado_data['puesto'] = puesto
                            
                            if len(row) > 8 and safe_str(row[8]):  # Email
                                empleado_data['email'] = safe_str(row[8])
                            
                            if len(row) > 9:  # Fecha nacimiento
                                empleado_data['fecha_nacimiento'] = parse_date(row[9])
                            
                            if len(row) > 10 and safe_str(row[10]):  # CURP
                                empleado_data['curp'] = safe_str(row[10])
                            
                            if len(row) > 11 and safe_str(row[11]):  # RFC
                                empleado_data['rfc'] = safe_str(row[11])
                            
                            if len(row) > 12 and safe_str(row[12]):  # NSS
                                empleado_data['nss'] = safe_str(row[12])
                            
                            if len(row) > 13 and safe_str(row[13]):  # Empresa
                                empresa_val = safe_str(row[13])
                                if empresa_val in ['MIGMAR', 'Migmar']:
                                    empleado_data['empresa'] = 'MIGMAR'
                                elif empresa_val in ['MARCO_MORALES', 'Marco Morales']:
                                    empleado_data['empresa'] = 'MARCO_MORALES'
                            
                            # Información de domicilio
                            if len(row) > 14: empleado_data['direccion'] = safe_str(row[14])
                            if len(row) > 15: empleado_data['colonia'] = safe_str(row[15])
                            if len(row) > 16: empleado_data['codigo_postal'] = safe_str(row[16])
                            if len(row) > 17: empleado_data['ciudad'] = safe_str(row[17])
                            if len(row) > 18: empleado_data['estado'] = safe_str(row[18])
                            if len(row) > 19: empleado_data['pais'] = safe_str(row[19])
                            if len(row) > 20: empleado_data['telefono_personal'] = safe_str(row[20])
                            
                            if len(row) > 21:  # Estado civil
                                estado_civil = safe_str(row[21])
                                for choice in Empleado.ESTADO_CIVIL_CHOICES:
                                    if estado_civil in choice:
                                        empleado_data['estado_civil'] = choice[0]
                                        break
                            
                            if len(row) > 22: empleado_data['nacionalidad'] = safe_str(row[22])
                            
                            # Supervisor
                            if len(row) > 23 and safe_str(row[23]):
                                supervisor_nombre = safe_str(row[23])
                                if supervisor_nombre:
                                    # Buscar por nombre completo
                                    nombres = supervisor_nombre.split()
                                    if len(nombres) >= 2:
                                        supervisor = Empleado.objects.filter(
                                            nombre__icontains=nombres[0],
                                            apellido__icontains=nombres[-1]
                                        ).first()
                                        if supervisor:
                                            empleado_data['supervisor'] = supervisor
                            
                            # Datos familiares
                            if len(row) > 24: empleado_data['nombre_conyuge'] = safe_str(row[24])
                            if len(row) > 25: empleado_data['telefono_conyuge'] = safe_str(row[25])
                            
                            # Datos bancarios
                            if len(row) > 26: empleado_data['banco'] = safe_str(row[26])
                            if len(row) > 27: empleado_data['clabe_interbancaria'] = safe_str(row[27])
                            if len(row) > 28: empleado_data['numero_cuenta'] = safe_str(row[28])
                            if len(row) > 29: empleado_data['numero_tarjeta'] = safe_str(row[29])
                            
                            # Referencias
                            if len(row) > 32: empleado_data['nombre_referencia_1'] = safe_str(row[32])
                            if len(row) > 33: empleado_data['telefono_referencia_1'] = safe_str(row[33])
                            if len(row) > 34: empleado_data['relacion_referencia_1'] = safe_str(row[34])
                            if len(row) > 35: empleado_data['nombre_referencia_2'] = safe_str(row[35])
                            if len(row) > 36: empleado_data['telefono_referencia_2'] = safe_str(row[36])
                            if len(row) > 37: empleado_data['relacion_referencia_2'] = safe_str(row[37])
                            
                            # Buscar o crear empleado
                            empleado, created = Empleado.objects.update_or_create(
                                origen_id=origen_id,
                                defaults=empleado_data
                            )
                            
                            if created:
                                resultados['empleados_activos']['creados'] += 1
                                print(f"  ✅ Creado: {empleado.nombre_completo}")
                            else:
                                resultados['empleados_activos']['actualizados'] += 1
                                print(f"  🔄 Actualizado: {empleado.nombre_completo}")
                                
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['empleados_activos']['errores'].append(error_msg)
                        print(f"  ❌ Error en fila {row_idx}: {str(e)}")
                        if row_idx <= 3:
                            import traceback
                            traceback.print_exc()
                        continue
            
            print(f"  ✅ Empleados activos: {resultados['empleados_activos']['creados']} creados, {resultados['empleados_activos']['actualizados']} actualizados")
            
            # ========== 3. IMPORTAR EMPLEADOS EN PAPELERA ==========
            if 'Empleados_Papelera' in wb.sheetnames:
                ws_papelera = wb['Empleados_Papelera']
                print("🗑️  Importando empleados en papelera...")
                
                for row_idx, row in enumerate(ws_papelera.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row[:4]):
                                continue
                            
                            origen_id = safe_str(row[0])
                            if not origen_id:
                                resultados['empleados_papelera']['errores'].append(
                                    f"Fila {row_idx}: origen_id vacío"
                                )
                                continue
                            
                            numero_empleado = safe_str(row[1])
                            nombre = safe_str(row[2])
                            apellido = safe_str(row[3])
                            
                            # ⚠️ CRÍTICO: Validar fecha de ingreso
                            fecha_ingreso_raw = row[6] if len(row) > 6 else None
                            fecha_ingreso = parse_date(fecha_ingreso_raw)
                            
                            if not fecha_ingreso:
                                fecha_ingreso = timezone.now().date()
                                print(f"⚠️  Fila {row_idx} (papelera): Fecha de ingreso inválida. Usando fecha actual: {fecha_ingreso}")
                            
                            if not nombre and not apellido:
                                continue
                            
                            # Buscar o crear empleado para papelera
                            empleado_data = {
                                'numero_empleado': numero_empleado or None,
                                'nombre': nombre,
                                'apellido': apellido,
                                'fecha_contratacion': fecha_ingreso,
                                'fecha_ingreso': fecha_ingreso,
                                'activo': False,
                                'eliminado': True,
                                'eliminado_por': request.user,
                                'puesto_anterior': ''
                            }
                            
                            # Fecha de eliminación
                            if len(row) > 38:
                                empleado_data['fecha_eliminacion'] = parse_date(row[38]) or timezone.now()
                            
                            # Motivo de eliminación
                            if len(row) > 39:
                                empleado_data['motivo_eliminacion'] = safe_str(row[39])
                            
                            empleado, created = Empleado.objects.update_or_create(
                                origen_id=origen_id,
                                defaults=empleado_data
                            )
                            
                            if created:
                                resultados['empleados_papelera']['creados'] += 1
                                print(f"  🗑️  Creado en papelera: {empleado.nombre_completo}")
                            else:
                                resultados['empleados_papelera']['actualizados'] += 1
                                print(f"  🔄 Actualizado en papelera: {empleado.nombre_completo}")
                                
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['empleados_papelera']['errores'].append(error_msg)
                        continue
            
            print(f"  ✅ Empleados papelera: {resultados['empleados_papelera']['creados']} creados, {resultados['empleados_papelera']['actualizados']} actualizados")
            
            # ========== 4. IMPORTAR BAJAS HISTÓRICAS ==========
            if 'Bajas_Historicas' in wb.sheetnames:
                ws_bajas_historicas = wb['Bajas_Historicas']
                print("📉 Importando bajas históricas...")
                
                for row_idx, row in enumerate(ws_bajas_historicas.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row[:4]):
                                continue
                            
                            origen_id = safe_str(row[0])
                            if not origen_id:
                                resultados['bajas_historicas']['errores'].append(
                                    f"Fila {row_idx}: origen_id vacío"
                                )
                                continue
                            
                            numero_empleado = safe_str(row[1])
                            nombre = safe_str(row[2])
                            apellido = safe_str(row[3])
                            
                            # ⚠️ CRÍTICO: Validar fecha de ingreso
                            fecha_ingreso_raw = row[6] if len(row) > 6 else None
                            fecha_ingreso = parse_date(fecha_ingreso_raw)
                            
                            if not fecha_ingreso:
                                fecha_ingreso = timezone.now().date()
                                print(f"⚠️  Fila {row_idx} (bajas): Fecha de ingreso inválida. Usando fecha actual: {fecha_ingreso}")
                            
                            if not nombre and not apellido:
                                continue
                            
                            # Buscar o crear empleado para baja histórica
                            empleado_data = {
                                'numero_empleado': numero_empleado or None,
                                'nombre': nombre,
                                'apellido': apellido,
                                'fecha_contratacion': fecha_ingreso,
                                'fecha_ingreso': fecha_ingreso,
                                'activo': False,
                                'eliminado': False,
                                'puesto_anterior': ''
                            }
                            
                            # Fecha de inactivación
                            if len(row) > 31:
                                empleado_data['fecha_inactivacion'] = parse_date(row[31])
                            
                            # Motivo de inactivación
                            if len(row) > 30 and safe_str(row[30]):
                                motivo_nombre = safe_str(row[30])
                                motivo_inactivacion, _ = MotivoInactivacion.objects.get_or_create(
                                    motivo=motivo_nombre[:200],
                                    defaults={'descripcion': f'Importado desde backup - {timezone.now().date()}'}
                                )
                                empleado_data['motivo_inactivacion'] = motivo_inactivacion
                            
                            empleado, created = Empleado.objects.update_or_create(
                                origen_id=origen_id,
                                defaults=empleado_data
                            )
                            
                            if created:
                                resultados['bajas_historicas']['creados'] += 1
                                print(f"  📉 Creado baja histórica: {empleado.nombre_completo}")
                            else:
                                resultados['bajas_historicas']['actualizados'] += 1
                                print(f"  🔄 Actualizado baja histórica: {empleado.nombre_completo}")
                                
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['bajas_historicas']['errores'].append(error_msg)
                        continue
            
            print(f"  ✅ Bajas históricas: {resultados['bajas_historicas']['creados']} creados, {resultados['bajas_historicas']['actualizados']} actualizados")
            
            # ========== 5. IMPORTAR REGISTROS DE BAJAS ==========
            if 'Registros_Bajas' in wb.sheetnames:
                ws_bajas_detalles = wb['Registros_Bajas']
                print("📋 Importando registros de bajas...")
                
                for row_idx, row in enumerate(ws_bajas_detalles.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            baja_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['registros_bajas']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Buscar motivos
                            motivo_principal = None
                            motivo_secundario = None
                            motivo_detalle = None
                            
                            if len(row) > 3 and safe_str(row[3]):
                                motivo_principal = MotivoBaja.objects.filter(
                                    nombre__iexact=safe_str(row[3])
                                ).first()
                            
                            if len(row) > 4 and safe_str(row[4]):
                                motivo_secundario = MotivoBaja.objects.filter(
                                    nombre__iexact=safe_str(row[4])
                                ).first()
                            
                            if len(row) > 5 and safe_str(row[5]):
                                motivo_detalle = MotivoBaja.objects.filter(
                                    nombre__iexact=safe_str(row[5])
                                ).first()
                            
                            # Datos de la baja
                            baja_data = {
                                'empleado': empleado,
                                'motivo_principal': motivo_principal,
                                'motivo_secundario': motivo_secundario,
                                'motivo_detalle': motivo_detalle,
                                'fecha_baja': parse_date(row[6]) if len(row) > 6 else None,
                                'comentario_baja': safe_str(row[7]) if len(row) > 7 else '',
                                'es_recontratable': safe_str(row[8]) == 'Sí' if len(row) > 8 else False,
                                'motivo_recontratable': safe_str(row[9]) if len(row) > 9 else '',
                                'fecha_posible_recontratacion': parse_date(row[10]) if len(row) > 10 else None,
                                'recontratado': safe_str(row[11]) == 'Sí' if len(row) > 11 else False,
                                'fecha_recontratacion': parse_date(row[12]) if len(row) > 12 else None,
                                'fue_conciliacion_arbitraje': safe_str(row[13]) == 'Sí' if len(row) > 13 else False,
                                'fecha_conciliacion_arbitraje': parse_date(row[14]) if len(row) > 14 else None,
                                'registrado_por': request.user,
                            }
                            
                            baja, created = BajaEmpleado.objects.update_or_create(
                                id=baja_id,
                                defaults=baja_data
                            )
                            
                            if created:
                                resultados['registros_bajas']['creados'] += 1
                                print(f"  ✅ Registro de baja creado para {empleado.nombre_completo}")
                            else:
                                resultados['registros_bajas']['actualizados'] += 1
                                print(f"  🔄 Registro de baja actualizado para {empleado.nombre_completo}")
                                
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['registros_bajas']['errores'].append(error_msg)
                        print(f"  ❌ Error en baja fila {row_idx}: {str(e)}")
            
            print(f"  ✅ Registros bajas: {resultados['registros_bajas']['creados']} creados, {resultados['registros_bajas']['actualizados']} actualizados")
            
            # ========== 6. IMPORTAR SALARIOS ==========
            if 'Salarios' in wb.sheetnames:
                ws_salarios = wb['Salarios']
                print("💰 Importando salarios...")
                
                for row_idx, row in enumerate(ws_salarios.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            salario_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['salarios']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            salario_data = {
                                'empleado': empleado,
                                'sueldo_diario': safe_float(row[3]),
                                'fecha_efectiva': parse_date(row[4]),
                                'observaciones': safe_str(row[5]) if len(row) > 5 else ''
                            }
                            
                            obj = Salario.objects.filter(id=salario_id).first()
                            if obj:
                                for k, v in salario_data.items():
                                    setattr(obj, k, v)
                                obj.save()
                                created = False
                            else:
                                obj = Salario.objects.create(**salario_data)
                                created = True
                            
                            resultados['salarios']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['salarios']['errores'].append(error_msg)
            
            print(f"  ✅ Salarios: {resultados['salarios']['creados']} creados, {resultados['salarios']['actualizados']} actualizados")
            
            # ========== 7. IMPORTAR CONTRATOS ==========
            if 'Contratos' in wb.sheetnames:
                ws_contratos = wb['Contratos']
                print("📝 Importando contratos...")
                
                for row_idx, row in enumerate(ws_contratos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            contrato_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['contratos']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Determinar tipo de contrato
                            tipo_contrato_val = safe_str(row[3])
                            tipo_contrato = 'INDEFINIDO'
                            if 'temporal' in tipo_contrato_val.lower():
                                tipo_contrato = 'TEMPORAL'
                            elif 'prueba' in tipo_contrato_val.lower():
                                tipo_contrato = 'PRUEBA'
                            elif 'obra' in tipo_contrato_val.lower():
                                tipo_contrato = 'OBRA'
                            
                            contrato_data = {
                                'empleado': empleado,
                                'tipo_contrato': tipo_contrato,
                                'fecha_inicio': parse_date(row[4]),
                                'fecha_fin': parse_date(row[5]) if len(row) > 5 and row[5] else None,
                                'comentarios': safe_str(row[6]) if len(row) > 6 else ''
                            }
                            
                            contrato = Contrato.objects.filter(id=contrato_id).first()
                            if contrato:
                                for k, v in contrato_data.items():
                                    setattr(contrato, k, v)
                                contrato.save()
                                created = False
                            else:
                                contrato = Contrato.objects.create(**contrato_data)
                                created = True

                            resultados['contratos']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['contratos']['errores'].append(error_msg)
            
            print(f"  ✅ Contratos: {resultados['contratos']['creados']} creados, {resultados['contratos']['actualizados']} actualizados")
            
            # ========== 8. IMPORTAR HISTORIAL LABORAL ==========
            if 'Historial_Laboral' in wb.sheetnames:
                ws_historial = wb['Historial_Laboral']
                print("📅 Importando historial laboral...")
                
                for row_idx, row in enumerate(ws_historial.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            historial_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['historial']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Determinar tipo de evento
                            tipo_evento_val = safe_str(row[3])
                            tipo_evento = 'CONTRATACION'
                            if 'promocion' in tipo_evento_val.lower():
                                tipo_evento = 'PROMOCION'
                            elif 'transferencia' in tipo_evento_val.lower():
                                tipo_evento = 'TRANSFERENCIA'
                            elif 'cambio' in tipo_evento_val.lower():
                                tipo_evento = 'CAMBIO_PUESTO'
                            elif 'baja' in tipo_evento_val.lower():
                                tipo_evento = 'BAJA'
                            
                            # Buscar reemplazo si existe
                            reemplazo = None
                            if len(row) > 11 and row[11]:
                                reemplazo_origen_id = safe_str(row[11])
                                if reemplazo_origen_id:
                                    reemplazo = get_empleado_by_origen(reemplazo_origen_id)
                            
                            historial_data = {
                                'empleado': empleado,
                                'tipo_evento': tipo_evento,
                                'fecha_inicio': parse_date(row[4]),
                                'fecha_fin': parse_date(row[5]) if len(row) > 5 and row[5] else None,
                                'puesto': safe_str(row[6]) if len(row) > 6 else '',
                                'departamento': safe_str(row[7]) if len(row) > 7 else '',
                                'motivo_salida': safe_str(row[8]) if len(row) > 8 else '',
                                'descripcion': safe_str(row[9]) if len(row) > 9 else '',
                                'estatus': safe_str(row[10]) if len(row) > 10 else 'ACTIVO',
                                'reemplazo': reemplazo,
                                'fecha_reemplazo': parse_date(row[13]) if len(row) > 13 and row[13] else None
                            }
                            
                            historial_obj = HistorialLaboral.objects.filter(id=historial_id).first()
                            if historial_obj:
                                for k, v in historial_data.items():
                                    setattr(historial_obj, k, v)
                                historial_obj.save()
                                created = False
                            else:
                                historial_obj = HistorialLaboral.objects.create(**historial_data)
                                created = True

                            resultados['historial']['creados' if created else 'actualizados'] += 1

                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['historial']['errores'].append(error_msg)
            
            print(f"  ✅ Historial: {resultados['historial']['creados']} creados, {resultados['historial']['actualizados']} actualizados")
            
            # ========== 9. IMPORTAR VACACIONES ==========
            if 'Vacaciones' in wb.sheetnames:
                ws_vacaciones = wb['Vacaciones']
                print("🏖️  Importando vacaciones...")
                
                for row_idx, row in enumerate(ws_vacaciones.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            vacacion_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['vacaciones']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Buscar aprobado_por si existe
                            aprobado_por = None
                            if len(row) > 11 and safe_str(row[11]):
                                aprobador_nombre = safe_str(row[11])
                                if aprobador_nombre:
                                    nombres = aprobador_nombre.split()
                                    if len(nombres) >= 2:
                                        aprobado_por = User.objects.filter(
                                            first_name__icontains=nombres[0],
                                            last_name__icontains=nombres[-1]
                                        ).first()
                            
                            vacacion_data = {
                                'empleado': empleado,
                                'tipo_vacacion': safe_str(row[3]),
                                'fecha_solicitud': parse_date(row[4]),
                                'fecha_inicio': parse_date(row[5]),
                                'fecha_fin': parse_date(row[6]),
                                'dias_solicitados': safe_int(row[7], 0),
                                'dias_reales': safe_int(row[8], 0) if len(row) > 8 else safe_int(row[7], 0),
                                'periodo_correspondiente': safe_str(row[9]) if len(row) > 9 else '',
                                'estado': safe_str(row[10]) if len(row) > 10 else 'SOLICITADO',
                                'aprobado_por': aprobado_por,
                                'fecha_aprobacion': parse_date(row[12]) if len(row) > 12 and row[12] else None,
                                'observaciones': safe_str(row[13]) if len(row) > 13 else ''
                            }
                            
                            vacacion = Vacacion.objects.filter(id=vacacion_id).first()
                            if vacacion:
                                for k, v in vacacion_data.items():
                                    setattr(vacacion, k, v)
                                vacacion.save()
                                created = False
                            else:
                                vacacion = Vacacion.objects.create(**vacacion_data)
                                created = True

                            resultados['vacaciones']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['vacaciones']['errores'].append(error_msg)
            
            print(f"  ✅ Vacaciones: {resultados['vacaciones']['creados']} creados, {resultados['vacaciones']['actualizados']} actualizados")
            
            # ========== 10. IMPORTAR PRÉSTAMOS ==========
            if 'Préstamos' in wb.sheetnames:
                ws_prestamos = wb['Préstamos']
                print("💳 Importando préstamos...")
                
                for row_idx, row in enumerate(ws_prestamos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            prestamo_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['prestamos']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Buscar aprobado_por si existe
                            aprobado_por = None
                            if len(row) > 14 and safe_str(row[14]):
                                aprobador_nombre = safe_str(row[14])
                                if aprobador_nombre:
                                    nombres = aprobador_nombre.split()
                                    if len(nombres) >= 2:
                                        aprobado_por = User.objects.filter(
                                            first_name__icontains=nombres[0],
                                            last_name__icontains=nombres[-1]
                                        ).first()
                            
                            prestamo_data = {
                                'empleado': empleado,
                                'tipo_prestamo': safe_str(row[3]),
                                'fecha_solicitud': parse_date(row[4]),
                                'fecha_aprobacion': parse_date(row[5]) if len(row) > 5 and row[5] else None,
                                'monto_total': safe_float(row[6]),
                                'monto_pagado': safe_float(row[7]) if len(row) > 7 else 0,
                                'saldo_pendiente': safe_float(row[8]) if len(row) > 8 else safe_float(row[6]),
                                'tasa_interes': safe_float(row[9]) if len(row) > 9 else 0,
                                'plazo_semanas': safe_int(row[10], 0),
                                'fecha_primer_pago': parse_date(row[11]),
                                'estado': safe_str(row[12]) if len(row) > 12 else 'SOLICITADO',
                                'concepto': safe_str(row[13]) if len(row) > 13 else '',
                                'aprobado_por': aprobado_por,
                                'observaciones': safe_str(row[15]) if len(row) > 15 else ''
                            }
                            
                            prestamo = Prestamo.objects.filter(id=prestamo_id).first()
                            if prestamo:
                                for k, v in prestamo_data.items():
                                    setattr(prestamo, k, v)
                                prestamo.save()
                                created = False
                            else:
                                prestamo = Prestamo.objects.create(**prestamo_data)
                                created = True

                            resultados['prestamos']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['prestamos']['errores'].append(error_msg)
            
            print(f"  ✅ Préstamos: {resultados['prestamos']['creados']} creados, {resultados['prestamos']['actualizados']} actualizados")
            
            # ========== 11. IMPORTAR PAGOS DE PRÉSTAMOS ==========
            if 'Pagos_Préstamos' in wb.sheetnames:
                ws_pagos = wb['Pagos_Préstamos']
                print("💵 Importando pagos de préstamos...")
                
                for row_idx, row in enumerate(ws_pagos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue

                            pago_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[2])
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['pagos_prestamos']['errores'].append(
                                    f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})"
                                )
                                continue
                            
                            prestamo = Prestamo.objects.filter(
                                empleado=empleado,
                                fecha_primer_pago=parse_date(row[11])
                            ).first()

                            if not prestamo:
                                resultados['pagos_prestamos']['errores'].append(
                                    f"Fila {row_idx}: Préstamo no encontrado para empleado {empleado_origen_id}"
                                )
                                continue
                            
                            pago_data = {
                                'prestamo': prestamo,
                                'numero_pago': safe_int(row[4], 1),
                                'fecha_pago': parse_date(row[5]),
                                'fecha_pago_real': parse_date(row[6]) if len(row) > 6 and row[6] else None,
                                'monto_programado': safe_float(row[7]),
                                'monto_pagado': safe_float(row[8]) if len(row) > 8 else safe_float(row[7]),
                                'estado': safe_str(row[9]) if len(row) > 9 else 'PENDIENTE',
                                'observaciones': safe_str(row[10]) if len(row) > 10 else ''
                            }
                            
                            pago = PagoPrestamo.objects.filter(
                                prestamo=prestamo,
                                numero_pago=safe_int(row[4], 1)
                            ).first()
                            if pago:
                                for k, v in pago_data.items():
                                    setattr(pago, k, v)
                                pago.save()
                                created = False
                            else:
                                pago = PagoPrestamo.objects.create(**pago_data)
                                created = True

                            resultados['pagos_prestamos']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['pagos_prestamos']['errores'].append(error_msg)
            
            print(f"  ✅ Pagos préstamos: {resultados['pagos_prestamos']['creados']} creados, {resultados['pagos_prestamos']['actualizados']} actualizados")
            
            # ========== 12. IMPORTAR DOCUMENTOS OPERADOR ==========
            if 'Documentos_Operador' in wb.sheetnames:
                ws_docs = wb['Documentos_Operador']
                print("📄 Importando documentos operador...")
                
                for row_idx, row in enumerate(ws_docs.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            documento_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['documentos']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            # Buscar tipo de documento
                            tipo_documento = None
                            if len(row) > 3 and safe_str(row[3]):
                                tipo_nombre = safe_str(row[3])
                                tipo_documento = TipoDocumentoOperador.objects.filter(
                                    nombre__iexact=tipo_nombre
                                ).first()
                            
                            documento_data = {
                                'empleado': empleado,
                                'tipo_documento': tipo_documento,
                                'numero_documento': safe_str(row[4]) if len(row) > 4 else '',
                                'fecha_expedicion': parse_date(row[5]) if len(row) > 5 and row[5] else None,
                                'fecha_vencimiento': parse_date(row[6]) if len(row) > 6 and row[6] else None,
                                'observaciones': safe_str(row[7]) if len(row) > 7 else ''
                            }
                            
                            documento = DocumentoOperador.objects.filter(id=documento_id).first()
                            if documento:
                                for k, v in documento_data.items():
                                    setattr(documento, k, v)
                                documento.save()
                                created = False
                            else:
                                documento = DocumentoOperador.objects.create(**documento_data)
                                created = True

                            resultados['documentos']['creados' if created else 'actualizados'] += 1
                            
                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['documentos']['errores'].append(error_msg)
            
            print(f"  ✅ Documentos: {resultados['documentos']['creados']} creados, {resultados['documentos']['actualizados']} actualizados")
            
            # ========== 13. IMPORTAR HIJOS ==========
            if 'Hijos' in wb.sheetnames:
                ws_hijos = wb['Hijos']
                print("👶 Importando hijos...")
                
                for row_idx, row in enumerate(ws_hijos.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        with transaction.atomic():
                            if not row or not any(row):
                                continue
                            
                            hijo_id = safe_int(row[0])
                            empleado_origen_id = safe_str(row[1])
                            
                            empleado = get_empleado_by_origen(empleado_origen_id)
                            if not empleado:
                                resultados['hijos']['errores'].append(f"Fila {row_idx}: Empleado no encontrado (origen_id={empleado_origen_id})")
                                continue
                            
                            hijo_data = {
                                'empleado': empleado,
                                'nombre': safe_str(row[3]),
                                'fecha_nacimiento': parse_date(row[4]),
                                'edad': safe_int(row[5]) if len(row) > 5 and row[5] else None
                            }
                            
                            hijo = Hijo.objects.filter(id=hijo_id).first()
                            if hijo:
                                for k, v in hijo_data.items():
                                    setattr(hijo, k, v)
                                hijo.save()
                                created = False
                            else:
                                hijo = Hijo.objects.create(**hijo_data)
                                created = True

                            resultados['hijos']['creados' if created else 'actualizados'] += 1

                    except Exception as e:
                        error_msg = f"Fila {row_idx}: {str(e)[:100]}"
                        resultados['hijos']['errores'].append(error_msg)
            
            print(f"  ✅ Hijos: {resultados['hijos']['creados']} creados, {resultados['hijos']['actualizados']} actualizados")
            
            # ========== RESUMEN FINAL ==========
            total_empleados = (
                resultados['empleados_activos']['creados'] + resultados['empleados_activos']['actualizados'] +
                resultados['empleados_papelera']['creados'] + resultados['empleados_papelera']['actualizados'] +
                resultados['bajas_historicas']['creados'] + resultados['bajas_historicas']['actualizados']
            )
            
            total_creados = sum(v['creados'] for v in resultados.values())
            total_actualizados = sum(v['actualizados'] for v in resultados.values())
            total_errores = sum(len(v['errores']) for v in resultados.values())
            
            mensaje = f"""
            ✅ IMPORTACIÓN COMPLETADA
            
            📊 RESULTADOS:
            
            • EMPLEADOS ACTIVOS: {resultados['empleados_activos']['creados']} creados, {resultados['empleados_activos']['actualizados']} actualizados
            • EMPLEADOS EN PAPELERA: {resultados['empleados_papelera']['creados']} creados, {resultados['empleados_papelera']['actualizados']} actualizados
            • BAJAS HISTÓRICAS: {resultados['bajas_historicas']['creados']} creados, {resultados['bajas_historicas']['actualizados']} actualizados
            • REGISTROS DE BAJAS: {resultados['registros_bajas']['creados']} creados, {resultados['registros_bajas']['actualizados']} actualizados
            • SALARIOS: {resultados['salarios']['creados']} creados, {resultados['salarios']['actualizados']} actualizados
            • CONTRATOS: {resultados['contratos']['creados']} creados, {resultados['contratos']['actualizados']} actualizados
            • HISTORIAL LABORAL: {resultados['historial']['creados']} creados, {resultados['historial']['actualizados']} actualizados
            • VACACIONES: {resultados['vacaciones']['creados']} creados, {resultados['vacaciones']['actualizados']} actualizados
            • PRÉSTAMOS: {resultados['prestamos']['creados']} creados, {resultados['prestamos']['actualizados']} actualizados
            • PAGOS PRÉSTAMOS: {resultados['pagos_prestamos']['creados']} creados, {resultados['pagos_prestamos']['actualizados']} actualizados
            • DOCUMENTOS: {resultados['documentos']['creados']} creados, {resultados['documentos']['actualizados']} actualizados
            • HIJOS: {resultados['hijos']['creados']} creados, {resultados['hijos']['actualizados']} actualizados
            • CATÁLOGOS: {resultados['catalogos']['creados']} creados, {resultados['catalogos']['actualizados']} actualizados
            
            📈 TOTALES:
            • Empleados procesados: {total_empleados}
            • Registros creados: {total_creados}
            • Registros actualizados: {total_actualizados}
            ⚠️  Errores: {total_errores}
            """
            
            if total_errores > 0:
                mensaje += f"\n\n🔍 Primeros 5 errores:"
                error_count = 0
                for categoria, datos in resultados.items():
                    for error in datos['errores'][:1]:
                        if error_count < 5:
                            mensaje += f"\n• {categoria}: {error}"
                            error_count += 1
                        if error_count >= 5:
                            break
                    if error_count >= 5:
                        break
            
            messages.success(request, mensaje)
            print(f"=== IMPORTACIÓN FINALIZADA ===")
            print(f"Total creados: {total_creados}")
            print(f"Total actualizados: {total_actualizados}")
            print(f"Total errores: {total_errores}")
            
            return redirect('rh:inicio_rh')
            
        except Exception as e:
            error_msg = f'❌ Error al importar archivo: {str(e)}'
            print(f"ERROR GLOBAL: {error_msg}")
            print(traceback.format_exc())
            messages.error(request, error_msg)
            return redirect('rh:importar_rh_completo')
    
    context = {'page_title': 'Importar Datos RH desde Excel'}
    return render(request, 'rh/importar_rh_completo.html', context)


@login_required
def lista_bajas(request):
    """
    Vista para listar todas las bajas de empleados
    """
    # Obtener parámetros de filtrado
    tipo_baja = request.GET.get('tipo_baja')
    estado = request.GET.get('estado')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Obtener todas las bajas
    bajas = BajaEmpleado.objects.select_related(
        'empleado', 'motivo_principal'
    ).order_by('-fecha_baja')
    
    # Aplicar filtros
    if tipo_baja:
        bajas = bajas.filter(tipo_baja=tipo_baja)
    
    if estado:
        if estado == 'recontratable':
            bajas = bajas.filter(es_recontratable=True, recontratado=False)
        elif estado == 'recontratado':
            bajas = bajas.filter(recontratado=True)
        elif estado == 'definitiva':
            bajas = bajas.filter(es_recontratable=False, recontratado=False)
    
    if fecha_inicio:
        bajas = bajas.filter(fecha_baja__gte=fecha_inicio)
    
    if fecha_fin:
        bajas = bajas.filter(fecha_baja__lte=fecha_fin)
    
    # Estadísticas
    total_bajas = bajas.count()
    recontratables = bajas.filter(es_recontratable=True, recontratado=False).count()
    recontratados = bajas.filter(recontratado=True).count()
    
    context = {
        'bajas': bajas,
        'total_bajas': total_bajas,
        'recontratables': recontratables,
        'recontratados': recontratados,
        'page_title': 'Lista de Bajas de Empleados',
        'filtros': {
            'tipo_baja': tipo_baja,
            'estado': estado,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    }
    
    return render(request, 'rh/lista_bajas.html', context)

@login_required
def editar_baja(request, pk):
    """
    Vista para editar una baja existente
    """
    baja = get_object_or_404(
        BajaEmpleado.objects.select_related('empleado'),
        pk=pk
    )
    
    # Verificar que no esté recontratada (no se debería editar si ya fue recontratada)
    if baja.recontratado:
        messages.warning(
            request,
            f"No se puede editar la baja de {baja.empleado.nombre_completo} porque ya fue recontratado."
        )
        return redirect('rh:detalle_baja', pk=pk)
    
    if request.method == 'POST':
        # Usar el mismo formulario que para dar de baja, o crear uno específico
        # Dependiendo de cómo tengas tu formulario, podría ser:
        form = BajaEmpleadoForm(request.POST, request.FILES, instance=baja)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    baja_editada = form.save(commit=False)
                    baja_editada.ultima_actualizacion = timezone.now()
                    baja_editada.save()
                    
                    messages.success(
                        request,
                        f"✅ La baja de {baja.empleado.nombre_completo} ha sido actualizada exitosamente."
                    )
                    logger.info(f"Baja editada: {baja.empleado.nombre_completo} por {request.user}")
                    
                    return redirect('rh:detalle_baja', pk=pk)
                    
            except Exception as e:
                messages.error(
                    request,
                    f"❌ Error al actualizar la baja: {str(e)}"
                )
                logger.error(f"Error editando baja {pk}: {str(e)}")
        else:
            messages.error(
                request,
                "❌ Por favor corrige los errores en el formulario."
            )
    else:
        # Inicializar el formulario con la instancia existente
        form = BajaEmpleadoForm(instance=baja)
    
    context = {
        'form': form,
        'baja': baja,
        'page_title': f'Editar Baja - {baja.empleado.nombre_completo}',
        'modo_edicion': True,
    }
    
    return render(request, 'rh/editar_baja.html', context)

@login_required
def eliminar_baja(request, pk):
    """
    Vista para eliminar un registro de baja
    """
    baja = get_object_or_404(
        BajaEmpleado.objects.select_related('empleado'),
        pk=pk
    )
    
    if request.method == 'POST':
        try:
            empleado_nombre = baja.empleado.nombre_completo
            baja_id = baja.id
            
            # Verificar que no esté recontratada
            if baja.recontratado:
                messages.error(
                    request,
                    f"No se puede eliminar la baja de {empleado_nombre} porque ya fue recontratado."
                )
                return redirect('rh:detalle_baja', pk=pk)
            
            # Eliminar el documento adjunto si existe
            if baja.documento_baja:
                baja.documento_baja.delete(save=False)
            
            # Eliminar la baja
            baja.delete()
            
            messages.success(
                request,
                f"✅ El registro de baja de {empleado_nombre} ha sido eliminado exitosamente."
            )
            logger.info(f"Baja eliminada: {empleado_nombre} (ID: {baja_id}) por {request.user}")
            
            return redirect('rh:historial_bajas')
            
        except Exception as e:
            messages.error(
                request,
                f"❌ Error al eliminar la baja: {str(e)}"
            )
            logger.error(f"Error eliminando baja {pk}: {str(e)}")
            return redirect('rh:detalle_baja', pk=pk)
    
    # Si no es POST, mostrar página de confirmación
    context = {
        'baja': baja,
        'page_title': f'Eliminar Baja - {baja.empleado.nombre_completo}',
    }
    

# ========== CONTROL DE VACANTES CRUD ==========

class ControlVacanteListView(ListView):
    model = ControlVacante
    template_name = 'rh/controlvacante_list.html'
    context_object_name = 'controles'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Gestión de Metas de Vacantes'
        return context

class ControlVacanteCreateView(CreateView):
    model = ControlVacante
    form_class = ControlVacanteForm
    template_name = 'rh/controlvacante_form.html'
    success_url = reverse_lazy('rh:lista_control_vacantes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Nueva Meta de Vacante'
        context['titulo_card'] = 'Registrar Nueva Meta'
        return context

    def form_valid(self, form):
        messages.success(self.request, "✅ Meta registrada exitosamente.")
        return super().form_valid(form)

class ControlVacanteUpdateView(UpdateView):
    model = ControlVacante
    form_class = ControlVacanteForm
    template_name = 'rh/controlvacante_form.html'
    success_url = reverse_lazy('rh:lista_control_vacantes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Editar Meta: {self.object}'
        context['titulo_card'] = 'Actualizar Meta'
        return context

    def form_valid(self, form):
        messages.success(self.request, "✅ Meta actualizada exitosamente.")
        return super().form_valid(form)

class ControlVacanteDeleteView(DeleteView):
    model = ControlVacante
    template_name = 'rh/tipo_documento_operador_confirm_delete.html' # Reusando plantilla de confirmación genérica
    success_url = reverse_lazy('rh:lista_control_vacantes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Eliminar Meta'
        context['objeto_nombre'] = str(self.object)
        return context

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "✅ Meta eliminada exitosamente.")
        return super().delete(request, *args, **kwargs)

@login_required
def exportar_faltantes_excel(request):
    """
    Exporta la tabla de faltantes de vacantes a un archivo Excel.
    """
    controles = ControlVacante.objects.select_related('puesto', 'division').all()
    faltantes_data = []
    
    for c in controles:
        actual_query = Empleado.objects.filter(activo=True, empresa=c.empresa, puesto=c.puesto)
        if c.division:
            actual_query = actual_query.filter(division_operativa=c.division)
            
        count_actual = actual_query.count()
        faltantes = c.cantidad_presupuestada - count_actual
        
        if faltantes > 0:
            faltantes_data.append({
                'Empresa': c.get_empresa_display(),
                'Puesto': c.puesto.nombre,
                'División': c.division.nombre if c.division else 'General',
                'Objetivo': c.cantidad_presupuestada,
                'Actual': count_actual,
                'Faltantes': faltantes
            })

    # Crear DataFrame
    df = pd.DataFrame(faltantes_data)
    
    if df.empty:
        messages.info(request, "No hay datos de faltantes para exportar.")
        return redirect('rh:vacantes_dashboard')

    # Configurar respuesta HTTP para descarga de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Faltantes_Vacantes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    # Escribir a Excel usando pandas
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Faltantes')
        
        # Ajustar ancho de columnas y estilización
        workbook = writer.book
        worksheet = writer.sheets['Faltantes']
        
        # Estilo de cabecera
        header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
            # Ajustar ancho
            column_letter = get_column_letter(col_num + 1)
            worksheet.column_dimensions[column_letter].width = 20

    return response

@login_required
def auditoria_datos_operativos(request):
    """
    Identifica empleados activos con datos operativos incompletos.
    """
    # Empleados activos sin división operativa
    sin_division = Empleado.objects.filter(activo=True, division_operativa__isnull=True).distinct()
    
    # Empleados activos sin tipo de carga
    sin_carga = Empleado.objects.filter(activo=True, tipo_carga__isnull=True).distinct()
    
    # Empleados activos sin tipo de viaje
    sin_viaje = Empleado.objects.filter(activo=True, tipo_viaje__isnull=True).distinct()
    
    # Consolidar lista única de empleados con al menos un faltante
    empleados_con_faltantes = (sin_division | sin_carga | sin_viaje).distinct()
    
    # Detallar qué le falta a cada uno
    reporte = []
    for emp in empleados_con_faltantes:
        faltantes = []
        if not emp.division_operativa.exists():
            faltantes.append('División Operativa')
        if not emp.tipo_carga.exists():
            faltantes.append('Tipo de Carga')
        if not emp.tipo_viaje.exists():
            faltantes.append('Tipo de Viaje')
            
        if faltantes:
            reporte.append({
                'empleado': emp,
                'faltantes': faltantes,
                'faltantes_str': ", ".join(faltantes)
            })

    context = {
        'reporte': reporte,
        'total_incompletos': len(reporte),
        'total_activos': Empleado.objects.filter(activo=True).count(),
        'page_title': 'Auditoría de Calidad de Datos'
    }
    
    return render(request, 'rh/auditoria_datos.html', context)


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE CHECADAS
# ============================================================================
def calculate_lateness(check_in_time_str):
    try:
        check_in_time = pd.to_datetime(str(check_in_time_str), errors='coerce').time()
    except Exception:
        return 0 

    if pd.isnull(check_in_time):
        return 0

    nine_am = pd.to_datetime('09:00').time()
    nine_ten_am = pd.to_datetime('09:10').time()

    if check_in_time > nine_ten_am:
        check_in_datetime = pd.Timestamp.combine(pd.Timestamp.today().date(), check_in_time)
        nine_am_datetime = pd.Timestamp.combine(pd.Timestamp.today().date(), nine_am)
        return int((check_in_datetime - nine_am_datetime).total_seconds() / 60)
    return 0

def calculate_exit_minutes(row):
    day_str = str(row.get('Día de Semana', '')).lower()
    exit_val = row.get('Última Checada')
    
    try:
        exit_time = pd.to_datetime(str(exit_val), errors='coerce').time()
    except Exception:
        return 0
        
    if pd.isnull(exit_time):
        return 0
        
    if 'sábado' in day_str or 'sabado' in day_str:
        limit = pd.to_datetime('14:00').time()
    else:
        limit = pd.to_datetime('18:00').time()
        
    if exit_time > limit:
        today = pd.Timestamp.today().date()
        dt_exit = pd.Timestamp.combine(today, exit_time)
        dt_limit = pd.Timestamp.combine(today, limit)
        return int((dt_exit - dt_limit).total_seconds() / 60)
    return 0

def convert_time_to_text(time_val):
    if pd.isna(time_val) or str(time_val).strip() == '':
        return ''
    try:
        val_float = float(time_val)
        hours = int(val_float)
        minutes = int(round((val_float - hours) * 60))
        return f"{hours}h y {minutes}m"
    except ValueError:
        pass
    try:
        time_str = str(time_val)
        if ':' in time_str:
            parts = time_str.split(':')
            if len(parts) >= 2:
                return f"{int(parts[0])}h y {int(parts[1])}m"
    except Exception:
        pass
    return time_val

def process_data(df):
    if df.empty:
        return pd.DataFrame(), {}, {}

    columns_to_keep = [
        'ID de Empleado', 'Nombre', 'Apellido', 'Departamento', 'Fecha', 
        'Día de Semana', 'Primera Checada', 'Última Checada', 'Tiempo Total'
    ]
    
    existing_cols = [col for col in columns_to_keep if col in df.columns]
    df_filtered = df[existing_cols].copy()
    
    if 'Tiempo Total' in df_filtered.columns:
        df_filtered['Tiempo Detallado'] = df_filtered['Tiempo Total'].apply(convert_time_to_text)

    if 'Primera Checada' in df_filtered.columns:
        df_filtered['Retardo_min'] = df_filtered['Primera Checada'].apply(calculate_lateness)
        employee_tardiness = df_filtered.groupby('Nombre')['Retardo_min'].sum().to_dict()
    else:
        employee_tardiness = {}

    if 'Última Checada' in df_filtered.columns and 'Día de Semana' in df_filtered.columns:
        df_filtered['Minutos Extra'] = df_filtered.apply(calculate_exit_minutes, axis=1)
        employee_overtime = df_filtered.groupby('Nombre')['Minutos Extra'].sum().to_dict()
    else:
        df_filtered['Minutos Extra'] = 0
        employee_overtime = {}
    
    return df_filtered, employee_tardiness, employee_overtime

def write_to_excel_sheet(writer, df, sheet_name, tardiness_dict, overtime_dict, headers):
    workbook = writer.book
    worksheet = workbook.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = worksheet

    header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DDEBF7'})
    center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
    yellow_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00'})
    green_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE'})

    current_row = 0
    employee_names = df['Nombre'].unique() if 'Nombre' in df.columns else []

    for name in employee_names:
        worksheet.write_row(current_row, 0, headers, header_format)
        current_row += 1
        
        employee_data = df[df['Nombre'] == name]
        
        for _, row_data in employee_data.iterrows():
            for col_idx, header_name in enumerate(headers):
                cell_value = row_data.get(header_name, '')
                if pd.isna(cell_value): cell_value = ''
                
                cell_format_to_use = center_format
                
                if header_name == 'Primera Checada':
                    try:
                        check_in_time = pd.to_datetime(str(cell_value), errors='coerce').time()
                        if pd.notnull(check_in_time) and check_in_time > pd.to_datetime('09:10').time():
                            cell_format_to_use = yellow_format
                    except: pass
                
                if header_name == 'Minutos Extra':
                    try:
                        if float(cell_value) > 0:
                            cell_format_to_use = green_format
                    except: pass

                worksheet.write(current_row, col_idx, cell_value, cell_format_to_use)
            current_row += 1
            
        value_col_index = len(headers) - 1
        
        # TOTAL RETARDOS
        total_minutes_late = tardiness_dict.get(name, 0)
        worksheet.merge_range(current_row, 0, current_row, value_col_index - 2, 'TOTAL RETARDOS (Minutos):', header_format)
        worksheet.write(current_row, value_col_index - 1, total_minutes_late, center_format)
        current_row += 1
        
        if total_minutes_late > 0:
            h = total_minutes_late // 60
            m = total_minutes_late % 60
            worksheet.merge_range(current_row, 0, current_row, value_col_index - 2, 'Equivalente a:', header_format)
            worksheet.write(current_row, value_col_index - 1, f"{h}h y {m}m", center_format)
            current_row += 1

        # TOTAL TIEMPO EXTRA
        total_minutes_extra = overtime_dict.get(name, 0)
        worksheet.merge_range(current_row, 0, current_row, value_col_index - 2, 'TOTAL TIEMPO EXTRA (Minutos):', header_format)
        worksheet.write(current_row, value_col_index - 1, total_minutes_extra, green_format)
        current_row += 1

        if total_minutes_extra > 0:
            h_ex = total_minutes_extra // 60
            m_ex = total_minutes_extra % 60
            worksheet.merge_range(current_row, 0, current_row, value_col_index - 2, 'Equivalente a:', header_format)
            worksheet.write(current_row, value_col_index - 1, f"{h_ex}h y {m_ex}m", green_format)
            current_row += 1
        
        current_row += 2 

    for i, col in enumerate(headers):
        max_len = df[col].astype(str).str.len().max() if col in df.columns else 0
        if pd.isna(max_len): max_len = 0
        worksheet.set_column(i, i, max(max_len, len(col)) + 3)

# ============================================================================
# VISTA PRINCIPAL
# ============================================================================
@login_required
def procesar_checadas(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        try:
            # Leemos el archivo saltando las primeras 2 filas como en tu script
            df_full = pd.read_excel(archivo, skiprows=2)
            
            output = io.BytesIO()

            df_admon = df_full[df_full['Departamento'] == 'ADMON. MTY'].copy() if 'Departamento' in df_full.columns else pd.DataFrame()
            df_taller = df_full[df_full['Departamento'] == 'TALLER MTY'].copy() if 'Departamento' in df_full.columns else pd.DataFrame()

            admon_data, admon_tardiness, admon_overtime = process_data(df_admon)
            taller_data, taller_tardiness, taller_overtime = process_data(df_taller)
            
            final_headers = [
                'ID de Empleado', 'Nombre', 'Apellido', 'Departamento', 'Fecha', 
                'Día de Semana', 'Primera Checada', 'Última Checada', 
                'Tiempo Total', 'Tiempo Detallado', 'Minutos Extra'
            ]
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Si ambos están vacíos porque no existen esos departamentos, devolvemos todo en una pestaña
                if admon_data.empty and taller_data.empty:
                    todas_data, todas_tardiness, todas_overtime = process_data(df_full.copy())
                    write_to_excel_sheet(writer, todas_data, 'REPORTE', todas_tardiness, todas_overtime, final_headers)
                else:
                    if not admon_data.empty:
                        write_to_excel_sheet(writer, admon_data, 'ADMI', admon_tardiness, admon_overtime, final_headers)
                    if not taller_data.empty:
                        write_to_excel_sheet(writer, taller_data, 'TALLER', taller_tardiness, taller_overtime, final_headers)

            output.seek(0)
            
            # Devolvemos el Excel directamente al usuario para su descarga
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fecha_str = timezone.now().strftime("%Y%m%d_%H%M")
            response['Content-Disposition'] = f'attachment; filename="Reporte_Checadas_{fecha_str}.xlsx"'
            
            return response
            
        except Exception as e:
            messages.error(request, f"❌ Ocurrió un error procesando las checadas: {str(e)}")
            return redirect('rh:inicio_rh')
            
    messages.warning(request, "⚠️ No se recibió ningún archivo.")
    return redirect('rh:inicio_rh')


# ============================================================================
# API REST — ENDPOINTS PARA EL FRONTEND NEXT.JS
# ============================================================================

def _require_auth(request):
    if not request.user.is_authenticated:
        from django.http import JsonResponse
        return JsonResponse({'error': 'No autenticado'}, status=401)
    return None


@require_GET
def api_rh_dashboard(request):
    """GET /rh/api/dashboard/ — KPIs, alertas, cumpleaños, distribución, vacaciones y préstamos."""
    err = _require_auth(request)
    if err:
        return err

    today = date.today()

    total_empleados = Empleado.objects.no_eliminados().count()
    empleados_activos = Empleado.objects.no_eliminados().filter(activo=True).count()
    empleados_inactivos = Empleado.objects.no_eliminados().filter(activo=False).count()
    total_departamentos = Departamento.objects.count()

    operadores_migmar = Empleado.objects.no_eliminados().filter(
        activo=True, empresa='MIGMAR', puesto__nombre__icontains='Operador'
    ).count()
    operadores_marco = Empleado.objects.no_eliminados().filter(
        activo=True, empresa='MARCO_MORALES', puesto__nombre__icontains='Operador'
    ).count()

    # Alertas
    alertas = []
    fecha_limite_contrato = today + timedelta(days=30)
    for c in Contrato.objects.filter(
        tipo_contrato='DETERMINADO',
        fecha_fin__range=[today, fecha_limite_contrato],
        empleado__activo=True, empleado__eliminado=False
    ).select_related('empleado'):
        dias = (c.fecha_fin - today).days
        alertas.append({
            'tipo': 'warning',
            'mensaje': f'Contrato vence en {dias} días',
            'empleado': c.empleado.nombre_completo,
            'fecha': str(c.fecha_fin),
        })

    fecha_limite_docs = today + timedelta(days=15)
    for d in DocumentoOperador.objects.filter(
        fecha_vencimiento__lte=fecha_limite_docs,
        empleado__activo=True, empleado__eliminado=False
    ).select_related('empleado', 'tipo_documento'):
        tipo = 'danger' if d.fecha_vencimiento < today else 'warning'
        alertas.append({
            'tipo': tipo,
            'mensaje': f'{d.tipo_documento.nombre} {"VENCIDO" if tipo == "danger" else f"vence en {(d.fecha_vencimiento - today).days} días"}',
            'empleado': d.empleado.nombre_completo,
            'fecha': str(d.fecha_vencimiento),
        })

    for v in Vacacion.objects.filter(
        estado='APROBADO',
        fecha_inicio__range=[today, today + timedelta(days=7)],
        empleado__eliminado=False
    ).select_related('empleado'):
        alertas.append({
            'tipo': 'info',
            'mensaje': f'Vacaciones en {(v.fecha_inicio - today).days} días ({v.dias_solicitados} días)',
            'empleado': v.empleado.nombre_completo,
            'fecha': str(v.fecha_inicio),
        })

    alertas.sort(key=lambda x: (x['tipo'] != 'danger', x['tipo'] != 'warning'))

    # Cumpleaños próximos (30 días)
    cumpleanos = []
    for emp in Empleado.objects.no_eliminados().filter(activo=True).exclude(fecha_nacimiento=None):
        try:
            cumple = emp.fecha_nacimiento.replace(year=today.year)
        except ValueError:
            continue
        if cumple < today:
            try:
                cumple = cumple.replace(year=today.year + 1)
            except ValueError:
                continue
        dias = (cumple - today).days
        if 0 <= dias <= 30:
            cumpleanos.append({'nombre': emp.nombre_completo, 'dias': dias})
    cumpleanos.sort(key=lambda x: x['dias'])

    # Distribución por departamento
    dist = Empleado.objects.no_eliminados().filter(activo=True).values(
        'departamento__nombre'
    ).annotate(total=Count('id')).order_by('-total')
    distribucion = [{'nombre': r['departamento__nombre'] or 'Sin Asignar', 'total': r['total']} for r in dist]

    # Vacaciones
    vac_pendientes = Vacacion.objects.filter(estado='PENDIENTE', empleado__eliminado=False).count()
    vac_aprobadas = Vacacion.objects.filter(
        estado='APROBADO', fecha_inicio__gt=today, empleado__eliminado=False
    ).count()
    vac_en_curso = Vacacion.objects.filter(
        estado='APROBADO', fecha_inicio__lte=today, fecha_fin__gte=today, empleado__eliminado=False
    ).count()
    historico = HistoricoVacaciones.objects.filter(empleado__eliminado=False, año=today.year)
    total_dias_tomados = historico.aggregate(t=Sum('dias_tomados'))['t'] or 0
    meta_dias = historico.aggregate(t=Sum('dias_correspondientes'))['t'] or 0

    # Préstamos
    prestamos_qs = Prestamo.objects.filter(estado='EN_CURSO', empleado__eliminado=False)
    prestamos_activos_count = prestamos_qs.count()
    monto_total = float(prestamos_qs.aggregate(t=Sum('monto_total'))['t'] or 0)
    monto_pendiente = float(prestamos_qs.aggregate(t=Sum('saldo_pendiente'))['t'] or 0)

    return JsonResponse({
        'total_empleados': total_empleados,
        'empleados_activos': empleados_activos,
        'empleados_inactivos': empleados_inactivos,
        'total_departamentos': total_departamentos,
        'operadores_migmar': operadores_migmar,
        'operadores_marco': operadores_marco,
        'total_operadores': operadores_migmar + operadores_marco,
        'alertas': alertas,
        'cumpleanos_proximos': cumpleanos,
        'distribucion_departamentos': distribucion,
        'vacaciones': {
            'pendientes': vac_pendientes,
            'aprobadas': vac_aprobadas,
            'en_curso': vac_en_curso,
            'total_dias_tomados': total_dias_tomados,
            'meta_dias': meta_dias,
        },
        'prestamos': {
            'activos': prestamos_activos_count,
            'monto_total': monto_total,
            'monto_pendiente': monto_pendiente,
        },
    })


@require_GET
def api_rh_departamentos(request):
    """GET /rh/api/departamentos/"""
    err = _require_auth(request)
    if err:
        return err
    data = list(Departamento.objects.values('id', 'nombre').order_by('nombre'))
    return JsonResponse(data, safe=False)


@require_GET
def api_rh_puestos(request):
    """GET /rh/api/puestos/"""
    err = _require_auth(request)
    if err:
        return err
    data = list(Puesto.objects.values('id', 'nombre').order_by('nombre'))
    return JsonResponse(data, safe=False)


@require_GET
def api_rh_empleados(request):
    """GET /rh/api/empleados/?search=&departamento=&activo=&empresa=&ordering=&page=&page_size="""
    err = _require_auth(request)
    if err:
        return err

    qs = Empleado.objects.no_eliminados().select_related('departamento', 'puesto')

    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(nombre__icontains=search) |
            Q(apellido__icontains=search) |
            Q(numero_empleado__icontains=search)
        )

    depto = request.GET.get('departamento', '')
    if depto:
        qs = qs.filter(departamento_id=depto)

    activo = request.GET.get('activo', '')
    if activo == 'true':
        qs = qs.filter(activo=True)
    elif activo == 'false':
        qs = qs.filter(activo=False)

    empresa = request.GET.get('empresa', '')
    if empresa:
        qs = qs.filter(empresa=empresa)

    ordering = request.GET.get('ordering', 'apellido')
    field_map = {
        'nombre_completo': 'apellido', '-nombre_completo': '-apellido',
        'departamento': 'departamento__nombre', '-departamento': '-departamento__nombre',
        'puesto': 'puesto__nombre', '-puesto': '-puesto__nombre',
        'fecha_contratacion': 'fecha_contratacion', '-fecha_contratacion': '-fecha_contratacion',
        'numero_empleado': 'numero_empleado', '-numero_empleado': '-numero_empleado',
        'fecha_ingreso': 'fecha_ingreso', '-fecha_ingreso': '-fecha_ingreso',
    }
    order_field = field_map.get(ordering, 'apellido')
    qs = qs.order_by(order_field)

    try:
        page_size = min(int(request.GET.get('page_size', 20)), 100)
        page = max(int(request.GET.get('page', 1)), 1)
    except ValueError:
        page_size, page = 20, 1

    total = qs.count()
    offset = (page - 1) * page_size
    empleados = qs[offset:offset + page_size]

    results = []
    for e in empleados:
        results.append({
            'id': str(e.id),
            'numero_empleado': e.numero_empleado or '',
            'nombre': e.nombre or '',
            'apellido': e.apellido or '',
            'nombre_completo': e.nombre_completo,
            'departamento': e.departamento.nombre if e.departamento else '',
            'puesto': e.puesto.nombre if e.puesto else '',
            'empresa': e.get_empresa_display() if e.empresa else '',
            'email': e.email or '',
            'telefono_personal': e.telefono_personal or '',
            'curp': e.curp or '',
            'rfc': e.rfc or '',
            'nss': e.nss or '',
            'fecha_contratacion': str(e.fecha_contratacion) if e.fecha_contratacion else '',
            'fecha_nacimiento': str(e.fecha_nacimiento) if e.fecha_nacimiento else '',
            'fecha_ingreso': str(e.fecha_ingreso) if e.fecha_ingreso else '',
            'foto_perfil': e.foto_perfil.url if getattr(e, 'foto_perfil', None) else None,
            'activo': e.activo,
            'eliminado': e.eliminado,
        })

    return JsonResponse({
        'count': total,
        'next': None,
        'previous': None,
        'results': results,
    })


def api_rh_eliminar_empleado(request, pk):
    """POST /rh/api/empleados/<uuid>/eliminar/"""
    err = _require_auth(request)
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        empleado = get_object_or_404(Empleado, pk=pk, eliminado=False)
        data = json.loads(request.body or '{}')
        motivo = data.get('motivo', '')
        empleado.eliminado = True
        empleado.fecha_eliminacion = timezone.now()
        if motivo:
            empleado.motivo_eliminacion = motivo
        empleado.save(update_fields=['eliminado', 'fecha_eliminacion'])
        return JsonResponse({'ok': True, 'mensaje': f'{empleado.nombre_completo} eliminado.'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def api_rh_restaurar_empleado(request, pk):
    """POST /rh/api/empleados/<uuid>/restaurar/"""
    err = _require_auth(request)
    if err:
        return err
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        empleado = get_object_or_404(Empleado, pk=pk, eliminado=True)
        empleado.eliminado = False
        empleado.fecha_eliminacion = None
        empleado.save(update_fields=['eliminado', 'fecha_eliminacion'])
        return JsonResponse({'ok': True, 'mensaje': f'{empleado.nombre_completo} restaurado.'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@require_GET
def api_rh_vacaciones(request):
    """GET /rh/api/vacaciones/?estado=&page=&page_size="""
    err = _require_auth(request)
    if err:
        return err

    qs = Vacacion.objects.filter(empleado__eliminado=False).select_related('empleado', 'empleado__departamento')

    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado.upper())

    qs = qs.order_by('-fecha_creacion')

    try:
        page_size = min(int(request.GET.get('page_size', 20)), 100)
        page = max(int(request.GET.get('page', 1)), 1)
    except ValueError:
        page_size, page = 20, 1

    total = qs.count()
    offset = (page - 1) * page_size
    items = qs[offset:offset + page_size]

    # Balance por empleado (año actual) — una sola query
    from datetime import date as _d
    from .models import HistoricoVacaciones
    anio_actual = _d.today().year
    emp_ids = {v.empleado_id for v in items}
    historicos = {
        h.empleado_id: h for h in HistoricoVacaciones.objects.filter(
            empleado_id__in=emp_ids, año=anio_actual,
        )
    }
    tomados_map = {}
    pendientes_map = {}
    for v in Vacacion.objects.filter(empleado_id__in=emp_ids, fecha_inicio__year=anio_actual):
        if v.estado in ('APROBADO', 'GOZADO'):
            tomados_map[v.empleado_id] = tomados_map.get(v.empleado_id, 0) + (v.dias_solicitados or 0)
        elif v.estado == 'PENDIENTE':
            pendientes_map[v.empleado_id] = pendientes_map.get(v.empleado_id, 0) + (v.dias_solicitados or 0)

    def _balance(emp_id):
        h = historicos.get(emp_id)
        asignados = ((h.dias_correspondientes or 0) + (h.dias_extra or 0)) if h else 0
        tomados = tomados_map.get(emp_id, 0)
        pendientes = pendientes_map.get(emp_id, 0)
        return {
            'asignados': asignados,
            'tomados': tomados,
            'pendientes': pendientes,
            'restantes': max(0, asignados - tomados - pendientes),
        }

    results = []
    for v in items:
        bal = _balance(v.empleado_id)
        results.append({
            'id': str(v.id),
            'empleado': v.empleado.nombre_completo,
            'empleado_id': str(v.empleado_id) if v.empleado_id else '',
            'departamento': v.empleado.departamento.nombre if v.empleado.departamento_id else '',
            'fecha_inicio': str(v.fecha_inicio),
            'fecha_fin': str(v.fecha_fin),
            'dias': v.dias_solicitados,
            'estado': v.estado.lower(),
            'tipo': v.get_tipo_vacacion_display() if hasattr(v, 'get_tipo_vacacion_display') else '',
            'periodo_correspondiente': v.periodo_correspondiente or '',
            'anio': v.fecha_inicio.year if v.fecha_inicio else None,
            'fecha_creacion': v.fecha_creacion.isoformat() if v.fecha_creacion else '',
            'fecha_solicitud': str(v.fecha_solicitud) if v.fecha_solicitud else '',
            'balance_anio': anio_actual,
            'balance_asignados': bal['asignados'],
            'balance_tomados': bal['tomados'],
            'balance_pendientes': bal['pendientes'],
            'balance_restantes': bal['restantes'],
        })
    return JsonResponse({'count': total, 'next': None, 'previous': None, 'results': results})


@require_GET
def api_rh_prestamos(request):
    """GET /rh/api/prestamos/?estado=&page=&page_size="""
    err = _require_auth(request)
    if err:
        return err

    qs = Prestamo.objects.filter(empleado__eliminado=False).select_related('empleado')

    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado.upper())

    qs = qs.order_by('-fecha_solicitud')

    try:
        page_size = min(int(request.GET.get('page_size', 20)), 100)
        page = max(int(request.GET.get('page', 1)), 1)
    except ValueError:
        page_size, page = 20, 1

    total = qs.count()
    offset = (page - 1) * page_size
    items = qs[offset:offset + page_size]

    results = [
        {
            'id': str(p.id),
            'empleado': p.empleado.nombre_completo,
            'monto': float(p.monto_total),
            'saldo_pendiente': float(p.saldo_pendiente),
            'fecha_prestamo': str(p.fecha_solicitud) if hasattr(p, 'fecha_solicitud') else '',
            'estado': p.estado.lower(),
            'num_pagos': getattr(p, 'plazo_semanas', 0),
            'pagos_realizados': getattr(p, 'pagos_realizados', 0),
        }
        for p in items
    ]
    return JsonResponse({'count': total, 'next': None, 'previous': None, 'results': results})