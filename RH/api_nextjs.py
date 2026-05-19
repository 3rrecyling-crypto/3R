"""
API views adicionales para Next.js — CRUD de empleados y catálogos.
Se importan en urls.py con: from .api_nextjs import *
"""
import json
from decimal import Decimal
from datetime import date, datetime

from django.http import JsonResponse
from django.http.multipartparser import MultiPartParser
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Sum


def _parse_put_multipart(request):
    """Django no puebla request.POST en PUT/PATCH con multipart/form-data;
    esta función parsea el body y devuelve (post, files) usables."""
    if request.method in ('PUT', 'PATCH') and request.content_type and request.content_type.startswith('multipart/'):
        parser = MultiPartParser(request.META, request, request.upload_handlers, request.encoding)
        post, files = parser.parse()
        return post, files
    return request.POST, request.FILES

from .models import (
    Empleado, Salario, TipoViaje, TipoCarga, DivisionOperativa,
    Departamento, Puesto, Vacacion, HistoricoVacaciones, Prestamo,
)


def _require_auth(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    return None


def _empleado_to_dict(e):
    supervisor = e.supervisor
    return {
        'id': str(e.id),
        'numero_empleado': e.numero_empleado or '',
        'nombre': e.nombre or '',
        'apellido': e.apellido or '',
        'nombre_completo': e.nombre_completo,
        'email': e.email or '',
        'telefono_personal': e.telefono_personal or '',
        'fecha_contratacion': str(e.fecha_contratacion) if e.fecha_contratacion else '',
        'fecha_ingreso': str(e.fecha_ingreso) if e.fecha_ingreso else '',
        'puesto_id': str(e.puesto_id) if e.puesto_id else '',
        'puesto': e.puesto.nombre if e.puesto else '',
        'departamento_id': str(e.departamento_id) if e.departamento_id else '',
        'departamento': e.departamento.nombre if e.departamento else '',
        'supervisor': {
            'id': str(supervisor.id),
            'nombre_completo': supervisor.nombre_completo,
            'puesto': supervisor.puesto.nombre if supervisor.puesto else '',
        } if supervisor else None,
        'activo': e.activo,
        'fecha_nacimiento': str(e.fecha_nacimiento) if e.fecha_nacimiento else '',
        'estado_civil': e.estado_civil or '',
        'nacionalidad': e.nacionalidad or '',
        'curp': e.curp or '',
        'rfc': e.rfc or '',
        'nss': e.nss or '',
        'nombre_conyuge': e.nombre_conyuge or '',
        'telefono_conyuge': e.telefono_conyuge or '',
        'direccion': e.direccion or '',
        'codigo_postal': e.codigo_postal or '',
        'colonia': e.colonia or '',
        'ciudad': e.ciudad or '',
        'estado': e.estado or '',
        'pais': e.pais or '',
        'nombre_referencia_1': e.nombre_referencia_1 or '',
        'telefono_referencia_1': e.telefono_referencia_1 or '',
        'relacion_referencia_1': e.relacion_referencia_1 or '',
        'nombre_referencia_2': e.nombre_referencia_2 or '',
        'telefono_referencia_2': e.telefono_referencia_2 or '',
        'relacion_referencia_2': e.relacion_referencia_2 or '',
        'empresa': e.empresa or '',
        'tipo_viaje': list(e.tipo_viaje.values_list('id', flat=True)),
        'tipo_carga': list(e.tipo_carga.values_list('id', flat=True)),
        'division_operativa': list(e.division_operativa.values_list('id', flat=True)),
        'banco': e.banco or '',
        'numero_cuenta': e.numero_cuenta or '',
        'numero_tarjeta': e.numero_tarjeta or '',
        'clabe_interbancaria': e.clabe_interbancaria or '',
        'foto_perfil': e.foto_perfil.url if e.foto_perfil else None,
        'salarios': [
            {
                'id': s.id,
                'sueldo_diario': str(s.sueldo_diario or 0),
                'fecha_efectiva': str(s.fecha_efectiva) if s.fecha_efectiva else '',
                'observaciones': s.observaciones or '',
            }
            for s in e.salarios.order_by('-fecha_efectiva')
        ],
        'contratos': [
            {
                'id': c.id,
                'tipo_contrato': c.tipo_contrato or '',
                'fecha_inicio': str(c.fecha_inicio) if c.fecha_inicio else '',
                'fecha_fin': str(c.fecha_fin) if c.fecha_fin else '',
                'comentarios': c.comentarios or '',
                'archivo_url': c.archivo.url if c.archivo else None,
            }
            for c in e.contratos.order_by('-fecha_inicio')
        ] if hasattr(e, 'contratos') else [],
        'hijos': [
            {
                'id': h.id,
                'nombre': h.nombre or '',
                'fecha_nacimiento': str(h.fecha_nacimiento) if h.fecha_nacimiento else '',
            }
            for h in e.hijos.all()
        ] if hasattr(e, 'hijos') else [],
        'historial_laboral': [
            {
                'id': hl.id,
                'tipo_evento': hl.tipo_evento or '',
                'fecha_inicio': str(hl.fecha_inicio) if hl.fecha_inicio else '',
                'fecha_fin': str(hl.fecha_fin) if hl.fecha_fin else '',
                'descripcion': hl.descripcion or '',
                'motivo_salida': hl.motivo_salida or '',
                'puesto': hl.puesto or '',
                'departamento_id': str(hl.departamento_id) if hl.departamento_id else '',
                'archivo_url': hl.archivo.url if hl.archivo else None,
            }
            for hl in e.historial_laboral.order_by('-fecha_inicio')
        ] if hasattr(e, 'historial_laboral') else [],
        'documentos_operador': [
            {
                'id': d.id,
                'tipo_documento': d.tipo_documento or '',
                'numero_documento': d.numero_documento or '',
                'fecha_expedicion': str(d.fecha_expedicion) if d.fecha_expedicion else '',
                'fecha_vencimiento': str(d.fecha_vencimiento) if d.fecha_vencimiento else '',
                'observaciones': d.observaciones or '',
                'archivo_url': d.archivo.url if d.archivo else None,
            }
            for d in e.documentos_operador.all()
        ] if hasattr(e, 'documentos_operador') else [],
    }


def _apply_empleado_fields(empleado, data):
    simple = [
        'nombre', 'apellido', 'email', 'telefono_personal',
        'fecha_contratacion', 'numero_empleado',
        'fecha_nacimiento', 'estado_civil', 'nacionalidad',
        'curp', 'rfc', 'nss',
        'nombre_conyuge', 'telefono_conyuge',
        'direccion', 'codigo_postal', 'colonia', 'ciudad', 'pais',
        'nombre_referencia_1', 'telefono_referencia_1', 'relacion_referencia_1',
        'nombre_referencia_2', 'telefono_referencia_2', 'relacion_referencia_2',
        'empresa', 'banco', 'numero_cuenta', 'numero_tarjeta', 'clabe_interbancaria',
    ]
    for field in simple:
        if field in data:
            val = data[field]
            setattr(empleado, field, val if val else None)

    if 'estado_addr' in data:
        empleado.estado = data['estado_addr'] or None

    if 'puesto_id' in data and data['puesto_id']:
        try:
            empleado.puesto_id = int(data['puesto_id'])
        except (ValueError, TypeError):
            pass

    if 'departamento_id' in data and data['departamento_id']:
        try:
            empleado.departamento_id = int(data['departamento_id'])
        except (ValueError, TypeError):
            pass

    if 'supervisor_id' in data:
        sv = data['supervisor_id']
        if sv:
            try:
                empleado.supervisor = Empleado.objects.get(pk=sv)
            except Empleado.DoesNotExist:
                empleado.supervisor = None
        else:
            empleado.supervisor = None

    if 'activo' in data:
        val = str(data['activo']).strip().lower()
        empleado.activo = val in ('true', '1', 'yes', 'si', 'sí', 'on')


@csrf_exempt
@require_http_methods(["POST"])
def api_rh_crear_empleado(request):
    """POST /rh/api/empleados/crear/ — crea un empleado nuevo."""
    err = _require_auth(request)
    if err:
        return err
    try:
        with transaction.atomic():
            empleado = Empleado()
            _apply_empleado_fields(empleado, request.POST)
            if request.FILES.get('foto_perfil'):
                empleado.foto_perfil = request.FILES['foto_perfil']
            empleado.save()

            tvs = request.POST.getlist('tipo_viaje')
            if tvs:
                empleado.tipo_viaje.set(TipoViaje.objects.filter(id__in=tvs))
            tcs = request.POST.getlist('tipo_carga')
            if tcs:
                empleado.tipo_carga.set(TipoCarga.objects.filter(id__in=tcs))
            divs = request.POST.getlist('division_operativa')
            if divs:
                empleado.division_operativa.set(DivisionOperativa.objects.filter(id__in=divs))

        return JsonResponse(_empleado_to_dict(empleado), status=201)
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


@csrf_exempt
@require_http_methods(["GET", "PUT", "PATCH"])
def api_rh_empleado_detail(request, pk):
    """GET/PUT /rh/api/empleados/<uuid>/ — detalle o actualización."""
    err = _require_auth(request)
    if err:
        return err
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'GET':
        return JsonResponse(_empleado_to_dict(empleado))
    try:
        post_data, files_data = _parse_put_multipart(request)
        with transaction.atomic():
            _apply_empleado_fields(empleado, post_data)
            if files_data.get('foto_perfil'):
                empleado.foto_perfil = files_data['foto_perfil']
            empleado.save()
            empleado.tipo_viaje.set(TipoViaje.objects.filter(id__in=post_data.getlist('tipo_viaje')))
            empleado.tipo_carga.set(TipoCarga.objects.filter(id__in=post_data.getlist('tipo_carga')))
            empleado.division_operativa.set(DivisionOperativa.objects.filter(id__in=post_data.getlist('division_operativa')))
        return JsonResponse(_empleado_to_dict(empleado))
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


def api_rh_supervisores(request):
    """GET /rh/api/supervisores/"""
    err = _require_auth(request)
    if err:
        return err
    qs = Empleado.objects.filter(activo=True).select_related('puesto').order_by('apellido', 'nombre')
    return JsonResponse({
        'results': [
            {'id': str(e.id), 'nombre': e.nombre_completo}
            for e in qs
        ]
    })


def api_rh_tipos_viaje(request):
    """GET /rh/api/tipos-viaje/"""
    err = _require_auth(request)
    if err:
        return err
    return JsonResponse({'results': list(TipoViaje.objects.values('id', 'nombre').order_by('nombre'))})


def api_rh_tipos_carga(request):
    """GET /rh/api/tipos-carga/"""
    err = _require_auth(request)
    if err:
        return err
    return JsonResponse({'results': list(TipoCarga.objects.values('id', 'nombre').order_by('nombre'))})


def api_rh_divisiones(request):
    """GET /rh/api/divisiones/"""
    err = _require_auth(request)
    if err:
        return err
    return JsonResponse({'results': list(DivisionOperativa.objects.values('id', 'nombre').order_by('nombre'))})


# ── Departamentos CRUD ──────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(["GET", "POST"])
def api_departamentos_list(request):
    """GET /rh/api/departamentos/crear/  POST /rh/api/departamentos/crear/"""
    err = _require_auth(request)
    if err:
        return err
    if request.method == 'GET':
        data = [
            {'id': d.id, 'nombre': d.nombre, 'descripcion': d.descripcion or '',
             'total_empleados': d.empleados.filter(activo=True).count()}
            for d in Departamento.objects.order_by('nombre').prefetch_related('empleados')
        ]
        return JsonResponse({'results': data})
    body = json.loads(request.body)
    d = Departamento.objects.create(
        nombre=body.get('nombre', '').strip(),
        descripcion=body.get('descripcion', '').strip() or None,
    )
    return JsonResponse({'id': d.id, 'nombre': d.nombre, 'descripcion': d.descripcion or ''}, status=201)


@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def api_departamento_detail(request, pk):
    """GET/PUT/DELETE /rh/api/departamentos/<pk>/"""
    err = _require_auth(request)
    if err:
        return err
    d = get_object_or_404(Departamento, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': d.id, 'nombre': d.nombre, 'descripcion': d.descripcion or ''})
    if request.method == 'DELETE':
        d.delete()
        return JsonResponse({'ok': True})
    body = json.loads(request.body)
    d.nombre = body.get('nombre', d.nombre).strip()
    d.descripcion = body.get('descripcion', d.descripcion or '').strip() or None
    d.save()
    return JsonResponse({'id': d.id, 'nombre': d.nombre, 'descripcion': d.descripcion or ''})


# ── Puestos CRUD ───────────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(["POST"])
def api_puestos_crear(request):
    """POST /rh/api/puestos/crear/"""
    err = _require_auth(request)
    if err:
        return err
    body = json.loads(request.body)
    salario = body.get('salario_base')
    p = Puesto.objects.create(
        nombre=body.get('nombre', '').strip(),
        descripcion=body.get('descripcion', '').strip() or None,
        salario_base=Decimal(salario) if salario else None,
    )
    return JsonResponse({'id': p.id, 'nombre': p.nombre, 'descripcion': p.descripcion or '',
                         'salario_base': str(p.salario_base) if p.salario_base else None}, status=201)


@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def api_puesto_detail(request, pk):
    """GET/PUT/DELETE /rh/api/puestos/<pk>/"""
    err = _require_auth(request)
    if err:
        return err
    p = get_object_or_404(Puesto, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': p.id, 'nombre': p.nombre, 'descripcion': p.descripcion or '',
                             'salario_base': str(p.salario_base) if p.salario_base else None})
    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'ok': True})
    body = json.loads(request.body)
    p.nombre = body.get('nombre', p.nombre).strip()
    p.descripcion = body.get('descripcion', p.descripcion or '').strip() or None
    salario = body.get('salario_base')
    p.salario_base = Decimal(salario) if salario else None
    p.save()
    return JsonResponse({'id': p.id, 'nombre': p.nombre, 'descripcion': p.descripcion or '',
                         'salario_base': str(p.salario_base) if p.salario_base else None})


# ── Vacaciones — crear desde Next.js ────────────────────────────────────────────

@csrf_exempt
@require_http_methods(["POST"])
def api_rh_vacacion_crear(request):
    """POST /rh/api/vacaciones/crear/  — crea una solicitud de vacaciones."""
    err = _require_auth(request)
    if err:
        return err
    try:
        data = request.POST
        empleado_id = data.get('empleado_id')
        if not empleado_id:
            return JsonResponse({'error': 'Falta el empleado.'}, status=400)
        empleado = get_object_or_404(Empleado, pk=empleado_id)

        fecha_inicio_str = data.get('fecha_inicio') or ''
        fecha_fin_str = data.get('fecha_fin') or ''
        if not fecha_inicio_str or not fecha_fin_str:
            return JsonResponse({'error': 'Fechas de inicio y fin son obligatorias.'}, status=400)
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido (usa AAAA-MM-DD).'}, status=400)
        if fecha_fin < fecha_inicio:
            return JsonResponse({'error': 'La fecha fin no puede ser anterior a la de inicio.'}, status=400)

        try:
            dias = int(data.get('dias_solicitados') or 0)
        except (ValueError, TypeError):
            dias = 0
        if dias <= 0:
            return JsonResponse({'error': 'Los días solicitados deben ser mayor a cero.'}, status=400)

        periodo = (data.get('periodo_correspondiente') or '').strip()
        if not periodo:
            from datetime import date as _d
            periodo = _d.today().strftime('%Y-%m')

        with transaction.atomic():
            v = Vacacion.objects.create(
                empleado=empleado,
                tipo_vacacion=data.get('tipo_vacacion', 'ORDINARIAS'),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                dias_solicitados=dias,
                dias_reales=dias,
                periodo_correspondiente=periodo,
                estado='APROBADO',
                aprobado_por=request.user,
                fecha_aprobacion=date.today(),
                observaciones=(data.get('observaciones') or '').strip() or None,
                creado_por=request.user,
            )
            if request.FILES.get('documento_solicitud'):
                v.documento_solicitud = request.FILES['documento_solicitud']
                v.save(update_fields=['documento_solicitud'])
        return JsonResponse({
            'id': v.id,
            'empleado': v.empleado.nombre_completo,
            'fecha_inicio': str(v.fecha_inicio),
            'fecha_fin': str(v.fecha_fin),
            'dias': v.dias_solicitados,
            'estado': v.estado.lower(),
        }, status=201)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'{type(exc).__name__}: {exc}'}, status=400)


@csrf_exempt
@require_http_methods(["GET", "DELETE"])
def api_rh_vacacion_detail(request, pk):
    """GET/DELETE /rh/api/vacaciones/<pk>/ — detalle o eliminación."""
    err = _require_auth(request)
    if err:
        return err
    v = get_object_or_404(Vacacion.objects.select_related('empleado', 'empleado__departamento'), pk=pk)
    if request.method == 'DELETE':
        v.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({
        'id': v.id,
        'empleado_id': str(v.empleado.id),
        'empleado': v.empleado.nombre_completo,
        'departamento': v.empleado.departamento.nombre if v.empleado.departamento_id else '',
        'puesto': v.empleado.puesto.nombre if getattr(v.empleado, 'puesto_id', None) else '',
        'tipo_vacacion': v.tipo_vacacion,
        'tipo': v.get_tipo_vacacion_display() if hasattr(v, 'get_tipo_vacacion_display') else '',
        'fecha_inicio': str(v.fecha_inicio) if v.fecha_inicio else '',
        'fecha_fin': str(v.fecha_fin) if v.fecha_fin else '',
        'dias_solicitados': v.dias_solicitados,
        'dias_reales': v.dias_reales,
        'periodo_correspondiente': v.periodo_correspondiente or '',
        'anio': v.fecha_inicio.year if v.fecha_inicio else None,
        'estado': v.estado.lower(),
        'estado_display': v.get_estado_display() if hasattr(v, 'get_estado_display') else v.estado,
        'observaciones': v.observaciones or '',
        'fecha_solicitud': str(v.fecha_solicitud) if v.fecha_solicitud else '',
        'fecha_aprobacion': str(v.fecha_aprobacion) if v.fecha_aprobacion else '',
        'fecha_creacion': v.fecha_creacion.isoformat() if v.fecha_creacion else '',
        'aprobado_por': getattr(v.aprobado_por, 'username', '') if v.aprobado_por_id else '',
        'creado_por': getattr(v.creado_por, 'username', '') if v.creado_por_id else '',
        'documento_solicitud_url': v.documento_solicitud.url if v.documento_solicitud else None,
        'documento_aprobacion_url': v.documento_aprobacion.url if v.documento_aprobacion else None,
    })


@csrf_exempt
@require_http_methods(["GET", "POST"])
def api_rh_empleado_balance_vacaciones(request, pk):
    """GET/POST /rh/api/empleados/<uuid>/vacaciones-balance/?anio=
    GET devuelve el balance del año (correspondientes/tomados/pendientes/restantes).
    POST permite fijar dias_correspondientes y dias_extra para ese año."""
    err = _require_auth(request)
    if err:
        return err
    from datetime import date as _d
    empleado = get_object_or_404(Empleado, pk=pk)
    try:
        anio = int(request.GET.get('anio') or _d.today().year)
    except (ValueError, TypeError):
        anio = _d.today().year

    historico, _ = HistoricoVacaciones.objects.get_or_create(
        empleado=empleado, año=anio,
        defaults={'dias_correspondientes': 0, 'dias_tomados': 0, 'dias_pendientes': 0},
    )

    if request.method == 'POST':
        try:
            body = json.loads(request.body or '{}')
        except ValueError:
            body = {}
        if 'dias_correspondientes' in body:
            try:
                historico.dias_correspondientes = max(0, int(body['dias_correspondientes']))
            except (ValueError, TypeError):
                return JsonResponse({'error': 'dias_correspondientes inválido'}, status=400)
        if 'dias_extra' in body:
            try:
                historico.dias_extra = max(0, int(body['dias_extra']))
            except (ValueError, TypeError):
                return JsonResponse({'error': 'dias_extra inválido'}, status=400)
        historico.save()

    vacaciones_qs = Vacacion.objects.filter(empleado=empleado, fecha_inicio__year=anio)
    dias_tomados = vacaciones_qs.filter(estado__in=['APROBADO', 'GOZADO']).aggregate(
        total=Sum('dias_solicitados'))['total'] or 0
    dias_pendientes_solicitud = vacaciones_qs.filter(estado='PENDIENTE').aggregate(
        total=Sum('dias_solicitados'))['total'] or 0

    total_asignados = (historico.dias_correspondientes or 0) + (historico.dias_extra or 0)
    restantes = max(0, total_asignados - dias_tomados - dias_pendientes_solicitud)

    historico.dias_tomados = dias_tomados
    historico.dias_pendientes = restantes
    historico.save(update_fields=['dias_tomados', 'dias_pendientes'])

    return JsonResponse({
        'empleado_id': str(empleado.id),
        'empleado': empleado.nombre_completo,
        'anio': anio,
        'dias_correspondientes': historico.dias_correspondientes or 0,
        'dias_extra': historico.dias_extra or 0,
        'dias_tomados': dias_tomados,
        'dias_pendientes_solicitud': dias_pendientes_solicitud,
        'restantes': restantes,
    })


# ── Préstamos — crear / detalle desde Next.js ───────────────────────────────────

@csrf_exempt
@require_http_methods(["POST"])
def api_rh_prestamo_crear(request):
    """POST /rh/api/prestamos/crear/ — crea un préstamo y lo deja APROBADO."""
    err = _require_auth(request)
    if err:
        return err
    try:
        data = request.POST
        empleado_id = data.get('empleado_id')
        if not empleado_id:
            return JsonResponse({'error': 'Falta el empleado.'}, status=400)
        empleado = get_object_or_404(Empleado, pk=empleado_id)

        try:
            monto = Decimal(str(data.get('monto_total') or '0'))
        except Exception:
            return JsonResponse({'error': 'Monto inválido.'}, status=400)
        if monto <= 0:
            return JsonResponse({'error': 'El monto debe ser mayor a cero.'}, status=400)

        try:
            plazo = int(data.get('plazo_semanas') or 0)
        except (ValueError, TypeError):
            plazo = 0
        if plazo < 1 or plazo > 52:
            return JsonResponse({'error': 'El plazo debe estar entre 1 y 52 semanas.'}, status=400)

        try:
            tasa = Decimal(str(data.get('tasa_interes') or '0'))
        except Exception:
            tasa = Decimal('0')

        fecha_primer_pago_str = (data.get('fecha_primer_pago') or '').strip()
        if not fecha_primer_pago_str:
            return JsonResponse({'error': 'Falta la fecha del primer pago.'}, status=400)
        try:
            fecha_primer_pago = datetime.strptime(fecha_primer_pago_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido (AAAA-MM-DD).'}, status=400)

        concepto = (data.get('concepto') or '').strip()
        if not concepto:
            return JsonResponse({'error': 'El concepto es obligatorio.'}, status=400)

        with transaction.atomic():
            p = Prestamo.objects.create(
                empleado=empleado,
                tipo_prestamo=data.get('tipo_prestamo', 'PERSONAL'),
                monto_total=monto,
                tasa_interes=tasa,
                plazo_semanas=plazo,
                fecha_primer_pago=fecha_primer_pago,
                concepto=concepto,
                estado='APROBADO',
                fecha_aprobacion=date.today(),
                observaciones=(data.get('observaciones') or '').strip() or None,
            )
            if request.FILES.get('documento_solicitud'):
                p.documento_solicitud = request.FILES['documento_solicitud']
                p.save(update_fields=['documento_solicitud'])
            if request.FILES.get('contrato'):
                p.contrato = request.FILES['contrato']
                p.save(update_fields=['contrato'])
        return JsonResponse({
            'id': p.id,
            'empleado': p.empleado.nombre_completo,
            'monto_total': str(p.monto_total),
            'plazo_semanas': p.plazo_semanas,
            'estado': p.estado.lower(),
        }, status=201)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'{type(exc).__name__}: {exc}'}, status=400)


@csrf_exempt
@require_http_methods(["GET", "DELETE"])
def api_rh_prestamo_detail(request, pk):
    """GET/DELETE /rh/api/prestamos/<pk>/ — detalle o eliminación."""
    err = _require_auth(request)
    if err:
        return err
    p = get_object_or_404(Prestamo.objects.select_related('empleado', 'empleado__departamento'), pk=pk)
    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({
        'id': p.id,
        'empleado_id': str(p.empleado.id),
        'empleado': p.empleado.nombre_completo,
        'departamento': p.empleado.departamento.nombre if p.empleado.departamento_id else '',
        'puesto': p.empleado.puesto.nombre if getattr(p.empleado, 'puesto_id', None) else '',
        'tipo_prestamo': p.tipo_prestamo,
        'tipo': p.get_tipo_prestamo_display() if hasattr(p, 'get_tipo_prestamo_display') else '',
        'monto_total': str(p.monto_total),
        'monto_pagado': str(p.monto_pagado),
        'saldo_pendiente': str(p.saldo_pendiente),
        'tasa_interes': str(p.tasa_interes),
        'plazo_semanas': p.plazo_semanas,
        'fecha_primer_pago': str(p.fecha_primer_pago) if p.fecha_primer_pago else '',
        'fecha_solicitud': str(p.fecha_solicitud) if p.fecha_solicitud else '',
        'fecha_aprobacion': str(p.fecha_aprobacion) if p.fecha_aprobacion else '',
        'estado': p.estado.lower(),
        'estado_display': p.get_estado_display() if hasattr(p, 'get_estado_display') else p.estado,
        'concepto': p.concepto or '',
        'observaciones': p.observaciones or '',
        'pago_semanal': str(getattr(p, 'pago_semanal', 0) or 0),
        'documento_solicitud_url': p.documento_solicitud.url if p.documento_solicitud else None,
        'contrato_url': p.contrato.url if p.contrato else None,
    })


# ── Importación masiva de empleados desde Excel/CSV ───────────────────────────

_IMPORT_FIELD_MAP = {
    'nombre': 'nombre_completo',
    'nombrecompleto': 'nombre_completo',
    'licencia': 'licencia',
    'cp': 'codigo_postal', 'codigopostal': 'codigo_postal',
    'calle': 'calle',
    'numeroexterior': 'numero_exterior', 'noext': 'numero_exterior',
    'colonia': 'colonia',
    'localidad': 'localidad',
    'municipio': 'municipio',
    'estado': 'estado',
    'pais': 'pais',
    'fechadenacimiento': 'fecha_nacimiento', 'fechanacimiento': 'fecha_nacimiento',
    'nacimiento': 'fecha_nacimiento',
    'nss': 'nss',
    'rfc': 'rfc',
    'curp': 'curp',
    'departamento': 'departamento',
    'puesto': 'puesto',
    'email': 'email', 'correo': 'email',
    'telefono': 'telefono_personal', 'celular': 'telefono_personal',
}


def _norm_header(h):
    if h is None:
        return ''
    import unicodedata
    s = str(h).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.replace(' ', '').replace('-', '').replace('_', '').replace('.', '')


def _split_name(full):
    parts = (full or '').strip().split()
    if len(parts) >= 4:
        return ' '.join(parts[:-2]), ' '.join(parts[-2:])
    if len(parts) == 3:
        return parts[0], ' '.join(parts[1:])
    if len(parts) == 2:
        return parts[0], parts[1]
    return (parts[0] if parts else ''), ''


def _parse_fecha(value):
    if not value:
        return None
    if hasattr(value, 'date') and callable(value.date):
        try:
            return value.date()
        except Exception:
            pass
    if hasattr(value, 'strftime'):
        return value
    s = str(value).strip().split(' ')[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _read_table(uploaded_file):
    name = (uploaded_file.name or '').lower()
    if name.endswith('.csv'):
        import csv as _csv
        import io as _io
        text = uploaded_file.read().decode('utf-8-sig', errors='ignore')
        reader = _csv.reader(_io.StringIO(text))
        all_rows = [r for r in reader if any((c or '').strip() for c in r)]
        if not all_rows:
            return [], []
        return all_rows[0], all_rows[1:]
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl no esta instalado. pip install openpyxl")
    wb = load_workbook(filename=uploaded_file, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0])
    data = [list(r) for r in rows[1:] if any(c not in (None, '') for c in r)]
    return headers, data


@csrf_exempt
@require_http_methods(["POST"])
def api_rh_empleados_importar(request):
    err = _require_auth(request)
    if err:
        return err

    f = request.FILES.get('archivo')
    if not f:
        return JsonResponse({'error': "Falta el archivo (campo 'archivo')."}, status=400)

    try:
        headers, rows = _read_table(f)
    except Exception as exc:
        return JsonResponse({'error': 'No se pudo leer el archivo: ' + str(exc)}, status=400)

    if not headers:
        return JsonResponse({'error': 'El archivo esta vacio.'}, status=400)

    col_to_field = {}
    for idx, h in enumerate(headers):
        key = _norm_header(h)
        if key in _IMPORT_FIELD_MAP:
            col_to_field[idx] = _IMPORT_FIELD_MAP[key]

    if not col_to_field:
        return JsonResponse({
            'error': 'Ninguna columna coincide con campos conocidos. Usa encabezados como Nombre, CURP, RFC, NSS, Departamento, Puesto, etc.',
            'headers_recibidos': [str(h) for h in headers],
        }, status=400)

    creados = 0
    omitidos_dup = 0
    errores = []
    detalles_creados = []
    deptos_cache = {}
    puestos_cache = {}

    def get_or_create_depto(nombre):
        nombre = (nombre or '').strip()
        if not nombre:
            return None
        key = nombre.upper()
        if key in deptos_cache:
            return deptos_cache[key]
        obj = Departamento.objects.filter(nombre__iexact=nombre).first()
        if not obj:
            obj = Departamento.objects.create(nombre=nombre)
        deptos_cache[key] = obj
        return obj

    def get_or_create_puesto(nombre):
        nombre = (nombre or '').strip()
        if not nombre:
            return None
        key = nombre.upper()
        if key in puestos_cache:
            return puestos_cache[key]
        obj = Puesto.objects.filter(nombre__iexact=nombre).first()
        if not obj:
            obj = Puesto.objects.create(nombre=nombre)
        puestos_cache[key] = obj
        return obj

    for row_idx, row in enumerate(rows, start=2):
        data = {}
        for col_idx, field in col_to_field.items():
            try:
                val = row[col_idx]
            except IndexError:
                val = None
            if val is None:
                continue
            v = str(val).strip()
            if v:
                data[field] = v

        curp = (data.get('curp') or '').upper().strip()
        rfc = (data.get('rfc') or '').upper().strip()
        if curp and Empleado.objects.filter(curp__iexact=curp).exists():
            omitidos_dup += 1
            continue
        if rfc and Empleado.objects.filter(rfc__iexact=rfc).exists():
            omitidos_dup += 1
            continue

        try:
            with transaction.atomic():
                nombre_full = data.get('nombre_completo', '')
                nombre, apellido = _split_name(nombre_full)
                if not nombre:
                    errores.append({'fila': row_idx, 'error': 'Falta nombre'})
                    continue

                emp = Empleado(
                    nombre=nombre.upper(),
                    apellido=apellido.upper(),
                    curp=curp or None,
                    rfc=rfc or None,
                    nss=(data.get('nss') or '').strip() or None,
                    fecha_nacimiento=_parse_fecha(data.get('fecha_nacimiento')),
                    codigo_postal=(data.get('codigo_postal') or '').strip() or None,
                    colonia=(data.get('colonia') or '').strip() or None,
                    ciudad=(data.get('localidad') or '').strip() or None,
                    estado=(data.get('estado') or '').strip() or None,
                    pais=(data.get('pais') or 'Mexico').strip() or 'Mexico',
                    email=(data.get('email') or '').strip() or None,
                    telefono_personal=(data.get('telefono_personal') or '').strip() or None,
                )
                partes_dir = []
                if data.get('calle'):
                    partes_dir.append(data['calle'].strip())
                if data.get('numero_exterior'):
                    partes_dir.append('#' + data['numero_exterior'].strip())
                if data.get('municipio'):
                    partes_dir.append(data['municipio'].strip())
                if partes_dir:
                    emp.direccion = ', '.join(partes_dir)

                depto = get_or_create_depto(data.get('departamento'))
                puesto = get_or_create_puesto(data.get('puesto'))
                if depto:
                    emp.departamento = depto
                if puesto:
                    emp.puesto = puesto

                emp.save()
                creados += 1
                detalles_creados.append({
                    'fila': row_idx,
                    'id': str(emp.id),
                    'nombre': (emp.nombre + ' ' + (emp.apellido or '')).strip(),
                    'curp': emp.curp or '',
                    'rfc': emp.rfc or '',
                })
        except Exception as exc:
            errores.append({'fila': row_idx, 'error': str(exc)})

    return JsonResponse({
        'creados': creados,
        'omitidos_duplicados': omitidos_dup,
        'errores_count': len(errores),
        'errores': errores[:50],
        'detalles_creados': detalles_creados[:20],
        'total_filas': len(rows),
        'columnas_detectadas': {str(headers[i]): f for i, f in col_to_field.items()},
    })
