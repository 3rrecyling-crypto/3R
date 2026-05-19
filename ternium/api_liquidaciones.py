"""
API endpoints para Liquidaciones de Operador + export Excel de Viajes.
"""
import json
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_GET
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q, Sum

from .models import Viaje, LiquidacionOperador, LiquidacionConcepto


def _require_auth(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    return None


# Limpia descripciones quitando los códigos OR000001 / DE000001 / AM000001 que
# históricamente se inyectaban en el texto, para mostrar solo nombres de lugares.
import re as _re
_CODIGO_LUGAR_RE = _re.compile(r"\b(?:OR|DE|AM)\d{6}\b\s*-?\s*")


def _clean_descripcion(text):
    if not text:
        return text
    cleaned = _CODIGO_LUGAR_RE.sub('', text)
    # Limpiar separadores duplicados y espacios extras
    cleaned = _re.sub(r"\s+→\s+", " → ", cleaned)
    cleaned = _re.sub(r"\s+·\s+", " · ", cleaned)
    cleaned = _re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned


# ─── Serializers ───────────────────────────────────────────────────────────────

def _concepto_dict(c):
    return {
        'id': c.id,
        'tipo': c.tipo,
        'descripcion': _clean_descripcion(c.descripcion),
        'monto': str(c.monto or 0),
        'viaje_id': c.viaje_id,
        'viaje_folio': c.viaje.folio_carga or c.viaje.id_viaje if c.viaje_id else '',
    }


def _liquidacion_dict(l, full=False):
    base = {
        'id': l.id,
        'folio': l.folio,
        'operador_id': str(l.operador_id),
        'operador': l.operador.nombre_completo if l.operador_id else '',
        'fecha_inicio': str(l.fecha_inicio) if l.fecha_inicio else '',
        'fecha_fin': str(l.fecha_fin) if l.fecha_fin else '',
        'estado': l.estado,
        'estado_display': l.get_estado_display(),
        'fecha_pago': str(l.fecha_pago) if l.fecha_pago else '',
        'observaciones': l.observaciones or '',
        'creado_en': l.creado_en.isoformat() if l.creado_en else '',
        'total_viajes': str(l.total_viajes),
        'total_extras': str(l.total_extras),
        'total_descuentos': str(l.total_descuentos),
        'total_pagar': str(l.total_pagar),
        'conceptos_count': l.conceptos.count(),
    }
    if full:
        base['operador_data'] = {
            'id': str(l.operador.id),
            'nombre': l.operador.nombre_completo,
            'rfc': getattr(l.operador, 'rfc', '') or '',
            'numero_licencia': getattr(l.operador, 'numero_licencia', '') or getattr(l.operador, 'licencia', '') or '',
        }
        base['conceptos'] = [_concepto_dict(c) for c in l.conceptos.select_related('viaje').all()]
    return base


# ─── CRUD de Liquidaciones ─────────────────────────────────────────────────────

def _liquidaciones_queryset_from_request(request):
    """Construye el QuerySet de Liquidaciones aplicando todos los filtros estándar."""
    qs = LiquidacionOperador.objects.select_related('operador').order_by('-creado_en')

    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado.upper())

    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(folio__icontains=search) |
            Q(operador__nombre__icontains=search) |
            Q(operador__apellido__icontains=search)
        )

    # Rango de fechas (periodo de la liquidación: cualquier solapamiento con el rango)
    fi = request.GET.get('fecha_desde', '')
    ff = request.GET.get('fecha_hasta', '')
    if fi:
        try: qs = qs.filter(fecha_fin__gte=datetime.strptime(fi, '%Y-%m-%d').date())
        except ValueError: pass
    if ff:
        try: qs = qs.filter(fecha_inicio__lte=datetime.strptime(ff, '%Y-%m-%d').date())
        except ValueError: pass

    # Operador específico
    operador_id = request.GET.get('operador_id', '')
    if operador_id:
        qs = qs.filter(operador_id=operador_id)

    # Filtros por origen / destino: liquidaciones que tengan conceptos VIAJE con viaje en ese lugar
    origen_id = request.GET.get('origen_id', '')
    if origen_id:
        try:
            qs = qs.filter(conceptos__viaje__origen_id=int(origen_id)).distinct()
        except (ValueError, TypeError): pass
    destino_id = request.GET.get('destino_id', '')
    if destino_id:
        try:
            qs = qs.filter(conceptos__viaje__destino_id=int(destino_id)).distinct()
        except (ValueError, TypeError): pass

    return qs


@require_GET
def api_liquidaciones_list(request):
    """GET /api/liquidaciones/?search=&estado=&fecha_desde=&fecha_hasta=&operador_id=&origen_id=&destino_id=&page=&page_size="""
    err = _require_auth(request)
    if err:
        return err
    qs = _liquidaciones_queryset_from_request(request)
    try:
        page_size = min(int(request.GET.get('page_size', 20)), 100)
        page = max(int(request.GET.get('page', 1)), 1)
    except ValueError:
        page_size, page = 20, 1
    total = qs.count()
    offset = (page - 1) * page_size
    return JsonResponse({
        'count': total,
        'results': [_liquidacion_dict(l) for l in qs[offset:offset + page_size]],
    })


@csrf_exempt
@require_http_methods(["POST"])
def api_liquidacion_crear(request):
    """POST /api/liquidaciones/crear/  — body JSON con operador_id, fecha_inicio, fecha_fin.
    Auto-genera conceptos VIAJE por cada viaje en el rango."""
    err = _require_auth(request)
    if err:
        return err
    try:
        body = json.loads(request.body or '{}')
        from RH.models import Empleado
        operador_id = body.get('operador_id')
        if not operador_id:
            return JsonResponse({'error': 'Falta el operador.'}, status=400)
        operador = get_object_or_404(Empleado, pk=operador_id)
        fi_str = (body.get('fecha_inicio') or '').strip()
        ff_str = (body.get('fecha_fin') or '').strip()
        if not fi_str or not ff_str:
            return JsonResponse({'error': 'Indica periodo (fecha_inicio y fecha_fin).'}, status=400)
        try:
            fi = datetime.strptime(fi_str, '%Y-%m-%d').date()
            ff = datetime.strptime(ff_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido (AAAA-MM-DD).'}, status=400)
        if ff < fi:
            return JsonResponse({'error': 'fecha_fin no puede ser anterior a fecha_inicio.'}, status=400)

        with transaction.atomic():
            l = LiquidacionOperador.objects.create(
                operador=operador,
                fecha_inicio=fi,
                fecha_fin=ff,
                estado='BORRADOR',
                observaciones=(body.get('observaciones') or '').strip() or None,
                creado_por=request.user if request.user.is_authenticated else None,
            )
            # Auto-jala viajes del operador en el rango (excluyendo CANCELADO)
            viajes_qs = Viaje.objects.filter(
                operador=operador,
                fecha_viaje__gte=fi,
                fecha_viaje__lte=ff,
            ).exclude(estado='CANCELADO').order_by('fecha_viaje')
            for v in viajes_qs:
                LiquidacionConcepto.objects.create(
                    liquidacion=l,
                    tipo='VIAJE',
                    descripcion=(
                        f"Viaje {v.folio_carga or v.id_viaje} · "
                        f"{v.origen.nombre if v.origen_id else ''} → "
                        f"{v.destino.nombre if v.destino_id else ''}"
                    ),
                    monto=v.sueldo_operador or Decimal('0'),
                    viaje=v,
                )
        return JsonResponse(_liquidacion_dict(l, full=True), status=201)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'{type(exc).__name__}: {exc}'}, status=400)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "DELETE"])
def api_liquidacion_detail(request, pk):
    err = _require_auth(request)
    if err:
        return err
    l = get_object_or_404(
        LiquidacionOperador.objects.select_related('operador').prefetch_related('conceptos__viaje'),
        pk=pk,
    )
    if request.method == 'GET':
        return JsonResponse(_liquidacion_dict(l, full=True))
    if request.method == 'DELETE':
        l.delete()
        return JsonResponse({'ok': True})
    # PATCH
    try:
        body = json.loads(request.body or '{}')
        for k in ('estado', 'observaciones'):
            if k in body:
                val = (body[k] or '').strip()
                setattr(l, k, val or None if k != 'estado' else (val or l.estado))
        if 'fecha_pago' in body:
            try:
                l.fecha_pago = datetime.strptime(body['fecha_pago'], '%Y-%m-%d').date() if body['fecha_pago'] else None
            except ValueError:
                pass
        if 'fecha_inicio' in body and body['fecha_inicio']:
            try: l.fecha_inicio = datetime.strptime(body['fecha_inicio'], '%Y-%m-%d').date()
            except ValueError: pass
        if 'fecha_fin' in body and body['fecha_fin']:
            try: l.fecha_fin = datetime.strptime(body['fecha_fin'], '%Y-%m-%d').date()
            except ValueError: pass
        l.save()
        return JsonResponse(_liquidacion_dict(l, full=True))
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_liquidacion_concepto_crear(request, liquidacion_id):
    err = _require_auth(request)
    if err:
        return err
    l = get_object_or_404(LiquidacionOperador, pk=liquidacion_id)
    try:
        body = json.loads(request.body or '{}')
    except ValueError:
        body = {}
    descripcion = (body.get('descripcion') or '').strip()
    if not descripcion:
        return JsonResponse({'error': 'La descripción es obligatoria.'}, status=400)
    try:
        monto = Decimal(str(body.get('monto') or '0'))
    except (InvalidOperation, TypeError):
        return JsonResponse({'error': 'Monto inválido.'}, status=400)
    c = LiquidacionConcepto.objects.create(
        liquidacion=l,
        tipo=body.get('tipo', 'EXTRA'),
        descripcion=descripcion,
        monto=monto,
        viaje_id=body.get('viaje_id') or None,
    )
    return JsonResponse(_concepto_dict(c), status=201)


@csrf_exempt
@require_http_methods(["PATCH", "DELETE"])
def api_liquidacion_concepto_detail(request, pk):
    err = _require_auth(request)
    if err:
        return err
    c = get_object_or_404(LiquidacionConcepto, pk=pk)
    if request.method == 'DELETE':
        c.delete()
        return JsonResponse({'ok': True})
    try:
        body = json.loads(request.body or '{}')
    except ValueError:
        body = {}
    if 'descripcion' in body:
        c.descripcion = (body['descripcion'] or '').strip() or c.descripcion
    if 'tipo' in body and body['tipo'] in ('VIAJE', 'EXTRA', 'DESCUENTO'):
        c.tipo = body['tipo']
    if 'monto' in body:
        try: c.monto = Decimal(str(body['monto']))
        except (InvalidOperation, TypeError): pass
    c.save()
    return JsonResponse(_concepto_dict(c))


# ─── Helper: viajes pendientes de liquidar (no asignados a ninguna VIAJE concepto) ─

@require_GET
def api_viajes_pendientes_liquidar(request):
    """GET /api/liquidaciones/viajes-pendientes/?operador_id=&fecha_inicio=&fecha_fin=
    Devuelve viajes del operador en el rango que NO están en ninguna liquidación activa."""
    err = _require_auth(request)
    if err:
        return err
    operador_id = request.GET.get('operador_id')
    fi_str = request.GET.get('fecha_inicio', '')
    ff_str = request.GET.get('fecha_fin', '')
    if not operador_id:
        return JsonResponse({'results': []})

    qs = Viaje.objects.filter(operador_id=operador_id).exclude(estado='CANCELADO')
    if fi_str:
        try: qs = qs.filter(fecha_viaje__gte=datetime.strptime(fi_str, '%Y-%m-%d').date())
        except ValueError: pass
    if ff_str:
        try: qs = qs.filter(fecha_viaje__lte=datetime.strptime(ff_str, '%Y-%m-%d').date())
        except ValueError: pass
    qs = qs.order_by('fecha_viaje')
    return JsonResponse({
        'results': [
            {
                'id': v.id,
                'id_viaje': v.id_viaje,
                'folio_carga': v.folio_carga or '',
                'fecha_viaje': str(v.fecha_viaje) if v.fecha_viaje else '',
                'origen': v.origen.nombre if v.origen_id else '',
                'destino': v.destino.nombre if v.destino_id else '',
                'sueldo_operador': str(v.sueldo_operador or 0),
                'estado': v.estado,
            } for v in qs
        ]
    })


# ─── Excel export de Viajes ────────────────────────────────────────────────────

@require_GET
def api_viajes_exportar_excel(request):
    """GET /api/viajes/exportar-excel/?<filtros>
    Devuelve un .xlsx con todos los viajes filtrados."""
    err = _require_auth(request)
    if err:
        return err
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return JsonResponse({'error': 'openpyxl no está instalado. pip install openpyxl'}, status=500)

    qs = Viaje.objects.select_related('operador', 'unidad', 'origen', 'destino', 'empresa').order_by('-numero_viaje')
    # Filtros opcionales
    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado.upper())
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(id_viaje__icontains=search) | Q(folio_carga__icontains=search) |
            Q(operador__nombre__icontains=search) | Q(operador__apellido__icontains=search) |
            Q(unidad__internal_id__icontains=search)
        )
    fi = request.GET.get('fecha_inicio', '') or request.GET.get('fecha_desde', '')
    ff = request.GET.get('fecha_fin', '')    or request.GET.get('fecha_hasta', '')
    if fi:
        try: qs = qs.filter(fecha_viaje__gte=datetime.strptime(fi, '%Y-%m-%d').date())
        except ValueError: pass
    if ff:
        try: qs = qs.filter(fecha_viaje__lte=datetime.strptime(ff, '%Y-%m-%d').date())
        except ValueError: pass
    origen_id = (request.GET.get('origen_id') or '').strip()
    if origen_id:
        try: qs = qs.filter(origen_id=int(origen_id))
        except (TypeError, ValueError): pass
    destino_id = (request.GET.get('destino_id') or '').strip()
    if destino_id:
        try: qs = qs.filter(destino_id=int(destino_id))
        except (TypeError, ValueError): pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    # Estilos — paleta neutra estándar (gris/blanco/negro)
    header_font = Font(bold=True, color='000000', size=10)
    header_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='9CA3AF')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    headers = [
        '#', 'ID Viaje', 'Folio Carga', 'Fecha',
        'Operador', 'RFC Operador', 'No. Licencia',
        'Unidad', 'Placa', 'Modelo',
        'Origen', 'ID Origen', 'Destino', 'ID Destino', 'Mismo lugar',
        'Estado', 'Kms totales', 'Sueldo operador (MXN)',
        'Empresa', 'Observaciones', 'Creado en',
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = box

    total_sueldos = Decimal('0')
    total_kms = Decimal('0')
    for v in qs:
        op = v.operador
        u = v.unidad
        kms = sum((p.kms or 0) for p in v.paradas.all())
        try:
            kms_val = Decimal(str(kms))
        except (InvalidOperation, TypeError):
            kms_val = Decimal('0')
        sueldo = v.sueldo_operador or Decimal('0')
        total_sueldos += sueldo
        total_kms += kms_val
        ws.append([
            v.numero_viaje or v.id,
            v.id_viaje,
            v.folio_carga or '',
            v.fecha_viaje.strftime('%d/%m/%Y') if v.fecha_viaje else '',
            op.nombre_completo if op else '',
            getattr(op, 'rfc', '') or '',
            getattr(op, 'numero_licencia', '') or getattr(op, 'licencia', '') or '',
            u.internal_id if u else '',
            u.license_plate or '' if u else '',
            u.make_model or '' if u else '',
            v.origen.nombre if v.origen_id else '',
            v.origen.id_ubicacion if v.origen_id else '',
            v.destino.nombre if v.destino_id else '',
            v.destino.id_ubicacion if v.destino_id else '',
            'Sí' if v.mismo_origen_destino else 'No',
            v.get_estado_display(),
            float(kms_val),
            float(sueldo),
            v.empresa.nombre if v.empresa_id else '',
            (v.observaciones or '').replace('\n', ' '),
            v.creado_en.strftime('%d/%m/%Y %H:%M') if v.creado_en else '',
        ])

    # Fila de totales (paleta neutra)
    row_total = ws.max_row + 1
    ws.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row_total, column=17, value=float(total_kms)).font = Font(bold=True)
    ws.cell(row=row_total, column=18, value=float(total_sueldos)).font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_total, column=col_idx)
        cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        cell.border = box

    # Anchos de columna razonables
    widths = [6, 12, 13, 11, 28, 14, 14, 12, 12, 22, 22, 11, 22, 11, 10, 14, 13, 18, 18, 30, 16]
    for i, w in enumerate(widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

    # Aplicar bordes, zebra y formato a todas las celdas de datos
    for r in range(2, row_total):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (r % 2) == 0:
                cell.fill = zebra_fill
    # Format numerico
    for r in range(2, row_total + 1):
        ws.cell(row=r, column=17).number_format = '#,##0.000'
        ws.cell(row=r, column=18).number_format = '"$"#,##0.00'

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="viajes-{date.today()}.xlsx"'
    return resp


# ─── PDF profesional de Liquidación ────────────────────────────────────────────

@require_GET
def api_liquidacion_pdf(request, pk):
    """GET /api/liquidaciones/<pk>/pdf/ — Devuelve PDF imprimible de la liquidación."""
    err = _require_auth(request)
    if err:
        return err
    l = get_object_or_404(
        LiquidacionOperador.objects.select_related('operador').prefetch_related('conceptos__viaje'),
        pk=pk,
    )

    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
        )
    except ImportError:
        return JsonResponse({'error': 'reportlab no está instalado. pip install reportlab'}, status=500)

    import os
    from django.conf import settings as dj_settings

    # ─── Paleta ───────────────────────────────────────────────────────────
    BRAND       = colors.HexColor('#0a7d3a')   # verde 3R (logo / banda)
    GRAY_BAND   = colors.HexColor('#d4d4d4')   # título de sección
    GRAY_HEADER = colors.HexColor('#ebebeb')   # header de columnas
    GRAY_LINE   = colors.HexColor('#b8b8b8')
    GRAY_LINE_LT = colors.HexColor('#d8d8d8')
    GRAY_LIGHT  = colors.HexColor('#fafafa')
    GREEN_BG    = colors.HexColor('#ecfdf5')
    RED_BG      = colors.HexColor('#fef2f2')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=10*mm, bottomMargin=14*mm,
        title=f"Liquidación {l.folio}",
        author="3R Recycling",
    )

    styles = getSampleStyleSheet()
    s_micro = ParagraphStyle('micro', parent=styles['Normal'], fontSize=6.5, leading=8.5,
                              fontName='Helvetica', textColor=colors.HexColor('#555555'))
    s_small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8.5, leading=11,
                              fontName='Helvetica', textColor=colors.black)
    s_section = ParagraphStyle('section', parent=styles['Normal'], fontSize=9.5, leading=12,
                                fontName='Helvetica-Bold', textColor=colors.black, alignment=TA_CENTER)
    s_cellh = ParagraphStyle('cellh', parent=styles['Normal'], fontSize=7, leading=9,
                              fontName='Helvetica-Bold', textColor=colors.black, alignment=TA_CENTER)
    s_cell  = ParagraphStyle('cell',  parent=styles['Normal'], fontSize=8, leading=10,
                              fontName='Helvetica', textColor=colors.black, alignment=TA_LEFT, wordWrap='CJK')
    s_cellc = ParagraphStyle('cellc', parent=styles['Normal'], fontSize=8, leading=10,
                              fontName='Helvetica', textColor=colors.black, alignment=TA_CENTER, wordWrap='CJK')
    s_cellr = ParagraphStyle('cellr', parent=styles['Normal'], fontSize=8, leading=10,
                              fontName='Helvetica', textColor=colors.black, alignment=TA_RIGHT)

    story = []

    # ─── Datos empresa ─────────────────────────────────────────────────────
    empresa_nombre = '3R recycling'
    empresa_rfc = 'RRE1801236N0'
    empresa_dir = 'GARZA SADA #3921, Col. Contry\nCP 64860, Monterrey, Nuevo León, México'

    # ─── Header con logo + bloque derecho ──────────────────────────────────
    logo_path = os.path.join(dj_settings.BASE_DIR, 'static', 'r3_recycling', 'img', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=32*mm, height=32*mm)
            logo.hAlign = 'CENTER'
        except Exception:
            logo = Spacer(32*mm, 32*mm)
    else:
        logo = Spacer(32*mm, 32*mm)

    empresa_block = Paragraph(
        f'<para align="center" leading="14">'
        f'<font color="#000000" size="16"><b>{empresa_nombre}</b></font><br/>'
        f'<font color="#555555" size="8"><b>RFC:</b> {empresa_rfc}</font><br/>'
        f'<font color="#000000" size="8.5">{empresa_dir.replace(chr(10), "<br/>")}</font>'
        f'</para>', s_small,
    )

    label_st = ParagraphStyle('hl', parent=s_cellc, fontSize=7, leading=8.5,
                               fontName='Helvetica-Bold', textColor=colors.HexColor('#444444'))
    value_st = ParagraphStyle('hv', parent=s_cellc, fontSize=10, leading=12,
                               fontName='Helvetica-Bold', textColor=colors.black)

    doc_block_data = [
        [Paragraph('LIQUIDACIÓN', label_st)],
        [Paragraph(l.folio, value_st)],
        [Paragraph('PERIODO', label_st)],
        [Paragraph(f"{l.fecha_inicio.strftime('%d/%m/%Y')} – {l.fecha_fin.strftime('%d/%m/%Y')}", value_st)],
        [Paragraph('ESTADO', label_st)],
        [Paragraph(l.get_estado_display(), value_st)],
        [Paragraph('FECHA DE PAGO', label_st)],
        [Paragraph(l.fecha_pago.strftime('%d/%m/%Y') if l.fecha_pago else '—', value_st)],
    ]
    doc_block = Table(doc_block_data, colWidths=[52*mm])
    doc_block.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), GRAY_HEADER),
        ('BACKGROUND', (0,2), (-1,2), GRAY_HEADER),
        ('BACKGROUND', (0,4), (-1,4), GRAY_HEADER),
        ('BACKGROUND', (0,6), (-1,6), GRAY_HEADER),
        ('BOX', (0,0), (-1,-1), 0.5, GRAY_LINE),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))

    head = Table([[logo, empresa_block, doc_block]], colWidths=[40*mm, 94*mm, 52*mm])
    head.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(head)
    story.append(Spacer(1, 4))

    divider = Table([['']], colWidths=[186*mm], rowHeights=[0.4])
    divider.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), GRAY_LINE)]))
    story.append(divider)
    story.append(Spacer(1, 6))

    PAGE_W = 186*mm
    def section_title(text):
        t = Table([[Paragraph(text, s_section)]], colWidths=[PAGE_W])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), GRAY_BAND),
            ('BOX', (0,0), (-1,-1), 0.5, GRAY_LINE),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        return t

    # ─── Datos del operador ────────────────────────────────────────────────
    op = l.operador
    story.append(section_title('Datos del Operador'))
    op_table = Table([
        [Paragraph('Nombre', s_cellh), Paragraph('RFC', s_cellh), Paragraph('No. Licencia', s_cellh), Paragraph('Empleado #', s_cellh)],
        [
            Paragraph(op.nombre_completo, s_cell),
            Paragraph(getattr(op, 'rfc', '') or '—', s_cellc),
            Paragraph(getattr(op, 'numero_licencia', '') or getattr(op, 'licencia', '') or '—', s_cellc),
            Paragraph(str(getattr(op, 'numero_empleado', '') or op.id), s_cellc),
        ],
    ], colWidths=[80*mm, 36*mm, 40*mm, 30*mm])
    op_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), GRAY_HEADER),
        ('LINEBELOW', (0,0), (-1,0), 0.5, GRAY_LINE),
        ('BOX', (0,0), (-1,-1), 0.5, GRAY_LINE),
        ('INNERGRID', (0,0), (-1,-1), 0.25, GRAY_LINE_LT),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(op_table)
    story.append(Spacer(1, 8))

    # ─── Helper para tablas de conceptos ───────────────────────────────────
    def conceptos_table(rows_data, bg_color=None):
        rows = [[
            Paragraph('Concepto', s_cellh),
            Paragraph('Folio Viaje', s_cellh),
            Paragraph('Importe (MXN)', s_cellh),
        ]]
        rows.extend(rows_data)
        t = Table(rows, colWidths=[110*mm, 36*mm, 40*mm], repeatRows=1)
        style = [
            ('BACKGROUND', (0,0), (-1,0), GRAY_HEADER),
            ('LINEBELOW', (0,0), (-1,0), 0.5, GRAY_LINE),
            ('BOX', (0,0), (-1,-1), 0.5, GRAY_LINE),
            ('INNERGRID', (0,0), (-1,-1), 0.25, GRAY_LINE_LT),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, GRAY_LIGHT]),
        ]
        if bg_color:
            style.append(('BACKGROUND', (0,1), (-1,-1), bg_color))
            # Remove zebra so color stays solid
            style = [s for s in style if s[0] != 'ROWBACKGROUNDS']
        t.setStyle(TableStyle(style))
        return t

    # Agrupamos conceptos
    viajes_c = [c for c in l.conceptos.all() if c.tipo == 'VIAJE']
    bonos_c  = [c for c in l.conceptos.all() if c.tipo == 'EXTRA']
    deduc_c  = [c for c in l.conceptos.all() if c.tipo == 'DESCUENTO']

    # ─── Viajes ────────────────────────────────────────────────────────────
    story.append(section_title('Viajes de la Bitácora'))
    if viajes_c:
        rows = []
        total_v = Decimal('0')
        for c in viajes_c:
            folio = c.viaje.folio_carga or c.viaje.id_viaje if c.viaje_id else '—'
            try:
                m = Decimal(str(c.monto or 0))
            except (InvalidOperation, TypeError):
                m = Decimal('0')
            total_v += m
            rows.append([
                Paragraph(_clean_descripcion(c.descripcion), s_cell),
                Paragraph(folio, s_cellc),
                Paragraph(f'${m:,.2f}', s_cellr),
            ])
        rows.append([
            Paragraph('<b>Subtotal Viajes</b>', s_cell),
            '',
            Paragraph(f'<b>${total_v:,.2f}</b>', s_cellr),
        ])
        t = conceptos_table(rows)
        # Resaltar fila de subtotal
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,-1), (-1,-1), GRAY_HEADER),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('LINEABOVE', (0,-1), (-1,-1), 0.5, GRAY_LINE),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('<font color="#888888" size="9"><i>Sin viajes en el periodo.</i></font>', s_small))
    story.append(Spacer(1, 6))

    # ─── Bonos ─────────────────────────────────────────────────────────────
    if bonos_c:
        story.append(section_title('Bonos'))
        rows = []
        total_b = Decimal('0')
        for c in bonos_c:
            try: m = Decimal(str(c.monto or 0))
            except (InvalidOperation, TypeError): m = Decimal('0')
            total_b += m
            folio = (c.viaje.folio_carga or c.viaje.id_viaje) if c.viaje_id else '—'
            rows.append([
                Paragraph(_clean_descripcion(c.descripcion), s_cell),
                Paragraph(folio, s_cellc),
                Paragraph(f'+${m:,.2f}', s_cellr),
            ])
        rows.append([
            Paragraph('<b>Subtotal Bonos</b>', s_cell), '',
            Paragraph(f'<b>+${total_b:,.2f}</b>', s_cellr),
        ])
        t = conceptos_table(rows)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,-1), (-1,-1), GREEN_BG),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('LINEABOVE', (0,-1), (-1,-1), 0.5, GRAY_LINE),
            ('TEXTCOLOR', (2,1), (2,-1), colors.HexColor('#047857')),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

    # ─── Deducciones ───────────────────────────────────────────────────────
    if deduc_c:
        story.append(section_title('Deducciones'))
        rows = []
        total_d = Decimal('0')
        for c in deduc_c:
            try: m = Decimal(str(c.monto or 0))
            except (InvalidOperation, TypeError): m = Decimal('0')
            total_d += m
            folio = (c.viaje.folio_carga or c.viaje.id_viaje) if c.viaje_id else '—'
            rows.append([
                Paragraph(_clean_descripcion(c.descripcion), s_cell),
                Paragraph(folio, s_cellc),
                Paragraph(f'-${m:,.2f}', s_cellr),
            ])
        rows.append([
            Paragraph('<b>Subtotal Deducciones</b>', s_cell), '',
            Paragraph(f'<b>-${total_d:,.2f}</b>', s_cellr),
        ])
        t = conceptos_table(rows)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,-1), (-1,-1), RED_BG),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('LINEABOVE', (0,-1), (-1,-1), 0.5, GRAY_LINE),
            ('TEXTCOLOR', (2,1), (2,-1), colors.HexColor('#b91c1c')),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

    # ─── Resumen / Total a pagar ──────────────────────────────────────────
    story.append(section_title('Resumen'))
    try:
        tv = Decimal(str(l.total_viajes or 0))
        tb = Decimal(str(l.total_extras or 0))
        td = Decimal(str(l.total_descuentos or 0))
        tp = Decimal(str(l.total_pagar or 0))
    except (InvalidOperation, TypeError):
        tv = tb = td = tp = Decimal('0')

    resumen = Table([
        [Paragraph('Subtotal Viajes', s_cell), Paragraph(f'${tv:,.2f}', s_cellr)],
        [Paragraph('Bonos', s_cell), Paragraph(f'+${tb:,.2f}', s_cellr)],
        [Paragraph('Deducciones', s_cell), Paragraph(f'-${td:,.2f}', s_cellr)],
        [Paragraph('<font size="11"><b>TOTAL A PAGAR</b></font>', s_cell),
         Paragraph(f'<font size="13" color="#047857"><b>${tp:,.2f}</b></font>', s_cellr)],
    ], colWidths=[146*mm, 40*mm])
    resumen.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, GRAY_LINE),
        ('INNERGRID', (0,0), (-1,-1), 0.25, GRAY_LINE_LT),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,1), (-1,1), GREEN_BG),
        ('BACKGROUND', (0,2), (-1,2), RED_BG),
        ('BACKGROUND', (0,3), (-1,3), GRAY_HEADER),
        ('FONTNAME', (0,3), (-1,3), 'Helvetica-Bold'),
        ('LINEABOVE', (0,3), (-1,3), 1.2, BRAND),
        ('TEXTCOLOR', (1,1), (1,1), colors.HexColor('#047857')),
        ('TEXTCOLOR', (1,2), (1,2), colors.HexColor('#b91c1c')),
    ]))
    story.append(resumen)
    story.append(Spacer(1, 12))

    # ─── Observaciones ────────────────────────────────────────────────────
    if l.observaciones:
        story.append(section_title('Observaciones'))
        obs_table = Table([[Paragraph(l.observaciones.replace(chr(10), '<br/>'), s_small)]],
                          colWidths=[PAGE_W])
        obs_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.25, GRAY_LINE),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(obs_table)
        story.append(Spacer(1, 12))

    # ─── Firmas ───────────────────────────────────────────────────────────
    firmas = Table([
        [' ', ' ', ' '],
        ['_______________________', '_______________________', '_______________________'],
        ['Operador', 'Autorizó', 'Pagó'],
        [op.nombre_completo, '', ''],
    ], colWidths=[60*mm, 60*mm, 60*mm])
    firmas.setStyle(TableStyle([
        ('FONTSIZE', (0,1), (-1,1), 7),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'),
        ('FONTSIZE', (0,2), (-1,2), 7),
        ('FONTSIZE', (0,3), (-1,3), 6.5),
        ('TEXTCOLOR', (0,3), (-1,3), colors.HexColor('#666666')),
        ('TOPPADDING', (0,0), (-1,0), 16),
    ]))
    story.append(firmas)

    # ─── Pie ──────────────────────────────────────────────────────────────
    from datetime import datetime as _dt
    story.append(Spacer(1, 10))
    pie_div = Table([['']], colWidths=[PAGE_W], rowHeights=[0.3])
    pie_div.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), GRAY_LINE)]))
    story.append(pie_div)
    story.append(Spacer(1, 3))
    pie = Paragraph(
        f'<para align="center"><font color="#666666" size="6.5">'
        f'<b>3R Recycling</b> · Liquidación de Operador · Generado el {_dt.now().strftime("%d/%m/%Y a las %H:%M")} · Folio interno: {l.folio}'
        f'</font></para>', s_micro,
    )
    story.append(pie)

    doc.build(story)
    buffer.seek(0)
    resp = HttpResponse(buffer.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="liquidacion-{l.folio}.pdf"'
    return resp


# ─── Excel export de Liquidaciones (lista + detalle de conceptos) ──────────────

@require_GET
def api_liquidaciones_exportar_excel(request):
    """GET /api/liquidaciones/exportar-excel/?<mismos filtros que el listado>
    Genera un .xlsx con 2 hojas: Resumen de liquidaciones + Detalle de conceptos."""
    err = _require_auth(request)
    if err:
        return err
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return JsonResponse({'error': 'openpyxl no está instalado. pip install openpyxl'}, status=500)

    qs = _liquidaciones_queryset_from_request(request).prefetch_related('conceptos__viaje')

    # ─── Paleta neutra estándar (gris/blanco/negro) ───────────────────────
    header_font = Font(bold=True, color='000000', size=10)
    header_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')   # gris neutro
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='9CA3AF')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')     # gris muy suave
    zebra_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    wb = Workbook()

    # ─── Hoja 1: Viajes con las columnas pedidas ──────────────────────────
    # FECHA DE VIAJE · NÚMERO DE VIAJE · NOMBRE DEL OPERADOR ·
    # NÚM. CARTA TRASLADO · ORIGEN · DESTINO · FLETE
    ws = wb.active
    ws.title = "Viajes"
    headers = [
        'FECHA DE VIAJE', 'NÚMERO DE VIAJE', 'NOMBRE DEL OPERADOR',
        'NÚM. CARTA TRASLADO', 'ORIGEN', 'DESTINO', 'FLETE',
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = box

    total_flete = Decimal('0')
    row_count = 0
    for l in qs:
        for c in l.conceptos.filter(tipo='VIAJE').select_related('viaje'):
            v = c.viaje
            if not v:
                continue
            try:
                m = Decimal(str(c.monto or 0))
            except (InvalidOperation, TypeError):
                m = Decimal('0')
            total_flete += m
            row_count += 1
            ws.append([
                v.fecha_viaje.strftime('%d/%m/%Y') if v.fecha_viaje else '',
                v.numero_viaje or v.id,
                l.operador.nombre_completo if l.operador_id else '',
                v.id_viaje,
                v.origen.nombre if v.origen_id else '',
                v.destino.nombre if v.destino_id else '',
                float(m),
            ])

    if row_count == 0:
        ws.append(['Sin viajes en los filtros aplicados', '', '', '', '', '', ''])

    row_total = ws.max_row + 1
    ws.cell(row=row_total, column=6, value='TOTAL FLETE').font = Font(bold=True)
    ws.cell(row=row_total, column=6).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=row_total, column=7, value=float(total_flete)).font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row_total, column=c)
        cell.fill = total_fill
        cell.border = box

    widths = [16, 18, 36, 22, 28, 28, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for r in range(2, row_total):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (r % 2) == 0:
                cell.fill = zebra_fill
    for r in range(2, row_total + 1):
        ws.cell(row=r, column=7).number_format = '"$"#,##0.00'
    ws.freeze_panes = 'A2'

    # ─── Hoja 2: Resumen de Liquidaciones (auxiliar) ──────────────────────
    ws2 = wb.create_sheet("Resumen Liquidaciones")
    headers2 = [
        '#', 'Folio', 'Operador', 'Periodo desde', 'Periodo hasta',
        'Estado', 'Fecha de pago', 'Subtotal Viajes', 'Bonos',
        'Deducciones', 'TOTAL A PAGAR', '# Viajes',
    ]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = box

    total_general = Decimal('0')
    for idx, l in enumerate(qs, start=1):
        try:
            tv = Decimal(str(l.total_viajes or 0))
            tb = Decimal(str(l.total_extras or 0))
            td = Decimal(str(l.total_descuentos or 0))
            tp = Decimal(str(l.total_pagar or 0))
        except (InvalidOperation, TypeError):
            tv = tb = td = tp = Decimal('0')
        total_general += tp
        n_viajes = l.conceptos.filter(tipo='VIAJE').count()
        ws2.append([
            idx, l.folio,
            l.operador.nombre_completo if l.operador_id else '',
            l.fecha_inicio.strftime('%d/%m/%Y') if l.fecha_inicio else '',
            l.fecha_fin.strftime('%d/%m/%Y') if l.fecha_fin else '',
            l.get_estado_display(),
            l.fecha_pago.strftime('%d/%m/%Y') if l.fecha_pago else '',
            float(tv), float(tb), float(td), float(tp), n_viajes,
        ])

    rt2 = ws2.max_row + 1
    ws2.cell(row=rt2, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=rt2, column=11, value=float(total_general)).font = Font(bold=True)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=rt2, column=c)
        cell.fill = total_fill
        cell.border = box

    widths2 = [5, 12, 32, 13, 13, 12, 13, 14, 12, 14, 16, 10]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    for r in range(2, rt2):
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = box
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (r % 2) == 0:
                cell.fill = zebra_fill
    for r in range(2, rt2 + 1):
        for col in (8, 9, 10, 11):
            ws2.cell(row=r, column=col).number_format = '"$"#,##0.00'
    ws2.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="liquidaciones-{date.today()}.xlsx"'
    return resp
