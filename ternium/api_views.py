import datetime as dt
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction, IntegrityError
from .models import Empresa, Material, Lugar, LineaTransporte, Operador, Unidad, Contenedor
import json
from django.shortcuts import get_object_or_404
from .models import (
    Remision, Empresa, Material, Lugar, ConfiguracionManifiesto,
    HistorialRemision, EvidenciaRemision, DetalleRemision,
    ConfiguracionAlertaMerma, DestinatarioAlertaMerma, RemisionAlertaMermaLog,
    Alerta, ChatMensaje, InventarioPatio,
)
import decimal

import datetime
from django.utils import timezone
from django.core.files.storage import default_storage
from .models import Remision

# =========================================================================
# UTILIDADES DE NEGOCIO: MEDLINE
# =========================================================================

from django.db.models import F
# Importa tu modelo de inventario real (ejemplo: from .models import Inventario)

# =========================================================================
# IMPORTS ADICIONALES PARA EXPORTS Y ACCIONES DE REMISIONES
# =========================================================================
import io
import os
import zipfile
import urllib.request
from io import BytesIO
import boto3
from botocore.exceptions import BotoCoreError, NoCredentialsError
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenXLImage
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.views.decorators.http import require_POST, require_http_methods
from django.conf import settings
from .models import PrecioMedline

def _update_inventory_from_remision(remision, revert=False):
    """
    Actualiza (o revierte) los inventarios basándose en los detalles de la remisión.
    """
    # Factor de multiplicación: 1 para sumar al destino/restar al origen, -1 para revertir
    factor = -1 if revert else 1

    for detalle in remision.detalles.all():
        if not detalle.material:
            continue
            
        peso_carga = float(detalle.peso_ld or 0)
        peso_descarga = float(detalle.peso_dlv or 0)
        peso_rechazado = float(detalle.peso_rechazado or 0)

        # 1. DESCONTAR DEL ORIGEN (Carga)
        # Solo descontamos si el origen es un patio propio (donde llevamos inventario)
        if remision.origen and getattr(remision.origen, 'es_patio', False) and peso_carga > 0:
            inv, _ = InventarioPatio.objects.get_or_create(patio=remision.origen, material=detalle.material)
            cant = decimal.Decimal(str(peso_carga))
            actual = decimal.Decimal(inv.cantidad)
            inv.cantidad = actual + cant if revert else actual - cant  # revert devuelve al patio
            inv.save()

        # 2. SUMAR AL DESTINO (Descarga)
        if remision.destino and getattr(remision.destino, 'es_patio', False) and peso_descarga > 0:
            inv, _ = InventarioPatio.objects.get_or_create(patio=remision.destino, material=detalle.material)
            cant = decimal.Decimal(str(peso_descarga))
            actual = decimal.Decimal(inv.cantidad)
            inv.cantidad = actual - cant if revert else actual + cant  # revert quita lo sumado
            inv.save()

        # 3. SUMAR RECHAZOS (Regresos a Patio)
        if detalle.patio_rechazo and getattr(detalle.patio_rechazo, 'es_patio', False) and peso_rechazado > 0:
            inv, _ = InventarioPatio.objects.get_or_create(patio=detalle.patio_rechazo, material=detalle.material)
            cant = decimal.Decimal(str(peso_rechazado))
            actual = decimal.Decimal(inv.cantidad)
            inv.cantidad = actual - cant if revert else actual + cant  # revert quita el rechazo sumado
            inv.save()


from django.core.mail import send_mail
from django.conf import settings

def enviar_alerta_merma(remision):
    """
    Evalúa la merma por cada material del detalle de la remisión.
    El umbral se lee de ConfiguracionAlertaMerma (default 1 % si no existe
    configuración para ese material). Si algún material supera su umbral,
    se envía UN correo de alerta agrupando todos los materiales afectados.

    Anti-duplicado: si ya existe `RemisionAlertaMermaLog` para esta remisión,
    no se vuelve a mandar. Esto garantiza que el correo se envía SOLO UNA
    VEZ por remisión, aunque la remisión se edite múltiples veces.
    """
    UMBRAL_DEFAULT = 1.0

    # ── Anti-duplicado ──────────────────────────────────────────────────
    # Si ya se mandó la alerta para esta remisión, no hacemos nada.
    if RemisionAlertaMermaLog.objects.filter(remision=remision).exists():
        return

    alertas = []

    for det in remision.detalles.all():
        if not det.material:
            continue

        peso_ld = float(det.peso_ld or 0)
        if peso_ld <= 0:
            continue

        peso_dlv = float(det.peso_dlv or 0)
        peso_rechazado = float(det.peso_rechazado or 0)
        merma_kg = peso_ld - (peso_dlv + peso_rechazado)

        if merma_kg <= 0:
            continue

        pct_merma = (merma_kg / peso_ld) * 100

        try:
            umbral = float(det.material.config_alerta_merma.porcentaje_umbral)
        except ConfiguracionAlertaMerma.DoesNotExist:
            umbral = UMBRAL_DEFAULT

        if pct_merma > umbral:
            alertas.append({
                'material': det.material.nombre,
                'peso_ld': peso_ld,
                'peso_dlv': peso_dlv,
                'peso_rechazado': peso_rechazado,
                'merma_kg': merma_kg,
                'pct_merma': pct_merma,
                'umbral': umbral,
            })

    if not alertas:
        return

    lineas = "\n".join(
        f"  • {a['material']}: {a['merma_kg']:.2f} Kg perdidos "
        f"({a['pct_merma']:.2f}% — umbral: {a['umbral']}%)"
        for a in alertas
    )

    # ── Helpers para formatear seguros ─────────────────────────────────
    def _txt(v, fallback='—'):
        if v is None: return fallback
        s = str(v).strip()
        return s if s else fallback

    def _fmt_fecha(d):
        try:
            return d.strftime('%d/%m/%Y') if d else '—'
        except Exception:
            return str(d) if d else '—'

    def _fmt_dt(d):
        try:
            return d.strftime('%d/%m/%Y %H:%M') if d else '—'
        except Exception:
            return str(d) if d else '—'

    def _fmt_hora(h):
        try:
            return h.strftime('%H:%M') if h else '—'
        except Exception:
            return str(h) if h else '—'

    def _fmt_num(n, dec=3):
        try:
            return f"{float(n):,.{dec}f}" if n is not None else '—'
        except Exception:
            return '—'

    # ── Datos ya resueltos para el correo ──────────────────────────────
    empresa_nombre   = _txt(remision.empresa.nombre if remision.empresa_id else None)
    cliente_nombre   = _txt(remision.cliente.nombre if remision.cliente_id else None)
    origen_nombre    = _txt(remision.origen.nombre if remision.origen_id else None)
    destino_nombre   = _txt(remision.destino.nombre if remision.destino_id else None)
    operador_nombre  = _txt(
        remision.operador.nombre if remision.operador_id else (remision.operador_manual or None)
    )
    linea_nombre     = _txt(remision.linea_transporte.nombre if remision.linea_transporte_id else None)
    unidad_nombre    = _txt(
        f"{remision.unidad.numero_economico or remision.unidad.id}"
        if remision.unidad_id else (remision.unidad_manual or None)
    )
    placas_unidad    = _txt(
        getattr(remision.unidad, 'placas', None) if remision.unidad_id else remision.placas_unidad_manual
    )
    contenedor_nom   = _txt(
        f"{remision.contenedor.numero or remision.contenedor.id}"
        if remision.contenedor_id else (remision.contenedor_manual or None)
    )
    placas_contened  = _txt(
        getattr(remision.contenedor, 'placas', None) if remision.contenedor_id else remision.placas_contenedor_manual
    )

    # Totales calculados
    total_ld   = remision.total_peso_ld
    total_dlv  = remision.total_peso_dlv
    total_rech = remision.total_peso_rechazado
    pct_global = remision.porcentaje_merma

    # Status display amigable
    status_display = dict(Remision.STATUS_CHOICES).get(remision.status, remision.status)

    # Línea con todos los detalles de cada material (no solo los que superaron)
    detalle_total_lineas = []
    for det in remision.detalles.all():
        mat = det.material.nombre if det.material else '—'
        pld  = float(det.peso_ld or 0)
        pdlv = float(det.peso_dlv or 0)
        prec = float(det.peso_rechazado or 0)
        mkg  = pld - (pdlv + prec)
        mpct = (mkg / pld * 100) if pld > 0 else 0
        detalle_total_lineas.append(
            f"  • {mat:25} Carga: {pld:>10,.2f} Kg | Descarga: {pdlv:>10,.2f} Kg | "
            f"Rechazo: {prec:>10,.2f} Kg | Merma: {mkg:>+10,.2f} Kg ({mpct:>+6.2f}%)"
        )
    detalle_total = "\n".join(detalle_total_lineas) if detalle_total_lineas else "  (sin detalles capturados)"

    asunto = f"⚠️ ALERTA DE MERMA: Remisión {remision.remision} ({len(alertas)} material(es))"
    mensaje = (
        f"════════════════════════════════════════════════════════════════\n"
        f"  ALERTA DE MERMA — REMISIÓN {remision.remision}\n"
        f"════════════════════════════════════════════════════════════════\n\n"
        f"Se detectó merma superior al umbral configurado en la siguiente\n"
        f"remisión. Estos son TODOS los datos del registro:\n\n"

        f"─── IDENTIFICACIÓN ──────────────────────────────────────────────\n"
        f"  Folio Remisión:    {_txt(remision.remision)}\n"
        f"  Folio Medline:     {_txt(remision.folio_medline)}\n"
        f"  Fecha:             {_fmt_fecha(remision.fecha)}\n"
        f"  Estatus:           {_txt(status_display)}\n"
        f"  Empresa:           {empresa_nombre}\n"
        f"  Cliente:           {cliente_nombre}\n\n"

        f"─── TRANSPORTE ──────────────────────────────────────────────────\n"
        f"  Operador:          {operador_nombre}\n"
        f"  Línea Transporte:  {linea_nombre}\n"
        f"  Unidad:            {unidad_nombre}\n"
        f"  Placas Unidad:     {placas_unidad}\n"
        f"  Contenedor:        {contenedor_nom}\n"
        f"  Placas Contenedor: {placas_contened}\n\n"

        f"─── RUTA Y TIEMPOS ──────────────────────────────────────────────\n"
        f"  Origen:            {origen_nombre}\n"
        f"  Destino:           {destino_nombre}\n"
        f"  Inicia Carga:      {_fmt_dt(remision.inicia_ld)}\n"
        f"  Termina Carga:     {_fmt_dt(remision.termina_ld)}\n"
        f"  Folio Carga:       {_txt(remision.folio_ld)}\n"
        f"  Inicia Descarga:   {_fmt_dt(remision.inicia_dlv)}\n"
        f"  Termina Descarga:  {_fmt_dt(remision.termina_dlv)}\n"
        f"  Folio Descarga:    {_txt(remision.folio_dlv)}\n"
        f"  Hora Entrada:      {_fmt_hora(remision.hora_entrada)}\n"
        f"  Hora Salida:       {_fmt_hora(remision.hora_salida)}\n\n"

        f"─── BÁSCULA Y FACTURACIÓN ───────────────────────────────────────\n"
        f"  Peso Báscula:      {_fmt_num(remision.peso_bascula)} Kg\n"
        f"  Folio Factura:     {_txt(remision.factura_nombre)}\n\n"

        f"─── TOTALES DE PESO ─────────────────────────────────────────────\n"
        f"  Total Carga (LD):     {_fmt_num(total_ld)} Kg\n"
        f"  Total Descarga (DLV): {_fmt_num(total_dlv)} Kg\n"
        f"  Total Rechazo:        {_fmt_num(total_rech)} Kg\n"
        f"  Merma Global:         {pct_global:+.2f}%\n\n"

        f"─── DETALLE POR MATERIAL ────────────────────────────────────────\n"
        f"{detalle_total}\n\n"

        f"⚠️ MATERIALES QUE SUPERAN EL UMBRAL CONFIGURADO ⚠️\n"
        f"{lineas}\n\n"

        f"─── DESTRUCCIÓN FISCAL ──────────────────────────────────────────\n"
        f"  Fecha Destrucción: {_fmt_fecha(remision.fecha_destruccion)}\n"
        f"  Material 1:        {_txt(remision.destruccion_material_1)}"
        f" ({_fmt_num(remision.destruccion_peso_1)} Kg)\n"
        f"  Material 2:        {_txt(remision.destruccion_material_2)}"
        f" ({_fmt_num(remision.destruccion_peso_2)} Kg)\n"
        f"  Comentarios:       {_txt(remision.comentarios_destruccion)}\n\n"

        f"─── OBSERVACIONES ───────────────────────────────────────────────\n"
        f"  Descripción:       {_txt(remision.descripcion)}\n"
        f"  Comentario:        {_txt(remision.comentario)}\n"
        f"  Notas TRANE:       {_txt(remision.trazabilidad_notas)}\n\n"

        f"─── AUDITORÍA ───────────────────────────────────────────────────\n"
        f"  Auditado por:      "
        f"{_txt(remision.auditado_por.get_full_name() or remision.auditado_por.username if remision.auditado_por_id else None)}\n"
        f"  Auditado en:       {_fmt_dt(remision.auditado_en)}\n"
        f"  Creado en:         {_fmt_dt(remision.creado_en)}\n"
        f"  Última actualiz.:  {_fmt_dt(remision.actualizado_en)}\n\n"

        f"────────────────────────────────────────────────────────────────\n"
        f"Por favor revisa esta remisión en el sistema:\n"
        f"https://app.3recycling.com.mx/detalle-remision/{remision.id}\n"
        f"────────────────────────────────────────────────────────────────\n"
    )

    destinatarios = list(DestinatarioAlertaMerma.objects.values_list('email', flat=True))
    if not destinatarios:
        print(f"⚠️ Sin destinatarios configurados para alertas de merma — remisión {remision.remision}")
        return

    try:
        send_mail(
            subject=asunto,
            message=mensaje,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=destinatarios,
            fail_silently=False,  # queremos saber si falla para NO registrar
        )
        # Registramos el envío para que no se vuelva a mandar otro correo
        # por la misma remisión (aunque luego se edite).
        try:
            RemisionAlertaMermaLog.objects.create(
                remision=remision,
                materiales_alertados=len(alertas),
                detalle=lineas,
            )
        except Exception as log_err:
            # Si falla el log por una race condition (otro proceso ya lo
            # creó), no pasa nada — el correo ya se envió y la presencia
            # del log evitará duplicados.
            print(f"⚠️ Alerta enviada pero no se pudo registrar el log: {log_err}")
        print(f"✅ Alerta de merma enviada para la remisión {remision.remision}")
    except Exception as e:
        print(f"❌ Error al enviar alerta de merma: {e}")

def asignar_folio_medline(remision):
    fecha_base = remision.fecha or timezone.now().date()
    
    # 1. Si la fecha viene como texto (string) desde el frontend, la convertimos a date
    if isinstance(fecha_base, str):
        try:
            # Convierte el formato 'YYYY-MM-DD' a objeto date
            fecha_base = datetime.datetime.strptime(fecha_base, '%Y-%m-%d').date()
        except ValueError:
            # Si falla el formato, usamos la fecha actual por seguridad
            fecha_base = timezone.now().date()
            
    # 2. Si viene como datetime, le extraemos solo la fecha
    elif isinstance(fecha_base, datetime.datetime):
        fecha_base = fecha_base.date()
        
    # Ahora sí podemos hacer la comparación matemática sin errores
    if fecha_base < datetime.date(2026, 4, 1): return
    if remision.folio_medline: return
    if not remision.origen or 'MEDLINE' not in remision.origen.nombre.upper(): return

    aplica_material = False
    for detalle in remision.detalles.all():
        if detalle.material:
            mat_nom = detalle.material.nombre.upper()
            if "CARTON" in mat_nom or "CARTÓN" in mat_nom or "ARCHIVO" in mat_nom:
                aplica_material = True
                break
                
    if not aplica_material: return

    year = fecha_base.year
    month = f"{fecha_base.month:02d}"
    prefix = f"3R-{year}-{month}-"

    # Contador GLOBAL (no por mes/año):
    # · Encontramos el número más alto que se haya asignado en CUALQUIER
    #   folio Medline previo, sin importar año o mes.
    # · Asignamos el siguiente número entero al actual.
    # · El prefijo siempre usa el año y mes de la remisión (fecha_base),
    #   pero el número consecutivo crece monotónicamente.
    with transaction.atomic():
        existentes = (
            Remision.objects
            .select_for_update()
            .filter(folio_medline__startswith='3R-')
            .values_list('folio_medline', flat=True)
        )
        max_num = 0
        for f in existentes:
            if not f:
                continue
            try:
                n = int(str(f).split('-')[-1])
                if n > max_num:
                    max_num = n
            except (ValueError, TypeError):
                continue

        next_num = max_num + 1
        remision.folio_medline = f"{prefix}{next_num:03d}"
        remision.save(update_fields=['folio_medline'])


def _renumerar_folios_medline(folio_cancelado):
    """
    Dado el folio_medline que acaba de liberarse, renumera hacia abajo
    todos los folios del mismo mes que tengan número mayor, llenando el hueco.
    Debe llamarse DENTRO de un transaction.atomic() existente.
    Ejemplo: si se cancela 3R-2026-04-003 y existen 004 y 005,
             quedan renumerados a 003 y 004.
    """
    if not folio_cancelado:
        return
    parts = folio_cancelado.split('-')  # ['3R', '2026', '04', '003']
    if len(parts) != 4:
        return
    try:
        num_cancelado = int(parts[-1])
    except ValueError:
        return
    prefix = f"{parts[0]}-{parts[1]}-{parts[2]}-"

    # Obtener todos los folios del mismo mes, ordenar numéricamente en Python
    todos_mes = list(
        Remision.objects
        .select_for_update()
        .filter(folio_medline__startswith=prefix)
    )

    def _folio_num(r):
        try:
            return int(r.folio_medline.split('-')[-1])
        except (ValueError, IndexError, AttributeError):
            return 0

    para_renumerar = sorted(
        [r for r in todos_mes if _folio_num(r) > num_cancelado],
        key=_folio_num,
    )

    for i, rem in enumerate(para_renumerar):
        nuevo_num = num_cancelado + i
        rem.folio_medline = f"{prefix}{nuevo_num:03d}"
        rem.save(update_fields=['folio_medline'])


# =========================================================================
# UTILIDADES DE AWS S3 (Manejo de Archivos)
# =========================================================================

def _subir_archivo_a_s3(archivo, s3_path):
    """
    Sube un archivo al almacenamiento configurado (S3 a través de django-storages).
    Retorna la ruta con la que se guardó para asignarla al campo del modelo.
    """
    try:
        # Verifica que el archivo sea válido
        if not archivo:
            return None
            
        # default_storage detecta si estás usando local o S3 (Boto3) 
        # y guarda el archivo en la ruta especificada.
        file_name = default_storage.save(s3_path, archivo)
        return file_name
        
    except Exception as e:
        print(f"❌ Error al subir archivo a S3 ({s3_path}): {e}")
        return None

def _eliminar_archivo_de_s3(s3_path):
    """
    Elimina un archivo del bucket de S3 si existe. 
    Útil para reemplazar fotos/archivos sin dejar archivos "basura" en la nube.
    """
    try:
        if s3_path and default_storage.exists(s3_path):
            default_storage.delete(s3_path)
            print(f"✅ Archivo eliminado de S3: {s3_path}")
    except Exception as e:
        print(f"❌ Error al intentar eliminar archivo en S3 ({s3_path}): {e}")

@csrf_exempt
def api_remisiones_lista(request):
    """
    API JSON para remisiones. Implementa exactamente la misma 
    lógica de filtros y seguridad que RemisionListView.
    """
    
    # --- AUTH CHECK ---
    if not request.user.is_authenticated:
        return JsonResponse(
            {'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'},
            status=401
        )

    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    # 1. QUERYSET BASE OPTIMIZADO
    # Las relaciones dependen de las columnas activas del usuario: sin esto,
    # una columna como Operador o Línea de transporte costaría una consulta
    # por fila.
    from .columnas_remisiones import (
        SELECT_RELATED_POR_CLAVE,
        columnas_de as _columnas_de,
        valor_texto as _valor_texto_col,
    )

    _columnas_usuario = _columnas_de(request.user)
    _relaciones = {'empresa', 'origen', 'destino', 'operador'}
    for _c in _columnas_usuario:
        _rel = SELECT_RELATED_POR_CLAVE.get(_c['clave'])
        if _rel:
            _relaciones.add(_rel)

    queryset = Remision.objects.select_related(
        *sorted(_relaciones)
    ).prefetch_related(
        'detalles__material',
        'evidencias',
        'facturas'
    ).order_by('-pk')

    # 2. FILTRADO POR PERMISOS DE EMPRESA
    if not request.user.is_superuser:
        perfil = getattr(request.user, 'ternium_profile', None)
        if perfil:
            mis_empresas = perfil.empresas_autorizadas.all()
            queryset = queryset.filter(empresa__in=mis_empresas)
        else:
            queryset = queryset.none()

    # 3. LÓGICA DE FILTROS BÁSICOS Y FECHAS
    params = request.GET.copy()
    q_remision = params.get('q_remision', '').strip()

    # --- CASO A: SI BUSCAN POR FOLIO PRINCIPAL (BÚSQUEDA GLOBAL) ---
    if q_remision:
        queryset = queryset.filter(remision__icontains=q_remision)
        
    # --- CASO B: SI NO HAY FOLIO PRINCIPAL (FILTROS NORMALES + FECHAS) ---
    else:
        filtros_activos = any(
            k.startswith('q_') and v for k, v in params.items()
        )
        
        q_fecha_desde = params.get('q_fecha_desde', '').strip()
        q_fecha_hasta = params.get('q_fecha_hasta', '').strip()
        
        if not filtros_activos:
            today = timezone.now().date()
            month_ago = today - dt.timedelta(days=30)
            q_fecha_desde = month_ago.strftime('%Y-%m-%d')
            q_fecha_hasta = today.strftime('%Y-%m-%d')
            
        from django.utils.dateparse import parse_date
        _fd = parse_date(q_fecha_desde) if q_fecha_desde else None
        _fh = parse_date(q_fecha_hasta) if q_fecha_hasta else None
        if _fd:
            queryset = queryset.filter(fecha__gte=_fd)
        if _fh:
            queryset = queryset.filter(fecha__lte=_fh)

    # 4. APLICAR RESTO DE FILTROS EXACTOS
    q_prefijo = params.get('q_prefijo', '').strip()
    q_material = params.get('q_material', '').strip()
    q_origen = params.get('q_origen', '').strip()
    q_destino = params.get('q_destino', '').strip()
    q_status = params.get('q_status', '').strip()

    if q_prefijo:
        queryset = queryset.filter(empresa__prefijo__icontains=q_prefijo)
    if q_material.isdigit():
        queryset = queryset.filter(detalles__material_id=q_material)
    if q_origen.isdigit():
        queryset = queryset.filter(origen_id=q_origen)
    if q_destino.isdigit():
        queryset = queryset.filter(destino_id=q_destino)
    # NOTA: el filtro por status se aplica más abajo (tras calcular los contadores),
    # para que los badges PENDIENTES/CANCELADOS muestren los totales reales aunque
    # haya un quick-filter de status activo.

    # --- NUEVO FILTRO DE OPERADOR LIBRE ---
    q_operador = params.get('q_operador', '').strip()
    if q_operador:
        queryset = queryset.filter(
            Q(operador__nombre__icontains=q_operador) | 
            Q(operador_manual__icontains=q_operador)
        )

    # --- NUEVOS FILTROS DE FOLIOS OPERATIVOS ---
    q_folio_ld = params.get('q_folio_ld', '').strip()
    q_folio_dlv = params.get('q_folio_dlv', '').strip()

    if q_folio_ld:
        queryset = queryset.filter(folio_ld__icontains=q_folio_ld)
    if q_folio_dlv:
        queryset = queryset.filter(folio_dlv__icontains=q_folio_dlv)

    # --- FILTRO DE DESTRUCCIÓN FISCAL (COMPLETO / PENDIENTE) ---
    q_destruccion = params.get('q_destruccion', '').strip()
    if q_destruccion:
        completo_q = Q(
            fecha_destruccion__isnull=False,
            destruccion_material_1__isnull=False,
            foto_ingreso__isnull=False,
            foto_vertido__isnull=False,
            foto_destruccion__isnull=False
        ) & ~Q(destruccion_material_1='') & ~Q(foto_ingreso='') & ~Q(foto_vertido='') & ~Q(foto_destruccion='')
        
        if q_destruccion == 'COMPLETO':
            queryset = queryset.filter(completo_q)
        elif q_destruccion == 'PENDIENTE':
            configuraciones = ConfiguracionManifiesto.objects.all()
            q_requiere = Q()
            for conf in configuraciones:
                q_requiere |= Q(origen=conf.origen, detalles__material=conf.material)
            
            if configuraciones.exists():
                queryset = queryset.filter(q_requiere).exclude(completo_q)
            else:
                queryset = queryset.none()

    # Evitar duplicados si se filtra por relaciones ManyToMany
    if q_material or q_destruccion:
        queryset = queryset.distinct()

    # 6. CONTADORES (sobre la base SIN el filtro de status → badges con totales reales)
    total_pendientes = queryset.filter(status='PENDIENTE').count()
    total_cancelados = queryset.filter(status='CANCELADO').count()

    # Se aplica AQUÍ el filtro de status (tras los contadores) antes de paginar.
    if q_status:
        queryset = queryset.filter(status=q_status)

    # 5. PAGINACIÓN (saneada: nunca revienta con valor no numérico/vacío/0/negativo)
    try:
        page_number = int(params.get('page', 1))
    except (TypeError, ValueError):
        page_number = 1
    if page_number < 1:
        page_number = 1
    try:
        page_size = int(params.get('page_size', 15))
    except (TypeError, ValueError):
        page_size = 15
    page_size = max(1, min(page_size, 200))  # mínimo 1, tope 200 (evita ZeroDivisionError y abusos)
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page_number)

    # 7. SERIALIZACIÓN DE LA RESPUESTA
    remisiones_data = []
    for rem in page_obj.object_list:
        detalles = []
        sum_ld = sum_dlv = sum_rech = 0.0
        for det in rem.detalles.all():  # prefetched → sin query extra
            pld = float(det.peso_ld or 0)
            pdlv = float(det.peso_dlv or 0)
            sum_ld += pld
            sum_dlv += pdlv
            sum_rech += float(det.peso_rechazado or 0)
            detalles.append({
                'material': det.material.nombre if det.material else '-',
                'peso_ld': pld,
                'peso_dlv': pdlv,
                'bultos': det.bultos if hasattr(det, 'bultos') and det.bultos else None,
            })

        evidencias_urls = []
        if rem.evidencia_documento and rem.evidencia_documento.name:
            evidencias_urls.append(rem.evidencia_documento.url)
        if hasattr(rem, 'boleta_salida_medline') and rem.boleta_salida_medline and rem.boleta_salida_medline.name:
            evidencias_urls.append(rem.boleta_salida_medline.url)
        if rem.manifiesto and rem.manifiesto.name:
            evidencias_urls.append(rem.manifiesto.url)
            
        for campo in ['foto_ingreso', 'foto_ingreso_2', 'foto_vertido', 'foto_vertido_2', 'foto_destruccion', 'foto_destruccion_2']:
            archivo = getattr(rem, campo, None)
            if archivo and hasattr(archivo, 'name') and archivo.name:
                evidencias_urls.append(archivo.url)
                
        for ev in rem.evidencias.all():
            if ev.archivo and ev.archivo.name:
                evidencias_urls.append(ev.archivo.url)

        facturas_data = [{'id': fac.id} for fac in rem.facturas.all()]

        # Totales desde los detalles YA prefetcheados (evita ~6 queries aggregate por
        # fila que hacían las @property del modelo, que ignoran el prefetch → N+1).
        # Fórmulas idénticas al modelo: Dif = (descarga + rechazado) − carga;
        # % merma = ((carga − (descarga + rechazado)) / carga) × 100.
        total_ld = sum_ld
        total_dlv = sum_dlv
        diff = (total_dlv + sum_rech) - total_ld
        porcentaje_merma = ((total_ld - (total_dlv + sum_rech)) / total_ld * 100) if total_ld > 0 else 0.0

        # Banderas para mostrar (o no) el botón de Reporte Destrucción (Word).
        # Reflejan exactamente las mismas condicionales que el template Django:
        #   {% if remision.permite_manifiesto_destruccion and
        #         remision.destruccion_fiscal_completa %}
        try:
            permite_word = bool(rem.permite_manifiesto_destruccion)
        except Exception:
            permite_word = False
        try:
            destruccion_completa = bool(rem.destruccion_fiscal_completa)
        except Exception:
            destruccion_completa = False

        # Valores de las columnas que el usuario tenga activas. El frontend
        # pinta con su propio marcado las columnas de siempre (badges, enlaces)
        # y usa estos textos para las que se agregaron desde el panel.
        celdas = {c['clave']: _valor_texto_col(rem, c) for c in _columnas_usuario}

        remisiones_data.append({
            'celdas': celdas,
            'id': rem.pk,
            'remision': rem.remision,
            'fecha': rem.fecha.strftime('%d/%m/%Y') if rem.fecha else '',
            'fecha_iso': rem.fecha.isoformat() if rem.fecha else '',
            'status': rem.status,
            'status_display': rem.get_status_display(),
            'origen': rem.origen.nombre if rem.origen else '-',
            'destino': rem.destino.nombre if rem.destino else '-',
            'folio_ld': rem.folio_ld or '-',
            'folio_dlv': rem.folio_dlv or '-',
            # Folio Medline (3R-YYYY-MM-NNN) — solo lo tienen las remisiones
            # MEDLINE con cartón/archivo. Vacío "" para las demás.
            'folio_medline': rem.folio_medline or '',
            'total_peso_ld': total_ld,
            'total_peso_dlv': total_dlv,
            'diff': round(diff, 3),
            'porcentaje_merma': round(porcentaje_merma, 1),
            'detalles': detalles,
            'evidencias_urls': evidencias_urls,
            'facturas': facturas_data,
            'permite_manifiesto_destruccion': permite_word,
            'destruccion_fiscal_completa': destruccion_completa,
            # Alias cortos (lo que ya consume el front Next.js).
            'permite_manifiesto': permite_word,
            'destruccion_completa': destruccion_completa,
        })

    # 8. CATÁLOGOS PARA LLENAR LOS SELECTS DEL FRONTEND
    #    Para un no-superusuario se limitan a sus empresas autorizadas (misma
    #    lógica de permisos que el filtrado de remisiones, arriba).
    _emp_qs = Empresa.objects.all()
    _mat_qs = Material.objects.all()
    _lug_qs = Lugar.objects.all()
    if not request.user.is_superuser:
        _perfil = getattr(request.user, 'ternium_profile', None)
        _mis = _perfil.empresas_autorizadas.all() if _perfil else Empresa.objects.none()
        _emp_qs = _emp_qs.filter(pk__in=_mis)
        _mat_qs = _mat_qs.filter(empresas__in=_mis).distinct()
        _lug_qs = _lug_qs.filter(empresas__in=_mis).distinct()
    prefijos = list(_emp_qs.values_list('prefijo', flat=True).distinct().order_by('prefijo'))
    materiales_list = list(_mat_qs.values('id', 'nombre').order_by('nombre'))
    origenes_list = list(_lug_qs.filter(tipo__in=['ORIGEN', 'AMBOS']).values('id', 'nombre').order_by('nombre'))
    destinos_list = list(_lug_qs.filter(tipo__in=['DESTINO', 'AMBOS']).values('id', 'nombre').order_by('nombre'))
    estatus_choices = [{'value': v, 'display': d} for v, d in Remision.STATUS_CHOICES]

    # 9. RETORNO JSON FINAL
    from .columnas_remisiones import (
        COLUMNAS as _CATALOGO_COLUMNAS,
        catalogo_agrupado as _catalogo_agrupado,
        columnas_por_defecto as _columnas_por_defecto,
        config_de as _config_de,
    )
    _claves_guardadas, _usa_personalizada = _config_de(request.user)
    _activas, _disponibles = _catalogo_agrupado(_claves_guardadas)

    return JsonResponse({
        'columnas': [
            {'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
            for c in _columnas_usuario
        ],
        'columnas_config': {
            # 'activas' es lo que el panel debe mostrar AHORA: con la vista por
            # defecto seleccionada, siempre las columnas de siempre.
            'activas': [
                {'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
                for c in (_activas if _usa_personalizada else _columnas_por_defecto())
            ],
            # La tabla que armó el usuario, para restaurarla al cambiar de vista.
            'mi_tabla': [
                {'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
                for c in _activas
            ],
            'por_defecto': [
                {'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
                for c in _columnas_por_defecto()
            ],
            'catalogo': [
                {'clave': c['clave'], 'etiqueta': c['etiqueta'], 'grupo': c['grupo'],
                 'fija': bool(c.get('fija'))}
                for c in _CATALOGO_COLUMNAS
            ],
            'disponibles': [
                {'nombre': g['nombre'],
                 'columnas': [{'clave': c['clave'], 'etiqueta': c['etiqueta']} for c in g['columnas']]}
                for g in _disponibles
            ],
            'personalizada': _usa_personalizada,
        },
        'remisiones': remisiones_data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        },
        'counters': {
            'total_pendientes': total_pendientes,
            'total_cancelados': total_cancelados,
        },
        'filters': {
            'prefijos': prefijos,
            'materiales': materiales_list,
            'origenes': origenes_list,
            'destinos': destinos_list,
            'estatus_choices': estatus_choices,
        },
        'permissions': {
            'can_add': request.user.has_perm('ternium.add_remision'),
            'can_change': request.user.has_perm('ternium.change_remision'),
            'can_view_plastico': request.user.has_perm('ternium.view_plastico'),
            'can_view_tarimas': request.user.has_perm('ternium.view_controltarima'),
        },
    })


# =========================================================================
# FUNCIONES DE APOYO PARA GENERACIÓN DE FOLIOS (VÍAS API)
# =========================================================================

def calcular_siguiente_folio(prefijo):
    """
    Calcula el siguiente folio basado en números enteros para evitar
    que 'MTY-999' sea mayor que 'MTY-1000' alfabéticamente.
    """
    prefix_with_dash = f"{prefijo.strip().upper()}-"
    
    remisiones_existentes = Remision.objects.filter(
        remision__startswith=prefix_with_dash
    ).values_list('remision', flat=True)

    max_num = 0
    
    for rem_str in remisiones_existentes:
        try:
            parts = rem_str.split('-')
            if len(parts) > 1:
                num = int(parts[-1])
                if num > max_num:
                    max_num = num
        except ValueError:
            continue

    next_num = max_num + 1
    return f"{prefix_with_dash}{str(next_num).zfill(3)}"


@login_required
def get_next_remision_number(request, empresa_id):
    """
    Obtiene el siguiente folio para una empresa, verificando permisos.
    """
    try:
        if not request.user.is_superuser:
            perfil = getattr(request.user, 'ternium_profile', None)
            if not perfil or not perfil.empresas_autorizadas.filter(pk=empresa_id).exists():
                return JsonResponse({'error': 'No tienes permiso para generar folios de esta empresa.'}, status=403)

        empresa = Empresa.objects.get(pk=empresa_id)
        
        if empresa.prefijo:
            next_remision = calcular_siguiente_folio(empresa.prefijo)
            return JsonResponse({'next_remision': next_remision, 'is_manual': False})
        else:
            return JsonResponse({'is_manual': True})

    except Empresa.DoesNotExist:
        return JsonResponse({'error': 'Empresa no encontrada'}, status=404)

@csrf_exempt
def api_crear_remision(request):
    """
    API endpoint para crear una remisión con archivos (vía FormData).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)
    
    if not request.user.has_perm('ternium.add_remision'):
        return JsonResponse({'error': 'FORBIDDEN', 'detail': 'No tienes permisos para crear remisiones.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    archivos_s3_subidos = []  # rutas subidas a S3 para limpiarlas si hay rollback
    try:
        data = request.POST
        empresa_id = data.get('empresa')
        empresa_seleccionada = None

        if not empresa_id:
            return JsonResponse({'error': 'BAD_REQUEST', 'detail': 'La empresa es requerida.'}, status=400)

        if empresa_id:
            empresa_seleccionada = get_object_or_404(Empresa, pk=empresa_id)

            # Seguridad por empresa
            if not request.user.is_superuser:
                perfil = getattr(request.user, 'ternium_profile', None)
                if not perfil or not perfil.empresas_autorizadas.filter(pk=empresa_seleccionada.pk).exists():
                    return JsonResponse({'error': 'FORBIDDEN', 'detail': 'No tienes permiso para esta empresa.'}, status=403)

        with transaction.atomic():
            remision = Remision()
            remision.empresa = empresa_seleccionada
            
            # --- DATOS GENERALES ---
            from django.utils.dateparse import parse_date
            _fecha_str = (data.get('fecha') or '').strip()
            try:
                _fecha_val = parse_date(_fecha_str) if _fecha_str else None
            except ValueError:
                _fecha_val = None
            if _fecha_val is None:
                return JsonResponse({'error': 'VALIDACION', 'detail': 'La fecha es obligatoria y debe tener formato YYYY-MM-DD.'}, status=400)
            remision.fecha = _fecha_val
            if data.get('origen'): remision.origen_id = data.get('origen')
            if data.get('destino'): remision.destino_id = data.get('destino')
            if data.get('linea_transporte'): remision.linea_transporte_id = data.get('linea_transporte')
            # Paridad con RemisionForm de Django: origen y destino son obligatorios en el servidor.
            if not remision.origen_id:
                return JsonResponse({'error': 'VALIDACION', 'detail': 'El origen es obligatorio.'}, status=400)
            if not remision.destino_id:
                return JsonResponse({'error': 'VALIDACION', 'detail': 'El destino es obligatorio.'}, status=400)

            remision.inicia_ld = data.get('inicia_ld') or None
            remision.termina_ld = data.get('termina_ld') or None
            remision.inicia_dlv = data.get('inicia_dlv') or None
            remision.termina_dlv = data.get('termina_dlv') or None
            
            remision.folio_ld = data.get('folio_ld', '').upper()
            remision.folio_dlv = data.get('folio_dlv', '').upper()
            # Paridad Django (RemisionForm.clean): folio de carga/descarga únicos por empresa.
            if remision.folio_ld and Remision.objects.filter(folio_ld__iexact=remision.folio_ld, empresa=empresa_seleccionada).exists():
                return JsonResponse({'error': 'VALIDACION', 'detail': f"El Folio de Carga «{remision.folio_ld}» ya existe en otra remisión de {empresa_seleccionada.nombre}."}, status=400)
            if remision.folio_dlv and Remision.objects.filter(folio_dlv__iexact=remision.folio_dlv, empresa=empresa_seleccionada).exists():
                return JsonResponse({'error': 'VALIDACION', 'detail': f"El Folio de Descarga «{remision.folio_dlv}» ya existe en otra remisión de {empresa_seleccionada.nombre}."}, status=400)
            remision.comentario = data.get('comentario', '')
            remision.trazabilidad_notas = data.get('trazabilidad_notas', '')

            # --- DESTRUCCIÓN FISCAL Y BÁSCULA ---
            remision.hora_entrada = data.get('hora_entrada') or None
            remision.hora_salida = data.get('hora_salida') or None
            remision.factura_nombre = data.get('factura_nombre', '').upper()
            remision.fecha_destruccion = data.get('fecha_destruccion') or None
            remision.comentarios_destruccion = data.get('comentarios_destruccion', '')

            def _peso_seguro(valor):
                if valor is None or str(valor).strip() == '':
                    return None
                try:
                    return float(str(valor).replace(',', '.').strip())
                except (TypeError, ValueError):
                    return None
            remision.destruccion_material_1 = data.get('destruccion_material_1')
            remision.destruccion_peso_1 = _peso_seguro(data.get('destruccion_peso_1'))
            remision.destruccion_material_2 = data.get('destruccion_material_2')
            remision.destruccion_peso_2 = _peso_seguro(data.get('destruccion_peso_2'))
            if 'peso_bascula' in data:
                remision.peso_bascula = _peso_seguro(data.get('peso_bascula'))

            # --- LÓGICA MANUAL (OPERADOR, UNIDAD, CONTENEDOR) ---
            op_manual = data.get('operador_texto', '').strip().upper()
            uni_manual = data.get('unidad_texto', '').strip().upper()
            placa_uni = data.get('placas_unidad_texto', '').strip().upper()
            cont_manual = data.get('contenedor_texto', '').strip().upper()
            placa_cont = data.get('placas_contenedor_texto', '').strip().upper()

            if op_manual:
                remision.operador_manual = op_manual
            elif data.get('operador'):
                remision.operador_id = data.get('operador')

            if uni_manual:
                remision.unidad_manual = uni_manual
                remision.placas_unidad_manual = placa_uni
            elif data.get('unidad'):
                remision.unidad_id = data.get('unidad')

            if cont_manual:
                remision.contenedor_manual = cont_manual
                remision.placas_contenedor_manual = placa_cont
            elif data.get('contenedor'):
                remision.contenedor_id = data.get('contenedor')

            # --- GENERACIÓN DE FOLIO ---
            if empresa_seleccionada and empresa_seleccionada.prefijo:
                # read-max-then-increment sin bloqueo → carrera con unique_together.
                # Reintentar recalculando el folio ante IntegrityError, con savepoint
                # anidado para no romper la transacción exterior.
                for _intento in range(6):
                    remision.remision = calcular_siguiente_folio(empresa_seleccionada.prefijo)
                    try:
                        with transaction.atomic():
                            remision.save()
                        break
                    except IntegrityError:
                        remision.pk = None  # fuerza un INSERT nuevo en el siguiente intento
                else:
                    raise IntegrityError("No se pudo asignar un folio único tras varios reintentos.")
            else:
                remision.save()

            # --- SUBIDA DE ARCHIVOS (S3) ---
            # 1. Fotos Destrucción
            fotos_destruccion = ['foto_ingreso', 'foto_ingreso_2', 'foto_vertido', 'foto_vertido_2', 'foto_destruccion', 'foto_destruccion_2']
            for campo_foto in fotos_destruccion:
                if campo_foto in request.FILES:
                    archivo = request.FILES[campo_foto]
                    nombre_limpio = archivo.name.replace(" ", "_")
                    s3_path = f"remisiones/{remision.remision}/{campo_foto}_{nombre_limpio}"
                    ruta_s3 = _subir_archivo_a_s3(archivo, s3_path)
                    if ruta_s3:
                        setattr(remision, campo_foto, ruta_s3)
                        archivos_s3_subidos.append(ruta_s3)

            # 2. Manifiesto
            if 'manifiesto' in request.FILES:
                archivo = request.FILES['manifiesto']
                nombre_limpio = archivo.name.replace(" ", "_")
                ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/manifiesto_{nombre_limpio}")
                if ruta_s3:
                    remision.manifiesto = ruta_s3
                    archivos_s3_subidos.append(ruta_s3)

            # 3. Boleta Medline
            if 'boleta_salida_medline' in request.FILES:
                import re
                archivo = request.FILES['boleta_salida_medline']
                nombre_limpio = re.sub(r'[^a-zA-Z0-9_\-\.]', '', archivo.name.replace(" ", "_"))
                ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/medline_{nombre_limpio}")
                if ruta_s3:
                    remision.boleta_salida_medline = ruta_s3
                    archivos_s3_subidos.append(ruta_s3)

            remision.save() # Guardar rutas de archivos

            # 4. Evidencias Múltiples
            files = request.FILES.getlist('evidencia_documento')
            for i, archivo in enumerate(files):
                nombre_limpio = archivo.name.replace(" ", "_")
                ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/evidencia_{i}_{nombre_limpio}")
                if ruta_s3:
                    EvidenciaRemision.objects.create(remision=remision, archivo=ruta_s3)
                    archivos_s3_subidos.append(ruta_s3)

            # --- PROCESAR DETALLES (MATERIALES) ---
            # Se espera que el frontend envíe un JSON stringificado en un campo 'detalles'
            detalles_json = request.POST.get('detalles', '[]')
            try:
                detalles_list = json.loads(detalles_json)
                for det in detalles_list:
                    if det.get('material'):
                        DetalleRemision.objects.create(
                            remision=remision,
                            material_id=det.get('material'),
                            unidad_medida=(det.get('unidad') or 'KG'),  # paridad con /remisiones/crear/ (evita default 'TON' → peso ×1000)
                            bultos=det.get('bultos') or None,
                            peso_ld=det.get('peso_ld') or 0,
                            peso_dlv=det.get('peso_dlv') or 0,
                            peso_rechazado=det.get('peso_rechazado') or 0,
                            patio_rechazo_id=det.get('patio_rechazo') or None,
                            cliente=remision.destino # Forzar cliente = destino
                        )
            except json.JSONDecodeError:
                pass # Manejar error si es necesario

            # --- HISTORIAL, INVENTARIO Y ALERTAS ---
            HistorialRemision.objects.create(
                remision=remision, usuario=request.user, cambio="Creación de la remisión vía API"
            )
            
            # Folio Medline MANUAL (lo captura el usuario si aplica). Vacío = pendiente. Debe ser único.
            if 'folio_medline' in request.POST:
                _folio_med = (request.POST.get('folio_medline') or '').strip() or None
                if _folio_med and Remision.objects.filter(folio_medline=_folio_med).exclude(pk=remision.pk).exists():
                    raise ValueError(f"El folio Medline «{_folio_med}» ya existe en otra remisión. Debe ser único.")
                if _folio_med != remision.folio_medline:
                    remision.folio_medline = _folio_med
                    remision.save(update_fields=['folio_medline'])
            _update_inventory_from_remision(remision, revert=False)
            enviar_alerta_merma(remision)

            return JsonResponse({'success': True, 'remision_id': remision.pk, 'folio': remision.remision}, status=201)

    except Exception as e:
        # Si la transacción hizo rollback, los archivos ya subidos a S3 quedan huérfanos: los borramos.
        for _ruta_huerfana in archivos_s3_subidos:
            try:
                _eliminar_archivo_de_s3(_ruta_huerfana)
            except Exception:
                pass
        return JsonResponse({'error': 'Error en servidor', 'detail': str(e)}, status=500)


@csrf_exempt
def api_editar_remision(request, pk):
    """
    API endpoint para editar una remisión existente.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)
    
    if not request.user.has_perm('ternium.change_remision'):
        return JsonResponse({'error': 'FORBIDDEN', 'detail': 'No tienes permisos para editar.'}, status=403)

    if request.method != 'POST': # Usamos POST para manejar FormData correctamente
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    remision = get_object_or_404(Remision, pk=pk)

    # Seguridad por empresa
    if not request.user.is_superuser:
        perfil = getattr(request.user, 'ternium_profile', None)
        if not perfil or not perfil.empresas_autorizadas.filter(pk=remision.empresa.pk).exists():
            return JsonResponse({'error': 'FORBIDDEN', 'detail': 'No tienes permiso para esta empresa.'}, status=403)

    if remision.status == 'AUDITADO':
        return JsonResponse({'error': 'CONFLICT', 'detail': 'No se puede editar una remisión auditada.'}, status=409)

    try:
        data = request.POST
        cambios_log = []

        with transaction.atomic():
            # Revertir inventario ANTES de aplicar cambios
            _update_inventory_from_remision(remision, revert=True)

            # --- LOGICA MANUAL Y LOGS ---
            nuevo_op = data.get('operador_texto', '').strip().upper()
            nuevo_uni = data.get('unidad_texto', '').strip().upper()
            nueva_placa_uni = data.get('placas_unidad_texto', '').strip().upper()
            nuevo_cont = data.get('contenedor_texto', '').strip().upper()
            nueva_placa_cont = data.get('placas_contenedor_texto', '').strip().upper()

            if (remision.operador_manual or '') != nuevo_op: cambios_log.append("Se modificó el Operador Manual")
            if (remision.unidad_manual or '') != nuevo_uni: cambios_log.append("Se modificó la Unidad Manual")
            if (remision.placas_unidad_manual or '') != nueva_placa_uni: cambios_log.append("Se modificaron Placas Unidad")
            if (remision.contenedor_manual or '') != nuevo_cont: cambios_log.append("Se modificó Contenedor Manual")
            if (remision.placas_contenedor_manual or '') != nueva_placa_cont: cambios_log.append("Se modificaron Placas Cont.")

            # Aplicar Manuales
            if nuevo_op: remision.operador_manual = nuevo_op; remision.operador = None
            elif data.get('operador'): remision.operador_id = data.get('operador'); remision.operador_manual = None
            
            if nuevo_uni: 
                remision.unidad_manual = nuevo_uni; remision.placas_unidad_manual = nueva_placa_uni; remision.unidad = None
            elif data.get('unidad'): remision.unidad_id = data.get('unidad'); remision.unidad_manual = None
            
            if nuevo_cont: 
                remision.contenedor_manual = nuevo_cont; remision.placas_contenedor_manual = nueva_placa_cont; remision.contenedor = None
            elif data.get('contenedor'): remision.contenedor_id = data.get('contenedor'); remision.contenedor_manual = None

            # --- ACTUALIZAR DATOS GENERALES ---
            # (Agrega comparaciones al `cambios_log` si lo consideras necesario para el auditor)
            remision.fecha = data.get('fecha', remision.fecha)
            if data.get('origen'): remision.origen_id = data.get('origen')
            if data.get('destino'): remision.destino_id = data.get('destino')
            if data.get('linea_transporte'): remision.linea_transporte_id = data.get('linea_transporte')
            # Paridad Django: origen y destino obligatorios.
            if not remision.origen_id:
                return JsonResponse({'error': 'VALIDACION', 'detail': 'El origen es obligatorio.'}, status=400)
            if not remision.destino_id:
                return JsonResponse({'error': 'VALIDACION', 'detail': 'El destino es obligatorio.'}, status=400)

            # Presente-en-POST = actualizar (vacío → NULL, permite BORRAR el tiempo al
            # editar). Ausente = conservar el valor actual (un submit parcial no lo pisa).
            if 'inicia_ld' in data: remision.inicia_ld = data.get('inicia_ld') or None
            if 'termina_ld' in data: remision.termina_ld = data.get('termina_ld') or None
            if 'inicia_dlv' in data: remision.inicia_dlv = data.get('inicia_dlv') or None
            if 'termina_dlv' in data: remision.termina_dlv = data.get('termina_dlv') or None

            remision.folio_ld = data.get('folio_ld', remision.folio_ld).upper()
            remision.folio_dlv = data.get('folio_dlv', remision.folio_dlv).upper()
            # Paridad Django: folio de carga/descarga únicos por empresa (excluyendo esta remisión).
            if remision.folio_ld and Remision.objects.filter(folio_ld__iexact=remision.folio_ld, empresa=remision.empresa).exclude(pk=remision.pk).exists():
                return JsonResponse({'error': 'VALIDACION', 'detail': f"El Folio de Carga «{remision.folio_ld}» ya existe en otra remisión de {remision.empresa.nombre}."}, status=400)
            if remision.folio_dlv and Remision.objects.filter(folio_dlv__iexact=remision.folio_dlv, empresa=remision.empresa).exclude(pk=remision.pk).exists():
                return JsonResponse({'error': 'VALIDACION', 'detail': f"El Folio de Descarga «{remision.folio_dlv}» ya existe en otra remisión de {remision.empresa.nombre}."}, status=400)
            remision.comentario = data.get('comentario', remision.comentario)
            remision.trazabilidad_notas = data.get('trazabilidad_notas', remision.trazabilidad_notas)

            # --- DESTRUCCIÓN FISCAL Y BÁSCULA (faltaban en edición: se perdían al guardar) ---
            if 'hora_entrada' in data: remision.hora_entrada = data.get('hora_entrada') or None
            if 'hora_salida' in data: remision.hora_salida = data.get('hora_salida') or None
            if 'factura_nombre' in data: remision.factura_nombre = (data.get('factura_nombre') or '').upper()
            if 'fecha_destruccion' in data: remision.fecha_destruccion = data.get('fecha_destruccion') or None
            if 'comentarios_destruccion' in data: remision.comentarios_destruccion = data.get('comentarios_destruccion', '')

            def _peso_seguro_edit(valor):
                if valor is None or str(valor).strip() == '':
                    return None
                try:
                    return float(str(valor).replace(',', '.').strip())
                except (TypeError, ValueError):
                    return None
            if 'destruccion_material_1' in data: remision.destruccion_material_1 = data.get('destruccion_material_1') or None
            if 'destruccion_peso_1' in data: remision.destruccion_peso_1 = _peso_seguro_edit(data.get('destruccion_peso_1'))
            if 'destruccion_material_2' in data: remision.destruccion_material_2 = data.get('destruccion_material_2') or None
            if 'destruccion_peso_2' in data: remision.destruccion_peso_2 = _peso_seguro_edit(data.get('destruccion_peso_2'))
            if 'peso_bascula' in data: remision.peso_bascula = _peso_seguro_edit(data.get('peso_bascula'))

            # --- ACTUALIZAR ARCHIVOS ---
            # Recolectamos las rutas VIEJAS y las borramos de S3 SOLO tras confirmar
            # la transacción (on_commit): así un rollback posterior (p.ej. al recrear
            # los detalles) no destruye evidencia que en la BD seguiría existiendo.
            # IMPORTANTE: solo se encola la ruta vieja si (a) la subida del nuevo
            # archivo tuvo éxito y (b) la key nueva es DISTINTA de la vieja. Con
            # AWS_S3_FILE_OVERWRITE=True (default), subir un archivo del mismo nombre
            # sobrescribe la misma key; borrar "la vieja" borraría el archivo recién
            # subido y se perdería el adjunto. Comparar las rutas evita esa pérdida.
            rutas_s3_a_borrar = []
            fotos_destruccion = ['foto_ingreso', 'foto_ingreso_2', 'foto_vertido', 'foto_vertido_2', 'foto_destruccion', 'foto_destruccion_2']
            for campo_foto in fotos_destruccion:
                if campo_foto in request.FILES:
                    foto_actual = getattr(remision, campo_foto)
                    ruta_vieja = foto_actual.name if (foto_actual and hasattr(foto_actual, 'name') and foto_actual.name) else None

                    archivo = request.FILES[campo_foto]
                    ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/{campo_foto}_{archivo.name.replace(' ', '_')}")
                    if ruta_s3:
                        setattr(remision, campo_foto, ruta_s3)
                        cambios_log.append(f"Se actualizó {campo_foto}")
                        if ruta_vieja and ruta_vieja != ruta_s3:
                            rutas_s3_a_borrar.append(ruta_vieja)

            if 'manifiesto' in request.FILES:
                ruta_vieja = remision.manifiesto.name if (remision.manifiesto and remision.manifiesto.name) else None
                archivo = request.FILES['manifiesto']
                ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/manifiesto_{archivo.name.replace(' ', '_')}")
                if ruta_s3:
                    remision.manifiesto = ruta_s3
                    cambios_log.append("Manifiesto actualizado")
                    if ruta_vieja and ruta_vieja != ruta_s3:
                        rutas_s3_a_borrar.append(ruta_vieja)

            if 'boleta_salida_medline' in request.FILES:
                ruta_vieja = remision.boleta_salida_medline.name if (remision.boleta_salida_medline and remision.boleta_salida_medline.name) else None
                archivo = request.FILES['boleta_salida_medline']
                ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/medline_{archivo.name.replace(' ', '_')}")
                if ruta_s3:
                    remision.boleta_salida_medline = ruta_s3
                    cambios_log.append("Boleta Medline actualizada")
                    if ruta_vieja and ruta_vieja != ruta_s3:
                        rutas_s3_a_borrar.append(ruta_vieja)

            # Borra los archivos viejos de S3 solo si la transacción confirma.
            if rutas_s3_a_borrar:
                transaction.on_commit(lambda rutas=list(rutas_s3_a_borrar): [_eliminar_archivo_de_s3(p) for p in rutas])

            remision.save()

            # Evidencias múltiples
            if request.FILES.getlist('evidencia_documento'):
                archivos = request.FILES.getlist('evidencia_documento')
                conteo = remision.evidencias.count()
                for i, archivo in enumerate(archivos):
                    ruta_s3 = _subir_archivo_a_s3(archivo, f"remisiones/{remision.remision}/evidencia_{conteo+i+1}_{archivo.name.replace(' ', '_')}")
                    if ruta_s3: EvidenciaRemision.objects.create(remision=remision, archivo=ruta_s3)
                cambios_log.append(f"Se agregaron {len(archivos)} evidencias")

            # --- PROCESAR DETALLES ---
            detalles_json = request.POST.get('detalles')
            if detalles_json:
                try:
                    detalles_list = json.loads(detalles_json)
                    # En una API, lo más sencillo es borrar los materiales antiguos y recrearlos
                    # (Como ya revertimos el inventario arriba, es seguro)
                    remision.detalles.all().delete()
                    cambios_log.append("Se modificaron los materiales")

                    for det in detalles_list:
                        if det.get('material'):
                            DetalleRemision.objects.create(
                                remision=remision,
                                material_id=det.get('material'),
                                unidad_medida=(det.get('unidad') or 'KG'),  # paridad con /remisiones/crear/ (evita default 'TON' → peso ×1000)
                                bultos=det.get('bultos') or None,
                                peso_ld=det.get('peso_ld') or 0,
                                peso_dlv=det.get('peso_dlv') or 0,
                                peso_rechazado=det.get('peso_rechazado') or 0,
                                patio_rechazo_id=det.get('patio_rechazo') or None,
                                cliente=remision.destino
                            )
                except json.JSONDecodeError:
                    pass

            # --- HISTORIAL, INVENTARIO Y ALERTAS ---
            if cambios_log:
                cambios_unicos = list(dict.fromkeys(cambios_log))
                HistorialRemision.objects.create(remision=remision, usuario=request.user, cambio=" | ".join(cambios_unicos))

            # Folio Medline MANUAL (lo captura el usuario si aplica). Vacío = pendiente. Debe ser único.
            if 'folio_medline' in request.POST:
                _folio_med = (request.POST.get('folio_medline') or '').strip() or None
                if _folio_med and Remision.objects.filter(folio_medline=_folio_med).exclude(pk=remision.pk).exists():
                    raise ValueError(f"El folio Medline «{_folio_med}» ya existe en otra remisión. Debe ser único.")
                if _folio_med != remision.folio_medline:
                    remision.folio_medline = _folio_med
                    remision.save(update_fields=['folio_medline'])
            _update_inventory_from_remision(remision, revert=False) # Re-aplicamos inventario
            enviar_alerta_merma(remision)

            return JsonResponse({'success': True, 'remision_id': remision.pk}, status=200)

    except Exception as e:
        return JsonResponse({'error': 'Error en servidor', 'detail': str(e)}, status=500)


@csrf_exempt
def api_obtener_catalogos(request, empresa_id):
    """ Retorna los catálogos estructurados {id, text} para los comboboxes de React """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)
    # Un no-superusuario solo ve catálogos de las empresas que tiene autorizadas.
    if not request.user.is_superuser:
        perfil = getattr(request.user, 'ternium_profile', None)
        if not perfil or not perfil.empresas_autorizadas.filter(pk=empresa_id).exists():
            return JsonResponse({'error': 'FORBIDDEN', 'detail': 'Sin acceso a esta empresa.'}, status=403)
    if request.method == 'GET':
        try:
            materiales = list(Material.objects.filter(empresas__id=empresa_id).annotate(text=F('nombre')).values('id', 'text'))
            origenes = list(Lugar.objects.filter(empresas__id=empresa_id).annotate(text=F('nombre')).values('id', 'text'))
            destinos = list(Lugar.objects.filter(empresas__id=empresa_id, tipo__in=['DESTINO', 'AMBOS']).annotate(text=F('nombre')).values('id', 'text'))
            
            lineas = list(LineaTransporte.objects.all().annotate(text=F('nombre')).values('id', 'text'))
            operadores = list(Operador.objects.all().annotate(text=F('nombre')).values('id', 'text'))
            unidades = list(Unidad.objects.all().annotate(text=F('numero_economico')).values('id', 'text'))
            contenedores = list(Contenedor.objects.all().annotate(text=F('numero')).values('id', 'text'))
            patios = list(Lugar.objects.filter(es_patio=True).annotate(text=F('nombre')).values('id', 'text'))
            # Mapeos (origen, material) que disparan Destrucción/Báscula — paridad con verificarDestruccion de Django.
            configs_destruccion = list(ConfiguracionManifiesto.objects.values('origen_id', 'material_id'))

            return JsonResponse({
                'materiales': materiales, 'lugares_origen': origenes, 'lugares_destino': destinos,
                'lineas_transporte': lineas, 'operadores': operadores, 'unidades': unidades,
                'contenedores': contenedores, 'patios': patios,
                'configs_destruccion': configs_destruccion,
            }, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

# 🔥 NUEVO ENDPOINT PARA OBTENER LAS EMPRESAS DINÁMICAMENTE
@csrf_exempt
def api_obtener_empresas(request):
    """ Retorna la lista de todas las empresas para el frontend """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)
    if request.method == 'GET':
        try:
            empresas = list(Empresa.objects.annotate(text=F('nombre')).values('id', 'text'))
            return JsonResponse({'empresas': empresas}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# =========================================================================
# AGREGAR ESTE ENDPOINT EN tu api_views.py de Django
# =========================================================================
# Y en urls.py agregar:
#   path('api/remisiones/<int:pk>/detalle/', api_views.api_remision_detalle, name='api_remision_detalle'),
# =========================================================================

def api_remision_detalle(request, pk):
    """
    Retorna TODOS los datos de una remisión para la vista de detalle
    y para pre-cargar el formulario de edición.
    """
    # JSON 401 en vez de @login_required (que redirige a HTML y rompe el fetch).
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)
    remision = get_object_or_404(
        Remision.objects.select_related(
            'empresa', 'origen', 'destino', 'operador', 'linea_transporte',
            'unidad', 'contenedor'
        ).prefetch_related('detalles__material', 'evidencias', 'facturas', 'historial__usuario'),
        pk=pk
    )

    # Seguridad por empresa
    if not request.user.is_superuser:
        perfil = getattr(request.user, 'ternium_profile', None)
        if not perfil or not perfil.empresas_autorizadas.filter(pk=remision.empresa_id).exists():
            return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    # Detalles (materiales)
    detalles = []
    for det in remision.detalles.all():
        detalles.append({
            'id': det.pk,
            'material': det.material_id,
            'material_nombre': det.material.nombre if det.material else '-',
            'bultos': det.bultos,
            'peso_ld': float(det.peso_ld or 0),
            'peso_dlv': float(det.peso_dlv or 0),
            'peso_rechazado': float(det.peso_rechazado or 0),
            'patio_rechazo': det.patio_rechazo_id,
        })

    # Evidencias
    evidencias = []
    for ev in remision.evidencias.all():
        if ev.archivo and ev.archivo.name:
            evidencias.append({'id': ev.pk, 'url': ev.archivo.url, 'nombre': ev.archivo.name.split('/')[-1]})

    # Fotos destrucción
    fotos_destruccion = {}
    for campo in ['foto_ingreso', 'foto_ingreso_2', 'foto_vertido', 'foto_vertido_2', 'foto_destruccion', 'foto_destruccion_2']:
        archivo = getattr(remision, campo, None)
        if archivo and hasattr(archivo, 'name') and archivo.name:
            fotos_destruccion[campo] = archivo.url

    # Historial
    historial = []
    for h in sorted(remision.historial.all(), key=lambda x: x.fecha, reverse=True):
        historial.append({
            'fecha': h.fecha.strftime('%d/%m/%Y %H:%M'),
            'usuario': h.usuario.get_full_name() or h.usuario.username if h.usuario else 'Sistema',
            'cambio': h.cambio,
        })

    # Facturas
    facturas = [{'id': f.pk} for f in remision.facturas.all()]

    total_ld = float(remision.total_peso_ld or 0)
    total_dlv = float(remision.total_peso_dlv or 0)

    data = {
        'id': remision.pk,
        'remision': remision.remision,
        'fecha': remision.fecha.isoformat() if remision.fecha else '',
        'fecha_display': remision.fecha.strftime('%d/%m/%Y') if remision.fecha else '',
        'status': remision.status,
        'status_display': remision.get_status_display(),
        'comentario': remision.comentario or '',
        'trazabilidad_notas': remision.trazabilidad_notas or '',

        # Empresa
        'empresa': remision.empresa_id,
        'empresa_nombre': remision.empresa.nombre if remision.empresa else '',
        'empresa_prefijo': remision.empresa.prefijo if remision.empresa else '',

        # Origen / Destino
        'origen': remision.origen_id,
        'origen_nombre': remision.origen.nombre if remision.origen else '',
        'destino': remision.destino_id,
        'destino_nombre': remision.destino.nombre if remision.destino else '',

        # Transporte
        'linea_transporte': remision.linea_transporte_id,
        'linea_transporte_nombre': remision.linea_transporte.nombre if remision.linea_transporte else '',
        'operador': remision.operador_id,
        'operador_nombre': remision.operador.nombre if remision.operador else '',
        'operador_manual': remision.operador_manual or '',
        'unidad': remision.unidad_id,
        'unidad_nombre': str(remision.unidad) if remision.unidad else '',
        'unidad_manual': remision.unidad_manual or '',
        'placas_unidad_manual': remision.placas_unidad_manual or '',
        'contenedor': remision.contenedor_id,
        'contenedor_nombre': str(remision.contenedor) if remision.contenedor else '',
        'contenedor_manual': remision.contenedor_manual or '',
        'placas_contenedor_manual': remision.placas_contenedor_manual or '',

        # Folios operativos
        'folio_ld': remision.folio_ld or '',
        'folio_dlv': remision.folio_dlv or '',
        'inicia_ld': str(remision.inicia_ld) if remision.inicia_ld else '',
        'termina_ld': str(remision.termina_ld) if remision.termina_ld else '',
        'inicia_dlv': str(remision.inicia_dlv) if remision.inicia_dlv else '',
        'termina_dlv': str(remision.termina_dlv) if remision.termina_dlv else '',

        # Destrucción fiscal
        'fecha_destruccion': remision.fecha_destruccion.isoformat() if remision.fecha_destruccion else '',
        'hora_entrada': str(remision.hora_entrada) if remision.hora_entrada else '',
        'hora_salida': str(remision.hora_salida) if remision.hora_salida else '',
        'factura_nombre': remision.factura_nombre or '',
        'comentarios_destruccion': remision.comentarios_destruccion or '',
        'destruccion_material_1': remision.destruccion_material_1 or '',
        'destruccion_peso_1': float(remision.destruccion_peso_1) if remision.destruccion_peso_1 else None,
        'destruccion_material_2': remision.destruccion_material_2 or '',
        'destruccion_peso_2': float(remision.destruccion_peso_2) if remision.destruccion_peso_2 else None,

        # Medline
        'folio_medline': remision.folio_medline or '',
        'boleta_salida_medline': remision.boleta_salida_medline.url if remision.boleta_salida_medline and remision.boleta_salida_medline.name else None,
        'manifiesto': remision.manifiesto.url if remision.manifiesto and remision.manifiesto.name else None,

        # Totales
        'total_peso_ld': total_ld,
        'total_peso_dlv': total_dlv,
        'diff': round(total_dlv - total_ld, 2),
        'porcentaje_merma': float(remision.porcentaje_merma) if remision.porcentaje_merma else 0,

        # Relaciones
        'detalles': detalles,
        'evidencias': evidencias,
        'fotos_destruccion': fotos_destruccion,
        'facturas': facturas,
        'historial': historial,

        # Permisos
        'can_edit': request.user.has_perm('ternium.change_remision') and remision.status != 'AUDITADO',
    }

    return JsonResponse(data)


# =========================================================================
# EXPORTS Y ACCIONES DE REMISIONES (MOVIDOS DESDE views.py)
# =========================================================================

@login_required
def export_remisiones_to_excel(request):
    # 1. Recuperar filtros del GET
    tipo_reporte = request.GET.get('tipo_reporte', 'normal')

    # --- NUEVO: Capturamos el material seleccionado para Medline ---
    material_medline = request.GET.get('material_medline', '')
    # ---------------------------------------------------------------

    # Nuevos parámetros para Medline (Cartón y Archivo Muerto)
    precio_carton_req = request.GET.get('precio_carton', '0')
    precio_archivo_req = request.GET.get('precio_archivo', '0')
    mes_medline = request.GET.get('mes', '')
    # Diálogo de descarga: con_tabla='1' incluye la tabla de remisiones; '0' = solo manifiesto (sin renglones).
    con_tabla = request.GET.get('con_tabla', '1')

    q_remision = request.GET.get('q_remision', '')
    q_prefijo = request.GET.get('q_prefijo', '')
    q_empresa = request.GET.get('q_empresa', '')
    q_material = request.GET.get('q_material', '')
    q_status = request.GET.get('q_status', '')
    q_origen = request.GET.get('q_origen', '')
    q_destino = request.GET.get('q_destino', '')
    q_operador = request.GET.get('q_operador', '')
    q_fecha_desde = request.GET.get('q_fecha_desde', '')
    q_fecha_hasta = request.GET.get('q_fecha_hasta', '')
    q_folio_ld = request.GET.get('q_folio_ld', '')
    q_folio_dlv = request.GET.get('q_folio_dlv', '')

    # Convertir precios a decimal de forma segura
    try:
        precio_carton_val = float(precio_carton_req)
        precio_archivo_val = float(precio_archivo_req)
    except ValueError:
        precio_carton_val = 0.0
        precio_archivo_val = 0.0

    # 2. Construir QuerySet Base
    queryset = Remision.objects.all().select_related(
        'empresa', 'origen', 'destino', 'linea_transporte', 'operador', 'unidad', 'contenedor'
    ).prefetch_related('detalles', 'detalles__material', 'evidencias')

    # =========================================================================
    # SI ES MEDLINE: solo filtrar por mes y origen Medline, SIN filtros generales
    # =========================================================================
    if tipo_reporte == 'medline':
        queryset = queryset.filter(origen__nombre__icontains='MEDLINE')

        if mes_medline:
            try:
                year, month = mes_medline.split('-')
                queryset = queryset.filter(fecha__year=year, fecha__month=month)
            except ValueError:
                pass

            PrecioMedline.objects.update_or_create(
                mes=mes_medline,
                defaults={
                    'precio_carton': precio_carton_val,
                    'precio_archivo': precio_archivo_val
                }
            )

    # =========================================================================
    # TODOS LOS DEMÁS REPORTES: aplican filtros generales normalmente
    # =========================================================================
    else:
        if q_remision: queryset = queryset.filter(remision__icontains=q_remision)
        if q_prefijo: queryset = queryset.filter(remision__istartswith=q_prefijo)
        if q_empresa: queryset = queryset.filter(empresa_id=q_empresa)
        if q_material: queryset = queryset.filter(detalles__material__id=q_material).distinct()
        if q_status: queryset = queryset.filter(status=q_status)
        if q_origen: queryset = queryset.filter(origen__id=q_origen)
        if q_destino: queryset = queryset.filter(destino__id=q_destino)
        if q_operador: queryset = queryset.filter(operador__id=q_operador)
        if q_fecha_desde: queryset = queryset.filter(fecha__gte=q_fecha_desde)
        if q_fecha_hasta: queryset = queryset.filter(fecha__lte=q_fecha_hasta)
        if q_folio_ld: queryset = queryset.filter(folio_ld__icontains=q_folio_ld)
        if q_folio_dlv: queryset = queryset.filter(folio_dlv__icontains=q_folio_dlv)

    queryset = queryset.order_by('fecha', 'pk')

    # 3. Crear Libro y Hoja
    wb = Workbook()
    ws = wb.active

    # --- ESTILOS COMUNES ---
    center_style = Alignment(horizontal="center", vertical="center", wrap_text=False)
    red_font = Font(color="FF0000", bold=True)
    green_font = Font(color="009900", bold=True)
    link_font = Font(color="0563C1", underline="single")
    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # =========================================================================
    # LÓGICA REPORTE MEDLINE — dos pestañas: Cartón y Archivo Muerto
    # =========================================================================
    if tipo_reporte == 'medline':
        pink_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_pink = PatternFill(start_color="FF66B2", end_color="FF66B2", fill_type="solid")
        yellow_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_header = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        medline_headers = [
            "Remisión", "Origen", "Material", "Peso Carga (Kg)",
            "Folio Carga", "Inicia Carga", "PRECIO", "VENTA", "BULTOS", "FOLIO"
        ]

        def _estilo_encabezados_medline(sheet):
            for col_idx, cell in enumerate(sheet[1], start=1):
                cell.alignment = center_style
                cell.font = Font(bold=True, color="FFFFFF" if col_idx <= 6 else "000000")
                if col_idx <= 6: cell.fill = header_pink
                elif col_idx in [7, 8]: cell.fill = yellow_header
                elif col_idx in [9, 10]: cell.fill = green_header

        def _poblar_hoja_medline(sheet, tipo_mat, solo_totales=False):
            """
            tipo_mat: 'carton' | 'archivo'
            Rellena la hoja con las remisiones que correspondan al material indicado.
            Devuelve (total_peso, total_venta).
            """
            total_peso = 0.0
            total_venta = 0.0
            for remision in queryset:
                val_inicia = timezone.localtime(remision.inicia_ld).date() if remision.inicia_ld else remision.fecha
                fecha_str = val_inicia.strftime("%d/%m/%Y") if val_inicia else ""
                for d in remision.detalles.all():
                    mat_nom = d.material.nombre.upper() if d.material else "CARTON"
                    es_archivo = "ARCHIVO" in mat_nom
                    es_carton = "CARTON" in mat_nom or "CARTÓN" in mat_nom
                    if tipo_mat == 'archivo' and not es_archivo:
                        continue
                    if tipo_mat == 'carton' and not es_carton:
                        continue
                    bultos = int(d.bultos or 0)
                    peso_ld = float(d.peso_ld or 0)
                    precio_mostrado = precio_archivo_val if es_archivo else precio_carton_val
                    venta_calculada = peso_ld * precio_mostrado
                    total_peso += peso_ld
                    total_venta += venta_calculada
                    if solo_totales:
                        continue  # "Solo manifiesto": sin renglones, solo la fila de totales.
                    folio_generado = remision.folio_medline or "N/A"
                    row_data = [
                        remision.remision,
                        remision.origen.nombre if remision.origen else 'MEDLINE',
                        mat_nom,
                        peso_ld,
                        remision.folio_ld or '',
                        fecha_str,
                        precio_mostrado,
                        venta_calculada,
                        bultos,
                        folio_generado
                    ]
                    sheet.append(row_data)
                    current_row = sheet.max_row
                    for col_idx in range(1, 11):
                        cell = sheet.cell(row=current_row, column=col_idx)
                        cell.alignment = center_style
                        if col_idx <= 6: cell.fill = pink_fill
                        if col_idx == 4: cell.number_format = '#,##0.000'
                        if col_idx == 7: cell.number_format = '"$"#,##0.00'
                        if col_idx == 8: cell.number_format = '"$"#,##0.00'
            # Fila de totales
            sheet.append(["", "", "", total_peso, "", "", "", total_venta, "", ""])
            last = sheet.max_row
            sheet.cell(row=last, column=4).font = Font(bold=True)
            sheet.cell(row=last, column=4).number_format = '#,##0.000'
            sheet.cell(row=last, column=8).font = Font(bold=True)
            sheet.cell(row=last, column=8).number_format = '"$"#,##0.00'
            return total_peso, total_venta

        # Opciones del diálogo de descarga: "solo manifiesto" (sin tabla) vs "con
        # tabla", y filtro por material (Cartón / Archivo / Ambos).
        _con_tabla = (con_tabla != '0')
        _mat = (material_medline or '').strip().lower()
        _incluir_carton = _mat in ('', 'ambos', 'carton', 'cartón')
        _incluir_archivo = _mat in ('', 'ambos', 'archivo')

        # --- Pestaña 1: Cartón ---
        ws.title = "Cartón"
        ws.append(medline_headers)
        _estilo_encabezados_medline(ws)
        if _incluir_carton:
            _poblar_hoja_medline(ws, 'carton', solo_totales=not _con_tabla)

        # --- Pestaña 2: Archivo Muerto ---
        ws_archivo = wb.create_sheet(title="Archivo Muerto")
        ws_archivo.append(medline_headers)
        _estilo_encabezados_medline(ws_archivo)
        if _incluir_archivo:
            _poblar_hoja_medline(ws_archivo, 'archivo', solo_totales=not _con_tabla)

    # =========================================================================
    # LÓGICA REPORTE CONCENTRADO
    # =========================================================================
    elif tipo_reporte == 'concentrado':
        ws.title = "Resumen Concentrado"
        headers = [
            "Remisión", "Fecha", "Estatus", "Empresa", "Origen", "Destino",
            "Transporte", "Operador", "Unidad", "Contenedor", "Materiales",
            "Desglose Carga (Kg)", "Total Carga (Kg)", "Folio Carga", "Inicia Carga",
            "Desglose Descarga (Kg)", "Total Descarga (Kg)", "Folio Descarga", "Termina Descarga",
            "Total Rechazo (Kg)", "Dif. Neta (Kg)", "% Merma Global",
            "Comentarios", "Links Evidencias"
        ]
        ws.append(headers)
        for cell in ws[1]: cell.alignment = center_style; cell.font = header_font

        for remision in queryset:
            is_pendiente = (remision.status == 'PENDIENTE')
            evidencias_urls = [request.build_absolute_uri(ev.archivo.url) for ev in remision.evidencias.all() if ev.archivo]
            texto_evidencias = "Descargar Evidencia" if len(evidencias_urls) == 1 else ("Múltiples Archivos" if len(evidencias_urls) > 1 else "Sin adjuntos")

            val_inicia = timezone.localtime(remision.inicia_ld).date() if remision.inicia_ld else ''
            val_termina = timezone.localtime(remision.termina_dlv).date() if remision.termina_dlv else ''

            detalles = remision.detalles.all()
            mat_nombres = " | ".join([d.material.nombre for d in detalles if d.material]) or "Sin Material"

            if detalles.count() > 1:
                desglose_ld = " | ".join([f"{float(d.peso_ld or 0):.2f}" for d in detalles])
                desglose_dlv = " | ".join([f"{float(d.peso_dlv or 0):.2f}" for d in detalles])
            else:
                desglose_ld = "N/A"
                desglose_dlv = "N/A"

            total_ld = sum([float(d.peso_ld or 0) for d in detalles])
            total_dlv = sum([float(d.peso_dlv or 0) for d in detalles])
            total_rechazo = sum([float(d.peso_rechazado or 0) for d in detalles])

            if is_pendiente:
                diff = 0; porcentaje_merma = 0
            else:
                diff = (total_dlv + total_rechazo) - total_ld
                porcentaje_merma = (diff / total_ld) if total_ld > 0 else 0

            rechazo_val = total_rechazo if total_rechazo > 0 else "N/A"

            row_data = [
                remision.remision, remision.fecha, remision.get_status_display(),
                remision.empresa.nombre if remision.empresa else '',
                remision.origen.nombre if remision.origen else '',
                remision.destino.nombre if remision.destino else '',
                remision.linea_transporte.nombre if remision.linea_transporte else '',
                remision.operador.nombre if remision.operador else (remision.operador_manual or ''),
                remision.unidad.internal_id if remision.unidad else (remision.unidad_manual or ''),
                remision.contenedor.nombre if remision.contenedor else (remision.contenedor_manual or ''),
                mat_nombres,
                desglose_ld, total_ld, remision.folio_ld or '', val_inicia,
                desglose_dlv, total_dlv, remision.folio_dlv or '', val_termina,
                rechazo_val, diff, porcentaje_merma,
                remision.comentario or '', texto_evidencias
            ]
            ws.append(row_data)
            current_row = ws.max_row

            for cell in ws[current_row]:
                cell.alignment = center_style
                if is_pendiente: cell.fill = yellow_fill

            for col_idx in [2, 15, 19]: ws.cell(row=current_row, column=col_idx).number_format = 'dd/mm/yyyy'
            for col_idx in [13, 17, 21]: ws.cell(row=current_row, column=col_idx).number_format = '#,##0.000'
            if rechazo_val != "N/A": ws.cell(row=current_row, column=20).number_format = '#,##0.000'

            if not is_pendiente:
                cell_diff = ws.cell(row=current_row, column=21)
                if diff < 0: cell_diff.font = red_font
                elif diff > 0: cell_diff.font = green_font

            ws.cell(row=current_row, column=22).number_format = '0.00%'

            if len(evidencias_urls) == 1:
                cell_link = ws.cell(row=current_row, column=24)
                cell_link.hyperlink = evidencias_urls[0]; cell_link.font = link_font

    # =========================================================================
    # LÓGICA REPORTE A DETALLE
    # =========================================================================
    elif tipo_reporte == 'detallado':
        ws.title = "Remisiones a Detalle"
        headers = [
            "Remisión", "Fecha", "Estatus", "Empresa", "Origen", "Destino",
            "Transporte", "Operador", "Unidad", "Contenedor", "Material",
            "Peso Carga (Kg)", "Folio Carga", "Inicia Carga",
            "Peso Descarga (Kg)", "Folio Descarga", "Termina Descarga",
            "Rechazo", "Dif. (Kg)", "% Merma",
            "Comentarios", "Links Evidencias"
        ]
        ws.append(headers)
        for cell in ws[1]: cell.alignment = center_style; cell.font = header_font

        for remision in queryset:
            is_pendiente = (remision.status == 'PENDIENTE')
            evidencias_urls = [request.build_absolute_uri(ev.archivo.url) for ev in remision.evidencias.all() if ev.archivo]
            texto_evidencias = "Descargar Evidencia" if len(evidencias_urls) == 1 else ("Múltiples Archivos" if len(evidencias_urls) > 1 else "Sin adjuntos")

            val_inicia = timezone.localtime(remision.inicia_ld).date() if remision.inicia_ld else ''
            val_termina = timezone.localtime(remision.termina_dlv).date() if remision.termina_dlv else ''

            detalles = remision.detalles.all()
            if not detalles:
                detalles_list = [{'material': 'Sin Material', 'peso_ld': 0.0, 'peso_dlv': 0.0, 'peso_rechazado': 0.0}]
            else:
                detalles_list = [{
                    'material': d.material.nombre if d.material else 'Sin Material',
                    'peso_ld': float(d.peso_ld or 0),
                    'peso_dlv': float(d.peso_dlv or 0),
                    'peso_rechazado': float(d.peso_rechazado or 0)
                } for d in detalles]

            for d in detalles_list:
                peso_ld = d['peso_ld']
                peso_dlv = d['peso_dlv']
                peso_rechazado = d['peso_rechazado']

                if is_pendiente:
                    diff = 0; porcentaje_merma = 0
                else:
                    diff = (peso_dlv + peso_rechazado) - peso_ld
                    porcentaje_merma = (diff / peso_ld) if peso_ld > 0 else 0

                rechazo_val = peso_rechazado if peso_rechazado > 0 else "N/A"

                row_data = [
                    remision.remision, remision.fecha, remision.get_status_display(),
                    remision.empresa.nombre if remision.empresa else '',
                    remision.origen.nombre if remision.origen else '',
                    remision.destino.nombre if remision.destino else '',
                    remision.linea_transporte.nombre if remision.linea_transporte else '',
                    remision.operador.nombre if remision.operador else (remision.operador_manual or ''),
                    remision.unidad.internal_id if remision.unidad else (remision.unidad_manual or ''),
                    remision.contenedor.nombre if remision.contenedor else (remision.contenedor_manual or ''),
                    d['material'],
                    peso_ld, remision.folio_ld or '', val_inicia,
                    peso_dlv, remision.folio_dlv or '', val_termina,
                    rechazo_val, diff, porcentaje_merma,
                    remision.comentario or '', texto_evidencias
                ]

                ws.append(row_data)
                current_row = ws.max_row

                for cell in ws[current_row]:
                    cell.alignment = center_style
                    if is_pendiente: cell.fill = yellow_fill

                for col_idx in [2, 14, 17]: ws.cell(row=current_row, column=col_idx).number_format = 'dd/mm/yyyy'
                for col_idx in [12, 15, 19]: ws.cell(row=current_row, column=col_idx).number_format = '#,##0.000'
                if rechazo_val != "N/A": ws.cell(row=current_row, column=18).number_format = '#,##0.000'

                if not is_pendiente:
                    cell_diff = ws.cell(row=current_row, column=19)
                    if diff < 0: cell_diff.font = red_font
                    elif diff > 0: cell_diff.font = green_font

                ws.cell(row=current_row, column=20).number_format = '0.00%'

                if len(evidencias_urls) == 1:
                    cell_link = ws.cell(row=current_row, column=22)
                    cell_link.hyperlink = evidencias_urls[0]; cell_link.font = link_font

    # =========================================================================
    # LÓGICA REPORTE NORMAL
    # =========================================================================
    else:
        ws.title = "Remisiones (Normal)"
        headers = [
            "Remisión", "Fecha", "Estatus", "Empresa", "Origen", "Destino",
            "Transporte", "Operador", "Unidad", "Contenedor", "Materiales",
            "Peso Carga (Kg)", "Folio Carga", "Inicia Carga",
            "Peso Descarga (Kg)", "Folio Descarga", "Termina Descarga",
            "Rechazo", "Dif. (Kg)", "% Merma",
            "Comentarios", "Links Evidencias"
        ]
        ws.append(headers)
        for cell in ws[1]: cell.alignment = center_style; cell.font = header_font

        for remision in queryset:
            is_pendiente = (remision.status == 'PENDIENTE')
            evidencias_urls = [request.build_absolute_uri(ev.archivo.url) for ev in remision.evidencias.all() if ev.archivo]
            texto_evidencias = "Descargar Evidencia" if len(evidencias_urls) == 1 else ("Múltiples Archivos" if len(evidencias_urls) > 1 else "Sin adjuntos")

            val_inicia = timezone.localtime(remision.inicia_ld).date() if remision.inicia_ld else ''
            val_termina = timezone.localtime(remision.termina_dlv).date() if remision.termina_dlv else ''

            detalles = remision.detalles.all()
            materiales = ", ".join([d.material.nombre for d in detalles if d.material]) or "Sin Material"

            peso_ld = sum([float(d.peso_ld or 0) for d in detalles])
            peso_dlv = sum([float(d.peso_dlv or 0) for d in detalles])
            peso_rechazado = sum([float(d.peso_rechazado or 0) for d in detalles])

            if is_pendiente:
                diff = 0; porcentaje_merma = 0
            else:
                diff = (peso_dlv + peso_rechazado) - peso_ld
                porcentaje_merma = (diff / peso_ld) if peso_ld > 0 else 0

            rechazo_val = peso_rechazado if peso_rechazado > 0 else "N/A"

            row_data = [
                remision.remision, remision.fecha, remision.get_status_display(),
                remision.empresa.nombre if remision.empresa else '',
                remision.origen.nombre if remision.origen else '',
                remision.destino.nombre if remision.destino else '',
                remision.linea_transporte.nombre if remision.linea_transporte else '',
                remision.operador.nombre if remision.operador else (remision.operador_manual or ''),
                remision.unidad.internal_id if remision.unidad else (remision.unidad_manual or ''),
                remision.contenedor.nombre if remision.contenedor else (remision.contenedor_manual or ''),
                materiales,
                peso_ld, remision.folio_ld or '', val_inicia,
                peso_dlv, remision.folio_dlv or '', val_termina,
                rechazo_val, diff, porcentaje_merma,
                remision.comentario or '', texto_evidencias
            ]

            ws.append(row_data)
            current_row = ws.max_row

            for cell in ws[current_row]:
                cell.alignment = center_style
                if is_pendiente: cell.fill = yellow_fill

            for col_idx in [2, 14, 17]: ws.cell(row=current_row, column=col_idx).number_format = 'dd/mm/yyyy'
            for col_idx in [12, 15, 19]: ws.cell(row=current_row, column=col_idx).number_format = '#,##0.000'
            if rechazo_val != "N/A": ws.cell(row=current_row, column=18).number_format = '#,##0.000'

            if not is_pendiente:
                cell_diff = ws.cell(row=current_row, column=19)
                if diff < 0: cell_diff.font = red_font
                elif diff > 0: cell_diff.font = green_font

            ws.cell(row=current_row, column=20).number_format = '0.00%'

            if len(evidencias_urls) == 1:
                cell_link = ws.cell(row=current_row, column=22)
                cell_link.hyperlink = evidencias_urls[0]; cell_link.font = link_font

    if ws.max_row > 1 and tipo_reporte != 'medline':
        last_col_letter = get_column_letter(len(headers))
        full_range = f"A1:{last_col_letter}{ws.max_row}"
        tab = Table(displayName="TablaRemisiones", ref=full_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def _auto_ancho(sheet):
        for col in sheet.columns:
            max_length = 0; column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column_letter].width = min(max_length + 3, 60)

    _auto_ancho(ws)
    if tipo_reporte == 'medline':
        _auto_ancho(ws_archivo)

    nombre_base = f"Remisiones_{tipo_reporte.capitalize()}"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_base}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_reportes_especificos(request):
    tipo_reporte = request.GET.get('tipo_reporte', '')

    # Nuevos parámetros de fechas
    tipo_filtro = request.GET.get('tipo_filtro', 'mes') # 'mes' o 'rango'
    mes_req = request.GET.get('mes', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    precio_hudson_req = request.GET.get('precio_hudson', '0')
    try:
        precio_hudson_val = float(precio_hudson_req)
    except ValueError:
        precio_hudson_val = 0.0

    # 1. Base Queryset (Solo Terminados y Auditados)
    queryset = Remision.objects.filter(status__in=['TERMINADO', 'AUDITADO']).select_related(
        'empresa', 'origen', 'destino', 'linea_transporte', 'operador', 'unidad'
    ).prefetch_related('detalles', 'detalles__material', 'evidencias')

    # 2. Filtrar por tipo de reporte (Origen)
    if tipo_reporte in ['huaxing', 'bachoco', 'trane']:
        queryset = queryset.filter(origen__nombre__icontains=tipo_reporte.upper())
    elif tipo_reporte == 'hudson':
        queryset = queryset.filter(origen__nombre__icontains='HUDSON')
    else:
        return HttpResponse("Tipo de reporte no válido.", status=400)

    # 3. Aplicar Filtro Dinámico de Fechas
    if tipo_filtro == 'mes' and mes_req:
        try:
            year, month = mes_req.split('-')
            queryset = queryset.filter(fecha__year=year, fecha__month=month)
        except ValueError:
            pass
    elif tipo_filtro == 'rango':
        if fecha_inicio:
            queryset = queryset.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__lte=fecha_fin)

    # 4. Preparar el Excel
    wb = Workbook()
    ws = wb.active
    center_style = Alignment(horizontal='center', vertical='center')

    # ==========================================
    # LÓGICA PARA INSERCIÓN DE IMAGEN DESDE S3
    # ==========================================
    # Dejamos 5 filas vacías para que la tabla comience en la fila 6
    start_row_table = 6
    for _ in range(start_row_table - 1):
        ws.append([])

    # Mapa de imágenes en S3 (Exactamente con las extensiones que proporcionaste)
    mapa_imagenes = {
        'bachoco': 'BACHOCO.png',
        'huaxing': 'HUAXING.jpg',
        'hudson': 'HUDSON.png',
        'trane': 'TRANE.png'
    }

    nombre_img = mapa_imagenes.get(tipo_reporte)
    if nombre_img:
        try:
            # Reemplaza '3rrecycling' por el nombre exacto de tu bucket si es distinto en la URL
            s3_url = f"https://3rrecycling.s3.amazonaws.com/static/Fotos_Reportes/{nombre_img}"
            req = urllib.request.Request(s3_url, headers={'User-Agent': 'Mozilla/5.0'})

            with urllib.request.urlopen(req) as response:
                img_data = BytesIO(response.read())

            logo = OpenXLImage(img_data)

            # Ajuste de tamaño manteniendo la proporción (Alto fijo en 75px)
            aspect_ratio = logo.width / logo.height
            logo.height = 75
            logo.width = 75 * aspect_ratio

            ws.add_image(logo, 'A1') # Se coloca en la celda A1
        except Exception as e:
            print(f"No se pudo cargar la imagen de S3: {e}") # Falla en silencio para que el Excel sí se descargue

    # =========================================================================
    # ESTRUCTURA PARA HUAXING, BACHOCO O TRANE
    # =========================================================================
    if tipo_reporte in ['huaxing', 'bachoco', 'trane']:
        ws.title = f"Reporte_{tipo_reporte.capitalize()}"
        headers = [
            "Remisión", "Origen", "Destino", "Fecha", "Folio Carga",
            "Folio Descarga", "Operador", "Material", "Peso Carga (Kg)",
            "Peso Descarga (Kg)", "Links Evidencias"
        ]
        ws.append(headers)

        # Estilo encabezados (que ahora están en ws[start_row_table])
        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[start_row_table]:
            cell.alignment = center_style
            cell.font = header_font
            cell.fill = yellow_fill

        for remision in queryset:
            evidencias_urls = [request.build_absolute_uri(ev.archivo.url) for ev in remision.evidencias.all() if ev.archivo]
            texto_evidencias = "Descargar Evidencia" if len(evidencias_urls) == 1 else ("Múltiples Archivos" if len(evidencias_urls) > 1 else "Sin adjuntos")

            for d in remision.detalles.all():
                row_data = [
                    remision.remision,
                    remision.origen.nombre if remision.origen else '',
                    remision.destino.nombre if remision.destino else '',
                    remision.fecha.strftime("%d/%m/%Y") if remision.fecha else "",
                    remision.folio_ld or '',
                    remision.folio_dlv or '',
                    remision.operador.nombre if remision.operador else (remision.operador_manual or ''),
                    d.material.nombre if d.material else "S/M",
                    float(d.peso_ld or 0),
                    float(d.peso_dlv or 0),
                    texto_evidencias
                ]
                ws.append(row_data)

                current_row = ws.max_row
                for col_idx in range(1, 12):
                    ws.cell(row=current_row, column=col_idx).alignment = center_style

                ws.cell(row=current_row, column=9).number_format = '#,##0.000'
                ws.cell(row=current_row, column=10).number_format = '#,##0.000'

                # Link de evidencia
                if len(evidencias_urls) == 1:
                    cell_link = ws.cell(row=current_row, column=11)
                    cell_link.hyperlink = evidencias_urls[0]
                    cell_link.font = Font(color="0000FF", underline="single")

    # =========================================================================
    # ESTRUCTURA PARA HUDSON
    # =========================================================================
    elif tipo_reporte == 'hudson':
        ws.title = "Reporte_Hudson"
        headers = [
            "Remisión", "Empresa", "Operador", "Unidad", "Material",
            "Peso Carga (Kg)", "Folio Carga", "Inicia Carga", "PRECIO", "VENTA"
        ]
        ws.append(headers)

        pink_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_pink = PatternFill(start_color="FF66B2", end_color="FF66B2", fill_type="solid")
        green_header = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        for col_idx, cell in enumerate(ws[start_row_table], start=1):
            cell.alignment = center_style
            cell.font = Font(bold=True, color="FFFFFF" if col_idx <= 8 else "000000")
            if col_idx <= 8: cell.fill = header_pink
            elif col_idx in [9, 10]: cell.fill = green_header

        total_peso = 0
        total_venta = 0

        for remision in queryset:
            val_inicia = timezone.localtime(remision.inicia_ld).date() if remision.inicia_ld else remision.fecha
            fecha_str = val_inicia.strftime("%d/%m/%Y") if val_inicia else ""

            for d in remision.detalles.all():
                mat_nom = d.material.nombre.upper() if d.material else "S/M"
                peso_ld = float(d.peso_ld or 0)

                venta_calculada = peso_ld * precio_hudson_val

                total_peso += peso_ld
                total_venta += venta_calculada

                row_data = [
                    remision.remision,
                    remision.empresa.nombre if remision.empresa else 'HUDSON',
                    remision.operador.nombre if remision.operador else (remision.operador_manual or ''),
                    remision.unidad.internal_id if remision.unidad else (remision.unidad_manual or ''),
                    mat_nom,
                    peso_ld,
                    remision.folio_ld or '',
                    fecha_str,
                    precio_hudson_val,
                    venta_calculada
                ]
                ws.append(row_data)

                current_row = ws.max_row
                for col_idx in range(1, 11):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.alignment = center_style
                    if col_idx <= 8: cell.fill = pink_fill
                    if col_idx == 6: cell.number_format = '#,##0.000'
                    if col_idx == 9: cell.number_format = '"$"#,##0.00'
                    if col_idx == 10: cell.number_format = '"$"#,##0.00'

        # Fila de Totales de Hudson
        ws.append(["", "", "", "", "TOTALES:", total_peso, "", "", "", total_venta])
        last_row = ws.max_row
        ws.cell(row=last_row, column=6).font = Font(bold=True)
        ws.cell(row=last_row, column=6).number_format = '#,##0.000'
        ws.cell(row=last_row, column=10).font = Font(bold=True)
        ws.cell(row=last_row, column=10).number_format = '"$"#,##0.00'

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length and cell.value:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # 5. Retornar el archivo Excel
    nombre_archivo = f"Reporte_{tipo_reporte.capitalize()}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response


@csrf_exempt
@require_POST
def exportar_zip_medline(request):
    """
    Exporta en un archivo ZIP las Boletas de Salida de remisiones cuyo
    origen sea MEDLINE, filtradas opcionalmente por:
      - rango de fechas: fecha_inicio, fecha_fin (sobre Remision.fecha)
      - rango de folios: folio_inicio, folio_fin (sobre Remision.folio_medline,
        con respaldo en Remision.remision si folio_medline está vacío)
    Devuelve siempre JSON en caso de error (consumido desde la SPA Next.js).
    """
    queryset = Remision.objects.filter(
        origen__nombre__icontains='MEDLINE'
    ).exclude(boleta_salida_medline__exact='').exclude(boleta_salida_medline__isnull=True)

    fecha_inicio = (request.POST.get('fecha_inicio') or request.GET.get('fecha_inicio') or '').strip()
    fecha_fin    = (request.POST.get('fecha_fin')    or request.GET.get('fecha_fin')    or '').strip()
    folio_inicio = (request.POST.get('folio_inicio') or request.GET.get('folio_inicio') or '').strip()
    folio_fin    = (request.POST.get('folio_fin')    or request.GET.get('folio_fin')    or '').strip()

    if fecha_inicio and fecha_fin:
        try:
            queryset = queryset.filter(fecha__range=(fecha_inicio, fecha_fin))
        except Exception:
            return JsonResponse({'error': 'Rango de fechas inválido. Usa formato YYYY-MM-DD.'}, status=400)

    if folio_inicio and folio_fin:
        queryset = queryset.filter(
            Q(folio_medline__gte=folio_inicio, folio_medline__lte=folio_fin) |
            Q(remision__gte=folio_inicio, remision__lte=folio_fin)
        )

    if not queryset.exists():
        return JsonResponse({
            'error': 'No se encontraron boletas MEDLINE con archivo adjunto para los filtros indicados.'
        }, status=404)

    buffer = BytesIO()
    archivos_agregados = 0
    errores_lectura = []

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for remision in queryset:
            archivo = remision.boleta_salida_medline
            if not (archivo and hasattr(archivo, 'name') and archivo.name):
                continue
            try:
                ext = archivo.name.split('.')[-1] if '.' in archivo.name else 'pdf'
                folio_ld_part   = remision.folio_ld  or 'SinFolioCarga'
                folio_dlv_part  = remision.folio_dlv or 'SinFolioDescarga'
                remision_part   = remision.remision or remision.folio_medline or str(remision.pk)
                nuevo_nombre    = f"{remision_part}_{folio_ld_part}_{folio_dlv_part}.{ext}"

                file_data = None
                try:
                    with archivo.open('rb') as f:
                        file_data = f.read()
                except Exception:
                    try:
                        req = urllib.request.Request(archivo.url, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=30) as response:
                            file_data = response.read()
                    except Exception as e_url:
                        errores_lectura.append(f"{nuevo_nombre}: {e_url}")
                        continue

                if file_data:
                    zip_file.writestr(nuevo_nombre, file_data)
                    archivos_agregados += 1
            except Exception as e:
                errores_lectura.append(f"id={remision.id}: {e}")
                continue

    if archivos_agregados == 0:
        return JsonResponse({
            'error': 'Se encontraron registros pero no fue posible leer los archivos en S3.',
            'detalles': errores_lectura[:10],
        }, status=502)

    buffer.seek(0)
    filename_parts = ['Boletas_MEDLINE']
    if fecha_inicio and fecha_fin:
        filename_parts.append(f'{fecha_inicio}_a_{fecha_fin}')
    elif folio_inicio and folio_fin:
        filename_parts.append(f'{folio_inicio}_a_{folio_fin}')
    filename = '_'.join(filename_parts) + '.zip'

    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['X-Archivos-Agregados'] = str(archivos_agregados)
    return response


@login_required
def descargar_reporte_destruccion(request, remision_id):
    try:
        remision = get_object_or_404(Remision, pk=remision_id)

        origen_nombre = remision.origen.nombre if remision.origen else "S/N"
        operador_nombre = remision.operador.nombre if remision.operador else (remision.operador_manual or "S/N")

        unidad_nombre = remision.unidad.internal_id if remision.unidad else (remision.unidad_manual or "S/N")
        placas_unidad = getattr(remision.unidad, 'license_plate', getattr(remision.unidad, 'placas', "S/N")) if remision.unidad else (getattr(remision, 'placas_unidad_manual', "S/N") or "S/N")

        contenedor_nombre = remision.contenedor.nombre if remision.contenedor else (remision.contenedor_manual or "S/N")
        placas_cont = getattr(remision.contenedor, 'placas', "S/N") if remision.contenedor else (getattr(remision, 'placas_contenedor_manual', "S/N") or "S/N")

        template_path = os.path.join(settings.BASE_DIR, 'templates_word', 'plantilla_destruccion.docx')
        from docxtpl import DocxTemplate, InlineImage
        from docx.shared import Mm
        doc = DocxTemplate(template_path)

        def get_inline_image(image_field, width_mm=55):
            if not image_field or not image_field.name:
                return ""
            try:
                s3_client = boto3.client('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID, aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY, region_name=settings.AWS_S3_REGION_NAME)
                s3_key = f"{settings.AWS_MEDIA_LOCATION}/{image_field.name}"
                file_stream = io.BytesIO()
                s3_client.download_fileobj(settings.AWS_STORAGE_BUCKET_NAME, s3_key, file_stream)
                file_stream.seek(0)
                return InlineImage(doc, file_stream, width=Mm(width_mm))
            except Exception as e:
                print(f"Error cargando imagen para Word: {e}")
                return ""

        # Verificamos si existe un Material 2 seleccionado
        tiene_mat2 = bool(remision.destruccion_material_2)

        # --- DICCIONARIO DE DATOS PARA EL ARCHIVO DE WORD ---
        context = {
            'fecha': remision.fecha_destruccion.strftime('%d/%m/%Y') if remision.fecha_destruccion else (remision.fecha.strftime('%d/%m/%Y') if remision.fecha else ""),
            'folio': remision.remision,
            'origen': origen_nombre,
            'operador': operador_nombre,

            # --- FILA 1 ---
            'unidad': unidad_nombre,
            'placas_unidad': placas_unidad,
            'caja': contenedor_nombre,
            'placas_caja': placas_cont,
            'hora_inicio': remision.hora_entrada.strftime('%H:%M') if remision.hora_entrada else "",
            'hora_fin': remision.hora_salida.strftime('%H:%M') if remision.hora_salida else "",
            'material1': remision.destruccion_material_1 or "",
            'peso': f"{remision.destruccion_peso_1}" if remision.destruccion_peso_1 else "",

            # --- FILA 2 (Solo se llenan si hay material 2) ---
            'unidad2': unidad_nombre if tiene_mat2 else "",
            'placas_unidad2': placas_unidad if tiene_mat2 else "",
            'caja2': contenedor_nombre if tiene_mat2 else "",
            'placas_caja2': placas_cont if tiene_mat2 else "",
            'hora_inicio2': remision.hora_entrada.strftime('%H:%M') if (remision.hora_entrada and tiene_mat2) else "",
            'hora_fin2': remision.hora_salida.strftime('%H:%M') if (remision.hora_salida and tiene_mat2) else "",
            'material2': remision.destruccion_material_2 or "",
            'peso2': f"{remision.destruccion_peso_2}" if remision.destruccion_peso_2 else "",

            'suma': {'peso': remision.peso_bascula if remision.peso_bascula else 0},
            'Comentarios': remision.comentarios_destruccion or "",

            'foto1': get_inline_image(remision.foto_ingreso, 55),
            'foto1_2': get_inline_image(remision.foto_ingreso_2, 55),
            'foto2': get_inline_image(remision.foto_vertido, 55),
            'foto2_1': get_inline_image(remision.foto_vertido_2, 55),
            'foto3': get_inline_image(remision.foto_destruccion, 55),
            'foto3_1': get_inline_image(remision.foto_destruccion_2, 55),
        }

        doc.render(context)

        buffer = io.BytesIO()
        doc.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Destruccion_{remision.remision}.docx"'
        return response

    except Exception as e:
        messages.error(request, f"Ocurrió un error al generar el Word: {str(e)}")
        return redirect('remision_lista')


@login_required
@require_POST
@permission_required('ternium.can_audit_remision', raise_exception=True)
def auditar_remision(request, pk):
    remision = get_object_or_404(Remision, pk=pk)

    if remision.status == 'TERMINADO':
        try:
            with transaction.atomic():
                remision.status = 'AUDITADO'
                remision.auditado_por = request.user
                remision.auditado_en = timezone.now()
                remision.save()

                HistorialRemision.objects.create(
                    remision=remision,
                    usuario=request.user,
                    cambio="ESTATUS CAMBIADO A: AUDITADO"
                )

            messages.success(request, f'La remisión {remision.remision} ha sido auditada.')
        except Exception as e:
            messages.error(request, f"Error al auditar: {e}")
    else:
        messages.error(request, 'Esta remisión no puede ser auditada (debe estar TERMINADA).')

    return redirect('detalle_remision', pk=pk)


@permission_required('ternium.change_remision', raise_exception=True)
def cancelar_remision(request, pk):
    """
    Cancela una remisión, revierte el movimiento de inventario
    y evita que aparezca en el Dashboard.
    """
    remision = get_object_or_404(Remision, pk=pk)

    if remision.status == 'AUDITADO':
        messages.error(request, 'No se puede cancelar una remisión que ya fue auditada.')
        return redirect('remision_lista')

    if remision.status == 'CANCELADO':
        messages.warning(request, 'Esta remisión ya estaba cancelada.')
        return redirect('remision_lista')

    try:
        with transaction.atomic():
            # 1. Revertir inventario
            _update_inventory_from_remision(remision, revert=True)

            # 2. Guardar folio Medline antes de liberarlo
            folio_liberado = remision.folio_medline
            remision.folio_medline = None

            # 3. Actualizar estatus
            remision.status = 'CANCELADO'

            usuario_nombre = request.user.username
            fecha_str = timezone.now().strftime("%d/%m/%Y %H:%M")
            remision.descripcion += f" [CANCELADA por {usuario_nombre} el {fecha_str}]"

            remision.save()

            # 4. (Deshabilitado) Renumerar folios Medline posteriores.
            # Con el contador GLOBAL monotónico, los huecos NO se rellenan:
            # cada nuevo folio toma el max(existentes)+1 sin importar
            # cancelaciones. Esto evita reorganizar registros históricos.
            # _renumerar_folios_medline(folio_liberado)
            _ = folio_liberado  # silencio el linter para variable usada antes

            # --- HISTORIAL: REGISTRAR CANCELACIÓN ---
            HistorialRemision.objects.create(
                remision=remision,
                usuario=request.user,
                cambio="ESTATUS CAMBIADO A: CANCELADO (Inventario revertido)"
            )

        messages.success(request, f'La remisión {remision.remision} ha sido CANCELADA.')
    except Exception as e:
        messages.error(request, f'Error al cancelar: {e}')

    return redirect('remision_lista')


@require_POST
def eliminar_evidencia_individual(request, pk):
    evidencia = get_object_or_404(EvidenciaRemision, pk=pk)
    remision_id = evidencia.remision.pk

    # Seguridad
    if evidencia.remision.status == 'AUDITADO':
        messages.error(request, "No se puede eliminar evidencia de una remisión auditada.")
    else:
        # Borrar de S3
        if evidencia.archivo and hasattr(evidencia.archivo, 'name'):
            _eliminar_archivo_de_s3(evidencia.archivo.name)

        # Borrar de BD
        evidencia.delete()
        messages.success(request, "Archivo eliminado correctamente.")

    return redirect('editar_remision', pk=remision_id)


# =========================================================================
# CATÁLOGOS CRUD API + PERFIL + PERMISOS  (appended to api_views.py)
# =========================================================================
from .models import Profile

def _json_body(request):
    try:
        return json.loads(request.body), None
    except Exception:
        return None, JsonResponse({'error': 'JSON invalido'}, status=400)

def _perm_check(request, perm, *, alt=None):
    """
    Verifica que el usuario tenga `perm`. Si se pasa `alt` (str o lista),
    también se aceptan esos permisos como bypass.
    Uso típico: dentro del módulo de Viajes queremos que un usuario con
    `ternium.acceso_viajes` pueda editar lugares/unidades/etc. al vuelo
    sin necesidad de tener cada permiso granular del catálogo.
    """
    if request.user.is_superuser:
        return None
    if request.user.has_perm(perm):
        return None
    if alt:
        alts = [alt] if isinstance(alt, str) else list(alt)
        if any(request.user.has_perm(p) for p in alts):
            return None
    return JsonResponse({'error': 'FORBIDDEN', 'detail': f'Permiso requerido: {perm}'}, status=403)

# ── EMPRESAS ─────────────────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_cat_empresas(request):
    if request.method == 'GET':
        qs = Empresa.objects.all().order_by('nombre')
        q = request.GET.get('q', '')
        if q:
            qs = qs.filter(nombre__icontains=q)
        return JsonResponse({'results': [{'id': e.id, 'nombre': e.nombre, 'prefijo': e.prefijo or ''} for e in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_empresa')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        prefijo = body.get('prefijo', '').strip().upper()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if Empresa.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe una empresa con ese nombre'}, status=400)
        e = Empresa.objects.create(nombre=nombre, prefijo=prefijo or None)
        return JsonResponse({'id': e.id, 'nombre': e.nombre, 'prefijo': e.prefijo or ''}, status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_empresa_detail(request, pk):
    e = get_object_or_404(Empresa, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': e.id, 'nombre': e.nombre, 'prefijo': e.prefijo or ''})
    if request.method in ('PUT', 'PATCH'):
        err = _perm_check(request, 'ternium.change_empresa')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        if 'nombre' in body: e.nombre = body['nombre'].strip()
        if 'prefijo' in body: e.prefijo = body['prefijo'].strip().upper() or None
        e.save()
        return JsonResponse({'id': e.id, 'nombre': e.nombre, 'prefijo': e.prefijo or ''})
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_empresa')
        if err: return err
        e.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── helper: set M2M empresas ──────────────────────────────────────────────
def _set_empresas(obj, body):
    if 'empresas' in body:
        ids = [int(i) for i in body['empresas'] if str(i).isdigit()]
        obj.empresas.set(Empresa.objects.filter(id__in=ids))

def _empresas_ids(obj):
    return list(obj.empresas.values_list('id', flat=True))

# ── LUGARES ──────────────────────────────────────────────────────────────
_LUGAR_STR_FIELDS = (
    'nombre', 'tipo', 'razon_social', 'rfc', 'regimen_fiscal', 'uso_cfdi',
    'calle', 'numero_exterior', 'numero_interior', 'colonia',
    'codigo_postal', 'localidad', 'municipio', 'estado', 'pais',
)

def _lugar_to_dict(l):
    return {
        'id': l.id, 'nombre': l.nombre, 'tipo': l.tipo, 'es_patio': l.es_patio,
        'razon_social': l.razon_social or '', 'rfc': l.rfc or '',
        'regimen_fiscal': l.regimen_fiscal or '', 'uso_cfdi': l.uso_cfdi or '',
        'calle': l.calle or '', 'numero_exterior': l.numero_exterior or '',
        'numero_interior': l.numero_interior or '', 'colonia': l.colonia or '',
        'codigo_postal': l.codigo_postal or '', 'localidad': l.localidad or '',
        'municipio': l.municipio or '', 'estado': l.estado or '',
        'pais': l.pais or 'México',
        'empresas': _empresas_ids(l),
    }

def _lugar_from_body(l, body):
    for field in _LUGAR_STR_FIELDS:
        if field in body:
            setattr(l, field, body[field] or None)
    if 'es_patio' in body:
        l.es_patio = bool(body['es_patio'])

@csrf_exempt
@login_required
def api_cat_lugares(request):
    if request.method == 'GET':
        qs = Lugar.objects.prefetch_related('empresas').all().order_by('nombre')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(nombre__icontains=q)
        return JsonResponse({'results': [_lugar_to_dict(l) for l in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_lugar')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if Lugar.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe un lugar con ese nombre'}, status=400)
        l = Lugar(nombre=nombre)
        _lugar_from_body(l, body)
        l.save()
        _set_empresas(l, body)
        return JsonResponse(_lugar_to_dict(l), status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_lugar_detail(request, pk):
    l = get_object_or_404(Lugar, pk=pk)
    if request.method == 'GET':
        return JsonResponse(_lugar_to_dict(l))
    if request.method in ('PUT', 'PATCH'):
        # Editar lugar también se hace desde dentro del detalle de un viaje
        # (al editar una parada). Aceptamos acceso_viajes como bypass.
        err = _perm_check(request, 'ternium.change_lugar', alt='ternium.acceso_viajes')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        _lugar_from_body(l, body)
        l.save()
        _set_empresas(l, body)
        return JsonResponse(_lugar_to_dict(l))
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_lugar')
        if err: return err
        l.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── LINEAS DE TRANSPORTE ─────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_cat_lineas(request):
    if request.method == 'GET':
        qs = LineaTransporte.objects.prefetch_related('empresas').all().order_by('nombre')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(nombre__icontains=q)
        return JsonResponse({'results': [{'id': lt.id, 'nombre': lt.nombre, 'empresas': _empresas_ids(lt)} for lt in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_lineatransporte')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if LineaTransporte.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe una linea con ese nombre'}, status=400)
        lt = LineaTransporte.objects.create(nombre=nombre)
        _set_empresas(lt, body)
        return JsonResponse({'id': lt.id, 'nombre': lt.nombre, 'empresas': _empresas_ids(lt)}, status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_linea_detail(request, pk):
    lt = get_object_or_404(LineaTransporte, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': lt.id, 'nombre': lt.nombre, 'empresas': _empresas_ids(lt)})
    if request.method in ('PUT', 'PATCH'):
        err = _perm_check(request, 'ternium.change_lineatransporte')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        if 'nombre' in body: lt.nombre = body['nombre'].strip()
        lt.save()
        _set_empresas(lt, body)
        return JsonResponse({'id': lt.id, 'nombre': lt.nombre, 'empresas': _empresas_ids(lt)})
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_lineatransporte')
        if err: return err
        lt.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── OPERADORES ───────────────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_cat_operadores(request):
    if request.method == 'GET':
        qs = Operador.objects.prefetch_related('empresas').all().order_by('nombre')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(nombre__icontains=q)
        return JsonResponse({'results': [{'id': o.id, 'nombre': o.nombre, 'folio': o.folio or '', 'empresas': _empresas_ids(o)} for o in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_operador')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if Operador.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe un operador con ese nombre'}, status=400)
        o = Operador.objects.create(nombre=nombre, folio=body.get('folio', '') or None)
        _set_empresas(o, body)
        return JsonResponse({'id': o.id, 'nombre': o.nombre, 'folio': o.folio or '', 'empresas': _empresas_ids(o)}, status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_operador_detail(request, pk):
    o = get_object_or_404(Operador, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': o.id, 'nombre': o.nombre, 'folio': o.folio or '', 'empresas': _empresas_ids(o)})
    if request.method in ('PUT', 'PATCH'):
        err = _perm_check(request, 'ternium.change_operador')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        if 'nombre' in body: o.nombre = body['nombre'].strip()
        if 'folio' in body: o.folio = body['folio'].strip() or None
        o.save()
        _set_empresas(o, body)
        return JsonResponse({'id': o.id, 'nombre': o.nombre, 'folio': o.folio or '', 'empresas': _empresas_ids(o)})
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_operador')
        if err: return err
        o.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── MATERIALES ───────────────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_cat_materiales(request):
    if request.method == 'GET':
        qs = Material.objects.prefetch_related('empresas').all().order_by('nombre')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(nombre__icontains=q)
        return JsonResponse({'results': [{'id': m.id, 'nombre': m.nombre,
            'clave_sat': m.clave_sat or '', 'clave_unidad_sat': m.clave_unidad_sat or '',
            'empresas': _empresas_ids(m)} for m in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_material')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if Material.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe un material con ese nombre'}, status=400)
        m = Material.objects.create(
            nombre=nombre,
            clave_sat=body.get('clave_sat', '') or None,
            clave_unidad_sat=body.get('clave_unidad_sat', 'KGM') or 'KGM',
        )
        _set_empresas(m, body)
        return JsonResponse({'id': m.id, 'nombre': m.nombre,
            'clave_sat': m.clave_sat or '', 'clave_unidad_sat': m.clave_unidad_sat or '',
            'empresas': _empresas_ids(m)}, status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_material_detail(request, pk):
    m = get_object_or_404(Material, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': m.id, 'nombre': m.nombre,
            'clave_sat': m.clave_sat or '', 'clave_unidad_sat': m.clave_unidad_sat or '',
            'empresas': _empresas_ids(m)})
    if request.method in ('PUT', 'PATCH'):
        err = _perm_check(request, 'ternium.change_material')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        if 'nombre' in body: m.nombre = body['nombre'].strip()
        if 'clave_sat' in body: m.clave_sat = body['clave_sat'].strip() or None
        if 'clave_unidad_sat' in body: m.clave_unidad_sat = body['clave_unidad_sat'].strip() or 'KGM'
        m.save()
        _set_empresas(m, body)
        return JsonResponse({'id': m.id, 'nombre': m.nombre,
            'clave_sat': m.clave_sat or '', 'clave_unidad_sat': m.clave_unidad_sat or '',
            'empresas': _empresas_ids(m)})
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_material')
        if err: return err
        m.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── UNIDADES ─────────────────────────────────────────────────────────────
def _unidad_to_dict(u):
    return {
        'id': u.id,
        'internal_id': u.internal_id,
        'license_plate': u.license_plate or '',
        'make_model': u.make_model or '',
        'year': u.year or '',
        'color': u.color or '',
        'vin': u.vin or '',
        'asset_type': u.asset_type,
        'ownership': u.ownership,
        'acquisition_date': str(u.acquisition_date) if u.acquisition_date else '',
        'operational_status': u.operational_status,
        'insurance_policy': u.insurance_policy or '',
        'insurance_due_date': str(u.insurance_due_date) if u.insurance_due_date else '',
        'circulation_license': u.circulation_license or '',
        'license_due_date': str(u.license_due_date) if u.license_due_date else '',
        # SCT / Carta Porte
        'permiso_sct': u.permiso_sct or '',
        'no_permiso_sct': u.no_permiso_sct or '',
        'nombre_aseguradora': u.nombre_aseguradora or '',
        'no_poliza_seguro': u.no_poliza_seguro or '',
        'eco_remolque_1': u.eco_remolque_1 or '',
        'placa_remolque_1': u.placa_remolque_1 or '',
        'eco_remolque_2': u.eco_remolque_2 or '',
        'placa_remolque_2': u.placa_remolque_2 or '',
        'notes': u.notes or '',
        'empresas': _empresas_ids(u),
    }

def _unidad_from_body(u, body):
    for field in ('internal_id', 'license_plate', 'make_model', 'color', 'vin',
                  'asset_type', 'ownership', 'operational_status',
                  'insurance_policy', 'circulation_license',
                  'permiso_sct', 'no_permiso_sct',
                  'nombre_aseguradora', 'no_poliza_seguro',
                  'eco_remolque_1', 'placa_remolque_1',
                  'eco_remolque_2', 'placa_remolque_2',
                  'notes'):
        if field in body:
            val = body[field]
            if isinstance(val, str):
                val = val.strip()
            setattr(u, field, val or None)
    for date_field in ('acquisition_date', 'insurance_due_date', 'license_due_date'):
        if date_field in body:
            setattr(u, date_field, body[date_field] or None)
    if 'year' in body:
        try:
            u.year = int(body['year']) if body['year'] else None
        except (ValueError, TypeError):
            u.year = None

@csrf_exempt
@login_required
def api_cat_unidades(request):
    if request.method == 'GET':
        qs = Unidad.objects.prefetch_related('empresas').all().order_by('internal_id')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(Q(internal_id__icontains=q) | Q(license_plate__icontains=q))
        return JsonResponse({'results': [_unidad_to_dict(u) for u in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_unidad')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        internal_id = body.get('internal_id', '').strip()
        if not internal_id:
            return JsonResponse({'error': 'El ID interno es requerido'}, status=400)
        if Unidad.objects.filter(internal_id__iexact=internal_id).exists():
            return JsonResponse({'error': 'Ya existe una unidad con ese ID'}, status=400)
        u = Unidad(internal_id=internal_id)
        _unidad_from_body(u, body)
        u.save()
        _set_empresas(u, body)
        return JsonResponse(_unidad_to_dict(u), status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_unidad_detail(request, pk):
    u = get_object_or_404(Unidad, pk=pk)
    if request.method == 'GET':
        return JsonResponse(_unidad_to_dict(u))
    if request.method in ('PUT', 'PATCH'):
        # Editar la unidad también se hace desde el detalle de un viaje
        # (tarjetas de Unidad / SCT). Aceptamos acceso_viajes como bypass.
        err = _perm_check(request, 'ternium.change_unidad', alt='ternium.acceso_viajes')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        _unidad_from_body(u, body)
        u.save()
        _set_empresas(u, body)
        return JsonResponse(_unidad_to_dict(u))
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_unidad')
        if err: return err
        u.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── CONTENEDORES ─────────────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_cat_contenedores(request):
    if request.method == 'GET':
        qs = Contenedor.objects.prefetch_related('empresas').all().order_by('nombre')
        q = request.GET.get('q', '')
        if q: qs = qs.filter(Q(nombre__icontains=q) | Q(placas__icontains=q))
        return JsonResponse({'results': [{'id': c.id, 'nombre': c.nombre, 'placas': c.placas or '', 'empresas': _empresas_ids(c)} for c in qs]})
    if request.method == 'POST':
        err = _perm_check(request, 'ternium.add_contenedor')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es requerido'}, status=400)
        if Contenedor.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'error': 'Ya existe un contenedor con ese nombre'}, status=400)
        c = Contenedor.objects.create(nombre=nombre, placas=body.get('placas', '').strip())
        _set_empresas(c, body)
        return JsonResponse({'id': c.id, 'nombre': c.nombre, 'placas': c.placas or '', 'empresas': _empresas_ids(c)}, status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_cat_contenedor_detail(request, pk):
    c = get_object_or_404(Contenedor, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'id': c.id, 'nombre': c.nombre, 'placas': c.placas or '', 'empresas': _empresas_ids(c)})
    if request.method in ('PUT', 'PATCH'):
        err = _perm_check(request, 'ternium.change_contenedor')
        if err: return err
        body, err = _json_body(request)
        if err: return err
        if 'nombre' in body: c.nombre = body['nombre'].strip()
        if 'placas' in body: c.placas = body['placas'].strip()
        c.save()
        _set_empresas(c, body)
        return JsonResponse({'id': c.id, 'nombre': c.nombre, 'placas': c.placas or '', 'empresas': _empresas_ids(c)})
    if request.method == 'DELETE':
        err = _perm_check(request, 'ternium.delete_contenedor')
        if err: return err
        c.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── CONTROL MANIFIESTO TRANE ─────────────────────────────────────────────
from .models import ControlManifiestoTrane

def _trane_to_dict(r):
    return {
        'id': r.id,
        'folio': r.folio or '',
        'fecha_captura': str(r.fecha_captura) if r.fecha_captura else '',
        'linea_transporte_id': r.linea_transporte_id,
        'linea_transporte': r.linea_transporte.nombre if r.linea_transporte else '',
        'operador_id': r.operador_id,
        'operador': r.operador.nombre if r.operador else '',
        'operador_manual': r.operador_manual or '',
        'destino_id': r.destino_id,
        'destino': r.destino.nombre if r.destino else '',
        'material_id': r.material_id,
        'material': r.material.nombre if r.material else '',
        'cantidad_kg': float(r.cantidad_kg) if r.cantidad_kg else 0,
        'unidad_manual': r.unidad_manual or '',
        'placas_unidad_manual': r.placas_unidad_manual or '',
        'contenedor_manual': r.contenedor_manual or '',
        'placas_contenedor_manual': r.placas_contenedor_manual or '',
        'manifiesto_url': r.manifiesto.url if r.manifiesto else '',
        'documento_trane_url': r.documento_trane.url if r.documento_trane else '',
    }

def _trane_from_post(r, request):
    d = request.POST
    if d.get('fecha_captura'): r.fecha_captura = d['fecha_captura']
    if 'folio' in d: r.folio = d['folio'].strip() or None
    for fk in ('destino_id', 'material_id', 'linea_transporte_id', 'operador_id', 'origen_id'):
        val = d.get(fk, '')
        setattr(r, fk, int(val) if val and val.isdigit() else None)
    for txt in ('operador_manual', 'unidad_manual', 'placas_unidad_manual',
                'contenedor_manual', 'placas_contenedor_manual'):
        if txt in d: setattr(r, txt, d[txt].strip() or None)
    if d.get('cantidad_kg'):
        try: r.cantidad_kg = float(d['cantidad_kg'])
        except ValueError: pass
    if 'manifiesto' in request.FILES: r.manifiesto = request.FILES['manifiesto']
    if 'documento_trane' in request.FILES: r.documento_trane = request.FILES['documento_trane']

@csrf_exempt
@login_required
def api_trane_manifiestos(request):
    if request.method == 'GET':
        qs = ControlManifiestoTrane.objects.select_related(
            'linea_transporte', 'operador', 'destino', 'material'
        ).order_by('-fecha_captura', '-id')
        fi = request.GET.get('fecha_inicio')
        ff = request.GET.get('fecha_fin')
        if fi: qs = qs.filter(fecha_captura__gte=fi)
        if ff: qs = qs.filter(fecha_captura__lte=ff)
        return JsonResponse({'results': [_trane_to_dict(r) for r in qs]})
    if request.method == 'POST':
        r = ControlManifiestoTrane()
        _trane_from_post(r, request)
        r.save()
        return JsonResponse(_trane_to_dict(r), status=201)
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

@csrf_exempt
@login_required
def api_trane_manifiesto_detail(request, pk):
    r = get_object_or_404(ControlManifiestoTrane, pk=pk)
    if request.method == 'GET':
        return JsonResponse(_trane_to_dict(r))
    if request.method in ('PUT', 'PATCH'):
        _trane_from_post(r, request)
        r.save()
        return JsonResponse(_trane_to_dict(r))
    if request.method == 'DELETE':
        r.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── COLUMNAS DE LA TABLA DE REMISIONES ───────────────────────────────────
@csrf_exempt
def api_columnas_remisiones(request):
    """Lee y guarda la configuración de columnas del usuario para el SPA.

    Comparte catálogo y reglas con la pantalla Django (columnas_remisiones.py),
    así que las dos versiones muestran siempre lo mismo. Devuelve 401 en JSON
    en vez de redirigir a HTML, que rompería el fetch del frontend.
    """
    from .columnas_remisiones import (
        catalogo_agrupado, columnas_por_defecto, config_de, sanear,
    )
    from .models import Profile

    if not request.user.is_authenticated:
        return JsonResponse({'error': 'NO_AUTORIZADO', 'detail': 'Sesión no válida.'}, status=401)

    def _payload(claves=None, personalizada=None):
        # Tras guardar hay que pasar los valores nuevos: request.user trae el
        # Profile cacheado y config_de devolvería la configuración anterior.
        if claves is None:
            claves, personalizada = config_de(request.user)
        activas, disponibles = catalogo_agrupado(claves)
        return {
            'activas': [{'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
                        for c in activas],
            'disponibles': [
                {'nombre': g['nombre'],
                 'columnas': [{'clave': c['clave'], 'etiqueta': c['etiqueta']} for c in g['columnas']]}
                for g in disponibles
            ],
            'personalizada': personalizada,
            'por_defecto': [{'clave': c['clave'], 'etiqueta': c['etiqueta'], 'fija': bool(c.get('fija'))}
                            for c in columnas_por_defecto()],
        }

    if request.method == 'GET':
        return JsonResponse(_payload())

    if request.method == 'POST':
        try:
            datos = json.loads(request.body or '{}')
        except (ValueError, TypeError):
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        recibidas = datos.get('columnas')
        if recibidas is not None and not isinstance(recibidas, list):
            return JsonResponse({'error': 'Se esperaba una lista de columnas'}, status=400)

        perfil, _ = Profile.objects.get_or_create(user=request.user)
        previas, _ = config_de(request.user)

        claves = sanear(recibidas) if recibidas else previas
        personalizada = bool(datos.get('personalizada', True))

        perfil.columnas_remisiones = {'columnas': claves, 'personalizada': personalizada}
        perfil.save(update_fields=['columnas_remisiones'])

        respuesta = _payload(claves, personalizada)
        respuesta['ok'] = True
        return JsonResponse(respuesta)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ── PERFIL DE USUARIO ────────────────────────────────────────────────────
@csrf_exempt
@login_required
def api_perfil(request):
    user = request.user
    profile, _ = Profile.objects.get_or_create(user=user)
    if request.method == 'GET':
        avatar_url = None
        try:
            if profile.avatar and profile.avatar.name:
                avatar_url = profile.avatar.url
        except Exception:
            pass
        # Build flat permission codename list (app_label.codename)
        if user.is_superuser:
            user_perm_codes = [f"ternium.{p[0].split('.')[1]}" for p in _MODULE_PERMS]
        else:
            user_perm_codes = []
            for p in user.user_permissions.select_related('content_type').all():
                user_perm_codes.append(f"{p.content_type.app_label}.{p.codename}")
        return JsonResponse({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'is_superuser': user.is_superuser,
            'is_staff': user.is_staff,
            'rol': profile.rol or '',
            'user_perms': user_perm_codes,
            'area': profile.area or '',
            'telefono': profile.telefono or '',
            'empresa': profile.empresa or '',
            'avatar_url': avatar_url,
            'grupos': [{'id': g.id, 'name': g.name} for g in user.groups.all()],
            'empresas_autorizadas': [{'id': e.id, 'nombre': e.nombre} for e in profile.empresas_autorizadas.all()],
        })
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
        except Exception:
            body = {}
        user.first_name = body.get('first_name', user.first_name)
        user.last_name = body.get('last_name', user.last_name)
        user.email = body.get('email', user.email)
        user.save()
        profile.area = body.get('area', profile.area)
        profile.telefono = body.get('telefono', profile.telefono)
        profile.empresa = body.get('empresa', profile.empresa)
        profile.save()
        return JsonResponse({'ok': True, 'message': 'Perfil actualizado correctamente'})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)

# ── GESTION DE PERMISOS POR USUARIO (ADMIN) ──────────────────────────────
_MODULE_PERMS = [
    # ── Acceso a módulos (nivel app) ─────────────────────────────────────────
    ('ternium.acceso_dashboard',   'Dashboard Principal'),
    ('ternium.acceso_remisiones',  'Remisiones'),
    ('ternium.acceso_trane',       'Portal Trane'),
    ('ternium.acceso_bancos',      'Flujo Bancario'),
    ('ternium.acceso_diesel',      'Control Diésel'),
    ('ternium.acceso_catalogos',   'Catálogos'),
    # ── CRUD Remisiones ──────────────────────────────────────────────────────
    ('ternium.view_remision',      'Ver Remisiones'),
    ('ternium.add_remision',       'Crear Remisiones'),
    ('ternium.change_remision',    'Editar Remisiones'),
    ('ternium.delete_remision',    'Eliminar Remisiones'),
    # ── CRUD Catálogos ───────────────────────────────────────────────────────
    ('ternium.view_empresa',       'Ver Empresas'),
    ('ternium.add_empresa',        'Agregar Empresas'),
    ('ternium.change_empresa',     'Editar Empresas'),
    ('ternium.add_lugar',          'Agregar Lugares'),
    ('ternium.change_lugar',       'Editar Lugares'),
    ('ternium.add_operador',       'Agregar Operadores'),
    ('ternium.change_operador',    'Editar Operadores'),
    ('ternium.add_material',       'Agregar Materiales'),
    ('ternium.change_material',    'Editar Materiales'),
    ('ternium.add_unidad',         'Agregar Unidades'),
    ('ternium.change_unidad',      'Editar Unidades'),
    ('ternium.add_contenedor',     'Agregar Contenedores'),
    ('ternium.change_contenedor',  'Editar Contenedores'),
]

@login_required
def api_admin_usuarios(request):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)
    from django.contrib.auth.models import User as AuthUser
    users = AuthUser.objects.all().select_related('ternium_profile').prefetch_related(
        'groups', 'user_permissions'
    ).order_by('username')
    data = []
    for u in users:
        profile = getattr(u, 'ternium_profile', None)
        data.append({
            'id': u.id, 'username': u.username,
            'full_name': u.get_full_name() or u.username,
            'email': u.email, 'is_active': u.is_active,
            'is_superuser': u.is_superuser, 'is_staff': u.is_staff,
            'rol': (profile.rol or '') if profile else '',
            'area': (profile.area or '') if profile else '',
            'empresa': (profile.empresa or '') if profile else '',
            'grupos': [g.name for g in u.groups.all()],
            'empresas_autorizadas': [{'id': e.id, 'nombre': e.nombre} for e in profile.empresas_autorizadas.all()] if profile else [],
            'last_login': u.last_login.isoformat() if u.last_login else None,
            'date_joined': u.date_joined.isoformat() if u.date_joined else None,
        })
    return JsonResponse({'users': data})

@csrf_exempt
@login_required
def api_admin_usuario_permisos(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)
    from django.contrib.auth.models import User as AuthUser, Permission
    from django.contrib.contenttypes.models import ContentType
    usuario = get_object_or_404(AuthUser, pk=pk)
    profile, _ = Profile.objects.get_or_create(user=usuario)
    if request.method == 'GET':
        user_perm_codes = set()
        for p in usuario.user_permissions.all():
            user_perm_codes.add(f"{p.content_type.app_label}.{p.codename}")
        all_empresas = [{'id': e.id, 'nombre': e.nombre} for e in Empresa.objects.all().order_by('nombre')]
        return JsonResponse({
            'id': usuario.id, 'username': usuario.username,
            'full_name': usuario.get_full_name(), 'email': usuario.email,
            'is_active': usuario.is_active, 'is_staff': usuario.is_staff,
            'is_superuser': usuario.is_superuser,
            'rol': profile.rol or '',
            'area': profile.area or '', 'empresa': profile.empresa or '', 'telefono': profile.telefono or '',
            'user_perms': list(user_perm_codes),
            'empresas_autorizadas': [e.id for e in profile.empresas_autorizadas.all()],
            'all_empresas': all_empresas,
            'available_perms': [{'codename': p[0], 'label': p[1]} for p in _MODULE_PERMS],
            'last_login': usuario.last_login.isoformat() if usuario.last_login else None,
            'date_joined': usuario.date_joined.isoformat() if usuario.date_joined else None,
        })
    if request.method in ('PUT', 'PATCH', 'POST'):
        body, err = _json_body(request)
        if err: return err
        if 'is_active' in body: usuario.is_active = bool(body['is_active'])
        if 'is_staff' in body and request.user.is_superuser: usuario.is_staff = bool(body['is_staff'])
        usuario.save()
        if 'area' in body: profile.area = body['area']
        if 'empresa' in body: profile.empresa = body['empresa']
        if 'telefono' in body: profile.telefono = body['telefono']
        if 'rol' in body and not usuario.is_superuser:
            allowed_roles = ('', 'flujos_bancos', 'ternium')
            if body['rol'] in allowed_roles:
                profile.rol = body['rol']
        profile.save()
        if 'empresas_autorizadas' in body:
            ids = [int(i) for i in body['empresas_autorizadas'] if str(i).isdigit()]
            profile.empresas_autorizadas.set(Empresa.objects.filter(id__in=ids))
        if 'user_perms' in body:
            perm_codes = body['user_perms']
            perms_to_set = []
            for code in perm_codes:
                parts = code.split('.')
                if len(parts) == 2:
                    try:
                        perm = Permission.objects.get(
                            content_type__app_label=parts[0],
                            codename=parts[1],
                        )
                        perms_to_set.append(perm)
                    except Permission.DoesNotExist:
                        pass
            usuario.user_permissions.set(perms_to_set)
        return JsonResponse({'ok': True, 'message': f'Usuario {usuario.username} actualizado'})
    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


# =========================================================================
# CONFIGURACIÓN ALERTAS DE MERMA (CRUD)
# ─────────────────────────────────────────────────────────────────────────
# GET  /api/admin/alertas-merma/?q=<material>   → lista con buscador
# POST /api/admin/alertas-merma/                → crear/actualizar umbral
# PUT  /api/admin/alertas-merma/<pk>/           → editar umbral
# DELETE /api/admin/alertas-merma/<pk>/         → eliminar config (vuelve al default 1%)
# =========================================================================

def _alerta_to_dict(cfg):
    return {
        'id': cfg.id,
        'material_id': cfg.material_id,
        'material_nombre': cfg.material.nombre,
        'porcentaje_umbral': float(cfg.porcentaje_umbral),
    }


@csrf_exempt
@login_required
def api_alertas_merma(request):
    """Lista de configuraciones + creación."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    if request.method == 'GET':
        q = request.GET.get('q', '').strip()
        qs = ConfiguracionAlertaMerma.objects.select_related('material').order_by('material__nombre')
        if q:
            qs = qs.filter(material__nombre__icontains=q)

        # Incluir también materiales sin configuración (con default 1%)
        include_all = request.GET.get('include_all', '') == '1'
        if include_all:
            configurados_ids = set(qs.values_list('material_id', flat=True))
            materiales_qs = Material.objects.order_by('nombre')
            if q:
                materiales_qs = materiales_qs.filter(nombre__icontains=q)
            resultados = list(qs.values('id', 'material_id', 'porcentaje_umbral'))
            for mat in materiales_qs:
                if mat.id not in configurados_ids:
                    resultados.append({
                        'id': None,
                        'material_id': mat.id,
                        'material_nombre': mat.nombre,
                        'porcentaje_umbral': 1.0,
                        'es_default': True,
                    })
            # Re-enrich with material_nombre for configured ones
            cfg_dict = {c.material_id: c for c in qs}
            data = []
            for mat in materiales_qs:
                if mat.id in {r['material_id'] for r in resultados if r.get('id')}:
                    cfg = cfg_dict[mat.id]
                    data.append({**_alerta_to_dict(cfg), 'es_default': False})
                else:
                    data.append({'id': None, 'material_id': mat.id, 'material_nombre': mat.nombre, 'porcentaje_umbral': 1.0, 'es_default': True})
            return JsonResponse({'results': data})

        return JsonResponse({'results': [_alerta_to_dict(c) for c in qs]})

    if request.method == 'POST':
        body, err = _json_body(request)
        if err:
            return err
        material_id = body.get('material_id') or body.get('material')
        porcentaje = body.get('porcentaje_umbral', 1.0)
        if not material_id:
            return JsonResponse({'error': 'material_id es requerido'}, status=400)
        try:
            porcentaje = float(porcentaje)
            if porcentaje < 0 or porcentaje > 100:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse({'error': 'porcentaje_umbral debe ser un número entre 0 y 100'}, status=400)
        material = get_object_or_404(Material, pk=material_id)
        cfg, _ = ConfiguracionAlertaMerma.objects.update_or_create(
            material=material,
            defaults={'porcentaje_umbral': porcentaje},
        )
        return JsonResponse(_alerta_to_dict(cfg), status=201)

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


@csrf_exempt
@login_required
def api_alerta_merma_detail(request, pk):
    """Editar o eliminar una configuración específica."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    cfg = get_object_or_404(ConfiguracionAlertaMerma, pk=pk)

    if request.method == 'GET':
        return JsonResponse(_alerta_to_dict(cfg))

    if request.method in ('PUT', 'PATCH'):
        body, err = _json_body(request)
        if err:
            return err
        porcentaje = body.get('porcentaje_umbral', cfg.porcentaje_umbral)
        try:
            porcentaje = float(porcentaje)
            if porcentaje < 0 or porcentaje > 100:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse({'error': 'porcentaje_umbral debe ser un número entre 0 y 100'}, status=400)
        cfg.porcentaje_umbral = porcentaje
        cfg.save()
        return JsonResponse(_alerta_to_dict(cfg))

    if request.method == 'DELETE':
        cfg.delete()
        return JsonResponse({'ok': True, 'detail': 'Configuración eliminada. El material usará el umbral default (1%).'})

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


# =========================================================================
# DESTINATARIOS DE ALERTAS DE MERMA (CRUD)
# GET    /api/admin/alertas-merma/correos/        → lista
# POST   /api/admin/alertas-merma/correos/        → agregar { email }
# DELETE /api/admin/alertas-merma/correos/<pk>/   → eliminar
# =========================================================================

@csrf_exempt
@login_required
def api_correos_alerta_merma(request):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    if request.method == 'GET':
        qs = DestinatarioAlertaMerma.objects.all()
        return JsonResponse({'results': [{'id': d.id, 'email': d.email} for d in qs]})

    if request.method == 'POST':
        body, err = _json_body(request)
        if err:
            return err
        email = (body.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return JsonResponse({'error': 'Correo inválido'}, status=400)
        if DestinatarioAlertaMerma.objects.filter(email=email).exists():
            return JsonResponse({'error': 'Este correo ya está registrado'}, status=400)
        dest = DestinatarioAlertaMerma.objects.create(email=email)
        return JsonResponse({'id': dest.id, 'email': dest.email}, status=201)

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


@csrf_exempt
@login_required
def api_correo_alerta_merma_detail(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    dest = get_object_or_404(DestinatarioAlertaMerma, pk=pk)

    if request.method == 'DELETE':
        dest.delete()
        return JsonResponse({'ok': True})

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


# =========================================================================
# ADMIN — DESTRUCCIÓN FISCAL (ConfiguracionManifiesto)
# =========================================================================

def _ser_config_manifiesto(c):
    return {
        'id': c.id,
        'origen_id': c.origen_id,
        'origen_nombre': c.origen.nombre,
        'material_id': c.material_id,
        'material_nombre': c.material.nombre,
    }


@csrf_exempt
@login_required
def api_destruccion_fiscal(request):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    if request.method == 'GET':
        qs = ConfiguracionManifiesto.objects.select_related('origen', 'material').order_by('origen__nombre', 'material__nombre')
        return JsonResponse({'results': [_ser_config_manifiesto(c) for c in qs]})

    if request.method == 'POST':
        body, err = _json_body(request)
        if err:
            return err
        origen_id = body.get('origen_id')
        material_id = body.get('material_id')
        if not origen_id or not material_id:
            return JsonResponse({'error': 'Se requieren origen_id y material_id'}, status=400)
        if ConfiguracionManifiesto.objects.filter(origen_id=origen_id, material_id=material_id).exists():
            return JsonResponse({'error': 'Esta combinación ya existe'}, status=400)
        origen = get_object_or_404(Lugar, pk=origen_id)
        material = get_object_or_404(Material, pk=material_id)
        c = ConfiguracionManifiesto.objects.create(origen=origen, material=material)
        c.refresh_from_db()
        return JsonResponse(_ser_config_manifiesto(c), status=201)

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


@csrf_exempt
@login_required
def api_destruccion_fiscal_detail(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    c = get_object_or_404(ConfiguracionManifiesto, pk=pk)

    if request.method == 'DELETE':
        c.delete()
        return JsonResponse({'ok': True})

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)


@login_required
def api_destruccion_fiscal_catalogos(request):
    if not request.user.is_superuser:
        return JsonResponse({'error': 'FORBIDDEN'}, status=403)

    origenes = Lugar.objects.filter(tipo__in=['ORIGEN', 'AMBOS']).order_by('nombre').values('id', 'nombre')
    materiales = Material.objects.order_by('nombre').values('id', 'nombre')
    return JsonResponse({
        'origenes': list(origenes),
        'materiales': list(materiales),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# CENTRO DE ALERTAS (Bell icon + módulo de administración)
# ═══════════════════════════════════════════════════════════════════════════════

def _alerta_to_dict(a, user=None):
    """Serializa una Alerta. `unread` es relativo al usuario que consulta."""
    unread = True
    if user and user.is_authenticated:
        unread = not a.leida_por.filter(pk=user.pk).exists()
    # Tiempo relativo amigable ("hace 5 min", "ayer", etc.)
    from django.utils.timesince import timesince
    try:
        relative = timesince(a.creada_en, timezone.now())
        relative = f"hace {relative.split(',')[0]}"
    except Exception:
        relative = a.creada_en.strftime('%d/%m/%Y %H:%M')
    return {
        'id':     a.id,
        'tipo':   a.tipo,
        'type':   a.tipo,  # alias para compat con front
        'title':  a.title,
        'desc':   a.desc,
        'message': a.desc,  # alias
        'creada_en': a.creada_en.isoformat() if a.creada_en else None,
        'time': relative,
        'creada_por': (a.creada_por.get_full_name() or a.creada_por.username) if a.creada_por_id else None,
        'unread': unread,
    }


@csrf_exempt
@require_http_methods(['GET', 'POST'])
def api_alertas_list(request):
    """
    GET  /api/alertas/  → lista todas las alertas (cualquier autenticado puede VER).
    POST /api/alertas/  → crea una alerta (solo Staff).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)

    if request.method == 'GET':
        qs = Alerta.objects.all().select_related('creada_por').prefetch_related('leida_por')
        only_unread = request.GET.get('unread', '').lower() in ('1', 'true', 'yes')
        if only_unread:
            qs = qs.exclude(leida_por=request.user)
        return JsonResponse({
            'results': [_alerta_to_dict(a, request.user) for a in qs],
            'total':   qs.count(),
        })

    # POST — crear alerta: SOLO staff/superuser
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'error': 'Permiso denegado. Solo Staff puede crear alertas.'}, status=403)

    try:
        body = json.loads(request.body or '{}')
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    tipo  = (body.get('tipo') or 'info').strip()
    title = (body.get('title') or '').strip()
    desc  = (body.get('desc') or body.get('message') or '').strip()

    if tipo not in dict(Alerta.TIPO_CHOICES):
        return JsonResponse({'error': f'Tipo inválido. Debe ser uno de: {list(dict(Alerta.TIPO_CHOICES).keys())}'}, status=400)
    if not title:
        return JsonResponse({'error': 'El título es obligatorio.'}, status=400)
    if not desc:
        return JsonResponse({'error': 'El mensaje es obligatorio.'}, status=400)

    a = Alerta.objects.create(tipo=tipo, title=title, desc=desc, creada_por=request.user)
    # El creador la marca automáticamente como leída para sí mismo.
    a.leida_por.add(request.user)
    return JsonResponse(_alerta_to_dict(a, request.user), status=201)


@csrf_exempt
@require_http_methods(['PATCH', 'POST', 'DELETE'])
def api_alerta_detail(request, pk):
    """
    PATCH/POST /api/alertas/<pk>/leida/  → marca como leída por el usuario actual.
    DELETE     /api/alertas/<pk>/        → borra la alerta (solo Staff).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)

    a = get_object_or_404(Alerta, pk=pk)

    if request.method in ('PATCH', 'POST'):
        a.leida_por.add(request.user)
        return JsonResponse(_alerta_to_dict(a, request.user))

    # DELETE
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'error': 'Permiso denegado. Solo Staff puede borrar alertas.'}, status=403)
    a.delete()
    return JsonResponse({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# CHAT IA — Asistente del sistema (por usuario)
# ═══════════════════════════════════════════════════════════════════════════════

def _chat_msg_to_dict(m):
    return {
        'id':       m.id,
        'rol':      m.rol,
        'role':     m.rol,  # alias para front
        'contenido': m.contenido,
        'content':  m.contenido,  # alias
        'creado_en': m.creado_en.isoformat() if m.creado_en else None,
    }


@csrf_exempt
@require_http_methods(['GET', 'POST', 'DELETE'])
def api_chat_ia(request):
    """
    GET    /api/chat/   → devuelve el historial del usuario actual (ordenado asc).
    POST   /api/chat/   → recibe {mensaje}, guarda + responde + guarda respuesta.
    DELETE /api/chat/   → borra todo el historial del usuario actual.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)

    if request.method == 'GET':
        msgs = ChatMensaje.objects.filter(user=request.user).order_by('creado_en')
        return JsonResponse({'results': [_chat_msg_to_dict(m) for m in msgs]})

    if request.method == 'DELETE':
        ChatMensaje.objects.filter(user=request.user).delete()
        return JsonResponse({'ok': True})

    # POST — preguntar
    try:
        body = json.loads(request.body or '{}')
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    mensaje_usuario = (body.get('mensaje') or body.get('message') or '').strip()
    if not mensaje_usuario:
        return JsonResponse({'error': 'El mensaje es obligatorio'}, status=400)

    # Historial PREVIO (sin el mensaje actual) para dar memoria conversacional.
    historial = list(
        ChatMensaje.objects.filter(user=request.user).order_by('creado_en')
        .values('rol', 'contenido')
    )

    # Guarda mensaje del usuario
    msg_user = ChatMensaje.objects.create(
        user=request.user, rol='user', contenido=mensaje_usuario,
    )

    # Genera la respuesta del asistente: IA real (Gemini gratis) con contexto del
    # sistema + memoria, filtrada por los PERMISOS del usuario. Si no hay
    # GEMINI_API_KEY o Gemini falla, cae automáticamente al motor por reglas.
    try:
        from .chat_ia import responder_ia
        respuesta_texto = responder_ia(mensaje_usuario, request.user, historial=historial)
    except Exception as exc:
        respuesta_texto = f"⚠️ Tuve un problema generando la respuesta: {exc}"

    msg_bot = ChatMensaje.objects.create(
        user=request.user, rol='bot', contenido=respuesta_texto,
    )

    return JsonResponse({
        'pregunta':  _chat_msg_to_dict(msg_user),
        'respuesta': _chat_msg_to_dict(msg_bot),
    })
