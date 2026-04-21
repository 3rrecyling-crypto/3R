from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q
from django.utils import timezone
from io import BytesIO # <--- IMPORTANTE AGREGAR ESTO
import re 
from django.db import transaction
from openpyxl.utils import get_column_letter
from django.conf import settings
from collections import defaultdict
from decimal import Decimal
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import csv
import io
import os
import requests # Necesario para Banxico si no estaba importado explícitamente

# --- IMPORTACIÓN DE SEGURIDAD ---
from django.contrib.auth.decorators import permission_required

# Importaciones de AWS S3
import boto3
from botocore.exceptions import BotoCoreError, NoCredentialsError

# Importaciones de Excel y XML
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importaciones locales (Modelos y Forms)
from .models import (
    Cuenta, Movimiento, SubCategoria, UnidadNegocio, 
    Operacion, Categoria, ComprobanteFiscal, Tercero
)
from .forms import (
    MovimientoForm, 
    TransferenciaForm, 
    CuentaForm, 
    TerceroForm, 
    ImportarTxtForm, 
    ImportarExcelForm, 
    ComprobanteForm,
    CategoriaForm, 
    SubCategoriaForm
)

# ---------------------------------------------------------
# UTILIDADES S3 (No requieren decorador, son internas)
# ---------------------------------------------------------
def _subir_archivo_a_s3(archivo_obj, s3_ruta_relativa):
    """
    Sube un archivo a S3.
    - `s3_ruta_relativa` es la ruta SIN 'media/' (ej: 'xmls/2024/archivo.xml').
    - Devuelve la misma ruta relativa si tiene éxito, para guardarla en la DB.
    """
    try:
        # Asegurar que el puntero esté al inicio
        archivo_obj.seek(0)
        
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Boto3 necesita la ruta completa (Key) dentro del bucket
        full_s3_path = f"{settings.AWS_MEDIA_LOCATION}/{s3_ruta_relativa}"

        s3_client.upload_fileobj(
            archivo_obj,
            settings.AWS_STORAGE_BUCKET_NAME,
            full_s3_path
        )
        return s3_ruta_relativa
        
    except (BotoCoreError, NoCredentialsError, Exception) as e:
        print(f"Error al subir el archivo a S3: {e}")
        return None

def _eliminar_archivo_de_s3(ruta_completa_s3):
    """
    Elimina un archivo de S3.
    - `ruta_completa_s3` es el nombre del campo FileField (ej: 'xmls/2024/foto.jpg').
    """
    if not ruta_completa_s3:
        return
    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        # Construimos la Key completa incluyendo la carpeta media
        full_key = f"{settings.AWS_MEDIA_LOCATION}/{str(ruta_completa_s3)}"
        
        s3_client.delete_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=full_key
        )
    except (BotoCoreError, NoCredentialsError, Exception) as e:
        print(f"Error al eliminar archivo antiguo de S3: {e}")
        
        
def procesar_datos_xml_desde_bytes(contenido_bytes):
    """
    Lee bytes del XML, extrae UUID y busca impuestos agresivamente
    (Funciona para tu XML A-2728.xml y otros formatos).
    """
    datos = {
        'uuid': None, 
        'iva': Decimal('0.00'), 
        'ret_iva': Decimal('0.00'), 
        'ret_isr': Decimal('0.00')
    }
    
    try:
        xml_buffer = BytesIO(contenido_bytes)
        tree = ET.parse(xml_buffer)
        root = tree.getroot()
        
        # Namespaces
        ns = {'cfdi': 'http://www.sat.gob.mx/cfd/4', 'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'}
        if 'http://www.sat.gob.mx/cfd/3' in root.tag:
            ns['cfdi'] = 'http://www.sat.gob.mx/cfd/3'

        # 1. UUID
        tfd = root.find('.//tfd:TimbreFiscalDigital', ns)
        if tfd is not None:
            datos['uuid'] = tfd.get('UUID')

        # 2. IMPUESTOS GLOBALES (Prioridad)
        impuestos_node = root.find('cfdi:Impuestos', ns)
        
        if impuestos_node is not None:
            # IVA Trasladado
            traslados_totales = impuestos_node.get('TotalImpuestosTrasladados')
            if traslados_totales:
                datos['iva'] = Decimal(traslados_totales)
            
            # Retenciones Globales
            retenciones = impuestos_node.findall('cfdi:Retenciones/cfdi:Retencion', ns)
            for ret in retenciones:
                imp = ret.get('Impuesto')
                importe = Decimal(ret.get('Importe') or 0)
                if imp == '002': datos['ret_iva'] += importe
                if imp == '001': datos['ret_isr'] += importe

        # 3. FALLBACK CONCEPTOS (Si no se halló en globales)
        if datos['iva'] == 0 and datos['ret_iva'] == 0 and datos['ret_isr'] == 0:
            conceptos = root.findall('cfdi:Conceptos/cfdi:Concepto', ns)
            for c in conceptos:
                # Traslados
                for t in c.findall('.//cfdi:Traslado', ns):
                    if t.get('Impuesto') == '002':
                        datos['iva'] += Decimal(t.get('Importe') or 0)
                # Retenciones
                for r in c.findall('.//cfdi:Retencion', ns):
                    imp = r.get('Impuesto')
                    val = Decimal(r.get('Importe') or 0)
                    if imp == '002': datos['ret_iva'] += val
                    if imp == '001': datos['ret_isr'] += val

    except Exception as e:
        print(f"Error procesando XML interno: {e}")
    
    return datos
        
def procesar_datos_xml(archivo_obj):
    """
    Lee un archivo, crea una copia en memoria (BytesIO) para no cerrar el original,
    extrae UUID, IVA y Retenciones.
    """
    datos = {
        'uuid': None, 
        'iva': Decimal('0.00'), 
        'ret_iva': Decimal('0.00'), 
        'ret_isr': Decimal('0.00')
    }
    
    try:
        # 1. Asegurar lectura desde inicio
        archivo_obj.seek(0)
        
        # 2. COPIA SEGURA EN MEMORIA
        contenido = archivo_obj.read()
        archivo_obj.seek(0) # Resetear archivo original para que Django/S3 lo pueda leer después
        
        # Stream en memoria independiente
        xml_buffer = BytesIO(contenido)
        
        # 3. Parsear
        tree = ET.parse(xml_buffer)
        root = tree.getroot()
        
        # 4. Namespaces
        ns = {'cfdi': 'http://www.sat.gob.mx/cfd/4', 'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'}
        if 'cfdi' not in root.tag:
            ns['cfdi'] = 'http://www.sat.gob.mx/cfd/3'

        # 5. UUID
        tfd = root.find('.//tfd:TimbreFiscalDigital', ns)
        if tfd is not None:
            datos['uuid'] = tfd.get('UUID')

        # 6. Impuestos
        impuestos_node = root.find('cfdi:Impuestos', ns)
        encontrado_global = False
        
        if impuestos_node is not None:
            total_traslados = impuestos_node.get('TotalImpuestosTrasladados')
            if total_traslados:
                datos['iva'] = Decimal(total_traslados)
                encontrado_global = True
            
            retenciones = impuestos_node.findall('cfdi:Retenciones/cfdi:Retencion', ns)
            if retenciones:
                for ret in retenciones:
                    imp = ret.get('Impuesto')
                    importe = Decimal(ret.get('Importe') or 0)
                    if imp == '002': datos['ret_iva'] += importe
                    if imp == '001': datos['ret_isr'] += importe
                encontrado_global = True

        if not encontrado_global or (datos['iva'] == 0 and datos['ret_iva'] == 0):
            conceptos = root.findall('cfdi:Conceptos/cfdi:Concepto', ns)
            for c in conceptos:
                for t in c.findall('.//cfdi:Traslado', ns):
                    if t.get('Impuesto') == '002':
                        datos['iva'] += Decimal(t.get('Importe') or 0)
                for r in c.findall('.//cfdi:Retencion', ns):
                    imp = r.get('Impuesto')
                    val = Decimal(r.get('Importe') or 0)
                    if imp == '002': datos['ret_iva'] += val
                    if imp == '001': datos['ret_isr'] += val

    except Exception as e:
        print(f"Error procesando estructura XML: {e}")
    
    return datos

def obtener_tipo_cambio_banxico():
    token = getattr(settings, 'BANXICO_API_TOKEN', None)
    
    if not token:
        # Si no hay token configurado, retornamos un valor default seguro
        return 20.50 

    series = "SF43718" # Serie FIX
    url = f"https://www.banxico.org.mx/SieAPIRest/service/v1/series/{series}/datos/oportuno"
    
    headers = {
        'Bmx-Token': token,
        'Accept': 'application/json'
    }

    try:
        response = requests.get(url, headers=headers, timeout=5)
        response.raise_for_status()
        
        data = response.json()
        dato_str = data['bmx']['series'][0]['datos'][0]['dato']
        return float(dato_str)
        
    except Exception as e:
        print(f"Error consultando Banxico: {e}")
        # En caso de error de conexión o API, retornamos fallback
        return 20.50

# ---------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def dashboard(request):
    # 1. FILTROS DE FECHA
    filtro_tiempo = request.GET.get('filtro', 'hoy')
    hoy = timezone.now().date()
    fecha_inicio = hoy
    fecha_fin = hoy

    if filtro_tiempo == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
    elif filtro_tiempo == 'mes':
        fecha_inicio = hoy.replace(day=1)
    elif filtro_tiempo == 'custom':
        f_ini = request.GET.get('fecha_inicio')
        f_fin = request.GET.get('fecha_fin')
        if f_ini and f_fin:
            try:
                fecha_inicio = datetime.strptime(f_ini, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(f_fin, '%Y-%m-%d').date()
            except ValueError:
                pass

    # 2. CÁLCULO DE KPIs
    movs_rango = Movimiento.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
    
    ingresos_periodo = movs_rango.aggregate(Sum('abono'))['abono__sum'] or 0
    egresos_periodo = movs_rango.aggregate(Sum('cargo'))['cargo__sum'] or 0
    balance_periodo = ingresos_periodo - egresos_periodo
    
    movimientos_ingresos_count = movs_rango.filter(abono__gt=0).count()
    movimientos_egresos_count = movs_rango.filter(cargo__gt=0).count()

    # 3. CUENTAS Y TIPO DE CAMBIO
    todas_cuentas = Cuenta.objects.all()
    
    raw_tc = obtener_tipo_cambio_banxico()
    try:
        tipo_cambio_actual = Decimal(str(raw_tc)) if raw_tc else Decimal('20.00')
    except Exception:
        tipo_cambio_actual = Decimal('20.00')

    total_mxn = sum(c.saldo_actual for c in todas_cuentas if c.moneda == 'MXN')
    total_usd = sum(c.saldo_actual for c in todas_cuentas if c.moneda == 'USD')
    total_usd_convertido = total_usd * tipo_cambio_actual
    saldo_total_consolidado = total_mxn + total_usd_convertido

    cuentas_ordenadas = sorted(todas_cuentas, key=lambda c: c.saldo_actual, reverse=True)
    
    for cuenta in cuentas_ordenadas:
        if cuenta.moneda == 'USD':
            cuenta.saldo_convertido_temp = cuenta.saldo_actual * tipo_cambio_actual
        else:
            cuenta.saldo_convertido_temp = 0
    
    cuentas_mxn = [c for c in todas_cuentas if c.moneda == 'MXN']
    cuentas_usd = [c for c in todas_cuentas if c.moneda == 'USD']

    # 4. MOVIMIENTOS RECIENTES
    movimientos_recientes = Movimiento.objects.select_related('cuenta').order_by('-fecha', '-id')[:5]

    importar_form = ImportarTxtForm()

    # 5. CONTEXTO
    context = {
        'filtro_actual': filtro_tiempo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        'importar_form': importar_form,
        
        'ingresos_periodo': ingresos_periodo,
        'egresos_periodo': egresos_periodo,
        'balance_periodo': balance_periodo,
        'movimientos_ingresos_count': movimientos_ingresos_count,
        'movimientos_egresos_count': movimientos_egresos_count,
        'cuentas': cuentas_ordenadas, 
        'cuentas_mxn': cuentas_mxn, 
        'cuentas_usd': cuentas_usd, 
        'total_mxn': total_mxn,
        'total_usd': total_usd,
        'saldo_total': saldo_total_consolidado, 
        'tipo_cambio': tipo_cambio_actual,
        'total_usd_convertido': total_usd_convertido, 
        'movimientos_recientes': movimientos_recientes,
    }
    return render(request, 'flujo_bancos/dashboard.html', context)

# ---------------------------------------------------------
# IMPORTAR MOVIMIENTOS (EXCEL)
# ---------------------------------------------------------
def importar_movimientos(request):
    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            cuenta = form.cleaned_data['cuenta_destino']
            archivo = request.FILES['archivo_excel']
            
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active 
                
                count_creados = 0
                saldo_banco_excel = None 
                
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    fecha_raw = row[0]
                    if not fecha_raw: continue 

                    # --- LECTURA DE SALDO POR RENGLÓN ---
                    # Leemos el saldo de CADA movimiento (Columna E / índice 4)
                    saldo_renglon = 0
                    if len(row) >= 5:
                        val_saldo = row[4]
                        if val_saldo is not None:
                            try:
                                if isinstance(val_saldo, str):
                                    val_saldo = val_saldo.replace(',', '').replace('$', '')
                                saldo_renglon = Decimal(str(val_saldo))
                                
                                # Si es la primera fila, guardamos para actualizar el saldo inicial de la cuenta después
                                if i == 0:
                                    saldo_banco_excel = saldo_renglon
                            except:
                                pass

                    concepto = row[1]
                    cargo_raw = row[2]
                    abono_raw = row[3]
                    
                    str_cargo = str(cargo_raw) if cargo_raw is not None else '0'
                    str_abono = str(abono_raw) if abono_raw is not None else '0'
                    
                    str_cargo = str_cargo.replace(',', '').replace('$', '')
                    str_abono = str_abono.replace(',', '').replace('$', '')

                    cargo_decimal = Decimal(str_cargo) if str_cargo else Decimal(0)
                    abono_decimal = Decimal(str_abono) if str_abono else Decimal(0)

                    existe = Movimiento.objects.filter(
                        cuenta=cuenta,
                        fecha=fecha_raw,
                        concepto=concepto,
                        cargo=cargo_decimal,
                        abono=abono_decimal
                    ).exists()
                    
                    if not existe:
                        Movimiento.objects.create(
                            cuenta=cuenta,
                            fecha=fecha_raw,
                            concepto=concepto or "Sin concepto",
                            cargo=cargo_decimal,
                            abono=abono_decimal,
                            saldo_banco=saldo_renglon, # <--- GUARDAMOS EL SALDO DEL EXCEL
                            estatus='PENDIENTE'
                        )
                        count_creados += 1
                
                # --- ACTUALIZACIÓN DE SALDO INICIAL ---
                if saldo_banco_excel is not None:
                    totales = Movimiento.objects.filter(cuenta=cuenta).aggregate(
                        sum_abono=Sum('abono'), 
                        sum_cargo=Sum('cargo')
                    )
                    total_abonos = totales['sum_abono'] or 0
                    total_cargos = totales['sum_cargo'] or 0
                    
                    nuevo_saldo_inicial = saldo_banco_excel - (total_abonos - total_cargos)
                    
                    cuenta.saldo_inicial = nuevo_saldo_inicial
                    cuenta.save()
                    
                    messages.success(request, f"Importados {count_creados} movimientos. Saldo alineado al corte: ${saldo_banco_excel:,.2f}")
                else:
                    messages.success(request, f"Importados {count_creados} movimientos.")

                return redirect('lista_movimientos')
                
            except Exception as e:
                messages.error(request, f"Error crítico al importar: {str(e)}")
    else:
        form = ImportarExcelForm()
    
    return render(request, 'flujo_bancos/importar_movimientos.html', {'form': form})
# ---------------------------------------------------------
# CANCELAR TRANSFERENCIA
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def cancelar_transferencia(request, pk):
    salida = get_object_or_404(Movimiento, pk=pk)
    
    if salida.auditado:
        messages.error(request, "No se puede cancelar una transferencia auditada.")
        return redirect('lista_transferencias')

    concepto_base = salida.concepto.split(' (Envío a')[0]
    entrada = Movimiento.objects.filter(
        fecha=salida.fecha,
        abono__gt=0,
        concepto__icontains=concepto_base
    ).exclude(id=salida.id).first()

    if entrada:
        if entrada.comprobante:
            _eliminar_archivo_de_s3(str(entrada.comprobante))
        entrada.delete()
    
    if salida.comprobante:
        _eliminar_archivo_de_s3(str(salida.comprobante))

    salida.delete()
    
    messages.success(request, "Transferencia cancelada y archivos eliminados.")
    return redirect('lista_transferencias')

# ---------------------------------------------------------
# LISTAR MOVIMIENTOS
# ---------------------------------------------------------
# En views.py

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def lista_movimientos(request):
    # --- 1. PREPARACIÓN DE CUENTAS (Para el filtro) ---
    cuentas = Cuenta.objects.all().order_by('nombre')
    
    # Obtener el ID de la cuenta desde la URL
    cuenta_id = request.GET.get('cuenta')

    # --- LÓGICA DEL DEFAULT (BBVA 15171) ---
    # Si no se seleccionó ninguna cuenta (es la primera carga), buscamos la default.
    if not cuenta_id:
        cuenta_default = Cuenta.objects.filter(nombre__icontains="BBVA 15171").first()
        if cuenta_default:
            cuenta_id = cuenta_default.id
        elif cuentas.exists():
            # Fallback: Si no existe "BBVA 15171", usar la primera que encuentre
            cuenta_id = cuentas.first().id

    # --- CONSULTA BASE ---
    qs = Movimiento.objects.all().select_related('cuenta', 'subcategoria', 'categoria')

    # --- APLICAR FILTROS ---
    
    # 1. Filtro de Cuenta (OBLIGATORIO)
    # Como "no se pueden elegir todas", siempre aplicamos el filtro si tenemos un ID
    if cuenta_id:
        qs = qs.filter(cuenta_id=cuenta_id)

    # Resto de filtros
    q = request.GET.get('q', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    estatus_filtro = request.GET.get('estatus', '')
    auditado_filtro = request.GET.get('auditado', '') 

    if q:
        qs = qs.filter(Q(concepto__icontains=q) | Q(tercero__icontains=q)) # Quitamos búsqueda por nombre de cuenta ya que está filtrada
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)
    if estatus_filtro:
        qs = qs.filter(estatus=estatus_filtro)
    if auditado_filtro == '1':
        qs = qs.filter(auditado=True)
    elif auditado_filtro == '0':
        qs = qs.filter(auditado=False)

    # Ordenamiento: Fecha ascendente, ID ascendente (cronológico por día)
    qs = qs.order_by('fecha', 'id')

    # --- PAGINACIÓN ---
    paginator = Paginator(qs, 27) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'movimientos': page_obj, 
        'importar_form': ImportarExcelForm(),
        'estatus_filtro': estatus_filtro,
        # Nuevas variables para el template
        'cuentas': cuentas,
        'cuenta_seleccionada': int(cuenta_id) if cuenta_id else None
    }
    
    return render(request, 'flujo_bancos/lista_movimientos.html', context)
# ---------------------------------------------------------
# CREAR MOVIMIENTO
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Guardar Movimiento
                    movimiento = form.save()

                    # 2. Procesar Archivos
                    archivos = request.FILES.getlist('archivos_comprobantes')
                    if archivos:
                        for archivo in archivos:
                            # Leemos en memoria para no gastar el puntero
                            contenido_bytes = archivo.read()
                            
                            ext = os.path.splitext(archivo.name)[1].lower()
                            fecha_hoy = timezone.now()
                            s3_path = f"{'xmls' if ext == '.xml' else 'pdfs'}/{fecha_hoy.year}/{fecha_hoy.month}/{archivo.name}"
                            
                            nuevo_comp = ComprobanteFiscal(movimiento=movimiento)

                            # A) Extracción de Datos (Solo si es XML)
                            if ext == '.xml':
                                datos_xml = procesar_datos_xml_desde_bytes(contenido_bytes)
                                # GUARDAMOS EN LA TABLA
                                nuevo_comp.uuid = datos_xml['uuid']
                                nuevo_comp.monto_iva = datos_xml['iva']
                                nuevo_comp.monto_ret_iva = datos_xml['ret_iva']
                                nuevo_comp.monto_ret_isr = datos_xml['ret_isr']

                            # B) Subida a S3
                            memoria = BytesIO(contenido_bytes)
                            ruta = _subir_archivo_a_s3(memoria, s3_path)
                            
                            if ruta:
                                if ext == '.xml': nuevo_comp.archivo_xml.name = ruta
                                else: nuevo_comp.archivo_pdf.name = ruta
                                nuevo_comp.save() # Commit a la tabla ComprobanteFiscal

                        # 3. Actualizar Movimiento con la suma de la tabla
                        recalcular_iva_movimiento(movimiento)

                    messages.success(request, "Movimiento registrado.")
                    return redirect('detalle_movimiento', pk=movimiento.pk)

            except Exception as e:
                messages.error(request, f"Error: {e}")
                return render(request, 'flujo_bancos/crear_movimiento.html', {'form': form})
    else:
        form = MovimientoForm(initial={'fecha': timezone.now().date()})

    return render(request, 'flujo_bancos/crear_movimiento.html', {'form': form})
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def actualizar_saldo_cuenta(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk)
    if request.method == 'POST':
        nuevo_saldo = request.POST.get('saldo_inicial')
        try:
            cuenta.saldo_inicial = float(nuevo_saldo)
            cuenta.save()
            messages.success(request, f"Saldo inicial de {cuenta.nombre} actualizado.")
        except ValueError:
            messages.error(request, "Valor inválido.")
    return redirect('dashboard_bancos')

# ---------------------------------------------------------
# EDITAR MOVIMIENTO
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def editar_movimiento(request, pk):
    movimiento_original = get_object_or_404(Movimiento, pk=pk)
    
    if request.method == 'POST':
        form = MovimientoForm(request.POST, request.FILES, instance=movimiento_original)
        if form.is_valid():
            try:
                with transaction.atomic():
                    movimiento = form.save()

                    # Eliminar viejos
                    ids_eliminar = request.POST.get('ids_eliminar', '')
                    if ids_eliminar:
                        for comp_id in ids_eliminar.split(','):
                            if comp_id:
                                c = ComprobanteFiscal.objects.filter(id=comp_id, movimiento=movimiento).first()
                                if c:
                                    if c.archivo_xml: _eliminar_archivo_de_s3(c.archivo_xml.name)
                                    if c.archivo_pdf: _eliminar_archivo_de_s3(c.archivo_pdf.name)
                                    c.delete()

                    # Subir nuevos
                    archivos = request.FILES.getlist('archivos_comprobantes')
                    if archivos:
                        for archivo in archivos:
                            contenido_bytes = archivo.read()
                            ext = os.path.splitext(archivo.name)[1].lower()
                            fecha_hoy = timezone.now()
                            s3_path = f"{'xmls' if ext == '.xml' else 'pdfs'}/{fecha_hoy.year}/{fecha_hoy.month}/{archivo.name}"
                            
                            nuevo_comp = ComprobanteFiscal(movimiento=movimiento)

                            if ext == '.xml':
                                datos_xml = procesar_datos_xml_desde_bytes(contenido_bytes)
                                nuevo_comp.uuid = datos_xml['uuid']
                                nuevo_comp.monto_iva = datos_xml['iva']
                                nuevo_comp.monto_ret_iva = datos_xml['ret_iva']
                                nuevo_comp.monto_ret_isr = datos_xml['ret_isr']

                            memoria = BytesIO(contenido_bytes)
                            ruta = _subir_archivo_a_s3(memoria, s3_path)
                            
                            if ruta:
                                if ext == '.xml': nuevo_comp.archivo_xml.name = ruta
                                else: nuevo_comp.archivo_pdf.name = ruta
                                nuevo_comp.save()

                    recalcular_iva_movimiento(movimiento)

                messages.success(request, 'Movimiento actualizado.')
                return redirect('detalle_movimiento', pk=movimiento.pk)
            except Exception as e:
                messages.error(request, f"Error: {e}")
    else:
        form = MovimientoForm(instance=movimiento_original)

    context = {
        'form': form,
        'movimiento': movimiento_original,
        'lista_conceptos': Movimiento.objects.values_list('concepto', flat=True).distinct()
    }
    return render(request, 'flujo_bancos/crear_movimiento.html', context)
# ---------------------------------------------------------
# CREAR TRANSFERENCIA
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def crear_transferencia(request):
    if request.method == 'POST':
        form = TransferenciaForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            origen = data['cuenta_origen']
            destino = data['cuenta_destino']
            monto_origen = data['monto']
            tc = data['tipo_cambio'] or 1.0
            fecha = data['fecha']
            concepto_base = data['concepto']

            monto_destino = monto_origen * tc

            saldo_previo_origen = origen.saldo_actual
            saldo_previo_destino = destino.saldo_actual
            
            saldo_final_origen = saldo_previo_origen - monto_origen
            saldo_final_destino = saldo_previo_destino + monto_destino

            Movimiento.objects.create(
                cuenta=origen,
                fecha=fecha,
                concepto=f"{concepto_base} (Envío a {destino.nombre})",
                cargo=monto_origen,
                abono=0,
                saldo_banco=saldo_final_origen,
                comentarios=f"Salida de fondos. TC aplicado: {tc}"
            )

            Movimiento.objects.create(
                cuenta=destino,
                fecha=fecha,
                concepto=f"{concepto_base} (Recepción de {origen.nombre})",
                cargo=0,
                abono=monto_destino,
                saldo_banco=saldo_final_destino,
                comentarios=f"Entrada de fondos. Monto original: {monto_origen} {origen.moneda}. TC: {tc}"
            )

            return redirect('dashboard_bancos')
    else:
        form = TransferenciaForm()
    
    return render(request, 'flujo_bancos/form_transferencia.html', {'form': form})

# ---------------------------------------------------------
# AJAX y OTROS
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def cargar_subcategorias(request):
    categoria_id = request.GET.get('categoria_id')
    subcategorias = SubCategoria.objects.filter(categoria_id=categoria_id).order_by('nombre').values('id', 'nombre')
    return JsonResponse(list(subcategorias), safe=False)

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def obtener_saldo_cuenta(request):
    cuenta_id = request.GET.get('cuenta_id')
    if cuenta_id:
        cuenta = get_object_or_404(Cuenta, id=cuenta_id)
        return JsonResponse({'saldo': cuenta.saldo_actual})
    return JsonResponse({'saldo': 0})

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def ajax_obtener_tc(request):
    tc = obtener_tipo_cambio_banxico()
    return JsonResponse({'tc': tc})

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def editar_cuenta(request, cuenta_id):
    cuenta = get_object_or_404(Cuenta, id=cuenta_id)
    if request.method == 'POST':
        form = CuentaForm(request.POST, instance=cuenta)
        if form.is_valid():
            form.save()
            return redirect('dashboard_bancos')
    else:
        form = CuentaForm(instance=cuenta)
    
    return render(request, 'flujo_bancos/editar_cuenta.html', {
        'form': form, 
        'cuenta': cuenta
    })
    
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def crear_tercero(request):
    if request.method == 'POST':
        form = TerceroForm(request.POST)
        if form.is_valid():
            form.save()
            return render(request, 'flujo_bancos/close_popup.html') 
    else:
        form = TerceroForm()
    
    return render(request, 'flujo_bancos/crear_tercero.html', {'form': form})

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def auditar_movimiento(request, pk):
    mov = get_object_or_404(Movimiento, pk=pk)
    mov.auditado = True
    mov.save()
    messages.success(request, 'Movimiento auditado y bloqueado correctamente.')
    return redirect('lista_movimientos')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def detalle_movimiento(request, pk):
    mov = get_object_or_404(Movimiento, pk=pk)
    # Leemos directo de la BD, NO de S3 (Rápido y Seguro)
    comprobantes = mov.comprobantes.all()
    
    lista_xmls = []
    for comp in comprobantes:
        datos = {
            'obj': comp,
            'uuid': comp.uuid or '---',
            'iva': comp.monto_iva,
            'ret_iva': comp.monto_ret_iva, 
            'ret_isr': comp.monto_ret_isr,
            'es_xml': bool(comp.archivo_xml)
        }
        lista_xmls.append(datos)
        
    totales = {
        'iva': mov.iva,
        'ret_iva': mov.ret_iva,
        'ret_isr': mov.ret_isr
    }

    context = {'mov': mov, 'lista_xmls': lista_xmls, 'totales_xml': totales}
    return render(request, 'flujo_bancos/detalle_movimiento.html', context)

# ---------------------------------------------------------
# REPORTES EXCEL
# ---------------------------------------------------------
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def exportar_movimientos_excel(request):
    # 1. OBTENER TODOS LOS DATOS EN UNA SOLA CONSULTA (OPTIMIZADO)
    # Usamos select_related para traer los nombres de las cuentas y categorías de un jalón
    movimientos = Movimiento.objects.select_related(
        'cuenta', 
        'unidad_negocio', 
        'operacion', 
        'categoria', 
        'subcategoria'
    ).order_by('cuenta__nombre', 'fecha', 'id') # Ordenamos por cuenta primero para agrupar fácil
    
    # --- APLICAR FILTROS (Igual que antes) ---
    q = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo = request.GET.get('tipo')
    estatus = request.GET.get('estatus')
    modo_exportacion = request.GET.get('modo', 'consolidado')

    if q:
        movimientos = movimientos.filter(Q(concepto__icontains=q) | Q(tercero__icontains=q))
    if fecha_inicio:
        movimientos = movimientos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha__lte=fecha_fin)
    if tipo == 'ingreso':
        movimientos = movimientos.filter(abono__gt=0)
    elif tipo == 'egreso':
        movimientos = movimientos.filter(cargo__gt=0)
    if estatus:
        movimientos = movimientos.filter(estatus=estatus)
    
    # Excluir movimientos huérfanos sin cuenta (para que no rompa el Excel)
    movimientos = movimientos.exclude(cuenta__isnull=True)

    # 2. CONFIGURACIÓN DEL ARCHIVO EXCEL
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "Estado_Cuenta_Segregado.xlsx" if modo_exportacion == 'segregado' else "Estado_Cuenta_Consolidado.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = openpyxl.Workbook()
    
    # --- ESTILOS (Definidos una sola vez para rendimiento) ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")
    currency_format = '"$"#,##0.00'
    date_format = 'DD/MM/YYYY'

    headers = [
        "FECHA", "CONCEPTO / REFERENCIA", "CARGO", "ABONO", "SALDO",
        "UNIDAD DE NEGOCIO", "OPERACIÓN", "CATEGORIA", "SUBCATEGORIA", 
        "TERCERO", "IVA", "RET IVA", "RET ISR", "COMENTARIO", "ESTATUS", "AUDITADO"
    ]

    # --- FUNCIÓN AUXILIAR PARA LIMPIAR NOMBRE DE PESTAÑA ---
    def limpiar_nombre_hoja(nombre):
        invalido = ['\\', '/', '*', '[', ']', ':', '?']
        nuevo_nombre = str(nombre)
        for char in invalido:
            nuevo_nombre = nuevo_nombre.replace(char, '')
        return nuevo_nombre[:30] # Excel solo permite 31 caracteres

    # --- FUNCIÓN PARA LLENAR DATOS EN UNA HOJA ---
    def llenar_hoja(ws, lista_movs, incluir_columna_cuenta=False):
        # Encabezados dinámicos
        encabezados_actuales = list(headers)
        if incluir_columna_cuenta:
            encabezados_actuales.insert(1, "CUENTA") # Si es consolidado, agregamos la columna Cuenta

        ws.append(encabezados_actuales)
        
        # Estilo Encabezados
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
            cell.border = thin_border
            # Ancho aproximado
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Llenado de Filas
        row_num = 2
        for mov in lista_movs:
            saldo_banco = mov.saldo_banco if mov.saldo_banco is not None else 0
            unidad = mov.unidad_negocio.nombre if mov.unidad_negocio else ""
            operacion = mov.operacion.nombre if mov.operacion else ""
            categoria = mov.categoria.nombre if mov.categoria else ""
            subcategoria = mov.subcategoria.nombre if mov.subcategoria else ""
            tercero = mov.tercero if mov.tercero else ""
            comentario = mov.comentarios if mov.comentarios else ""

            datos_fila = [
                mov.fecha, mov.concepto, mov.cargo, mov.abono, saldo_banco,
                unidad, operacion, categoria, subcategoria, tercero,
                mov.iva, mov.ret_iva, mov.ret_isr, comentario,
                mov.get_estatus_display(), 'SI' if mov.auditado else 'NO'
            ]
            
            if incluir_columna_cuenta:
                datos_fila.insert(1, mov.cuenta.nombre)

            ws.append(datos_fila)

            # Aplicar Estilos a la fila (optimizado)
            # Detectamos índices de columnas numéricas basándonos en si insertamos cuenta o no
            offset = 1 if incluir_columna_cuenta else 0
            
            # Índices relativos de columnas con dinero (Cargo, Abono, Saldo, Impuestos)
            money_cols = [3, 4, 5, 11, 12, 13] 
            money_cols = [x + offset for x in money_cols]

            for col_idx, cell in enumerate(ws[row_num], start=1):
                cell.border = thin_border
                
                if col_idx == 1: # Fecha
                    cell.number_format = date_format
                    cell.alignment = center_aligned
                elif col_idx in money_cols: # Dinero
                    cell.number_format = currency_format
                    cell.alignment = center_aligned
                elif (incluir_columna_cuenta and col_idx == 2) or (not incluir_columna_cuenta and col_idx == 2):
                    # Concepto (o Cuenta si está insertada)
                    cell.alignment = left_aligned
                else:
                    cell.alignment = center_aligned
            
            row_num += 1

        # Ajuste automático de ancho
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width

    # 3. LÓGICA DE AGRUPACIÓN (AQUÍ ESTÁ LA MAGIA DE LA VELOCIDAD)
    if modo_exportacion == 'segregado':
        # Eliminamos la hoja por defecto
        wb.remove(wb.active)
        
        # Agrupamos en memoria (Python Dictionary)
        # Esto evita ir a la base de datos mil veces
        grupos = defaultdict(list)
        for mov in movimientos:
            grupos[mov.cuenta].append(mov)
        
        if not grupos:
            wb.create_sheet("Sin Datos")
        
        # Iteramos sobre el diccionario en memoria
        for cuenta_obj, lista_movs in grupos.items():
            nombre_limpio = limpiar_nombre_hoja(cuenta_obj.nombre)
            # Si el nombre ya existe (por el recorte de 30 chars), añadimos un sufijo
            if nombre_limpio in wb.sheetnames:
                nombre_limpio = f"{nombre_limpio[:25]}_{cuenta_obj.id}"
                
            ws = wb.create_sheet(title=nombre_limpio)
            llenar_hoja(ws, lista_movs, incluir_columna_cuenta=False)
            
    else:
        # Modo Consolidado (Todo en una hoja)
        ws = wb.active
        ws.title = "Consolidado"
        llenar_hoja(ws, movimientos, incluir_columna_cuenta=True)

    wb.save(response)
    return response

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def exportar_transferencias_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Transferencias"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    currency_format = '"$"#,##0.00_-'

    headers = [
        "Fecha", "Concepto Base", 
        "Cuenta Origen", "Saldo Inicial (Origen)", "Salida ($)", "Saldo Final (Origen)",
        "Cuenta Destino", "Saldo Inicial (Destino)", "Entrada ($)", "Saldo Final (Destino)",
        "Comentarios"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border

    salidas = Movimiento.objects.filter(cargo__gt=0).filter(
        Q(concepto__icontains='(Envío a') | Q(concepto__icontains='Transferencia')
    ).select_related('cuenta').order_by('fecha', 'id')

    row_num = 2

    for salida in salidas:
        saldo_final_org = salida.saldo_banco if salida.saldo_banco is not None else 0
        saldo_inicial_org = saldo_final_org + salida.cargo

        concepto_full = salida.concepto
        concepto_base = concepto_full
        nombre_destino = "---"
        monto_entrada = 0
        saldo_final_dest = 0
        saldo_inicial_dest = 0

        match = re.search(r'\(Envío a (.*?)\)', concepto_full)
        if match:
            base_search = concepto_full.split(' (Envío a')[0]
            concepto_base = base_search
            posible_entrada = Movimiento.objects.filter(
                fecha=salida.fecha,
                abono__gt=0,
                concepto__icontains=base_search
            ).exclude(id=salida.id).first()

            if posible_entrada:
                nombre_destino = posible_entrada.cuenta.nombre
                monto_entrada = posible_entrada.abono
                saldo_final_dest = posible_entrada.saldo_banco if posible_entrada.saldo_banco is not None else 0
                saldo_inicial_dest = saldo_final_dest - monto_entrada

        row = [
            salida.fecha,
            concepto_base,
            salida.cuenta.nombre,
            saldo_inicial_org,
            salida.cargo,
            saldo_final_org,
            nombre_destino,
            saldo_inicial_dest,
            monto_entrada,
            saldo_final_dest,
            salida.comentarios
        ]
        ws.append(row)

        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.alignment = center_aligned
            cell.border = thin_border
            if col_idx in [4, 5, 6, 8, 9, 10]:
                cell.number_format = currency_format
            if col_idx == 1:
                cell.number_format = 'DD/MM/YYYY'

        row_num += 1

    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Transferencias_Pro.xlsx"'
    wb.save(response)
    return response

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def gestion_categorias_view(request):
    # 1. Obtener las categorías de la base de datos
    categorias = Categoria.objects.prefetch_related('subcategorias').all().order_by('nombre')
    
    # 2. Crear los formularios vacíos
    cat_form = CategoriaForm()
    sub_form = SubCategoriaForm()
    
    # 3. DEFINIR LA VARIABLE CONTEXT
    context = {
        'categorias': categorias,
        'cat_form': cat_form,
        'sub_form': sub_form
    }
    
    # 4. Renderizar usando el contexto
    return render(request, 'flujo_bancos/gestion_categorias.html', context)

# 2. FUNCIONES PARA CATEGORÍAS
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría creada.")
    return redirect('bancos_categorias_lista')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría actualizada.")
    return redirect('bancos_categorias_lista')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def eliminar_categoria(request, pk):
    try:
        categoria = get_object_or_404(Categoria, pk=pk)
        categoria.delete()
        messages.success(request, "Categoría eliminada.")
    except Exception as e:
        messages.error(request, "No se puede eliminar porque tiene movimientos asociados.")
    return redirect('bancos_categorias_lista')

# 3. FUNCIONES PARA SUBCATEGORÍAS
@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def crear_subcategoria(request):
    if request.method == 'POST':
        form = SubCategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subcategoría creada correctamente.")
    return redirect('bancos_categorias_lista')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def editar_subcategoria(request, pk):
    sub = get_object_or_404(SubCategoria, pk=pk)
    if request.method == 'POST':
        # Permitimos cambiar nombre pero preservamos la categoría si no se envía
        form = SubCategoriaForm(request.POST, instance=sub)
        if form.is_valid():
            form.save()
            messages.success(request, "Subcategoría actualizada.")
    return redirect('bancos_categorias_lista')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def eliminar_subcategoria(request, pk):
    sub = get_object_or_404(SubCategoria, pk=pk)
    sub.delete()
    messages.success(request, "Subcategoría eliminada.")
    return redirect('bancos_categorias_lista')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def eliminar_movimiento(request, pk):
    mov = get_object_or_404(Movimiento, pk=pk)

    # 1. Seguridad: No borrar si ya está auditado
    if mov.auditado:
        messages.error(request, "No es posible eliminar un movimiento que ya ha sido auditado.")
        return redirect('lista_movimientos')

    # 2. Limpieza S3: Borrar archivo adjunto si existe
    if mov.comprobante:
        try:
            _eliminar_archivo_de_s3(str(mov.comprobante))
        except Exception as e:
            print(f"Advertencia: No se pudo borrar archivo S3: {e}")

    # 3. Borrar registro
    mov.delete()
    messages.success(request, "El movimiento ha sido eliminado correctamente.")
    return redirect('lista_movimientos')

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def lista_transferencias(request):
    transferencias = Movimiento.objects.filter(
        cargo__gt=0
    ).filter(
        Q(concepto__icontains='(Envío a') | Q(concepto__icontains='Transferencia')
    ).select_related('cuenta').order_by('-fecha', '-id') # <--- CAMBIO AQUÍ: '-fecha'

    paginator = Paginator(transferencias, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'transferencias': page_obj,
    }
    return render(request, 'flujo_bancos/lista_transferencias.html', context)

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def subir_comprobante(request, movimiento_id):
    movimiento = get_object_or_404(Movimiento, pk=movimiento_id)
    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES)
        if form.is_valid():
            comp = form.save(commit=False)
            comp.movimiento = movimiento
            
            # XML
            if 'archivo_xml' in request.FILES:
                f = request.FILES['archivo_xml']
                contenido = f.read()
                
                path = f"xmls/{timezone.now().year}/{timezone.now().month}/{f.name}"
                
                datos = procesar_datos_xml_desde_bytes(contenido)
                comp.uuid = datos['uuid']
                comp.monto_iva = datos['iva']
                comp.monto_ret_iva = datos['ret_iva']
                comp.monto_ret_isr = datos['ret_isr']
                
                memoria = BytesIO(contenido)
                ruta = _subir_archivo_a_s3(memoria, path)
                if ruta: comp.archivo_xml.name = ruta

            # PDF
            if 'archivo_pdf' in request.FILES:
                f = request.FILES['archivo_pdf']
                contenido = f.read()
                path = f"pdfs/{timezone.now().year}/{timezone.now().month}/{f.name}"
                memoria = BytesIO(contenido)
                ruta = _subir_archivo_a_s3(memoria, path)
                if ruta: comp.archivo_pdf.name = ruta

            comp.save()
            recalcular_iva_movimiento(movimiento)
            
            messages.success(request, "Comprobante subido.")
            return redirect('detalle_movimiento', pk=movimiento.pk)
    return redirect('detalle_movimiento', pk=movimiento.pk)
    
def recalcular_iva_movimiento(movimiento):
    """
    Suma los datos guardados en la BD (Tabla ComprobanteFiscal)
    y actualiza el Movimiento padre. NO lee archivos.
    """
    comps = movimiento.comprobantes.all()
    
    total_iva = sum(c.monto_iva for c in comps)
    total_ret_iva = sum(c.monto_ret_iva for c in comps)
    total_ret_isr = sum(c.monto_ret_isr for c in comps)
    
    movimiento.iva = total_iva
    movimiento.ret_iva = total_ret_iva
    movimiento.ret_isr = total_ret_isr
    movimiento.iva_total_xml = total_iva
    
    movimiento.save()

@permission_required('flujo_bancos.acceso_flujo_bancos', raise_exception=True)
def eliminar_comprobante(request, comprobante_id):
    comp = get_object_or_404(ComprobanteFiscal, pk=comprobante_id)
    mov_id = comp.movimiento.pk
    
    # Eliminar de S3 usando la función segura
    if comp.archivo_xml: _eliminar_archivo_de_s3(comp.archivo_xml.name)
    if comp.archivo_pdf: _eliminar_archivo_de_s3(comp.archivo_pdf.name)
    
    comp.delete()
    
    # Recalcular movimiento padre
    recalcular_iva_movimiento(Movimiento.objects.get(pk=mov_id))
    
    messages.success(request, "Comprobante eliminado.")
    return redirect('detalle_movimiento', pk=mov_id)

